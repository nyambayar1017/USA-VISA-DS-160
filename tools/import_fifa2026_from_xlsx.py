from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


def normalize_stage(value: str) -> str:
    text = str(value or "").strip()
    mapped = {
        "OPENING CEREMONY MEX": "Opening Ceremony",
        "OPENING CEREMONY USA": "Opening Ceremony",
        "Group stage": "Group Stage",
        "ROUND 32": "Round of 32",
        "ROUND 16": "Round of 16",
        "Quarter/Final": "Quarter Final",
        "Semi final": "Semi Final",
        "FINAL": "Final",
    }
    return mapped.get(text, text.title() if text else "")


def normalize_city(value: str) -> str:
    text = str(value or "").strip()
    mapped = {
        "NY": "New York",
        "NEWYORK": "New York",
        "LOS ANGELES": "Los Angeles",
        "MEXICO": "Mexico City",
        "GUADALAJARA": "Guadalajara",
        "DALLAS": "Dallas",
        "HOUSTON": "Houston",
        "SEATTLE": "Seattle",
        "KANSAS": "Kansas",
        "SAN FRANCISCO": "San Francisco",
        "BOSTON": "Boston",
        "PHILADELPHIA": "Philadelphia",
        "ATLANTA": "Atlanta",
        "VANCOUVER": "Vancouver",
        "MIAMI": "Miami",
        "LA stadium": "Los Angeles",
        "Los angeles": "Los Angeles",
        "NEW YORK": "New York",
    }
    return mapped.get(text, text.title() if text else "")


def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def category_name_from_seat(seat_details: str) -> str:
    seat = str(seat_details or "").strip().lstrip("-").strip()
    if not seat:
      return "Seat details from source sheet"
    if "Seat will be assigned later" in seat:
      return "Seat will be assigned later"
    parts = [part.strip() for part in seat.split(" - ") if part.strip()]
    return " - ".join(parts[:2]) if len(parts) >= 2 else seat


def seat_section_from_seat(seat_details: str) -> str:
    seat = str(seat_details or "").strip().lstrip("-").strip()
    if not seat:
        return "Seat details from source sheet"
    parts = [part.strip() for part in seat.split(" - ") if part.strip()]
    if len(parts) >= 4:
        return " - ".join(parts[2:4])
    if len(parts) >= 2:
        return " - ".join(parts[-2:])
    return seat


def build_records(xlsx_path: Path) -> dict:
    workbook = load_workbook(xlsx_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    current = {}
    previous_price = {"1": None, "2": None, "3": None}
    lots = []
    matches_seen = {}
    lot_counters = defaultdict(int)

    for row_index in range(3, 55):
        row = [sheet.cell(row_index, column).value for column in range(1, 17)]
        if row[2]:
            current = {
                "stage": normalize_stage(row[1] or current.get("stage", "")),
                "matchNumber": str(row[2]).strip(),
                "matchDate": row[3].strftime("%Y-%m-%d") if hasattr(row[3], "strftime") else "",
                "teamA": str(row[4] or "").strip(),
                "teamB": str(row[5] or "").strip(),
                "city": normalize_city(row[6]),
            }
            matches_seen[current["matchNumber"]] = dict(current)
            previous_price = {"1": None, "2": None, "3": None}

        for category_code, qty_col, seat_col, price_col in (("1", 7, 8, 9), ("2", 10, 11, 12), ("3", 13, 14, 15)):
            quantity = row[qty_col]
            seat_details = row[seat_col]
            price = row[price_col]
            if quantity is None and seat_details is None and price is None:
                continue
            if price is not None:
                previous_price[category_code] = int(price)
            if quantity is None:
                continue

            quantity = int(quantity or 0)
            if quantity <= 0:
                continue

            lot_counters[current["matchNumber"]] += 1
            seat_text = str(seat_details or "").strip()
            match_key = f"{current['matchNumber']}-{category_code}-{lot_counters[current['matchNumber']]}"
            lots.append(
                {
                    "id": f"fifa-{slugify(match_key)}",
                    "stage": current["stage"],
                    "matchNumber": current["matchNumber"],
                    "matchLabel": f"{current['teamA']} vs {current['teamB']}",
                    "matchDate": current["matchDate"],
                    "teamA": current["teamA"],
                    "teamB": current["teamB"],
                    "city": current["city"],
                    "venue": f"{current['city']} Stadium" if current["city"] else "",
                    "categoryCode": category_code,
                    "categoryName": category_name_from_seat(seat_text),
                    "seatSection": seat_section_from_seat(seat_text),
                    "seatDetails": seat_text or "Seat details from source sheet",
                    "seatAssignedLater": "assigned later" in seat_text.lower(),
                    "price": int(previous_price[category_code] or 0),
                    "currency": "USD",
                    "totalQuantity": quantity,
                    "visibility": "public",
                    "status": "active",
                    "notes": "Imported from DTX 2026 WC - Sales Registration.xlsx",
                    "createdAt": "2026-04-13T00:00:00+08:00",
                    "updatedAt": "2026-04-13T00:00:00+08:00",
                }
            )

    for match_number, match in matches_seen.items():
        if any(item["matchNumber"] == match_number for item in lots):
            continue
        lots.append(
            {
                "id": f"fifa-{slugify(match_number)}-placeholder",
                "stage": match["stage"],
                "matchNumber": match_number,
                "matchLabel": f"{match['teamA']} vs {match['teamB']}",
                "matchDate": match["matchDate"],
                "teamA": match["teamA"],
                "teamB": match["teamB"],
                "city": match["city"],
                "venue": f"{match['city']} Stadium" if match["city"] else "",
                "categoryCode": "1",
                "categoryName": "No current inventory",
                "seatSection": "No seat inventory",
                "seatDetails": "No ticket lots listed in the source sheet for this match.",
                "seatAssignedLater": False,
                "price": 0,
                "currency": "USD",
                "totalQuantity": 0,
                "visibility": "private",
                "status": "hidden",
                "notes": "Placeholder imported from workbook because the match exists without sellable rows.",
                "createdAt": "2026-04-13T00:00:00+08:00",
                "updatedAt": "2026-04-13T00:00:00+08:00",
            }
        )

    lots.sort(key=lambda item: (item["matchDate"], item["matchNumber"], item["categoryCode"], item["seatDetails"]))
    return {"tickets": lots, "sales": []}


def main() -> int:
    if len(sys.argv) != 3:
        print("Usage: python3 tools/import_fifa2026_from_xlsx.py <input.xlsx> <output.json>")
        return 1
    input_path = Path(sys.argv[1]).expanduser()
    output_path = Path(sys.argv[2]).expanduser()
    payload = build_records(input_path)
    output_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    total_quantity = sum(item.get("totalQuantity", 0) for item in payload["tickets"])
    matches = len({item.get("matchNumber") for item in payload["tickets"]})
    print(f"Wrote {len(payload['tickets'])} ticket lots for {matches} matches and {total_quantity} total tickets.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
