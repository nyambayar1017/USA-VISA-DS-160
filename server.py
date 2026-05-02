import copy
import html
import hashlib
import hmac
import json
import base64
import mimetypes
import os
import re
import secrets
import sys
import time
import urllib.error
import urllib.request
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path
from urllib.parse import parse_qs, unquote, quote
from urllib import request as urlrequest
from uuid import uuid4
import xml.etree.ElementTree as ET
from wsgiref.simple_server import make_server


BASE_DIR = Path(__file__).resolve().parent
PUBLIC_DIR = BASE_DIR / "public"
DATA_DIR = Path(os.environ.get("DATA_DIR", str(BASE_DIR / "data"))).expanduser()
GENERATED_DIR = DATA_DIR / "generated"
BACKUP_DIR = DATA_DIR / "backups"
PORT = int(os.environ.get("PORT", "3000"))
HOST = os.environ.get("HOST", "0.0.0.0")
ADMIN_TOKEN = os.environ.get("ADMIN_TOKEN", "change-me")
TEMPLATE_FILE = Path(os.environ.get("CONTRACT_TEMPLATE", DATA_DIR / "GEREE-template.docx"))
TEMPLATE_FALLBACK = Path(__file__).resolve().parent / "data" / "GEREE-template.docx"
ALT_TEMPLATE = Path(__file__).resolve().parent / "data" / "Гэрээ.docx"
CONTRACTS_FILE = DATA_DIR / "contracts.json"
DS160_FILE = DATA_DIR / "ds160_applications.json"
FINANCE_FILE = DATA_DIR / "finance_entries.json"
BOOKINGS_FILE = DATA_DIR / "hotel_bookings.json"
RESERVATIONS_FILE = DATA_DIR / "reservations.json"
CAMP_RESERVATIONS_FILE = DATA_DIR / "camp_reservations.json"
FLIGHT_RESERVATIONS_FILE = DATA_DIR / "flight_reservations.json"
TRANSFER_RESERVATIONS_FILE = DATA_DIR / "transfer_reservations.json"
CAMP_TRIPS_FILE = DATA_DIR / "camp_trips.json"
TRIP_UPLOADS_DIR = DATA_DIR / "trip-uploads"
ANNOUNCEMENT_UPLOADS_DIR = DATA_DIR / "announcement-uploads"
MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10 MB
ALLOWED_UPLOAD_EXTENSIONS = {".pdf", ".doc", ".docx", ".xls", ".xlsx", ".jpg", ".jpeg", ".png", ".gif", ".txt"}
CAMP_SETTINGS_FILE = DATA_DIR / "camp_settings.json"
SETTINGS_FILE = DATA_DIR / "settings.json"
ANNOUNCEMENTS_FILE = DATA_DIR / "announcements.json"
CAMP_CONTRACTS_DIR = DATA_DIR / "camp-contracts"
MAIL_FOLLOWUPS_FILE = DATA_DIR / "mail_followups.json"
MAIL_DOSSIERS_FILE = DATA_DIR / "mail_dossiers.json"
NOTES_FILE = DATA_DIR / "notes.json"
FIFA2026_FILE = DATA_DIR / "fifa2026.json"
FIFA2026_RESET_MARKER_FILE = DATA_DIR / "fifa2026_manual_reset_v3.txt"
MANAGER_DASHBOARD_FILE = DATA_DIR / "manager_dashboard.json"
TASK_ATTACHMENTS_DIR = DATA_DIR / "task_attachments"
TASK_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".heic", ".heif"}
USERS_FILE = DATA_DIR / "users.json"
SESSIONS_FILE = DATA_DIR / "sessions.json"
NOTIFICATIONS_FILE = DATA_DIR / "notifications.json"
NOTIFICATIONS_MAX = 200
GROUPS_FILE = DATA_DIR / "tourist_groups.json"
TOURISTS_FILE = DATA_DIR / "tourists.json"
INVOICES_FILE = DATA_DIR / "invoices.json"
PAYMENT_REQUESTS_FILE = DATA_DIR / "payment_requests.json"
TRIP_TEMPLATES_FILE = DATA_DIR / "trip_templates.json"
SERVICE_TEMPLATES_FILE = DATA_DIR / "service_templates.json"
TRIP_CREATORS_FILE = DATA_DIR / "trip_creators.json"
GALLERY_FILE = DATA_DIR / "gallery.json"
GALLERY_FOLDERS_FILE = DATA_DIR / "gallery_folders.json"
GALLERY_UPLOADS_DIR = DATA_DIR / "gallery-uploads"
CONTENT_FILE = DATA_DIR / "content.json"
MEAL_TEMPLATES_FILE = DATA_DIR / "meal_templates.json"
LOCATIONS_FILE = DATA_DIR / "locations.json"
WORD_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
XML_NS = "http://www.w3.org/XML/1998/namespace"
SESSION_COOKIE = "travelx_session"
SESSION_SECRET = os.environ.get("SESSION_SECRET", ADMIN_TOKEN)
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "").strip().lower()
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "").strip()
BACKUP_WEBHOOK_URL = os.environ.get("BACKUP_WEBHOOK_URL", "").strip()
BACKUP_TOKEN = os.environ.get("BACKUP_TOKEN", "").strip()
OCRSPACE_API_KEY = os.environ.get("OCRSPACE_API_KEY", "").strip()
TOURIST_PASSPORT_TEMP_DIR = DATA_DIR / "tourist-passport-temp"
MAIL_ACCOUNTS_PATH = DATA_DIR / "mail-accounts.json"
MAIL_MESSAGES_DIR = DATA_DIR / "mail-messages"
MAIL_UPLOADS_DIR = DATA_DIR / "mail-uploads"
MAIL_TEMPLATES_PATH = DATA_DIR / "mail-templates.json"
MAIL_SIGNATURES_PATH = DATA_DIR / "mail-signatures.json"
STEPPE_COMPANY_NAME = "“АНЛОК СТЕП МОНГОЛИА” ХХК"
STEPPE_CITY = "Улаанбаатар хот"
STEPPE_ADDRESS_LINES = [
    "Хан-Уул дүүрэг, 17-р хороо,",
    "Их Монгол Улсын гудамж, Кинг Товер, 121-102 тоот",
    "Улаанбаатар хот, Монгол улс.",
]
STEPPE_MANAGER = "Г. Бужинлхам"
STEPPE_PHONES = "72007722, 85178822"
STEPPE_CONTACT_PHONES = "72007722, 85178822"
STEPPE_EMAIL = "booking@steppe-mongolia.com"
DEFAULT_ROOM_CHOICES = [
    "Double Standard Ger",
    "Twin Standard Ger",
    "Standard Ger 3/4 pax",
    "Luxury Ger with bathroom",
    "Luxury Ger without bathroom",
]
DEFAULT_LANGUAGES = [
    "English",
    "French",
    "Mongolian",
    "Korean",
    "Spanish",
    "Italian",
    "Other",
]
RESERVATION_TYPE_LABELS = {
    "camp": "Баазын захиалга",
    "tent": "Майхны захиалга",
    "hotel": "Буудлын захиалга",
    "herder": "Малчин айлын захиалга",
}
MONGOLIA_TZ = timezone(timedelta(hours=8))

ET.register_namespace("w", WORD_NS)


def ensure_data_store():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    GENERATED_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    TRIP_UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    ANNOUNCEMENT_UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    GALLERY_UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    TOURIST_PASSPORT_TEMP_DIR.mkdir(parents=True, exist_ok=True)
    MAIL_MESSAGES_DIR.mkdir(parents=True, exist_ok=True)
    MAIL_UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    TASK_ATTACHMENTS_DIR.mkdir(parents=True, exist_ok=True)
    for file_path in [
        CONTRACTS_FILE,
        DS160_FILE,
        FINANCE_FILE,
        BOOKINGS_FILE,
        RESERVATIONS_FILE,
        CAMP_RESERVATIONS_FILE,
        FLIGHT_RESERVATIONS_FILE,
        TRANSFER_RESERVATIONS_FILE,
        CAMP_TRIPS_FILE,
        USERS_FILE,
        SESSIONS_FILE,
        NOTIFICATIONS_FILE,
        GROUPS_FILE,
        TOURISTS_FILE,
        INVOICES_FILE,
    ]:
        if not file_path.exists():
            file_path.write_text("[]", encoding="utf-8")
    if not MANAGER_DASHBOARD_FILE.exists():
        MANAGER_DASHBOARD_FILE.write_text(
            json.dumps(default_manager_dashboard(), indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
    if not CAMP_SETTINGS_FILE.exists():
        CAMP_SETTINGS_FILE.write_text(
            json.dumps(
                {
                    "campNames": ["Khustai camp"],
                    "staffAssignments": [STEPPE_MANAGER],
                    "roomChoices": DEFAULT_ROOM_CHOICES,
                },
                indent=2,
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
    if not FIFA2026_FILE.exists():
        FIFA2026_FILE.write_text(
            json.dumps(default_fifa2026_data(), indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
    bootstrap_admin_user()


def read_json_list(file_path):
    """Read a JSON-list-on-disk store. Tolerates the file not existing yet,
    being empty, or being briefly truncated mid-write — those return [] rather
    than 500 every endpoint that touches it. Surface the underlying problem
    in the worker log so we still notice corruption."""
    ensure_data_store()
    try:
        text = file_path.read_text(encoding="utf-8")
    except FileNotFoundError:
        return []
    except OSError as exc:
        print(f"[read_json_list] OSError on {file_path}: {exc}", flush=True)
        return []
    text = text.strip()
    if not text:
        return []
    try:
        data = json.loads(text)
    except json.JSONDecodeError as exc:
        # "Extra data" specifically means the parser saw a complete, valid JSON
        # value followed by trailing garbage — i.e. a concurrent-write
        # collision where two workers raced on the same file. The valid prefix
        # IS the recoverable data; tear off the trailing junk and persist the
        # clean version (atomically) instead of returning an empty list.
        if "Extra data" in str(exc):
            try:
                decoder = json.JSONDecoder()
                data, idx = decoder.raw_decode(text)
                print(
                    f"[read_json_list] recovered {file_path} prefix: {idx} of {len(text)} bytes",
                    flush=True,
                )
                # Snapshot the corrupted copy first in case manual review is wanted,
                # then atomically rewrite the clean version.
                try:
                    backup_path = file_path.with_suffix(file_path.suffix + ".corrupt")
                    backup_path.write_text(text, encoding="utf-8")
                except Exception:
                    pass
                try:
                    tmp_path = file_path.with_suffix(file_path.suffix + ".tmp")
                    tmp_path.write_text(
                        json.dumps(data, indent=2, ensure_ascii=False),
                        encoding="utf-8",
                    )
                    os.replace(tmp_path, file_path)
                except Exception as werr:
                    print(
                        f"[read_json_list] could not persist recovery for {file_path}: {werr}",
                        flush=True,
                    )
                if isinstance(data, list):
                    return data
                return []
            except Exception:
                pass
        print(f"[read_json_list] JSON decode failed on {file_path}: {exc}", flush=True)
        # Stash the bad copy so we can inspect it later, then start clean.
        try:
            backup_path = file_path.with_suffix(file_path.suffix + ".corrupt")
            backup_path.write_text(text, encoding="utf-8")
        except Exception:
            pass
        return []
    if not isinstance(data, list):
        print(f"[read_json_list] Non-list payload in {file_path}; got {type(data).__name__}", flush=True)
        return []
    return data


def write_json_list(file_path, records):
    """Atomic write: render to a sibling temp file then rename, so a worker
    crash mid-write can't leave the canonical file half-written (which would
    500 every read endpoint downstream).

    The temp filename is per-write unique (pid + random suffix) so two
    concurrent saves of the same file don't share a .tmp and clobber each
    other's rename — that was producing FileNotFoundError on tourists.json
    when the user clicked move-down quickly."""
    ensure_data_store()
    tmp_path = file_path.with_suffix(
        f"{file_path.suffix}.{os.getpid()}.{uuid4().hex[:8]}.tmp"
    )
    try:
        tmp_path.write_text(
            json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8"
        )
        os.replace(tmp_path, file_path)
    except Exception:
        # If the rename fails for any reason, don't leave the unique .tmp
        # behind to accumulate.
        try:
            if tmp_path.exists():
                tmp_path.unlink()
        except Exception:
            pass
        raise


def list_backup_sources():
    return [
        CONTRACTS_FILE,
        DS160_FILE,
        FINANCE_FILE,
        BOOKINGS_FILE,
        RESERVATIONS_FILE,
        CAMP_RESERVATIONS_FILE,
        FLIGHT_RESERVATIONS_FILE,
        TRANSFER_RESERVATIONS_FILE,
        CAMP_TRIPS_FILE,
        CAMP_SETTINGS_FILE,
        FIFA2026_FILE,
        MANAGER_DASHBOARD_FILE,
        USERS_FILE,
        SESSIONS_FILE,
    ]


def notify_backup_webhook(payload):
    if not BACKUP_WEBHOOK_URL:
        return
    try:
        data = json.dumps(payload).encode("utf-8")
        req = urlrequest.Request(
            BACKUP_WEBHOOK_URL,
            data=data,
            headers={"Content-Type": "application/json"},
        )
        urlrequest.urlopen(req, timeout=8)
    except Exception:
        return


def create_backup_archive():
    ensure_data_store()
    timestamp = now_mongolia().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"travelx-backup-{timestamp}.zip"
    archive_path = BACKUP_DIR / filename
    manifest = {
        "createdAt": now_mongolia().isoformat(),
        "files": [],
    }
    with zipfile.ZipFile(archive_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for file_path in list_backup_sources():
            if not file_path.exists():
                continue
            arcname = file_path.name
            archive.write(file_path, arcname)
            manifest["files"].append(arcname)
        archive.writestr("manifest.json", json.dumps(manifest, indent=2, ensure_ascii=False))
    notify_backup_webhook(
        {
            "filename": filename,
            "createdAt": manifest["createdAt"],
            "fileCount": len(manifest["files"]),
        }
    )
    return archive_path


def list_backup_archives():
    ensure_data_store()
    backups = []
    for path in BACKUP_DIR.glob("travelx-backup-*.zip"):
        stat = path.stat()
        backups.append(
            {
                "filename": path.name,
                "size": stat.st_size,
                "createdAt": datetime.fromtimestamp(stat.st_mtime, MONGOLIA_TZ).isoformat(),
            }
        )
    backups.sort(key=lambda item: item["createdAt"], reverse=True)
    return backups


def restore_json_from_latest_backup(filename, normalizer):
    ensure_data_store()
    for item in list_backup_archives():
        archive_path = BACKUP_DIR / item["filename"]
        if not archive_path.exists():
            continue
        try:
            with zipfile.ZipFile(archive_path, "r") as archive:
                if filename not in archive.namelist():
                    continue
                payload = json.loads(archive.read(filename).decode("utf-8"))
        except Exception:
            continue
        normalized = normalizer(payload)
        if normalized:
            return normalized
    return None


def read_json_object(file_path, default):
    ensure_data_store()
    if not file_path.exists():
        file_path.write_text(json.dumps(default, indent=2, ensure_ascii=False), encoding="utf-8")
        return copy.deepcopy(default)
    try:
        data = json.loads(file_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        data = copy.deepcopy(default)
    if not isinstance(data, dict):
        data = copy.deepcopy(default)
    return data


def write_json_object(file_path, payload):
    ensure_data_store()
    file_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def default_manager_dashboard():
    return {
        "tasks": [],
        "reminders": [],
        "contacts": [],
    }


def read_contracts():
    return read_json_list(CONTRACTS_FILE)


def write_contracts(contracts):
    write_json_list(CONTRACTS_FILE, contracts)


# Manager-created contract templates. Each template is a structured
# JSON {id, name, sections: [{title, paragraphs: [text, ...]}]}.
# At contract render time, when data["templateId"] is set we replace
# the body's sections with the template's; paragraphs go through the
# same sentence-replacement table the DOCX flow uses, so legacy data
# tokens like "Энэхүү гэрээгээр Дэлхий Трэвел Икс нь 2026/..." still
# get substituted with actual dates/names.
CONTRACT_TEMPLATES_FILE = DATA_DIR / "contract_templates.json"


def read_contract_templates():
    return read_json_list(CONTRACT_TEMPLATES_FILE)


def write_contract_templates(records):
    write_json_list(CONTRACT_TEMPLATES_FILE, records)


def read_ds160_applications():
    records = read_json_list(DS160_FILE)
    normalized = [normalize_ds160_record(record) for record in records if isinstance(record, dict)]
    normalized.sort(
        key=lambda item: item.get("submittedAt") or item.get("updatedAt") or item.get("createdAt") or "",
        reverse=True,
    )
    return normalized


def write_ds160_applications(records):
    write_json_list(DS160_FILE, [normalize_ds160_record(record) for record in records if isinstance(record, dict)])


def read_finance_entries():
    return read_json_list(FINANCE_FILE)


def write_finance_entries(records):
    write_json_list(FINANCE_FILE, records)


def read_bookings():
    return read_json_list(BOOKINGS_FILE)


def write_bookings(records):
    write_json_list(BOOKINGS_FILE, records)


def read_reservations():
    return read_json_list(RESERVATIONS_FILE)


def write_reservations(records):
    write_json_list(RESERVATIONS_FILE, records)


def read_camp_reservations():
    return read_json_list(CAMP_RESERVATIONS_FILE)


def write_camp_reservations(records):
    write_json_list(CAMP_RESERVATIONS_FILE, records)


def read_flight_reservations():
    return read_json_list(FLIGHT_RESERVATIONS_FILE)


def write_flight_reservations(records):
    write_json_list(FLIGHT_RESERVATIONS_FILE, records)


def read_transfer_reservations():
    return read_json_list(TRANSFER_RESERVATIONS_FILE)


def write_transfer_reservations(records):
    write_json_list(TRANSFER_RESERVATIONS_FILE, records)


def read_camp_trips():
    records = read_json_list(CAMP_TRIPS_FILE)
    if any(not r.get("serial") for r in records):
        records = backfill_trip_serials(records)
        write_json_list(CAMP_TRIPS_FILE, records)
    return records


def write_camp_trips(records):
    write_json_list(CAMP_TRIPS_FILE, records)


def backfill_trip_serials(records):
    counters = {}
    sorted_records = sorted(
        records,
        key=lambda r: (
            r.get("createdAt") or "",
            r.get("startDate") or "",
            r.get("id") or "",
        ),
    )
    for record in sorted_records:
        if record.get("serial"):
            continue
        company = normalize_company(record.get("company"))
        # USM uses S- (Steppe) and everything else uses T-. Mirrors
        # next_trip_serial below so backfill matches new-trip numbering.
        prefix = "S-" if company == "USM" else "T-"
        counters.setdefault(company, 0)
        counters[company] += 1
        # Find a unique serial that doesn't collide with existing ones
        while True:
            candidate = f"{prefix}{counters[company]:04d}"
            if not any(r.get("serial") == candidate for r in records):
                record["serial"] = candidate
                break
            counters[company] += 1
    return records


def migrate_usm_serials_t_to_s_once():
    # One-shot migration: USM trips originally got T- serials (or were
    # back-filled with T-) before USM switched to the S- prefix. This walks
    # camp_trips, groups, tourists exactly once and rewrites the dependent
    # serials so the UI shows S-XXXX consistently. The marker file makes
    # it idempotent so deploys / restarts don't re-run it.
    marker = DATA_DIR / ".migration_usm_serial_t_to_s_done"
    if marker.exists():
        return
    try:
        trips = read_json_list(CAMP_TRIPS_FILE)
    except Exception:
        return
    rename_map = {}
    used_serials = {(t.get("serial") or "").upper() for t in trips}
    for trip in trips:
        if normalize_company(trip.get("company")) != "USM":
            continue
        old = (trip.get("serial") or "").upper()
        if not old.startswith("T-"):
            continue
        new = "S-" + old[2:]
        if new in used_serials:
            continue
        trip["serial"] = new
        used_serials.discard(old)
        used_serials.add(new)
        rename_map[old] = new
    if not rename_map:
        try:
            marker.write_text("")
        except Exception:
            pass
        return
    try:
        write_json_list(CAMP_TRIPS_FILE, trips)
    except Exception:
        return
    # Update group records that point at a renamed trip serial.
    try:
        groups = read_json_list(GROUPS_FILE)
        for g in groups:
            old_ts = (g.get("tripSerial") or "").upper()
            if old_ts not in rename_map:
                continue
            new_ts = rename_map[old_ts]
            g["tripSerial"] = new_ts
            g_serial = g.get("serial") or ""
            if g_serial.startswith(old_ts + "-"):
                g["serial"] = new_ts + g_serial[len(old_ts):]
        write_json_list(GROUPS_FILE, groups)
    except Exception:
        pass
    # Same for tourist records.
    try:
        tourists = read_json_list(TOURISTS_FILE)
        for t in tourists:
            old_ts = (t.get("tripSerial") or "").upper()
            if old_ts not in rename_map:
                continue
            new_ts = rename_map[old_ts]
            t["tripSerial"] = new_ts
            t_serial = t.get("serial") or ""
            if t_serial.startswith(old_ts + "-"):
                t["serial"] = new_ts + t_serial[len(old_ts):]
        write_json_list(TOURISTS_FILE, tourists)
    except Exception:
        pass
    try:
        marker.write_text("")
    except Exception:
        pass


def next_trip_serial(company):
    company = normalize_company(company)
    prefix = "S-" if company == "USM" else "T-"
    existing = read_json_list(CAMP_TRIPS_FILE)
    max_seq = 0
    for record in existing:
        if normalize_company(record.get("company")) != company:
            continue
        serial = (record.get("serial") or "").upper()
        # Read either prefix when scanning so legacy USM trips (T-prefix)
        # don't collide with the new S- numbering.
        for known in ("T-", "S-"):
            if serial.startswith(known):
                try:
                    num = int(serial[len(known):])
                    if num > max_seq:
                        max_seq = num
                except ValueError:
                    pass
                break
    return f"{prefix}{max_seq + 1:04d}"


def read_tourist_groups():
    return read_json_list(GROUPS_FILE)


def write_tourist_groups(records):
    write_json_list(GROUPS_FILE, records)


def read_tourists():
    return read_json_list(TOURISTS_FILE)


def write_tourists(records):
    write_json_list(TOURISTS_FILE, records)


def read_invoices():
    try:
        if not INVOICES_FILE.exists():
            INVOICES_FILE.write_text("[]", encoding="utf-8")
        return json.loads(INVOICES_FILE.read_text(encoding="utf-8"))
    except Exception:
        return []


def write_invoices(records):
    INVOICES_FILE.parent.mkdir(parents=True, exist_ok=True)
    INVOICES_FILE.write_text(json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8")


def read_payment_requests():
    try:
        if not PAYMENT_REQUESTS_FILE.exists():
            PAYMENT_REQUESTS_FILE.write_text("[]", encoding="utf-8")
        return json.loads(PAYMENT_REQUESTS_FILE.read_text(encoding="utf-8"))
    except Exception:
        return []


def write_payment_requests(records):
    PAYMENT_REQUESTS_FILE.parent.mkdir(parents=True, exist_ok=True)
    PAYMENT_REQUESTS_FILE.write_text(
        json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8"
    )


def read_trip_templates():
    try:
        if not TRIP_TEMPLATES_FILE.exists():
            TRIP_TEMPLATES_FILE.write_text("[]", encoding="utf-8")
        return json.loads(TRIP_TEMPLATES_FILE.read_text(encoding="utf-8"))
    except Exception:
        return []


def write_trip_templates(records):
    TRIP_TEMPLATES_FILE.parent.mkdir(parents=True, exist_ok=True)
    TRIP_TEMPLATES_FILE.write_text(
        json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8"
    )


def read_service_templates():
    try:
        if not SERVICE_TEMPLATES_FILE.exists():
            SERVICE_TEMPLATES_FILE.write_text("[]", encoding="utf-8")
        return json.loads(SERVICE_TEMPLATES_FILE.read_text(encoding="utf-8"))
    except Exception:
        return []


def write_service_templates(records):
    SERVICE_TEMPLATES_FILE.parent.mkdir(parents=True, exist_ok=True)
    SERVICE_TEMPLATES_FILE.write_text(
        json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8"
    )


def _normalize_trip_template_lines(raw):
    if not isinstance(raw, list):
        return []
    out = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        try:
            day_offset = int(item.get("dayOffset") or 1)
        except (TypeError, ValueError):
            day_offset = 1
        try:
            amount = float(item.get("amount") or 0)
        except (TypeError, ValueError):
            amount = 0.0
        out.append({
            "dayOffset": max(0, day_offset),
            "category": str(item.get("category") or "").strip(),
            "payeeName": str(item.get("payeeName") or "").strip(),
            "amount": amount,
            "currency": (str(item.get("currency") or "MNT").strip().upper() or "MNT"),
            "note": str(item.get("note") or "").strip(),
        })
    return out


def read_trip_creators():
    """Trip-creator records: presentation + quotation per trip, keyed by tripId.
    Stored as a JSON object {tripId: doc} so the per-trip lookup is O(1)."""
    try:
        if not TRIP_CREATORS_FILE.exists():
            TRIP_CREATORS_FILE.write_text("{}", encoding="utf-8")
        data = json.loads(TRIP_CREATORS_FILE.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def write_trip_creators(store):
    TRIP_CREATORS_FILE.parent.mkdir(parents=True, exist_ok=True)
    TRIP_CREATORS_FILE.write_text(json.dumps(store, indent=2, ensure_ascii=False), encoding="utf-8")


def next_invoice_serial(company=None):
    """Legacy fallback used when an invoice has no group context. Invoice
    serials are workspace-scoped: USM uses S-NNNNNN, DTX uses T-NNNNNN.
    Counters are shared across both prefixes (and legacy plain-digit
    serials) so we never clash with what's already in storage.

    Normal flow uses next_invoice_serial_from_group() instead so the
    serial matches the group it bills."""
    company = normalize_company(company) if company else ""
    prefix = "S-" if company == "USM" else ("T-" if company == "DTX" else "")
    existing = read_invoices()
    max_seq = 0
    for record in existing:
        s = str(record.get("serial") or "")
        core = s
        for known in ("S-", "T-"):
            if s.startswith(known):
                core = s[len(known):]
                break
        if core.isdigit():
            max_seq = max(max_seq, int(core))
    return f"{prefix}{(max_seq + 1):06d}"


def next_invoice_serial_from_group(group_serial, fallback_company=None):
    """Invoice serial = group serial (e.g. 'T-0001-G1'). If that serial
    is already taken by another invoice (revisions, accidental duplicate
    invoice for the same group), append -2 / -3 / ... until free.

    Falls back to the legacy company counter when no group serial is
    available so old code paths still work."""
    base = (group_serial or "").strip()
    if not base:
        return next_invoice_serial(fallback_company)
    used = {str(r.get("serial") or "") for r in read_invoices()}
    if base not in used:
        return base
    n = 2
    while f"{base}-{n}" in used:
        n += 1
    return f"{base}-{n}"


def find_tourist_group_serial(group_id):
    """Look up a group's serial by id. Returns "" if not found."""
    if not group_id:
        return ""
    for g in read_tourist_groups():
        if g.get("id") == group_id:
            return str(g.get("serial") or "")
    return ""


def _trip_company_for_invoice(payload):
    """Resolve a trip's company from an invoice payload's tripId, falling back
    to whatever the payload already carried so legacy/manual rows aren't
    re-prefixed accidentally."""
    trip_id = (payload or {}).get("tripId")
    if not trip_id:
        return ""
    trip = find_camp_trip(trip_id)
    return normalize_company((trip or {}).get("company")) if trip else ""


def build_invoice(payload, actor=None):
    items = []
    for raw in payload.get("items") or []:
        desc = normalize_text(raw.get("description"))
        if not desc:
            continue
        try:
            qty = float(raw.get("qty") or 1)
        except Exception:
            qty = 1
        try:
            price = float(raw.get("price") or 0)
        except Exception:
            price = 0
        items.append({
            "description": desc,
            "qty": qty,
            "price": price,
            "total": round(qty * price, 2),
        })
    installments = []
    for raw in payload.get("installments") or []:
        desc = normalize_text(raw.get("description")) or "Installment"
        try:
            amount = float(raw.get("amount") or 0)
        except Exception:
            amount = 0
        installments.append({
            "description": desc,
            "amount": amount,
            "issueDate": normalize_text(raw.get("issueDate")),
            "dueDate": normalize_text(raw.get("dueDate")),
            "status": normalize_text(raw.get("status")) or "pending",
        })
    total = round(sum(i["total"] for i in items), 2)
    # Per-invoice FX rate to MNT, frozen at save time so later rate
    # moves don't rewrite past income. MNT invoices default to 1;
    # USD/EUR/etc default to whatever the form sent (the form
    # auto-suggests USD=3500, EUR=4100 but the user can override).
    currency_for_invoice = (normalize_text(payload.get("currency")) or "MNT").upper()
    try:
        fx_rate = float(payload.get("fxRate") or 0)
    except Exception:
        fx_rate = 0
    if currency_for_invoice == "MNT":
        fx_rate = 1.0
    if fx_rate <= 0:
        fx_rate = 1.0
    mnt_total = round(total * fx_rate, 2)
    # Snapshot the chosen bank account onto the invoice itself so the
    # invoice render survives later edits in Settings.
    bank_id = normalize_text(payload.get("bankAccountId"))
    bank_snapshot = None
    if bank_id:
        for b in read_settings().get("bankAccounts") or []:
            if b.get("id") == bank_id:
                bank_snapshot = b
                break
    if not bank_snapshot and payload.get("bankAccount"):
        bank_snapshot = payload.get("bankAccount")
    group_id = normalize_text(payload.get("groupId"))
    return {
        "id": str(uuid4()),
        "serial": next_invoice_serial_from_group(
            find_tourist_group_serial(group_id),
            _trip_company_for_invoice(payload),
        ),
        "tripId": normalize_text(payload.get("tripId")),
        "groupId": group_id,
        "payerId": normalize_text(payload.get("payerId")),
        "payerName": normalize_text(payload.get("payerName")),
        "payerAddress": normalize_text(payload.get("payerAddress")),
        "participantIds": [normalize_text(x) for x in (payload.get("participantIds") or []) if x],
        "items": items,
        "total": total,
        "installments": installments,
        "currency": normalize_text(payload.get("currency")) or "MNT",
        "fxRate": fx_rate,
        "mntTotal": mnt_total,
        "bankAccountId": bank_id,
        "bankAccount": bank_snapshot,
        "status": "draft",
        "createdAt": now_mongolia().isoformat(),
        "createdBy": actor_snapshot(actor),
        "updatedAt": "",
        "updatedBy": actor_snapshot(actor),
    }


def validate_invoice(data):
    if not data.get("tripId"):
        return "tripId is required"
    if not data.get("groupId"):
        return "groupId is required"
    if not data.get("payerName") and not data.get("payerId"):
        return "Payer is required"
    if not data.get("items"):
        return "At least one item is required"
    return None


def reconcile_trip_serial_with_company(trip):
    """If a trip's stored serial uses a prefix that doesn't match its
    current company (e.g. company switched from USM→DTX after the serial
    was minted, or it was migrated incorrectly), rewrite the serial in
    place so the prefix lines up with the active company. Persists the
    fix to disk and cascades to existing group/tourist serials that
    embed the old trip serial.

    Returns the corrected serial string. No-op when already consistent.
    """
    if not trip:
        return ""
    serial = (trip.get("serial") or "").strip()
    if not serial:
        return serial
    company = normalize_company(trip.get("company"))
    expected_prefix = "S-" if company == "USM" else "T-"
    cur_prefix = serial[:2] if serial[:2] in ("S-", "T-") else ""
    if not cur_prefix or cur_prefix == expected_prefix:
        return serial
    new_serial = expected_prefix + serial[len(cur_prefix):]
    # Persist the trip update.
    try:
        trips = read_camp_trips()
        for rec in trips:
            if rec.get("id") == trip.get("id") and (rec.get("serial") or "") == serial:
                rec["serial"] = new_serial
                break
        write_json_list(CAMP_TRIPS_FILE, trips)
    except Exception:
        return serial
    trip["serial"] = new_serial
    # Cascade: existing groups + tourists store serials prefixed with the
    # old trip serial; rewrite them so navigation links + display stay in
    # sync.
    try:
        groups = read_json_list(GROUPS_FILE)
        changed = False
        for g in groups:
            if (g.get("tripSerial") or "") == serial:
                g["tripSerial"] = new_serial
                changed = True
            g_serial = g.get("serial") or ""
            if g_serial.startswith(serial + "-"):
                g["serial"] = new_serial + g_serial[len(serial):]
                changed = True
        if changed:
            write_json_list(GROUPS_FILE, groups)
    except Exception:
        pass
    try:
        tourists = read_json_list(TOURISTS_FILE)
        changed = False
        for t in tourists:
            if (t.get("tripSerial") or "") == serial:
                t["tripSerial"] = new_serial
                changed = True
            t_serial = t.get("serial") or ""
            if t_serial.startswith(serial + "-"):
                t["serial"] = new_serial + t_serial[len(serial):]
                changed = True
        if changed:
            write_json_list(TOURISTS_FILE, tourists)
    except Exception:
        pass
    return new_serial


def next_group_serial(trip_serial, trip_id):
    if not trip_serial:
        return ""
    existing = read_tourist_groups()
    max_seq = 0
    prefix = f"{trip_serial}-G"
    for record in existing:
        if record.get("tripId") != trip_id:
            continue
        serial = record.get("serial") or ""
        if serial.startswith(prefix):
            try:
                num = int(serial[len(prefix):])
                if num > max_seq:
                    max_seq = num
            except ValueError:
                pass
    return f"{prefix}{max_seq + 1}"


def next_tourist_serial(group_serial, group_id):
    if not group_serial:
        return ""
    existing = read_tourists()
    max_seq = 0
    prefix = f"{group_serial}-"
    for record in existing:
        if record.get("groupId") != group_id:
            continue
        serial = record.get("serial") or ""
        if serial.startswith(prefix):
            try:
                num = int(serial[len(prefix):])
                if num > max_seq:
                    max_seq = num
            except ValueError:
                pass
    return f"{prefix}{max_seq + 1:02d}"


def next_promo_tourist_serial():
    """Serial for trip-less promo contacts: PR-0001, PR-0002, … global to the
    tourists store (no per-group prefix since they don't belong to a group)."""
    prefix = "PR-"
    max_seq = 0
    for record in read_tourists():
        serial = record.get("serial") or ""
        if serial.startswith(prefix):
            try:
                n = int(serial[len(prefix):])
                if n > max_seq:
                    max_seq = n
            except ValueError:
                pass
    return f"{prefix}{max_seq + 1:04d}"


def normalize_manager_dashboard(payload):
    if not isinstance(payload, dict):
        return default_manager_dashboard()
    return {
        "tasks": payload.get("tasks") if isinstance(payload.get("tasks"), list) else [],
        "reminders": payload.get("reminders") if isinstance(payload.get("reminders"), list) else [],
        "contacts": payload.get("contacts") if isinstance(payload.get("contacts"), list) else [],
    }


def read_manager_dashboard():
    payload = read_json_object(MANAGER_DASHBOARD_FILE, default_manager_dashboard())
    return normalize_manager_dashboard(payload)


def write_manager_dashboard(payload):
    write_json_object(MANAGER_DASHBOARD_FILE, normalize_manager_dashboard(payload))


def normalize_option_list(values):
    cleaned = []
    for value in values or []:
        text = normalize_text(value)
        if text and text not in cleaned:
            cleaned.append(text)
    return cleaned


DEFAULT_AIRLINE_ALIASES = [
    {"name": "Turkish Airlines", "alias": "TK"},
    {"name": "MIAT", "alias": "MIAT"},
    {"name": "Hunnu Air", "alias": "Hunnu"},
    {"name": "Aero Mongolia", "alias": "Aero"},
]


def default_camp_settings():
    return {
        "campNames": ["Khustai camp"],
        "locationNames": ["Khustai"],
        "staffAssignments": [STEPPE_MANAGER],
        "roomChoices": DEFAULT_ROOM_CHOICES,
        "campLocations": {"Khustai camp": "Khustai"},
        "transferPlaces": [],
        "transferDrivers": [],
        "airlineAliases": list(DEFAULT_AIRLINE_ALIASES),
    }


def normalize_airline_aliases(payload):
    """Airline alias records: { name, alias }. Empty names are dropped, alias
    falls back to the airline name if missing. Names are case-insensitively
    deduped so the picker doesn't show duplicates."""
    out = []
    if not isinstance(payload, list):
        return out
    seen = set()
    for item in payload:
        if not isinstance(item, dict):
            continue
        name = normalize_text(item.get("name"))
        if not name:
            continue
        key = name.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append({
            "name": name,
            "alias": normalize_text(item.get("alias")) or name,
        })
    return out


def normalize_transfer_drivers(payload):
    """Driver records: { id, name, carType, plateNumber, phoneNumber, salary }.
    Salary is stored as integer MNT. Existing IDs are preserved; new entries
    get a generated UUID so transfer-reservation snapshots stay stable."""
    out = []
    if not isinstance(payload, list):
        return out
    seen = set()
    for item in payload:
        if not isinstance(item, dict):
            continue
        name = normalize_text(item.get("name"))
        if not name:
            continue
        rec = {
            "id": normalize_text(item.get("id")) or str(uuid4()),
            "name": name,
            "carType": normalize_text(item.get("carType")),
            "plateNumber": normalize_text(item.get("plateNumber")),
            "phoneNumber": normalize_text(item.get("phoneNumber")),
            "salary": parse_int(item.get("salary")),
        }
        if rec["id"] in seen:
            continue
        seen.add(rec["id"])
        out.append(rec)
    return out


def normalize_camp_location_map(payload, camp_names=None):
    normalized = {}
    if isinstance(payload, dict):
        items = payload.items()
    else:
        items = []
    for camp_name, location_name in items:
        camp_text = normalize_text(camp_name)
        location_text = normalize_text(location_name)
        if camp_text:
            normalized[camp_text] = location_text
    for camp_name in camp_names or []:
        normalized.setdefault(camp_name, "")
    return normalized


def read_camp_settings():
    payload = read_json_object(CAMP_SETTINGS_FILE, default_camp_settings())
    camp_names = normalize_option_list(payload.get("campNames")) or ["Khustai camp"]
    camp_locations = normalize_camp_location_map(payload.get("campLocations"), camp_names)
    location_names = normalize_option_list(payload.get("locationNames")) or ["Khustai"]
    location_names = normalize_option_list(location_names + [value for value in camp_locations.values() if value])
    raw_details = payload.get("campDetails") or {}
    camp_details = {}
    if isinstance(raw_details, dict):
        for name, value in raw_details.items():
            key = normalize_text(name)
            if not key or not isinstance(value, dict):
                continue
            camp_details[key] = {
                "price": normalize_text(value.get("price")),
                "contractPath": normalize_text(value.get("contractPath")),
            }
    for name in camp_names:
        camp_details.setdefault(name, {"price": "", "contractPath": ""})
    return {
        "campNames": camp_names,
        "locationNames": location_names,
        "staffAssignments": normalize_option_list(payload.get("staffAssignments")) or [STEPPE_MANAGER],
        "roomChoices": normalize_option_list(payload.get("roomChoices")) or DEFAULT_ROOM_CHOICES,
        "campLocations": camp_locations,
        "campDetails": camp_details,
        # Pickup and dropoff used to be two lists — merge legacy data into the
        # unified transferPlaces list so a saved place is usable on both sides.
        "transferPlaces": normalize_option_list(
            (payload.get("transferPlaces") or [])
            + (payload.get("transferPickups") or [])
            + (payload.get("transferDropoffs") or [])
        ),
        "transferDrivers": normalize_transfer_drivers(payload.get("transferDrivers")),
        "airlineAliases": normalize_airline_aliases(payload.get("airlineAliases")) or list(DEFAULT_AIRLINE_ALIASES),
    }


def write_camp_settings(payload):
    write_json_object(CAMP_SETTINGS_FILE, payload)


# Pre-seeded destinations from Bataa's reference list. Settings page lets
# admins add/remove on top of this; the seed only fires when settings.json
# doesn't exist yet.
DEFAULT_DESTINATIONS = [
    "Ази", "Африк", "Европ", "Араб", "Гэр бүл", "Фу Куок", "Хайнан",
    "Дубай", "Вьетнам", "Круйз", "Сингапур", "Малайз", "Бали", "Жэжү",
    "Египет", "Турк", "Хятад", "Бангкок Паттаяа", "Пукет", "Тайланд",
    "Шанхай", "Америк",
]


def default_settings():
    # Seed with the two long-standing TravelX bank accounts so the invoice
    # bank picker has something to render the moment a new workspace boots.
    # Keys ("state" / "golomt") match the legacy INVOICE_BANK_ACCOUNTS keys
    # so old invoices keep resolving to the same option.
    return {
        "destinations": list(DEFAULT_DESTINATIONS),
        "bankAccounts": [
            {
                "id": "state",
                "label": "Төрийн Банк · MNT",
                "bankName": "Төрийн Банк",
                "accountName": "Дэлхий Трэвел Икс ХХК",
                "accountNumber": "MN030034 3432 7777 9999",
                "currency": "MNT",
                "swift": "",
                "notes": "",
            },
            {
                "id": "golomt",
                "label": "Голомт Банк · MNT",
                "bankName": "Голомт Банк",
                "accountName": "Дэлхий Трэвел Икс ХХК",
                "accountNumber": "MN80001500 3675114666",
                "currency": "MNT",
                "swift": "",
                "notes": "",
            },
        ],
    }


def _normalize_bank_accounts(value):
    """Each bank account is {id, label, bankName, accountName, accountNumber,
    currency, swift, notes, company}. company is "DTX" / "USM" / "" — empty
    means the account shows in both companies' bank dropdowns. Tolerates
    missing fields."""
    if not isinstance(value, list):
        return []
    out = []
    for item in value:
        if not isinstance(item, dict):
            continue
        label = normalize_text(item.get("label"))
        bank = normalize_text(item.get("bankName"))
        if not label and not bank:
            continue
        company = normalize_text(item.get("company")).upper()
        if company not in ("DTX", "USM"):
            company = ""
        out.append({
            "id": normalize_text(item.get("id")) or uuid4().hex,
            "label": label or bank,
            "bankName": bank,
            "accountName": normalize_text(item.get("accountName")),
            "accountNumber": normalize_text(item.get("accountNumber")),
            "currency": (normalize_text(item.get("currency")) or "MNT").upper(),
            "swift": normalize_text(item.get("swift")),
            "notes": normalize_text(item.get("notes")),
            "company": company,
        })
    return out


_DEFAULT_EXPENSE_CATEGORIES = [
    # Trip-related (outgoing payments tied to a specific trip)
    "Camp deposit", "Camp payment", "Hotel deposit", "Hotel payment",
    "Flight deposit", "Flight payment", "Transfer payment",
    "Driver salary", "Cook salary", "Guide salary",
    "Insurance", "Visa fees", "Restaurant", "Activities",
    # Office / overhead
    "Office rent", "Electricity", "СӨХ", "Internet", "Television",
    "Google Ads", "Meta Ads", "Chatbot", "Manager salary",
    "Director salary", "Accountant salary", "Bonus",
    # Catch-alls
    "Other",
]


def _normalize_categories(raw):
    if not isinstance(raw, list):
        return []
    seen = set()
    out = []
    for value in raw:
        v = str(value or "").strip()
        if not v:
            continue
        key = v.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(v)
    return out


def read_settings():
    payload = read_json_object(SETTINGS_FILE, default_settings())
    destinations = normalize_option_list(payload.get("destinations"))
    if not destinations:
        destinations = list(DEFAULT_DESTINATIONS)
    bank_accounts = _normalize_bank_accounts(payload.get("bankAccounts"))
    # First-run + legacy installations don't have bankAccounts persisted yet
    # — fall back to the seeded state+golomt entries so the invoice bank
    # picker has options the moment the page loads.
    if not bank_accounts:
        bank_accounts = _normalize_bank_accounts(default_settings()["bankAccounts"])
    expense_categories = _normalize_categories(payload.get("expenseCategories"))
    if not expense_categories:
        expense_categories = list(_DEFAULT_EXPENSE_CATEGORIES)
    payees = _normalize_categories(payload.get("expensePayees"))
    return {
        "destinations": destinations,
        "bankAccounts": bank_accounts,
        "expenseCategories": expense_categories,
        "expensePayees": payees,
    }


def write_settings(payload):
    write_json_object(SETTINGS_FILE, payload)


def default_fifa2026_data():
    return {"tickets": [], "sales": []}


def normalize_fifa2026_store(payload):
    default_payload = default_fifa2026_data()
    if not isinstance(payload, dict):
        return default_payload
    tickets = payload.get("tickets") if isinstance(payload.get("tickets"), list) else default_payload["tickets"]
    sales = payload.get("sales") if isinstance(payload.get("sales"), list) else []
    return {
        "tickets": tickets,
        "sales": sales,
    }


def read_fifa2026_store():
    payload = read_json_object(FIFA2026_FILE, default_fifa2026_data())
    return normalize_fifa2026_store(payload)


def write_fifa2026_store(payload):
    write_json_object(FIFA2026_FILE, normalize_fifa2026_store(payload))


def reset_fifa2026_store_from_seed():
    payload = normalize_fifa2026_store(default_fifa2026_data())
    write_fifa2026_store(payload)
    return payload


def fifa_store_totals(store):
    tickets = store.get("tickets", [])
    return {
        "lots": len(tickets),
        "matches": len({normalize_text(ticket.get("matchNumber")) for ticket in tickets if normalize_text(ticket.get("matchNumber"))}),
        "quantity": sum(max(parse_int(ticket.get("totalQuantity")), 0) for ticket in tickets),
    }


def ensure_fifa2026_manual_inventory():
    current = read_fifa2026_store()
    if current.get("tickets") or current.get("sales"):
        return current
    restored = restore_json_from_latest_backup(
        FIFA2026_FILE.name,
        lambda payload: (
            normalized
            if (normalized := normalize_fifa2026_store(payload)) and (normalized.get("tickets") or normalized.get("sales"))
            else None
        ),
    )
    if restored:
        write_fifa2026_store(restored)
        return restored
    return current


def json_response(start_response, status, payload, extra_headers=None):
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    headers = [
        ("Content-Type", "application/json; charset=utf-8"),
        ("Content-Length", str(len(body))),
    ]
    if extra_headers:
        headers.extend(extra_headers)
    start_response(status, headers)
    return [body]


def text_response(start_response, status, text, extra_headers=None):
    body = text.encode("utf-8")
    headers = [
        ("Content-Type", "text/plain; charset=utf-8"),
        ("Content-Length", str(len(body))),
    ]
    if extra_headers:
        headers.extend(extra_headers)
    start_response(status, headers)
    return [body]


def bytes_response(start_response, status, body, content_type, extra_headers=None):
    headers = [
        ("Content-Type", content_type),
        ("Content-Length", str(len(body))),
    ]
    if extra_headers:
        headers.extend(extra_headers)
    start_response(status, headers)
    return [body]


def file_response(start_response, file_path, extra_headers=None):
    mime_type, _ = mimetypes.guess_type(file_path.name)
    body = file_path.read_bytes()
    headers = [
        ("Content-Type", mime_type or "application/octet-stream"),
        ("Content-Length", str(len(body))),
    ]
    # Always revalidate HTML, JS, and CSS so cache-busting query strings
    # take effect immediately across browsers. Heavier static assets
    # (images, fonts, PDFs) are still browser-cacheable.
    if file_path.suffix.lower() in (".html", ".htm", ".js", ".css"):
        headers.append(("Cache-Control", "no-cache, no-store, must-revalidate"))
        headers.append(("Pragma", "no-cache"))
        headers.append(("Expires", "0"))
    if extra_headers:
        headers.extend(extra_headers)
    start_response("200 OK", headers)
    return [body]


def generated_download_headers(file_path):
    if file_path.suffix.lower() != ".pdf":
        return None
    return [("Content-Disposition", f'inline; filename="{file_path.name}"')]


def _serve_html_with_inline_data(start_response, html_path, data):
    """Serve an HTML file but replace the <!--__INITIAL_DATA__--> marker
    with a <script type="application/json"> tag holding `data`. Lets public
    pages render without a round-trip fetch on the first paint."""
    body = html_path.read_text(encoding="utf-8")
    if data is None:
        replacement = ""
    else:
        # Escape `</` so a literal "</script>" inside the JSON payload can't
        # break out of the script tag. ensure_ascii=False keeps Mongolian
        # characters readable in view-source.
        encoded = json.dumps(data, ensure_ascii=False).replace("</", "<\\/")
        replacement = f'<script type="application/json" id="__initial_data">{encoded}</script>'
    body = body.replace("<!--__INITIAL_DATA__-->", replacement)
    encoded_body = body.encode("utf-8")
    headers = [
        ("Content-Type", "text/html; charset=utf-8"),
        ("Content-Length", str(len(encoded_body))),
        # The doc is per-request; don't let intermediaries cache stale data.
        ("Cache-Control", "no-cache, no-store, must-revalidate"),
        ("Pragma", "no-cache"),
        ("Expires", "0"),
    ]
    start_response("200 OK", headers)
    return [encoded_body]


def collect_json(environ):
    try:
        content_length = int(environ.get("CONTENT_LENGTH") or "0")
        raw_body = environ["wsgi.input"].read(content_length)
        return json.loads(raw_body.decode("utf-8"))
    except (ValueError, json.JSONDecodeError):
        return None


def is_authorized(environ):
    query = environ.get("QUERY_STRING", "")
    params = {}
    for pair in query.split("&"):
        if "=" not in pair:
            continue
        key, value = pair.split("=", 1)
        params[key] = value
    return params.get("token") == ADMIN_TOKEN or environ.get("HTTP_X_ADMIN_TOKEN") == ADMIN_TOKEN


def build_user_account(payload):
    return {
        "id": str(uuid4()),
        "createdAt": datetime.now(timezone.utc).isoformat(),
        "approvedAt": "",
        "resetRequestedAt": "",
        "lastLoginAt": "",
        "fullName": normalize_text(payload.get("fullName")),
        "contractLastName": normalize_text(payload.get("contractLastName")),
        "contractFirstName": normalize_text(payload.get("contractFirstName")),
        "contractEmail": normalize_text(payload.get("contractEmail")).lower(),
        "contractPhone": normalize_text(payload.get("contractPhone")),
        "contractSignaturePath": normalize_text(payload.get("contractSignaturePath")),
        "email": normalize_text(payload.get("email")).lower(),
        "passwordHash": hash_password(payload.get("password") or ""),
        "role": "staff",
        "status": "pending",
    }


def validate_user_account(payload):
    email = normalize_text(payload.get("email")).lower()
    password = str(payload.get("password") or "")
    if not email or "@" not in email:
        return "Valid email is required"
    if len(password) < 6:
        return "Password must be at least 6 characters"
    if "confirmPassword" in payload and str(payload.get("confirmPassword") or "") != password:
        return "Passwords do not match"
    return None


def create_session(user_id):
    sessions = read_sessions()
    token = secrets.token_urlsafe(32)
    sessions.insert(0, {
        "token": token,
        "userId": user_id,
        "createdAt": datetime.now(timezone.utc).isoformat(),
    })
    write_sessions(sessions)
    return token


def handle_auth_register(environ, start_response):
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    error = validate_user_account(payload)
    if error:
        return json_response(start_response, "400 Bad Request", {"error": error})
    if find_user_by_email(payload.get("email")):
        return json_response(start_response, "400 Bad Request", {"error": "This email is already registered"})
    users = read_users()
    users.insert(0, build_user_account(payload))
    write_users(users)
    return json_response(start_response, "201 Created", {"ok": True, "message": "Your account request was sent. Wait for admin approval."})


def handle_auth_request_reset(environ, start_response):
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    email = normalize_text(payload.get("email")).lower()
    if not email or "@" not in email:
        return json_response(start_response, "400 Bad Request", {"error": "Valid email is required"})
    users = read_users()
    for user in users:
        if user.get("email") == email:
            user["resetRequestedAt"] = datetime.now(timezone.utc).isoformat()
            break
    write_users(users)
    return json_response(start_response, "200 OK", {"ok": True, "message": "If the account exists, a reset request was sent to admin."})


def handle_auth_bootstrap_admin(environ, start_response):
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    if payload.get("adminToken") != ADMIN_TOKEN:
        return json_response(start_response, "401 Unauthorized", {"error": "Invalid admin token"})
    if any(user.get("role") == "admin" for user in read_users()):
        return json_response(start_response, "400 Bad Request", {"error": "Admin account already exists"})
    error = validate_user_account(payload)
    if error:
        return json_response(start_response, "400 Bad Request", {"error": error})
    user = build_user_account(payload)
    user["role"] = "admin"
    user["status"] = "approved"
    user["approvedAt"] = datetime.now(timezone.utc).isoformat()
    users = read_users()
    users.insert(0, user)
    write_users(users)
    token = create_session(user["id"])
    return json_response(
        start_response,
        "201 Created",
        {"ok": True, "user": sanitize_user(user)},
        extra_headers=[("Set-Cookie", make_session_cookie(token))],
    )


def handle_auth_login(environ, start_response):
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    user = find_user_by_email(payload.get("email"))
    if not user or user.get("passwordHash") != hash_password(payload.get("password") or ""):
        return json_response(start_response, "401 Unauthorized", {"error": "Incorrect email or password"})
    if user.get("status") != "approved":
        return json_response(start_response, "403 Forbidden", {"error": f"Your account is {user.get('status', 'pending')}. Wait for admin approval."})
    token = create_session(user["id"])
    users = read_users()
    for item in users:
        if item["id"] == user["id"]:
            item["lastLoginAt"] = datetime.now(timezone.utc).isoformat()
    write_users(users)
    return json_response(
        start_response,
        "200 OK",
        {"ok": True, "user": sanitize_user(find_user_by_id(user["id"]))},
        extra_headers=[("Set-Cookie", make_session_cookie(token))],
    )


def handle_auth_logout(environ, start_response):
    token = parse_cookies(environ).get(SESSION_COOKIE)
    if token:
        sessions = [item for item in read_sessions() if item.get("token") != token]
        write_sessions(sessions)
    return json_response(
        start_response,
        "200 OK",
        {"ok": True},
        extra_headers=[("Set-Cookie", clear_session_cookie())],
    )


def handle_auth_me(environ, start_response):
    user = current_user(environ)
    if not user:
        return json_response(start_response, "401 Unauthorized", {"error": "Not logged in"})
    return json_response(start_response, "200 OK", {"user": user})


def handle_auth_profile_update(environ, start_response):
    user = require_login(environ, start_response)
    if not user:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})

    full_name = normalize_text(payload.get("fullName"))
    contract_last_name = normalize_text(payload.get("contractLastName"))
    contract_first_name = normalize_text(payload.get("contractFirstName"))
    contract_email = normalize_text(payload.get("contractEmail")).lower()
    contract_phone = normalize_text(payload.get("contractPhone"))
    signature_data = payload.get("contractSignatureData")
    avatar_data = payload.get("avatarData")
    remove_avatar = bool(payload.get("removeAvatar"))
    if len(full_name) < 2:
        return json_response(start_response, "400 Bad Request", {"error": "Name must be at least 2 characters"})
    if contract_last_name and len(contract_last_name) < 2:
        return json_response(start_response, "400 Bad Request", {"error": "Contract surname must be at least 2 characters"})
    if contract_first_name and len(contract_first_name) < 2:
        return json_response(start_response, "400 Bad Request", {"error": "Contract given name must be at least 2 characters"})
    if contract_email and "@" not in contract_email:
        return json_response(start_response, "400 Bad Request", {"error": "Contract email must be valid"})

    users = read_users()
    for record in users:
        if record.get("id") != user.get("id"):
            continue
        if signature_data:
            signature_path = save_manager_signature_image(signature_data, record["id"])
            if not signature_path:
                return json_response(start_response, "400 Bad Request", {"error": "Invalid manager signature"})
            record["contractSignaturePath"] = signature_path
        if avatar_data:
            avatar_path = save_user_avatar_image(avatar_data, record["id"])
            if not avatar_path:
                return json_response(start_response, "400 Bad Request", {"error": "Invalid profile picture"})
            record["avatarPath"] = avatar_path
        elif remove_avatar:
            record["avatarPath"] = ""
        record["fullName"] = full_name
        record["contractLastName"] = contract_last_name
        record["contractFirstName"] = contract_first_name
        record["contractEmail"] = contract_email
        record["contractPhone"] = contract_phone
        write_users(users)
        return json_response(start_response, "200 OK", {"ok": True, "user": sanitize_user(record)})

    return json_response(start_response, "404 Not Found", {"error": "User not found"})


def handle_list_users(environ, start_response):
    admin = require_admin(environ, start_response)
    if not admin:
        return []
    users = [sanitize_user(user) for user in read_users()]
    return json_response(start_response, "200 OK", {"entries": users})


def handle_list_notifications(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    try:
        scan_task_reminders()
    except Exception as exc:
        print(f"[task-reminder] scan error: {exc}", flush=True)
    params = parse_qs(environ.get("QUERY_STRING", ""))
    try:
        limit = int(params.get("limit", ["50"])[0])
    except (TypeError, ValueError):
        limit = 50
    limit = max(1, min(limit, NOTIFICATIONS_MAX))
    since = normalize_text(params.get("since", [""])[0])
    entries = read_notifications()
    if since:
        filtered = [e for e in entries if e.get("createdAt", "") > since]
        entries = filtered if filtered else entries
    last_read_iso = ""
    users = read_users()
    for record in users:
        if record.get("id") == actor.get("id"):
            last_read_iso = record.get("notificationsLastReadAt", "")
            break
    unread = [e for e in entries if e.get("createdAt", "") > last_read_iso] if last_read_iso else list(entries)
    return json_response(
        start_response,
        "200 OK",
        {
            "entries": entries[:limit],
            "unread": len(unread),
            "lastReadAt": last_read_iso,
        },
    )


def handle_mark_notifications_read(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    users = read_users()
    now_iso = datetime.now(timezone.utc).isoformat()
    for record in users:
        if record.get("id") == actor.get("id"):
            record["notificationsLastReadAt"] = now_iso
            write_users(users)
            break
    return json_response(start_response, "200 OK", {"ok": True, "lastReadAt": now_iso})


def handle_list_team_members(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    entries = []
    for user in read_users():
        if user.get("status") != "approved":
            continue
        display_name = normalize_text(user.get("fullName")) or normalize_text(user.get("name")) or normalize_text(user.get("email"))
        if not display_name:
            continue
        entries.append(
            {
                "id": user.get("id"),
                "fullName": display_name,
                "contractLastName": normalize_text(user.get("contractLastName")),
                "contractFirstName": normalize_text(user.get("contractFirstName")),
                "contractEmail": normalize_text(user.get("contractEmail")).lower(),
                "contractPhone": normalize_text(user.get("contractPhone")),
                "contractSignaturePath": normalize_text(user.get("contractSignaturePath")),
                "avatarPath": normalize_text(user.get("avatarPath")),
                "email": normalize_text(user.get("email")),
                "role": normalize_text(user.get("role")) or "staff",
            }
        )
    entries.sort(key=lambda item: item["fullName"].lower())
    return json_response(start_response, "200 OK", {"entries": entries})


# ── Admin broadcast announcements ───────────────────────────────────
# Admin posts a message in /admin → it shows up as a centred modal on
# every other manager's next page load. Each user can dismiss it once,
# admin can archive it to retire it for everyone.

def read_announcements():
    return read_json_list(ANNOUNCEMENTS_FILE)


def write_announcements(records):
    write_json_list(ANNOUNCEMENTS_FILE, records)


def _announcement_attachment_view(rec):
    att = rec.get("attachment") or None
    if not att:
        return None
    return {
        "originalName": att.get("originalName"),
        "size": att.get("size"),
        "mimeType": att.get("mimeType"),
        "downloadUrl": f"/api/announcements/{rec.get('id')}/attachment",
    }


def handle_list_announcements_active(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    user_id = actor.get("id")
    out = []
    for rec in read_announcements():
        if rec.get("archived"):
            continue
        if user_id in (rec.get("dismissedBy") or []):
            continue
        out.append({
            "id": rec.get("id"),
            "title": rec.get("title", ""),
            "body": rec.get("body", ""),
            "createdAt": rec.get("createdAt", ""),
            "createdBy": rec.get("createdBy", {}),
            "attachment": _announcement_attachment_view(rec),
        })
    out.sort(key=lambda r: r.get("createdAt") or "", reverse=True)
    return json_response(start_response, "200 OK", {"entries": out})


def handle_list_announcements(environ, start_response):
    admin = require_admin(environ, start_response)
    if not admin:
        return []
    users = read_users()
    user_name_by_id = {u.get("id"): (u.get("fullName") or u.get("email") or "") for u in users}
    out = []
    for rec in read_announcements():
        dismissed = rec.get("dismissedBy") or []
        out.append({
            "id": rec.get("id"),
            "title": rec.get("title", ""),
            "body": rec.get("body", ""),
            "createdAt": rec.get("createdAt", ""),
            "createdBy": rec.get("createdBy", {}),
            "archived": bool(rec.get("archived")),
            "dismissedCount": len(dismissed),
            "dismissedNames": [user_name_by_id.get(uid, uid) for uid in dismissed if uid],
            "attachment": _announcement_attachment_view(rec),
        })
    out.sort(key=lambda r: r.get("createdAt") or "", reverse=True)
    return json_response(start_response, "200 OK", {"entries": out})


def handle_create_announcement(environ, start_response):
    admin = require_admin(environ, start_response)
    if not admin:
        return []
    content_type = environ.get("CONTENT_TYPE", "")
    title = ""
    body = ""
    upload = None
    if "multipart/form-data" in content_type:
        fields, files = parse_multipart(environ)
        title = normalize_text(fields.get("title"))
        body = str(fields.get("body") or "").strip()
        upload = files.get("file")
    else:
        payload = collect_json(environ)
        if payload is None:
            return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
        title = normalize_text(payload.get("title"))
        body = str(payload.get("body") or "").strip()
    if not title or not body:
        return json_response(start_response, "400 Bad Request", {"error": "Title and message are both required"})
    ann_id = str(uuid4())
    attachment = None
    if upload and upload.get("data"):
        original_name = upload.get("filename") or "file"
        ext = Path(original_name).suffix.lower()
        if ext not in ALLOWED_UPLOAD_EXTENSIONS:
            return json_response(start_response, "400 Bad Request", {"error": f"File type {ext} not allowed"})
        data = upload["data"]
        if len(data) > MAX_UPLOAD_BYTES:
            return json_response(start_response, "400 Bad Request", {"error": "File too large (max 10 MB)"})
        ensure_data_store()
        stored_name = ann_id + ext
        (ANNOUNCEMENT_UPLOADS_DIR / stored_name).write_bytes(data)
        attachment = {
            "originalName": original_name,
            "storedName": stored_name,
            "mimeType": upload.get("content_type") or "application/octet-stream",
            "size": len(data),
        }
    rec = {
        "id": ann_id,
        "title": title,
        "body": body,
        "createdAt": now_mongolia().isoformat(),
        "createdBy": actor_snapshot(admin),
        "archived": False,
        "dismissedBy": [],
        "attachment": attachment,
    }
    records = read_announcements()
    records.append(rec)
    write_announcements(records)
    return json_response(start_response, "200 OK", {"ok": True, "entry": rec})


def handle_download_announcement_attachment(environ, start_response, announcement_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    rec = next((r for r in read_announcements() if r.get("id") == announcement_id), None)
    if not rec or not rec.get("attachment"):
        return json_response(start_response, "404 Not Found", {"error": "Attachment not found"})
    att = rec["attachment"]
    file_path = (ANNOUNCEMENT_UPLOADS_DIR / (att.get("storedName") or "")).resolve()
    if not str(file_path).startswith(str(ANNOUNCEMENT_UPLOADS_DIR.resolve())) or not file_path.exists():
        return json_response(start_response, "404 Not Found", {"error": "File missing"})
    data = file_path.read_bytes()
    headers = [
        ("Content-Type", att.get("mimeType") or "application/octet-stream"),
        ("Content-Length", str(len(data))),
        (
            "Content-Disposition",
            'attachment; filename="' + (att.get("originalName") or "file").replace('"', "") + '"',
        ),
        ("Cache-Control", "private, max-age=0, no-cache"),
    ]
    start_response("200 OK", headers)
    return [data]


# ── Trip Creator (presentation + quotation per trip) ──────────────────
# A "trip creator" doc is the public-facing brochure + price quote attached
# to one trip. Stored as {tripId: doc} so per-trip read/write is O(1).
TRIP_CREATOR_DEFAULT = {
    "title": "",
    "subtitle": "",  # e.g. "Family-friendly · 7 days · Central Mongolia"
    "language": "mn",
    "currency": "MNT",
    "totalDays": "",
    "totalKm": "",
    "priceFrom": "",
    "tripType": "TRIP",
    "offerType": "range",
    "internationalFlight": "included",
    "themes": [],
    "rate": 0,
    "comfort": 0,
    "difficulty": 0,
    "intro": "",
    "coverIds": [],
    "program": [],
    # Highlights bullets shown above day cards.
    "highlights": [],
    # Accommodation summary lines like "Ulaanbaatar (2 nights) – 3★ Hotel".
    "accommSummary": [],   # list[{nights, label, hotel}]
    # What the price covers / does not.
    "included": [],
    "notIncluded": [],
    # Per-trip primary contact shown in the sidebar.
    "manager": {"name": "", "role": "", "phone": "", "email": "", "avatar": ""},
    # International flight legs for the brochure flight table.
    "flightLegs": [],     # list[{n, date, dep, depFrom, arr, arrTo, flight}]
    # Free-text Mongolia / region guide shown in the public-page Guide tab.
    "mongoliaGuide": "",
    "quotation": {"rows": [], "note": ""},
}


def _ensure_id(rec):
    """Stable id helper. Days, accommSummary rows and flight legs get a uuid
    on first save so future edits, drag-and-drop reorders, and template
    instancing can target them by id rather than array index."""
    cur = (rec.get("id") or "").strip()
    if cur:
        return cur
    return str(uuid4())


def _trip_creator_normalize(payload):
    """Coerce client payload to the storage shape — protects the JSON file
    against junk types if the editor ever sends a wrong field."""
    out = dict(TRIP_CREATOR_DEFAULT)
    if not isinstance(payload, dict):
        return out
    for key in ("title", "subtitle", "language", "currency", "totalDays", "totalKm",
                "priceFrom", "tripType", "offerType", "internationalFlight",
                "intro", "mongoliaGuide"):
        if key in payload and payload[key] is not None:
            out[key] = str(payload[key])
    if isinstance(payload.get("themes"), list):
        out["themes"] = [str(t) for t in payload["themes"] if str(t).strip()]
    if isinstance(payload.get("coverIds"), list):
        out["coverIds"] = [str(t).strip() for t in payload["coverIds"] if str(t).strip()]
    for key in ("rate", "comfort", "difficulty"):
        try:
            out[key] = int(payload.get(key) or 0)
        except (TypeError, ValueError):
            out[key] = 0
    if isinstance(payload.get("program"), list):
        prog = []
        for entry in payload["program"]:
            if not isinstance(entry, dict):
                continue
            day_image_ids = entry.get("imageIds")
            if isinstance(day_image_ids, list):
                day_image_ids = [str(i).strip() for i in day_image_ids if str(i).strip()]
            else:
                day_image_ids = []
            meals_raw = entry.get("meals") or {}
            if not isinstance(meals_raw, dict):
                meals_raw = {}
            def _meal_venue(v):
                # Legacy schema stored bools (B/L/D checkboxes). Drop those
                # — the new schema is venue strings ("Hotel", "Restaurant").
                if isinstance(v, bool):
                    return ""
                return str(v or "").strip()
            prog.append({
                "id": _ensure_id(entry),
                "templateId": str(entry.get("templateId") or "").strip(),
                "locationId": str(entry.get("locationId") or "").strip(),
                "day": str(entry.get("day") or "").strip(),
                "title": str(entry.get("title") or "").strip(),
                "date": str(entry.get("date") or "").strip(),
                "fromName": str(entry.get("fromName") or "").strip(),
                "toName": str(entry.get("toName") or "").strip(),
                "distance": str(entry.get("distance") or "").strip(),
                "drive": str(entry.get("drive") or "").strip(),
                "accommodation": str(entry.get("accommodation") or "").strip(),
                "meals": {
                    "breakfast": _meal_venue(meals_raw.get("breakfast")),
                    "lunch": _meal_venue(meals_raw.get("lunch")),
                    "dinner": _meal_venue(meals_raw.get("dinner")),
                },
                "body": str(entry.get("body") or ""),
                "imageIds": day_image_ids,
            })
        out["program"] = prog
    # Free-form bullet lists.
    for key in ("highlights", "included", "notIncluded"):
        if isinstance(payload.get(key), list):
            out[key] = [str(s).strip() for s in payload[key] if str(s).strip()]
    if isinstance(payload.get("accommSummary"), list):
        summary = []
        for row in payload["accommSummary"]:
            if not isinstance(row, dict):
                continue
            summary.append({
                "id": _ensure_id(row),
                "templateId": str(row.get("templateId") or "").strip(),
                "nights": str(row.get("nights") or "").strip(),
                "label": str(row.get("label") or "").strip(),
                "hotel": str(row.get("hotel") or "").strip(),
            })
        out["accommSummary"] = summary
    if isinstance(payload.get("manager"), dict):
        m = payload["manager"]
        out["manager"] = {
            "name": str(m.get("name") or "").strip(),
            "role": str(m.get("role") or "").strip(),
            "phone": str(m.get("phone") or "").strip(),
            "email": str(m.get("email") or "").strip(),
            "avatar": str(m.get("avatar") or "").strip(),
        }
    if isinstance(payload.get("flightLegs"), list):
        legs = []
        for leg in payload["flightLegs"]:
            if not isinstance(leg, dict):
                continue
            legs.append({
                "id": _ensure_id(leg),
                "templateId": str(leg.get("templateId") or "").strip(),
                "n": str(leg.get("n") or "").strip(),
                "date": str(leg.get("date") or "").strip(),
                "dep": str(leg.get("dep") or "").strip(),
                "depFrom": str(leg.get("depFrom") or "").strip(),
                "arr": str(leg.get("arr") or "").strip(),
                "arrTo": str(leg.get("arrTo") or "").strip(),
                "flight": str(leg.get("flight") or "").strip(),
            })
        out["flightLegs"] = legs
    if isinstance(payload.get("quotation"), dict):
        q = payload["quotation"]
        rows = []
        if isinstance(q.get("rows"), list):
            for row in q["rows"]:
                if not isinstance(row, dict):
                    continue
                rows.append({
                    "label": str(row.get("label") or "").strip(),
                    "qty": str(row.get("qty") or "").strip(),
                    "unitPrice": str(row.get("unitPrice") or "").strip(),
                    "currency": str(row.get("currency") or "").strip(),
                })
        out["quotation"] = {
            "rows": rows,
            "note": str(q.get("note") or ""),
        }
    return out


def handle_get_trip_creator(environ, start_response, trip_id):
    if not require_login(environ, start_response):
        return []
    store = read_trip_creators()
    doc = store.get(trip_id) or {}
    # Merge with defaults so the editor always sees every key.
    merged = dict(TRIP_CREATOR_DEFAULT)
    merged.update(doc)
    return json_response(start_response, "200 OK", {"trip_id": trip_id, "doc": merged})


def build_public_trip_view(trip_id):
    """Reusable build of the public trip-creator brochure shape. Returns
    the dict (or None if the trip isn't published). Used both by the API
    endpoint and by the inline injection on /trip/<id>."""
    store = read_trip_creators()
    doc = store.get(trip_id)
    if not doc:
        return None
    trip = next((t for t in read_camp_trips() if t.get("id") == trip_id), None) or {}
    # Manager block: prefer the explicit doc.manager that the brochure
    # editor sets; for any missing field fall back to the last editor's
    # login profile (name, email, phone, avatar) so the public sidebar
    # always shows a real contact + face.
    manager = dict(doc.get("manager") or {})
    editor_snap = doc.get("updatedBy") or doc.get("createdBy") or {}
    editor_user = find_user_by_id(editor_snap.get("id")) if editor_snap.get("id") else None
    if not manager.get("name"):
        manager["name"] = (editor_user.get("fullName") if editor_user else "") or editor_snap.get("name") or ""
    if not manager.get("email"):
        manager["email"] = (editor_user.get("email") if editor_user else "") or editor_snap.get("email") or ""
    if not manager.get("phone") and editor_user:
        manager["phone"] = editor_user.get("contractPhone") or ""
    if not manager.get("avatar") and editor_user:
        manager["avatar"] = editor_user.get("avatarPath") or ""
    # Map waypoints: dedup'd from each program row's fromName/toName,
    # in order, so the public sidebar can render a Google-Maps directions
    # embed that traces the trip route.
    waypoints = []
    seen = set()
    for row in (doc.get("program") or []):
        for key in ("fromName", "toName"):
            value = (row.get(key) or "").strip()
            if not value:
                continue
            low = value.lower()
            if low in seen:
                continue
            seen.add(low)
            waypoints.append(value)
    return {
        "tripId": trip_id,
        "title": doc.get("title") or trip.get("tripName") or "",
        "subtitle": doc.get("subtitle") or "",
        "totalDays": doc.get("totalDays") or trip.get("totalDays") or "",
        "totalKm": doc.get("totalKm") or "",
        "priceFrom": doc.get("priceFrom") or "",
        "language": doc.get("language") or "mn",
        "tripType": doc.get("tripType") or trip.get("tripType") or "TRIP",
        "currency": doc.get("currency") or "MNT",
        "offerType": doc.get("offerType") or "range",
        "internationalFlight": doc.get("internationalFlight") or "included",
        "themes": doc.get("themes") or [],
        "rate": doc.get("rate") or 0,
        "comfort": doc.get("comfort") or 0,
        "difficulty": doc.get("difficulty") or 0,
        "intro": doc.get("intro") or "",
        "coverIds": doc.get("coverIds") or [],
        "program": doc.get("program") or [],
        "highlights": doc.get("highlights") or [],
        "accommSummary": doc.get("accommSummary") or [],
        "included": doc.get("included") or [],
        "notIncluded": doc.get("notIncluded") or [],
        "manager": manager,
        "mapWaypoints": waypoints,
        "flightLegs": doc.get("flightLegs") or [],
        "mongoliaGuide": doc.get("mongoliaGuide") or "",
        "quotation": doc.get("quotation") or {"rows": [], "note": ""},
        "trip": {
            "serial": trip.get("serial"),
            "startDate": trip.get("startDate"),
            "endDate": trip.get("endDate"),
            "company": normalize_company(trip.get("company")),
        },
    }


def handle_get_public_trip_creator(environ, start_response, trip_id):
    """No auth required — this is the read-only client-facing brochure
    fetched by /trip/<id>. Returns only fields that are safe to share
    publicly (no internal Mongolian notes the editor might use)."""
    public = build_public_trip_view(trip_id)
    if not public:
        return json_response(start_response, "404 Not Found", {"error": "Trip not published"})
    return json_response(start_response, "200 OK", public)


def handle_save_trip_creator(environ, start_response, trip_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    # Validate the trip exists — saving against a nonexistent tripId would
    # leave orphan rows in the store.
    if not any(t.get("id") == trip_id for t in read_camp_trips()):
        return json_response(start_response, "404 Not Found", {"error": "Trip not found"})
    doc = _trip_creator_normalize(payload)
    doc["updatedAt"] = now_mongolia().isoformat()
    doc["updatedBy"] = actor_snapshot(actor)
    store = read_trip_creators()
    store[trip_id] = doc
    write_trip_creators(store)
    return json_response(start_response, "200 OK", {"ok": True, "doc": doc})


# ── Gallery (shared media library: images + video URLs) ───────────────
GALLERY_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
# Pre-generated size variants. We keep client-compressed full image (≤1600px)
# as the canonical, and lazy-generate two smaller copies for grid consumers
# so a phone listing the gallery downloads ~30 KB thumbnails not 400 KB
# full photos. Skipped for animated GIFs (Pillow can't write multi-frame
# resampled GIFs cleanly without ImageMagick).
GALLERY_VARIANT_DIMS = {"thumb": 400, "medium": 1200}


def _gallery_variant_path(rec, size):
    """Return on-disk path for a size variant of a gallery image. Lazily
    generates the variant on first request and caches it. Returns the full
    image's path if size is unknown or generation fails."""
    stored_name = rec.get("storedName") or ""
    if not stored_name:
        return None
    full_path = (GALLERY_UPLOADS_DIR / stored_name).resolve()
    if not str(full_path).startswith(str(GALLERY_UPLOADS_DIR.resolve())) or not full_path.exists():
        return None
    if not size or size not in GALLERY_VARIANT_DIMS:
        return full_path
    suffix = full_path.suffix.lower()
    if suffix == ".gif":
        # Don't try to resize animated GIFs — fall back to original.
        return full_path
    variant_name = f"{full_path.stem}-{size}{suffix}"
    variant_path = GALLERY_UPLOADS_DIR / variant_name
    if variant_path.exists():
        return variant_path
    try:
        from PIL import Image
        max_dim = GALLERY_VARIANT_DIMS[size]
        with Image.open(full_path) as img:
            img.load()
            # Don't upscale — if the source is already smaller than the target
            # variant, we'd just be re-encoding for no reason. Hand back the
            # original instead.
            if max(img.size) <= max_dim:
                return full_path
            img.thumbnail((max_dim, max_dim))
            save_kwargs = {"optimize": True}
            if suffix in {".jpg", ".jpeg"}:
                if img.mode != "RGB":
                    img = img.convert("RGB")
                save_kwargs["quality"] = 82
                save_kwargs["progressive"] = True
            elif suffix == ".png":
                # Keep transparency for PNG; optimize crunches the size.
                pass
            elif suffix == ".webp":
                save_kwargs["quality"] = 82
                save_kwargs["method"] = 6
            img.save(variant_path, **save_kwargs)
        return variant_path
    except Exception as exc:
        print(f"[gallery] variant gen failed for {stored_name} → {size}: {exc}", flush=True)
        return full_path


def read_gallery():
    return read_json_list(GALLERY_FILE)


def write_gallery(records):
    write_json_list(GALLERY_FILE, records)


def read_gallery_folders():
    """Registered (named) folders. Items still carry the folder name in their
    own `folder` string field — this list just lets a user create an empty
    folder up front so they can set it as the destination at upload time."""
    return read_json_list(GALLERY_FOLDERS_FILE)


def write_gallery_folders(records):
    write_json_list(GALLERY_FOLDERS_FILE, records)


def _all_gallery_folder_names():
    """Union of registered folder names + names derived from gallery items."""
    names = {(rec.get("name") or "").strip() for rec in read_gallery_folders()}
    for item in read_gallery():
        f = (item.get("folder") or "").strip()
        if f:
            names.add(f)
    return sorted({n for n in names if n}, key=str.lower)


def _normalize_gallery_alt(value):
    """Coerce alt into the 9-language object shape.

    - string  → {"en": value, ...other langs blank}
    - dict    → keep recognised keys, blank the rest
    - other   → all blank

    Always returns a dict with every key in LOCATION_LANGUAGES so callers
    don't need to defensively check key presence.
    """
    out = {code: "" for code in LOCATION_LANGUAGES}
    if isinstance(value, dict):
        for code in LOCATION_LANGUAGES:
            v = value.get(code)
            if isinstance(v, str):
                out[code] = v.strip()
    elif isinstance(value, str):
        out["en"] = value.strip()
    return out


def _gallery_view(rec, request_origin=""):
    """Trim down to the public-safe shape + add a downloadable URL."""
    alt_obj = _normalize_gallery_alt(rec.get("alt"))
    return {
        "id": rec.get("id"),
        "kind": rec.get("kind") or "image",
        "originalName": rec.get("originalName") or "",
        "alt": alt_obj,
        "altText": alt_obj.get("en", ""),  # back-compat: legacy callers expecting flat string
        "mimeType": rec.get("mimeType") or "",
        "size": rec.get("size") or 0,
        "tags": rec.get("tags") or [],
        "folder": rec.get("folder") or "",
        "videoUrl": rec.get("videoUrl") or "",
        "uploadedAt": rec.get("uploadedAt") or "",
        "uploadedBy": rec.get("uploadedBy") or {},
        "url": f"/api/gallery/{rec.get('id')}/file" if rec.get("kind") == "image" else (rec.get("videoUrl") or ""),
    }


def handle_list_gallery(environ, start_response):
    if not require_login(environ, start_response):
        return []
    qs = parse_qs(environ.get("QUERY_STRING", ""))
    tag = (qs.get("tag") or [""])[0].strip().lower()
    kind = (qs.get("kind") or [""])[0].strip().lower()
    folder = (qs.get("folder") or [""])[0].strip()
    q = (qs.get("q") or [""])[0].strip().lower()
    rows = read_gallery()
    if kind:
        rows = [r for r in rows if (r.get("kind") or "image") == kind]
    if folder:
        # "__none__" matches items without a folder, anything else is exact match.
        if folder == "__none__":
            rows = [r for r in rows if not (r.get("folder") or "").strip()]
        else:
            rows = [r for r in rows if (r.get("folder") or "").lower() == folder.lower()]
    if tag:
        rows = [r for r in rows if any(tag == t.lower() for t in (r.get("tags") or []))]
    if q:
        rows = [
            r for r in rows
            if q in (r.get("originalName") or "").lower()
            or any(q in t.lower() for t in (r.get("tags") or []))
        ]
    rows.sort(key=lambda r: r.get("uploadedAt") or "", reverse=True)
    folder_names = _all_gallery_folder_names()
    counts = {}
    none_count = 0
    for item in read_gallery():
        f = (item.get("folder") or "").strip()
        if f:
            counts[f] = counts.get(f, 0) + 1
        else:
            none_count += 1
    folders_with_counts = [{"name": n, "count": counts.get(n, 0)} for n in folder_names]
    out = [_gallery_view(r) for r in rows]
    return json_response(start_response, "200 OK", {
        "entries": out,
        "count": len(out),
        "folders": folder_names,
        "folderStats": folders_with_counts,
        "noFolderCount": none_count,
        "totalCount": len(read_gallery()),
    })


def handle_update_gallery_item(environ, start_response, item_id):
    """Patch fields on a gallery record. Currently used for folder + tags."""
    if not require_login(environ, start_response):
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    records = read_gallery()
    rec = next((r for r in records if r.get("id") == item_id), None)
    if not rec:
        return json_response(start_response, "404 Not Found", {"error": "Item not found"})
    if "folder" in payload:
        rec["folder"] = str(payload["folder"] or "").strip()
    if "tags" in payload:
        raw = payload["tags"]
        if isinstance(raw, list):
            rec["tags"] = [str(t).strip() for t in raw if str(t).strip()]
        else:
            rec["tags"] = [t.strip() for t in str(raw).split(",") if t.strip()]
    if "originalName" in payload:
        new_name = str(payload["originalName"] or "").strip()
        if new_name:
            rec["originalName"] = new_name
    if "alt" in payload:
        rec["alt"] = _normalize_gallery_alt(payload["alt"])
    write_gallery(records)
    return json_response(start_response, "200 OK", {"ok": True, "entry": _gallery_view(rec)})


def handle_upload_gallery_image(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    fields, files = parse_multipart(environ)
    upload = files.get("file")
    if not upload or not upload.get("data"):
        return json_response(start_response, "400 Bad Request", {"error": "No file provided"})
    original_name = upload.get("filename") or "image"
    ext = Path(original_name).suffix.lower()
    if ext not in GALLERY_IMAGE_EXTS:
        return json_response(start_response, "400 Bad Request", {"error": f"Image type {ext} not allowed"})
    data = upload["data"]
    if len(data) > MAX_UPLOAD_BYTES:
        return json_response(start_response, "400 Bad Request", {"error": "Image too large (max 10 MB)"})
    ensure_data_store()
    rec_id = str(uuid4())
    stored_name = rec_id + ext
    (GALLERY_UPLOADS_DIR / stored_name).write_bytes(data)
    raw_tags = fields.get("tags") or ""
    tag_list = [t.strip() for t in raw_tags.split(",") if t.strip()]
    folder = (fields.get("folder") or "").strip()
    # Upload form posts a JSON-encoded alt object (one key per language)
    # under the "alt" field. Fall back to plain string for older clients.
    raw_alt = fields.get("alt") or ""
    try:
        alt_payload = json.loads(raw_alt) if raw_alt.startswith("{") else raw_alt
    except (ValueError, TypeError):
        alt_payload = raw_alt
    alt_obj = _normalize_gallery_alt(alt_payload)
    rec = {
        "id": rec_id,
        "kind": "image",
        "originalName": original_name,
        "alt": alt_obj,
        "storedName": stored_name,
        "mimeType": upload.get("content_type") or "application/octet-stream",
        "size": len(data),
        "tags": tag_list,
        "folder": folder,
        "uploadedAt": now_mongolia().isoformat(),
        "uploadedBy": actor_snapshot(actor),
    }
    records = read_gallery()
    records.append(rec)
    write_gallery(records)
    return json_response(start_response, "201 Created", {"ok": True, "entry": _gallery_view(rec)})


def handle_create_gallery_video(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    url = (payload.get("videoUrl") or "").strip()
    if not url.startswith("http"):
        return json_response(start_response, "400 Bad Request", {"error": "videoUrl must be a full http(s) URL"})
    raw_tags = payload.get("tags") or []
    tag_list = (
        [str(t).strip() for t in raw_tags if str(t).strip()]
        if isinstance(raw_tags, list)
        else [t.strip() for t in str(raw_tags).split(",") if t.strip()]
    )
    rec = {
        "id": str(uuid4()),
        "kind": "video",
        "originalName": (payload.get("title") or url).strip(),
        "videoUrl": url,
        "tags": tag_list,
        "uploadedAt": now_mongolia().isoformat(),
        "uploadedBy": actor_snapshot(actor),
    }
    records = read_gallery()
    records.append(rec)
    write_gallery(records)
    return json_response(start_response, "201 Created", {"ok": True, "entry": _gallery_view(rec)})


def handle_delete_gallery_item(environ, start_response, item_id):
    if not require_login(environ, start_response):
        return []
    records = read_gallery()
    rec = next((r for r in records if r.get("id") == item_id), None)
    if not rec:
        return json_response(start_response, "404 Not Found", {"error": "Item not found"})
    stored = rec.get("storedName")
    if stored:
        path = (GALLERY_UPLOADS_DIR / stored).resolve()
        if str(path).startswith(str(GALLERY_UPLOADS_DIR.resolve())) and path.exists():
            try:
                path.unlink()
            except Exception:
                pass
    write_gallery([r for r in records if r.get("id") != item_id])
    return json_response(start_response, "200 OK", {"ok": True})


def handle_create_gallery_folder(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    name = (payload.get("name") or "").strip()
    if not name:
        return json_response(start_response, "400 Bad Request", {"error": "Folder name required"})
    if len(name) > 60:
        return json_response(start_response, "400 Bad Request", {"error": "Folder name too long"})
    folders = read_gallery_folders()
    if any((f.get("name") or "").strip().lower() == name.lower() for f in folders):
        return json_response(start_response, "400 Bad Request", {"error": "Folder already exists"})
    folders.append({
        "name": name,
        "createdAt": now_mongolia().isoformat(),
        "createdBy": actor_snapshot(actor),
    })
    write_gallery_folders(folders)
    return json_response(start_response, "201 Created", {"ok": True, "name": name})


def handle_rename_gallery_folder(environ, start_response, old_name):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    new_name = (payload.get("name") or "").strip()
    if not new_name:
        return json_response(start_response, "400 Bad Request", {"error": "New name required"})
    folders = read_gallery_folders()
    # Update the registered list
    found = False
    for rec in folders:
        if (rec.get("name") or "").strip().lower() == old_name.strip().lower():
            rec["name"] = new_name
            rec["renamedAt"] = now_mongolia().isoformat()
            found = True
    if not any((rec.get("name") or "").strip().lower() == new_name.lower() for rec in folders):
        folders.append({
            "name": new_name,
            "createdAt": now_mongolia().isoformat(),
            "createdBy": actor_snapshot(actor),
        }) if not found else None
    write_gallery_folders(folders)
    # Move items that were in the old folder
    items = read_gallery()
    moved = 0
    for item in items:
        if (item.get("folder") or "").strip().lower() == old_name.strip().lower():
            item["folder"] = new_name
            moved += 1
    if moved:
        write_gallery(items)
    return json_response(start_response, "200 OK", {"ok": True, "moved": moved})


def handle_delete_gallery_folder(environ, start_response, name):
    """Remove the folder. With ?cascade=true also delete every photo
    inside; otherwise the photos are kept and just un-foldered."""
    if not require_login(environ, start_response):
        return []
    qs = parse_qs(environ.get("QUERY_STRING", ""))
    cascade = (qs.get("cascade") or [""])[0].strip().lower() in ("1", "true", "yes")
    folders = read_gallery_folders()
    target = name.strip().lower()
    new_folders = [f for f in folders if (f.get("name") or "").strip().lower() != target]
    write_gallery_folders(new_folders)
    items = read_gallery()
    if cascade:
        # Remove the actual records + their stored files for items in this folder.
        keep = []
        deleted = 0
        for item in items:
            if (item.get("folder") or "").strip().lower() == target:
                stored = item.get("storedName") or ""
                if stored:
                    fp = GALLERY_UPLOADS_DIR / stored
                    try:
                        if fp.exists():
                            fp.unlink()
                    except OSError:
                        pass
                deleted += 1
            else:
                keep.append(item)
        if deleted:
            write_gallery(keep)
        return json_response(start_response, "200 OK", {"ok": True, "deleted": deleted, "cleared": 0})
    cleared = 0
    for item in items:
        if (item.get("folder") or "").strip().lower() == target:
            item["folder"] = ""
            cleared += 1
    if cleared:
        write_gallery(items)
    return json_response(start_response, "200 OK", {"ok": True, "cleared": cleared, "deleted": 0})


def handle_serve_gallery_image(environ, start_response, item_id):
    """Public — gallery images may appear on /trip/<id> brochures so they
    must be reachable without auth. The item id is unguessable (UUID4).
    Supports ?size=thumb|medium|full (default full)."""
    rec = next((r for r in read_gallery() if r.get("id") == item_id), None)
    if not rec or rec.get("kind") != "image":
        return json_response(start_response, "404 Not Found", {"error": "Image not found"})
    qs = parse_qs(environ.get("QUERY_STRING", ""))
    size = (qs.get("size") or [""])[0].strip().lower()
    path = _gallery_variant_path(rec, size if size in GALLERY_VARIANT_DIMS else "")
    if not path:
        return json_response(start_response, "404 Not Found", {"error": "File missing"})
    data = path.read_bytes()
    # Variants are immutable for a given (id, size) pair, so cache aggressively.
    cache_max = 31536000 if size in GALLERY_VARIANT_DIMS else 86400
    headers = [
        ("Content-Type", rec.get("mimeType") or "application/octet-stream"),
        ("Content-Length", str(len(data))),
        ("Cache-Control", f"public, max-age={cache_max}"),
    ]
    start_response("200 OK", headers)
    return [data]


# ── Content library (Attractions / Hotels / Activities / etc.) ───────
CONTENT_TYPES = {"attraction", "accommodation", "activity", "destination", "supplier", "location"}


def read_content():
    return read_json_list(CONTENT_FILE)


def write_content(records):
    write_json_list(CONTENT_FILE, records)


def _slugify(value):
    """Conservative slug — keep ASCII letters/digits, replace anything else
    with underscores. Used as a fallback when the user doesn't provide one."""
    s = re.sub(r"[^A-Za-z0-9]+", "_", str(value or "").lower()).strip("_")
    return s or "item"


def _content_normalize(payload, existing=None):
    base = dict(existing or {})
    if not isinstance(payload, dict):
        return base
    out = dict(base)
    for key in ("slug", "type", "country", "title", "summary", "videoUrl", "location", "publishStatus"):
        if key in payload and payload[key] is not None:
            out[key] = str(payload[key]).strip()
    out["type"] = (out.get("type") or "").lower() or "attraction"
    if out["type"] not in CONTENT_TYPES:
        out["type"] = "attraction"
    out["publishStatus"] = (out.get("publishStatus") or "published").lower()
    if out["publishStatus"] not in {"published", "draft"}:
        out["publishStatus"] = "draft"
    out["slug"] = _slugify(out.get("slug") or out.get("title") or "")
    if isinstance(payload.get("imageIds"), list):
        out["imageIds"] = [str(i) for i in payload["imageIds"] if str(i).strip()]
    elif "imageIds" not in out:
        out["imageIds"] = []
    if isinstance(payload.get("bulletGroups"), list):
        groups = []
        for g in payload["bulletGroups"]:
            if not isinstance(g, dict):
                continue
            items = g.get("items") or []
            if not isinstance(items, list):
                continue
            groups.append({
                "heading": str(g.get("heading") or "").strip(),
                "items": [str(i).strip() for i in items if str(i).strip()],
            })
        out["bulletGroups"] = groups
    elif "bulletGroups" not in out:
        out["bulletGroups"] = []
    # Per-language translations of the title and summary. English is
    # the source (lives directly on title/summary). Other 8 languages
    # under translations: { mn: {title, summary}, fr: {...}, ... }.
    # Filter to known language codes so the JSON file doesn't grow
    # arbitrarily.
    if isinstance(payload.get("translations"), dict):
        valid_codes = {"mn", "fr", "it", "es", "ko", "zh", "ja", "ru"}
        cleaned = {}
        for code, val in payload["translations"].items():
            code = str(code or "").lower().strip()
            if code not in valid_codes or not isinstance(val, dict):
                continue
            title = str(val.get("title") or "").strip()
            summary = str(val.get("summary") or "").strip()
            if title or summary:
                cleaned[code] = {"title": title, "summary": summary}
        out["translations"] = cleaned
    elif "translations" not in out:
        out["translations"] = {}
    return out


def handle_list_content(environ, start_response):
    if not require_login(environ, start_response):
        return []
    qs = parse_qs(environ.get("QUERY_STRING", ""))
    type_filter = (qs.get("type") or [""])[0].strip().lower()
    country = (qs.get("country") or [""])[0].strip()
    publish = (qs.get("publish") or [""])[0].strip().lower()
    q = (qs.get("q") or [""])[0].strip().lower()
    rows = read_content()
    if type_filter:
        rows = [r for r in rows if (r.get("type") or "") == type_filter]
    if country:
        rows = [r for r in rows if (r.get("country") or "").lower() == country.lower()]
    if publish:
        rows = [r for r in rows if (r.get("publishStatus") or "") == publish]
    if q:
        rows = [
            r for r in rows
            if q in (r.get("slug") or "").lower()
            or q in (r.get("title") or "").lower()
            or q in (r.get("summary") or "").lower()
        ]
    rows.sort(key=lambda r: (r.get("slug") or "").lower())
    return json_response(start_response, "200 OK", {"entries": rows, "count": len(rows)})


def handle_create_content(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    rec = _content_normalize(payload)
    if not rec.get("title"):
        return json_response(start_response, "400 Bad Request", {"error": "Title is required"})
    records = read_content()
    if any(r.get("slug") == rec["slug"] for r in records):
        # Slug collision — append -2, -3 …
        n = 2
        base_slug = rec["slug"]
        while any(r.get("slug") == f"{base_slug}_{n}" for r in records):
            n += 1
        rec["slug"] = f"{base_slug}_{n}"
    rec["id"] = str(uuid4())
    rec["createdAt"] = now_mongolia().isoformat()
    rec["createdBy"] = actor_snapshot(actor)
    rec["updatedAt"] = rec["createdAt"]
    records.append(rec)
    write_content(records)
    return json_response(start_response, "201 Created", {"ok": True, "entry": rec})


def handle_get_content(environ, start_response, item_id):
    if not require_login(environ, start_response):
        return []
    rec = next((r for r in read_content() if r.get("id") == item_id), None)
    if not rec:
        return json_response(start_response, "404 Not Found", {"error": "Content not found"})
    return json_response(start_response, "200 OK", rec)


def handle_update_content(environ, start_response, item_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    records = read_content()
    idx = next((i for i, r in enumerate(records) if r.get("id") == item_id), None)
    if idx is None:
        return json_response(start_response, "404 Not Found", {"error": "Content not found"})
    merged = _content_normalize(payload, existing=records[idx])
    # Keep slug uniqueness — if the user changed the slug to an existing one,
    # block the update so the public lookup table stays clean.
    if any(r.get("slug") == merged["slug"] and r.get("id") != item_id for r in records):
        return json_response(start_response, "400 Bad Request", {"error": f"Slug '{merged['slug']}' already exists"})
    merged["id"] = item_id
    merged["createdAt"] = records[idx].get("createdAt") or now_mongolia().isoformat()
    merged["updatedAt"] = now_mongolia().isoformat()
    merged["updatedBy"] = actor_snapshot(actor)
    records[idx] = merged
    write_content(records)
    return json_response(start_response, "200 OK", {"ok": True, "entry": merged})


def handle_delete_content(environ, start_response, item_id):
    if not require_login(environ, start_response):
        return []
    records = read_content()
    new_records = [r for r in records if r.get("id") != item_id]
    if len(new_records) == len(records):
        return json_response(start_response, "404 Not Found", {"error": "Content not found"})
    write_content(new_records)
    return json_response(start_response, "200 OK", {"ok": True})


CONTENT_VIDEO_DIR = DATA_DIR / "content-videos"
ALLOWED_VIDEO_EXT = {".mp4", ".webm", ".mov", ".m4v"}
MAX_VIDEO_BYTES = 50 * 1024 * 1024  # 50 MB — keep storage tame on Render

def handle_upload_content_video(environ, start_response, item_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    records = read_content()
    idx = next((i for i, r in enumerate(records) if r.get("id") == item_id), None)
    if idx is None:
        return json_response(start_response, "404 Not Found", {"error": "Content not found"})
    fields, files = parse_multipart(environ)
    upload = files.get("file")
    if not upload:
        return json_response(start_response, "400 Bad Request", {"error": "No file uploaded"})
    original = upload.get("filename") or "video"
    ext = Path(original).suffix.lower()
    if ext not in ALLOWED_VIDEO_EXT:
        return json_response(start_response, "400 Bad Request", {"error": f"Video type {ext} not supported. Use MP4, WebM, or MOV."})
    data = upload.get("data") or b""
    if len(data) > MAX_VIDEO_BYTES:
        return json_response(start_response, "400 Bad Request", {"error": "Video too large (max 50 MB)."})
    CONTENT_VIDEO_DIR.mkdir(parents=True, exist_ok=True)
    # Replace any existing video file for this content (different ext OK).
    for existing in CONTENT_VIDEO_DIR.glob(f"{item_id}.*"):
        try: existing.unlink()
        except Exception: pass
    stored_name = f"{item_id}{ext}"
    (CONTENT_VIDEO_DIR / stored_name).write_bytes(data)
    rec = dict(records[idx])
    rec["videoFile"] = stored_name
    rec["updatedAt"] = now_mongolia().isoformat()
    rec["updatedBy"] = actor_snapshot(actor)
    records[idx] = rec
    write_content(records)
    return json_response(start_response, "200 OK", {"ok": True, "videoFile": stored_name})


def handle_delete_content_video(environ, start_response, item_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    records = read_content()
    idx = next((i for i, r in enumerate(records) if r.get("id") == item_id), None)
    if idx is None:
        return json_response(start_response, "404 Not Found", {"error": "Content not found"})
    rec = dict(records[idx])
    stored = rec.get("videoFile") or ""
    if stored:
        path = CONTENT_VIDEO_DIR / stored
        if path.exists():
            try: path.unlink()
            except Exception: pass
    rec["videoFile"] = ""
    rec["updatedAt"] = now_mongolia().isoformat()
    records[idx] = rec
    write_content(records)
    return json_response(start_response, "200 OK", {"ok": True})


CONTENT_LANGS = ["en", "mn", "fr", "it", "es", "ko", "zh", "ja", "ru"]


def _pick_content_lang(environ, lang_hint=None):
    """Pick a language code for the public popup. Order:
    1. Explicit ?lang= query parameter (overrides everything).
    2. Accept-Language header — first 2-letter code we support.
    3. Fallback to the source language ('en')."""
    if lang_hint:
        code = str(lang_hint).strip().lower()[:2]
        if code in CONTENT_LANGS:
            return code
    qs = parse_qs(environ.get("QUERY_STRING", ""))
    q_lang = (qs.get("lang", [""])[0] or "").strip().lower()[:2]
    if q_lang in CONTENT_LANGS:
        return q_lang
    accept = (environ.get("HTTP_ACCEPT_LANGUAGE") or "").lower()
    for piece in accept.split(","):
        code = piece.split(";")[0].strip()[:2]
        if code in CONTENT_LANGS:
            return code
    return "en"


def build_public_content_view(slug, environ=None, lang=None):
    rec = next((r for r in read_content() if r.get("slug") == slug), None)
    if not rec or rec.get("publishStatus") == "draft":
        return None
    chosen_lang = _pick_content_lang(environ or {}, lang_hint=lang)
    # Resolve title/summary from the chosen language; fall back to the
    # English source if that language is empty (and finally to whatever
    # is on the record).
    translations = rec.get("translations") or {}
    if chosen_lang == "en":
        loc_title = rec.get("title") or ""
        loc_summary = rec.get("summary") or ""
    else:
        slot = translations.get(chosen_lang) or {}
        loc_title = slot.get("title") or rec.get("title") or ""
        loc_summary = slot.get("summary") or rec.get("summary") or ""
    images = [
        {"id": img_id, "url": f"/api/gallery/{img_id}/file"}
        for img_id in (rec.get("imageIds") or [])
    ]
    # SEO hreflang siblings — every language version of this page.
    # Public consumers (search engines and language switcher) use this
    # to find the right URL per locale.
    base_path = f"/c/{rec.get('slug')}"
    hreflangs = []
    for code in CONTENT_LANGS:
        if code == "en":
            hreflangs.append({"lang": "en", "href": base_path, "label": "English"})
        else:
            slot = translations.get(code) or {}
            if slot.get("title") or slot.get("summary"):
                hreflangs.append({"lang": code, "href": f"{base_path}?lang={code}", "label": code})
    video_file = rec.get("videoFile") or ""
    return {
        "slug": rec.get("slug"),
        "type": rec.get("type"),
        "country": rec.get("country"),
        "title": loc_title,
        "summary": loc_summary,
        "lang": chosen_lang,
        "videoUrl": rec.get("videoUrl"),
        # Manager-uploaded video file. Public popup prefers this over
        # videoUrl when both are set.
        "videoFile": (f"/content-videos/{video_file}" if video_file else ""),
        "location": rec.get("location") or "",
        "images": images,
        "bulletGroups": rec.get("bulletGroups") or [],
        "hreflangs": hreflangs,
    }


def handle_get_public_content(environ, start_response, slug):
    """No auth — clients hit this through trip brochures."""
    public = build_public_content_view(slug, environ=environ)
    if not public:
        return json_response(start_response, "404 Not Found", {"error": "Content not found"})
    return json_response(start_response, "200 OK", public)


# ──────────────────────────────────────────────────────────────────────────
# Meal templates
# Reusable venue strings ("Hotel", "Local restaurant") that the trip-
# creator surfaces as datalist suggestions for breakfast/lunch/dinner.
# Schema: {id, name, createdAt}.
# ──────────────────────────────────────────────────────────────────────────

def read_meal_templates():
    return read_json_list(MEAL_TEMPLATES_FILE)


def write_meal_templates(records):
    write_json_list(MEAL_TEMPLATES_FILE, records)


MEAL_TEMPLATE_CATEGORIES = {"breakfast", "lunch", "dinner"}


def _meal_template_normalize(payload, existing=None):
    base = dict(existing or {})
    if not isinstance(payload, dict):
        return base
    out = dict(base)
    if "name" in payload:
        out["name"] = str(payload.get("name") or "").strip()
    if "category" in payload:
        cat = str(payload.get("category") or "").strip().lower()
        out["category"] = cat if cat in MEAL_TEMPLATE_CATEGORIES else ""
    return out


def handle_list_meal_templates(environ, start_response):
    if not require_login(environ, start_response):
        return []
    rows = read_meal_templates()
    rows.sort(key=lambda r: (r.get("name") or "").lower())
    return json_response(start_response, "200 OK", {"entries": rows, "count": len(rows)})


def _meal_template_collision(records, rec, ignore_id=None):
    name_lc = (rec.get("name") or "").lower()
    cat = rec.get("category") or ""
    for r in records:
        if r.get("id") == ignore_id:
            continue
        if (r.get("name") or "").lower() != name_lc:
            continue
        # Same name only collides when both are in the same category.
        if (r.get("category") or "") == cat:
            return r
    return None


def handle_create_meal_template(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    rec = _meal_template_normalize(payload)
    if not rec.get("name"):
        return json_response(start_response, "400 Bad Request", {"error": "Name is required"})
    records = read_meal_templates()
    if _meal_template_collision(records, rec):
        return json_response(start_response, "400 Bad Request", {"error": f"'{rec['name']}' already exists in this category"})
    rec["id"] = str(uuid4())
    rec["createdAt"] = now_mongolia().isoformat()
    records.append(rec)
    write_meal_templates(records)
    return json_response(start_response, "201 Created", {"ok": True, "entry": rec})


def handle_update_meal_template(environ, start_response, item_id):
    if not require_login(environ, start_response):
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    records = read_meal_templates()
    idx = next((i for i, r in enumerate(records) if r.get("id") == item_id), None)
    if idx is None:
        return json_response(start_response, "404 Not Found", {"error": "Template not found"})
    merged = _meal_template_normalize(payload, existing=records[idx])
    if not merged.get("name"):
        return json_response(start_response, "400 Bad Request", {"error": "Name is required"})
    if _meal_template_collision(records, merged, ignore_id=item_id):
        return json_response(start_response, "400 Bad Request", {"error": f"'{merged['name']}' already exists in this category"})
    merged["id"] = item_id
    records[idx] = merged
    write_meal_templates(records)
    return json_response(start_response, "200 OK", {"ok": True, "entry": merged})


def handle_delete_meal_template(environ, start_response, item_id):
    if not require_login(environ, start_response):
        return []
    records = read_meal_templates()
    new_records = [r for r in records if r.get("id") != item_id]
    if len(new_records) == len(records):
        return json_response(start_response, "404 Not Found", {"error": "Template not found"})
    write_meal_templates(new_records)
    return json_response(start_response, "200 OK", {"ok": True})


# ──────────────────────────────────────────────────────────────────────────
# Locations (template type)
# A reusable place record (e.g. "Ulaanbaatar", "Kharkhorin") with multi-
# language names, optional lat/lon for the sidebar map, and a list of
# gallery image ids that the trip-creator pulls from to auto-fill day
# hero images. Schema: {id, name, comment, latlonEnabled, latitude,
# longitude, names: {mn,en,fr,it,es,ko,zh,ja,ru}, imageIds[]}.
# ──────────────────────────────────────────────────────────────────────────

LOCATION_LANGUAGES = ["mn", "en", "fr", "it", "es", "ko", "zh", "ja", "ru"]


def read_locations():
    return read_json_list(LOCATIONS_FILE)


def write_locations(records):
    write_json_list(LOCATIONS_FILE, records)


def _location_normalize(payload, existing=None):
    base = dict(existing or {})
    if not isinstance(payload, dict):
        return base
    out = dict(base)
    if "name" in payload:
        out["name"] = str(payload.get("name") or "").strip()
    if "comment" in payload:
        out["comment"] = str(payload.get("comment") or "").strip()
    if "latlonEnabled" in payload:
        out["latlonEnabled"] = bool(payload.get("latlonEnabled"))
    if "latitude" in payload:
        out["latitude"] = str(payload.get("latitude") or "").strip()
    if "longitude" in payload:
        out["longitude"] = str(payload.get("longitude") or "").strip()
    if "names" in payload and isinstance(payload["names"], dict):
        names = dict(out.get("names") or {})
        for lang in LOCATION_LANGUAGES:
            if lang in payload["names"]:
                names[lang] = str(payload["names"].get(lang) or "").strip()
        out["names"] = names
    elif "names" not in out:
        out["names"] = {lang: "" for lang in LOCATION_LANGUAGES}
    # Images schema: list of {id, alt}. Older records used imageIds[]
    # without alt text — accept both shapes on input, always emit the
    # richer "images" form. Alt text is critical for SEO so we expose
    # it on every uploaded photo.
    if "images" in payload and isinstance(payload["images"], list):
        cleaned = []
        for entry in payload["images"]:
            if isinstance(entry, dict):
                img_id = str(entry.get("id") or "").strip()
                alt = str(entry.get("alt") or "").strip()
            else:
                img_id = str(entry or "").strip()
                alt = ""
            if img_id:
                cleaned.append({"id": img_id, "alt": alt})
        out["images"] = cleaned
    elif "imageIds" in payload and isinstance(payload["imageIds"], list):
        out["images"] = [
            {"id": str(i).strip(), "alt": ""}
            for i in payload["imageIds"]
            if str(i).strip()
        ]
    elif "images" not in out:
        # Migrate from legacy imageIds on existing rows.
        legacy = out.get("imageIds") or []
        out["images"] = [{"id": str(i).strip(), "alt": ""} for i in legacy if str(i).strip()]
    # Keep imageIds as a derived convenience for any consumers that
    # still want a flat id list.
    out["imageIds"] = [img["id"] for img in out.get("images") or []]
    return out


def handle_list_locations(environ, start_response):
    if not require_login(environ, start_response):
        return []
    rows = read_locations()
    rows.sort(key=lambda r: (r.get("name") or "").lower())
    return json_response(start_response, "200 OK", {"entries": rows, "count": len(rows)})


def handle_create_location(environ, start_response):
    if not require_login(environ, start_response):
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    rec = _location_normalize(payload)
    if not rec.get("name"):
        return json_response(start_response, "400 Bad Request", {"error": "Name is required"})
    records = read_locations()
    if any((r.get("name") or "").lower() == rec["name"].lower() for r in records):
        return json_response(start_response, "400 Bad Request", {"error": f"'{rec['name']}' already exists"})
    rec["id"] = str(uuid4())
    rec["createdAt"] = now_mongolia().isoformat()
    records.append(rec)
    write_locations(records)
    return json_response(start_response, "201 Created", {"ok": True, "entry": rec})


def handle_get_location(environ, start_response, item_id):
    if not require_login(environ, start_response):
        return []
    rec = next((r for r in read_locations() if r.get("id") == item_id), None)
    if not rec:
        return json_response(start_response, "404 Not Found", {"error": "Location not found"})
    return json_response(start_response, "200 OK", rec)


def handle_update_location(environ, start_response, item_id):
    if not require_login(environ, start_response):
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    records = read_locations()
    idx = next((i for i, r in enumerate(records) if r.get("id") == item_id), None)
    if idx is None:
        return json_response(start_response, "404 Not Found", {"error": "Location not found"})
    merged = _location_normalize(payload, existing=records[idx])
    if not merged.get("name"):
        return json_response(start_response, "400 Bad Request", {"error": "Name is required"})
    if any((r.get("name") or "").lower() == merged["name"].lower() and r.get("id") != item_id for r in records):
        return json_response(start_response, "400 Bad Request", {"error": f"'{merged['name']}' already exists"})
    merged["id"] = item_id
    records[idx] = merged
    write_locations(records)
    return json_response(start_response, "200 OK", {"ok": True, "entry": merged})


def handle_delete_location(environ, start_response, item_id):
    if not require_login(environ, start_response):
        return []
    records = read_locations()
    new_records = [r for r in records if r.get("id") != item_id]
    if len(new_records) == len(records):
        return json_response(start_response, "404 Not Found", {"error": "Location not found"})
    write_locations(new_records)
    return json_response(start_response, "200 OK", {"ok": True})


def handle_dismiss_announcement(environ, start_response, announcement_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    user_id = actor.get("id")
    records = read_announcements()
    for rec in records:
        if rec.get("id") != announcement_id:
            continue
        dismissed = rec.get("dismissedBy") or []
        if user_id and user_id not in dismissed:
            dismissed.append(user_id)
            rec["dismissedBy"] = dismissed
            write_announcements(records)
        return json_response(start_response, "200 OK", {"ok": True})
    return json_response(start_response, "404 Not Found", {"error": "Announcement not found"})


def handle_archive_announcement(environ, start_response, announcement_id):
    admin = require_admin(environ, start_response)
    if not admin:
        return []
    records = read_announcements()
    for rec in records:
        if rec.get("id") != announcement_id:
            continue
        rec["archived"] = True
        rec["archivedAt"] = now_mongolia().isoformat()
        rec["archivedBy"] = actor_snapshot(admin)
        write_announcements(records)
        return json_response(start_response, "200 OK", {"ok": True, "entry": rec})
    return json_response(start_response, "404 Not Found", {"error": "Announcement not found"})


def handle_update_user(environ, start_response, user_id):
    admin = require_admin(environ, start_response)
    if not admin:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    users = read_users()
    for user in users:
        if user["id"] != user_id:
            continue
        if "status" in payload:
            user["status"] = normalize_text(payload.get("status")).lower()
            if user["status"] == "approved":
                user["approvedAt"] = datetime.now(timezone.utc).isoformat()
        if "role" in payload:
            user["role"] = normalize_text(payload.get("role")).lower() or user["role"]
        if "fullName" in payload:
            user["fullName"] = normalize_text(payload.get("fullName"))
        if "contractLastName" in payload:
            user["contractLastName"] = normalize_text(payload.get("contractLastName"))
        if "contractFirstName" in payload:
            user["contractFirstName"] = normalize_text(payload.get("contractFirstName"))
        if "contractEmail" in payload:
            user["contractEmail"] = normalize_text(payload.get("contractEmail")).lower()
        if "contractPhone" in payload:
            user["contractPhone"] = normalize_text(payload.get("contractPhone"))
        if "contractSignaturePath" in payload:
            user["contractSignaturePath"] = normalize_text(payload.get("contractSignaturePath"))
        if "password" in payload:
            password = str(payload.get("password") or "")
            if len(password) < 6:
                return json_response(start_response, "400 Bad Request", {"error": "Password must be at least 6 characters"})
            user["passwordHash"] = hash_password(password)
            user["resetRequestedAt"] = ""
        write_users(users)
        return json_response(start_response, "200 OK", {"ok": True, "user": sanitize_user(user)})
    return json_response(start_response, "404 Not Found", {"error": "User not found"})


def handle_delete_user(environ, start_response, user_id):
    admin = require_admin(environ, start_response)
    if not admin:
        return []
    if user_id == admin.get("id"):
        return json_response(start_response, "400 Bad Request", {"error": "You cannot delete your own admin account"})

    users = read_users()
    before = len(users)
    users = [user for user in users if user.get("id") != user_id]
    if len(users) == before:
        return json_response(start_response, "404 Not Found", {"error": "User not found"})

    sessions = [session for session in read_sessions() if session.get("userId") != user_id]
    write_users(users)
    write_sessions(sessions)
    return json_response(start_response, "200 OK", {"ok": True})


def request_host(environ):
    host = environ.get("HTTP_HOST") or environ.get("SERVER_NAME") or ""
    return host.split(":", 1)[0].strip().lower()


def format_money(value):
    digits = re.sub(r"[^\d]", "", str(value or "0"))
    if not digits:
        digits = "0"
    return f"{int(digits):,}"


def normalize_text(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_person_name(value):
    text = normalize_text(value)
    if not text:
        return ""
    if " " not in text and len(text) % 2 == 0:
        mid = len(text) // 2
        if text[:mid] == text[mid:]:
            text = text[:mid]
    parts = text.split()

    deduped = []
    for part in parts:
        if not deduped or deduped[-1] != part:
            deduped.append(part)
    parts = deduped

    changed = True
    while changed:
        changed = False
        for chunk_size in range(1, (len(parts) // 2) + 1):
            if len(parts) % chunk_size != 0:
                continue
            repeat_count = len(parts) // chunk_size
            if repeat_count <= 1:
                continue
            first = parts[:chunk_size]
            if all(parts[index * chunk_size:(index + 1) * chunk_size] == first for index in range(1, repeat_count)):
                parts = first
                changed = True
                break
    if parts and len(set(parts)) == 1:
        return parts[0]
    return " ".join(parts)


def dedupe_full_name(last_name, first_name):
    last_clean = normalize_person_name(last_name)
    first_clean = normalize_person_name(first_name)
    if last_clean and first_clean and last_clean == first_clean:
        return last_clean
    combined = " ".join(part for part in [last_clean, first_clean] if part).strip()
    return normalize_person_name(combined)


def get_manager_display_name(data):
    full_name = normalize_person_name(data.get("managerFullName"))
    combined_name = dedupe_full_name(data.get("managerLastName"), data.get("managerFirstName"))

    if combined_name and full_name:
        if full_name == combined_name:
            return combined_name

        full_parts = full_name.split()
        combined_parts = combined_name.split()
        if combined_parts:
            if len(full_parts) >= len(combined_parts) and (
                full_parts[: len(combined_parts)] == combined_parts
                or full_parts[-len(combined_parts) :] == combined_parts
            ):
                return combined_name

    return combined_name or full_name or "Ч.Нямбаяр"


def get_manager_contract_formal_name(data):
    last_name = normalize_person_name(data.get("managerLastName"))
    first_name = normalize_person_name(data.get("managerFirstName"))
    if last_name and first_name:
        return f"{last_name} овогтой {first_name}"
    return get_manager_display_name(data)


def get_manager_signature_name(data):
    last_name = normalize_person_name(data.get("managerLastName"))
    first_name = normalize_person_name(data.get("managerFirstName"))
    if last_name and first_name:
        return f"{last_name[:1]}.{first_name}"
    if first_name:
        return first_name
    return get_manager_display_name(data)


def get_manager_contract_email(data):
    return normalize_text(data.get("managerEmail")) or "nyambayar@travelx.mn"


def get_manager_contract_phone(data):
    return normalize_text(data.get("managerPhone")) or "85178877"


def slugify(value):
    cleaned = re.sub(r"[^a-zA-Z0-9а-яА-ЯөүӨҮёЁ]+", "-", normalize_text(value))
    cleaned = re.sub(r"-{2,}", "-", cleaned).strip("-")
    return cleaned or "contract"


def parse_int(value):
    digits = re.sub(r"[^\d]", "", str(value or "0"))
    return int(digits or "0")


def parse_number(value):
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value or "").strip().replace(",", "")
    text = re.sub(r"[^\d.\-]", "", text)
    if text.count(".") > 1:
        parts = text.split(".")
        text = parts[0] + "." + "".join(parts[1:])
    try:
        return float(text or "0")
    except ValueError:
        return 0.0


def now_mongolia():
    return datetime.now(MONGOLIA_TZ)


def today_mongolia():
    return now_mongolia().strftime("%Y-%m-%d")


def parse_date_input(value):
    text = normalize_text(value)
    if not text:
        return None
    for pattern in ("%Y-%m-%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def normalize_stay_fields(check_in, nights, check_out):
    check_in_text = normalize_text(check_in)
    check_out_text = normalize_text(check_out)
    check_in_date = parse_date_input(check_in_text)
    check_out_date = parse_date_input(check_out_text)
    stay_count = parse_int(nights) or 0

    if not check_in_date:
        return check_in_text, max(stay_count, 1), check_out_text

    if stay_count > 0:
        check_out_date = check_in_date + timedelta(days=stay_count)
    elif check_out_date and check_out_date > check_in_date:
        stay_count = (check_out_date - check_in_date).days
    else:
        stay_count = 1
        check_out_date = check_in_date + timedelta(days=stay_count)

    return check_in_text, stay_count, check_out_date.strftime("%Y-%m-%d")


def read_users():
    return read_json_list(USERS_FILE)


def write_users(records):
    write_json_list(USERS_FILE, records)


def read_sessions():
    return read_json_list(SESSIONS_FILE)


def write_sessions(records):
    write_json_list(SESSIONS_FILE, records)


def read_notifications():
    try:
        return read_json_list(NOTIFICATIONS_FILE)
    except Exception:
        return []


def write_notifications(records):
    write_json_list(NOTIFICATIONS_FILE, records[:NOTIFICATIONS_MAX])


# ── Mail account storage ────────────────────────────────────────────
# App passwords are obfuscated (not encrypted) so they aren't trivially
# readable in the JSON file. The persistent disk is private; if an
# attacker reaches it, they have the SESSION_SECRET too. This is just
# a defense-in-depth layer against casual disk inspection.
def _mail_obfuscate(plaintext):
    if not plaintext:
        return ""
    raw = plaintext.encode("utf-8")
    key_seed = (SESSION_SECRET + ":mail").encode("utf-8")
    key = hashlib.sha256(key_seed).digest()
    out = bytes(b ^ key[i % len(key)] for i, b in enumerate(raw))
    return base64.b64encode(out).decode("ascii")


def _mail_deobfuscate(token):
    if not token:
        return ""
    try:
        raw = base64.b64decode(token.encode("ascii"))
    except Exception:
        return ""
    key_seed = (SESSION_SECRET + ":mail").encode("utf-8")
    key = hashlib.sha256(key_seed).digest()
    return bytes(b ^ key[i % len(key)] for i, b in enumerate(raw)).decode("utf-8", errors="replace")


def read_mail_accounts():
    try:
        if not MAIL_ACCOUNTS_PATH.exists():
            return []
        with MAIL_ACCOUNTS_PATH.open("r", encoding="utf-8") as fh:
            data = json.load(fh)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def write_mail_accounts(accounts):
    MAIL_ACCOUNTS_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp = MAIL_ACCOUNTS_PATH.with_suffix(".json.tmp")
    with tmp.open("w", encoding="utf-8") as fh:
        json.dump(accounts, fh, ensure_ascii=False, indent=2)
    tmp.replace(MAIL_ACCOUNTS_PATH)


def public_mail_account(account):
    """Strip the obfuscated password before sending to client."""
    return {k: v for k, v in account.items() if k != "appPassword"}


def get_mail_account_password(account):
    """Decode the stored app password for IMAP/SMTP use."""
    return _mail_deobfuscate(account.get("appPassword") or "")


def log_notification(kind, actor, title, detail="", meta=None):
    try:
        entries = read_notifications()
    except Exception:
        entries = []
    actor_info = actor_snapshot(actor) if actor else {"id": "", "email": "system", "name": "System"}
    entry = {
        "id": uuid4().hex,
        "kind": kind,
        "title": title,
        "detail": detail or "",
        "actor": actor_info,
        "actorAvatar": (actor.get("avatarPath") if isinstance(actor, dict) else "") or "",
        "createdAt": datetime.now(timezone.utc).isoformat(),
        "meta": meta or {},
    }
    entries.insert(0, entry)
    write_notifications(entries)
    return entry


def hash_password(password):
    return hashlib.sha256(f"{SESSION_SECRET}:{password}".encode("utf-8")).hexdigest()


def parse_cookies(environ):
    raw = environ.get("HTTP_COOKIE", "")
    cookies = {}
    for part in raw.split(";"):
        if "=" not in part:
            continue
        key, value = part.split("=", 1)
        cookies[key.strip()] = value.strip()
    return cookies


def make_session_cookie(token, max_age=60 * 60 * 24 * 30):
    return f"{SESSION_COOKIE}={token}; Path=/; HttpOnly; SameSite=Lax; Max-Age={max_age}"


def clear_session_cookie():
    return f"{SESSION_COOKIE}=; Path=/; HttpOnly; SameSite=Lax; Max-Age=0"


WORKSPACE_COOKIE = "activeWorkspace"
VALID_COMPANIES = {"DTX", "USM"}
DEFAULT_COMPANY = "USM"


def active_workspace(environ):
    cookies = parse_cookies(environ)
    value = (cookies.get(WORKSPACE_COOKIE) or "").strip().upper()
    return value if value in VALID_COMPANIES else ""


def normalize_company(value):
    v = (value or "").strip().upper()
    return v if v in VALID_COMPANIES else DEFAULT_COMPANY


def filter_by_company(records, environ):
    workspace = active_workspace(environ)
    if not workspace:
        return records
    return [r for r in records if normalize_company(r.get("company")) == workspace]


def sanitize_user(user):
    return {
        "id": user["id"],
        "email": user["email"],
        "fullName": user.get("fullName", ""),
        "contractLastName": user.get("contractLastName", ""),
        "contractFirstName": user.get("contractFirstName", ""),
        "contractEmail": user.get("contractEmail", ""),
        "contractPhone": user.get("contractPhone", ""),
        "contractSignaturePath": user.get("contractSignaturePath", ""),
        "avatarPath": user.get("avatarPath", ""),
        "role": user.get("role", "staff"),
        "status": user.get("status", "pending"),
        "createdAt": user.get("createdAt"),
        "approvedAt": user.get("approvedAt"),
        "resetRequestedAt": user.get("resetRequestedAt"),
        "lastLoginAt": user.get("lastLoginAt"),
    }


def actor_snapshot(user):
    if not user:
        return {"id": "", "email": "system", "name": "System"}
    return {
        "id": user.get("id", ""),
        "email": user.get("email", ""),
        "name": user.get("fullName") or user.get("email", ""),
    }


# The synthetic actor used by the in-app agent. Every record the agent
# creates (tasks, contacts, tourists, contracts, etc.) and every
# notification it triggers shows up under "Бата" with the panda avatar,
# so the team can see at a glance which actions came from the assistant
# versus which manager. The avatar is served as a static asset.
BATAA_ACTOR_ID = "bataa-agent"
BATAA_ACTOR_NAME = "Бата"
BATAA_ACTOR_EMAIL = "bataa@travelx.mn"
BATAA_AVATAR_PATH = "/assets/bataa-avatar.webp"


def _bataa_agent_actor(human_actor):
    """Wrap the synthetic Бата identity around a real logged-in user so
    permission checks (require_login etc.) still see something user-shaped,
    while record-stamping reads "Бата". The human actor is preserved on
    `_humanActor` for audit purposes."""
    return {
        "id": BATAA_ACTOR_ID,
        "email": BATAA_ACTOR_EMAIL,
        "fullName": BATAA_ACTOR_NAME,
        "name": BATAA_ACTOR_NAME,
        "avatarPath": BATAA_AVATAR_PATH,
        "role": (human_actor or {}).get("role") or "staff",
        "status": "approved",
        "_humanActor": actor_snapshot(human_actor) if human_actor else None,
    }


def find_user_by_email(email):
    needle = normalize_text(email).lower()
    for user in read_users():
        if user.get("email") == needle:
            return user
    return None


def find_user_by_id(user_id):
    for user in read_users():
        if user.get("id") == user_id:
            return user
    return None


def bootstrap_admin_user():
    if not ADMIN_EMAIL or not ADMIN_PASSWORD:
        return
    users = json.loads(USERS_FILE.read_text(encoding="utf-8"))
    if any(user.get("role") == "admin" for user in users):
        return
    users.insert(0, {
        "id": str(uuid4()),
        "createdAt": datetime.now(timezone.utc).isoformat(),
        "approvedAt": datetime.now(timezone.utc).isoformat(),
        "resetRequestedAt": "",
        "lastLoginAt": "",
        "fullName": "Admin",
        "contractLastName": "",
        "contractFirstName": "",
        "contractEmail": "",
        "contractPhone": "",
        "contractSignaturePath": "",
        "email": ADMIN_EMAIL,
        "passwordHash": hash_password(ADMIN_PASSWORD),
        "role": "admin",
        "status": "approved",
    })
    USERS_FILE.write_text(json.dumps(users, indent=2, ensure_ascii=False), encoding="utf-8")


def current_user(environ):
    token = parse_cookies(environ).get(SESSION_COOKIE)
    if not token:
        return None
    sessions = read_sessions()
    match = next((item for item in sessions if item.get("token") == token), None)
    if not match:
        return None
    user = find_user_by_id(match.get("userId"))
    if not user or user.get("status") != "approved":
        return None
    return sanitize_user(user)


def require_login(environ, start_response):
    user = current_user(environ)
    if user:
        return user
    json_response(start_response, "401 Unauthorized", {"error": "Login required"})
    return None


def require_admin(environ, start_response):
    user = current_user(environ)
    if not user:
        json_response(start_response, "401 Unauthorized", {"error": "Login required"})
        return None
    if user.get("role") != "admin":
        json_response(start_response, "403 Forbidden", {"error": "Admin access required"})
        return None
    return user


def backup_token_from_request(environ):
    params = parse_qs(environ.get("QUERY_STRING", ""))
    query_token = (params.get("token", [""])[0] or "").strip()
    header_token = (environ.get("HTTP_X_BACKUP_TOKEN", "") or "").strip()
    return query_token or header_token


def require_backup_admin(environ, start_response):
    user = current_user(environ)
    if user and user.get("role") == "admin":
        return user
    token = backup_token_from_request(environ)
    if BACKUP_TOKEN and token == BACKUP_TOKEN:
        return {"role": "admin", "email": "backup-token", "fullName": "Backup Token"}
    json_response(start_response, "401 Unauthorized", {"error": "Admin access required"})
    return None


def split_date_parts(value):
    raw = str(value or "").split("T", 1)[0]
    parts = raw.split("-")
    if len(parts) == 3:
        return {
            "year": parts[0],
            "month": parts[1],
            "day": parts[2],
        }
    return {"year": "", "month": "", "day": ""}


def format_contract_header_date(value):
    parts = split_date_parts(value)
    return f"{parts['year']} оны {parts['month']} сарын {parts['day']} өдөр"


def format_due_date_ordinal(value):
    parts = split_date_parts(value)
    return f"{parts['year']} оны {parts['month']}-р сарын {int(parts['day'] or '0')}-ий"


def format_balance_due_date(value):
    parts = split_date_parts(value)
    return f"{parts['year']} оны {parts['month']} сарын {int(parts['day'] or '0')}-ий"


def format_iso_date_display(value):
    raw = str(value or "").split("T", 1)[0]
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw):
        return raw
    return "-"


def parse_date_safe(value):
    raw = str(value or "").split("T", 1)[0]
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw).date()
    except ValueError:
        return None


def format_trip_range_phrase(start_value, end_value):
    start = normalize_text(start_value)
    end = normalize_text(end_value)
    if start and end:
        return f"{start} өдрөөс {end} хооронд"
    if start:
        return f"{start} өдрөөс"
    return end


def build_traveler_count_sentence(data):
    parts = []
    adult_count = parse_int(data.get("adultCount"))
    child_count = parse_int(data.get("childCount"))
    infant_count = parse_int(data.get("infantCount"))
    if adult_count:
        parts.append(f"{adult_count} том хүн")
    if child_count:
        parts.append(f"{child_count} хүүхэд")
    if infant_count:
        parts.append(f"{infant_count} нярай")
    if not parts:
        return "Энэхүү аялалд аялагчийн тоо тусгайлан тохиролцоно."
    return f"Энэхүү аялалд нийт {', '.join(parts)} оролцоно."


def build_traveler_representation_phrase(data):
    parts = []
    adult_count = parse_int(data.get("adultCount"))
    child_count = parse_int(data.get("childCount"))
    infant_count = parse_int(data.get("infantCount"))
    if adult_count:
        parts.append(f"{adult_count} том хүн")
    if child_count:
        parts.append(f"{child_count} хүүхэд")
    if infant_count:
        parts.append(f"{infant_count} нярай")
    total_count = parse_int(data.get("travelerCount"))
    if not parts and total_count:
        return f"/нийт {total_count} жуулчинг/"
    if not parts:
        return "/жуулчинг/"
    if adult_count and not child_count and not infant_count:
        return f"/нийт {total_count} жуулчинг/"
    return f"/{', '.join(parts)} нийт {total_count} жуулчинг/"


def build_contract_data(payload):
    today = datetime.now(timezone.utc).astimezone(MONGOLIA_TZ).date().isoformat()
    contract_date = payload.get("contractDate") or today

    tourist_last_name = normalize_text(payload.get("touristLastName"))
    tourist_first_name = normalize_text(payload.get("touristFirstName"))

    adult_count = parse_int(payload.get("adultCount"))
    child_count = parse_int(payload.get("childCount"))
    infant_count = parse_int(payload.get("infantCount"))
    ticket_only_count = parse_int(payload.get("ticketOnlyCount"))
    land_only_count = parse_int(payload.get("landOnlyCount"))
    custom_count = parse_int(payload.get("customCount"))

    traveler_count = parse_int(payload.get("travelerCount")) or adult_count + child_count + infant_count + ticket_only_count + land_only_count + custom_count
    adult_price_raw = parse_int(payload.get("adultPrice"))
    child_price_raw = parse_int(payload.get("childPrice"))
    infant_price_raw = parse_int(payload.get("infantPrice"))
    ticket_only_price_raw = parse_int(payload.get("ticketOnlyPrice"))
    land_only_price_raw = parse_int(payload.get("landOnlyPrice"))
    custom_price_raw = parse_int(payload.get("customPrice"))
    total_price_raw = parse_int(payload.get("totalPrice"))
    if not total_price_raw:
        total_price_raw = (
            adult_count * adult_price_raw
            + child_count * child_price_raw
            + infant_count * infant_price_raw
            + ticket_only_count * ticket_only_price_raw
            + land_only_count * land_only_price_raw
            + custom_count * custom_price_raw
        )
    deposit_raw = parse_int(payload.get("depositAmount"))
    balance_raw = max(total_price_raw - deposit_raw, 0)

    contract_serial = normalize_text(payload.get("contractSerial")).strip().strip("_")
    serial_matches = re.findall(r"DTX-\d{2}[A-ZА-Я]?-?\d{2}-\d+|DTX-\d{2}[A-ZА-Я]?-\d{2}-\d+", contract_serial)
    if serial_matches:
        contract_serial = serial_matches[-1]
    # Manager opted out of numbering for this one contract — store
    # an empty serial so the rendered page omits the "Дугаар: …"
    # line. The auto-numbering loop only looks at non-empty
    # serials with the matching prefix, so the next normal contract
    # still continues the sequence (no gap, no re-use).
    skip_serial = str(payload.get("skipSerial") or "").strip().lower() in ("1", "true", "yes", "on")
    if skip_serial:
        contract_serial = ""
    elif not contract_serial:
        now = datetime.now(timezone.utc).astimezone(MONGOLIA_TZ)
        year = str(now.year)[-2:]
        prefix = f"DTX-09A-{year}-"
        serials = [
            c.get("data", {}).get("contractSerial", "")
            for c in read_contracts()
            if str(c.get("data", {}).get("contractSerial", "")).startswith(prefix)
        ]
        numbers = []
        for item in serials:
            try:
                numbers.append(int(str(item).replace(prefix, "")))
            except ValueError:
                continue
        base_number = 46 if year == "26" else 0
        next_num = max(numbers or [base_number]) + 1
        contract_serial = f"{prefix}{next_num:03d}"

    manager_last = normalize_text(payload.get("managerLastName"))
    manager_first = normalize_text(payload.get("managerFirstName"))
    manager_full = dedupe_full_name(manager_last, manager_first)
    manager_email = normalize_text(payload.get("managerEmail")).lower()
    manager_phone = normalize_text(payload.get("managerPhone"))

    trip_start = normalize_text(payload.get("tripStartDate"))
    trip_end = normalize_text(payload.get("tripEndDate"))
    trip_duration = normalize_text(payload.get("tripDuration"))
    if not trip_duration and trip_start and trip_end:
        try:
            start_date = datetime.fromisoformat(trip_start).date()
            end_date = datetime.fromisoformat(trip_end).date()
            diff_days = (end_date - start_date).days + 1
            if diff_days > 0:
                trip_duration = f"{diff_days} өдөр {max(diff_days - 1, 0)} шөнө"
        except ValueError:
            trip_duration = trip_duration

    data = {
        "contractSerial": contract_serial,
        "contractDate": contract_date,
        "managerLastName": manager_last,
        "managerFirstName": manager_first,
        "managerFullName": normalize_person_name(manager_full),
        "managerEmail": manager_email,
        "managerPhone": manager_phone,
        "managerSignaturePath": normalize_text(payload.get("managerSignaturePath")),
        "touristLastName": tourist_last_name,
        "touristFirstName": tourist_first_name,
        "touristRegister": normalize_text(payload.get("touristRegister")),
        "clientPhone": normalize_text(payload.get("clientPhone")),
        "emergencyContactName": normalize_text(payload.get("emergencyContactName")),
        "emergencyContactRelation": normalize_text(payload.get("emergencyContactRelation")),
        "emergencyContactPhone": normalize_text(payload.get("emergencyContactPhone")),
        "destination": normalize_text(payload.get("destination")),
        "templateId": normalize_text(payload.get("templateId")),
        "tripStartDate": trip_start,
        "tripEndDate": trip_end,
        "tripDuration": trip_duration,
        "adultCount": adult_count,
        "childCount": child_count,
        "infantCount": infant_count,
        "ticketOnlyCount": ticket_only_count,
        "landOnlyCount": land_only_count,
        "customCount": custom_count,
        "travelerCount": traveler_count,
        "adultPrice": format_money(adult_price_raw),
        "childPrice": format_money(child_price_raw),
        "infantPrice": format_money(infant_price_raw),
        "ticketOnlyPrice": format_money(ticket_only_price_raw),
        "landOnlyPrice": format_money(land_only_price_raw),
        "customPriceLabel": normalize_text(payload.get("customPriceLabel")) or "Нэмэлт үйлчилгээ",
        "customPrice": format_money(custom_price_raw),
        "totalPrice": format_money(total_price_raw),
        "depositAmount": format_money(payload.get("depositAmount")),
        "balanceAmount": format_money(balance_raw),
        "depositDueDate": normalize_text(payload.get("depositDueDate")),
        "balanceDueDate": normalize_text(payload.get("balanceDueDate")),
    }
    data["touristSignature"] = (
        f"{data['touristLastName'][:1]}.{data['touristFirstName']}" if data["touristLastName"] else data["touristFirstName"]
    )

    parts = []
    if adult_count:
        parts.append(f"1 том хүний {data['adultPrice']} төгрөг")
    if child_count:
        parts.append(f"1 хүүхдийн {data['childPrice']} төгрөг")
    if infant_count:
        parts.append(f"1 нярай хүүхдийн {data['infantPrice']} төгрөг")
    if ticket_only_count:
        parts.append(f"1 зочин зөвхөн билеттэй {data['ticketOnlyPrice']} төгрөг")
    if land_only_count:
        parts.append(f"1 зочин газрын үйлчилгээтэй {data['landOnlyPrice']} төгрөг")
    if custom_count:
        parts.append(f"1 зочин {data['customPriceLabel']} {data['customPrice']} төгрөг")
    price_breakdown = ", ".join(part for part in parts if part)
    data["priceBreakdown"] = price_breakdown or ""
    data["paymentParagraph"] = (
        f"Энэхүү гэрээгээр аялагчийн төлбөр нь {price_breakdown}, нийт {data['travelerCount']} хүний {data['totalPrice']} төгрөг байхаар харилцан тохиролцож гэрээ байгуулав. Аялал зохион байгуулагч нь НӨАТ төлөгч биш болно."
        if price_breakdown
        else f"Энэхүү гэрээгээр аялагчийн төлбөр нь нийт {data['travelerCount']} хүний {data['totalPrice']} төгрөг байхаар харилцан тохиролцож гэрээ байгуулав. Аялал зохион байгуулагч нь НӨАТ төлөгч биш болно."
    )
    # Bank account chosen on the contract form — also reflect it in
    # the contract body so the deposit / balance clauses show the
    # actual account the client should wire to. Falls back to the
    # default Төрийн Банк wording when no bank is picked, so legacy
    # behaviour is preserved.
    bank_id = normalize_text(payload.get("bankAccountId"))
    bank_phrase = "Төрийн Банкны MN03 0034 3432 7777 9999"
    if bank_id:
        for b in read_settings().get("bankAccounts") or []:
            if b.get("id") == bank_id:
                name_parts = []
                if b.get("bankName"):
                    name_parts.append(f"{b['bankName']} банкны")
                elif b.get("label"):
                    name_parts.append(f"{b['label']} банкны")
                if b.get("accountNumber"):
                    name_parts.append(b["accountNumber"])
                if name_parts:
                    bank_phrase = " ".join(name_parts)
                break
    data["bankPhrase"] = bank_phrase
    data["depositParagraph"] = (
        f"Аяллын төлбөр дараах байдлаар хийгдэнэ.\n5.3.1. Аяллын урьдчилгаа төлбөр болох {data['depositAmount']} төгрөгийг {format_balance_due_date(data['depositDueDate'])} өдөр “Дэлхий Трэвел Икс” ХХК-ний {bank_phrase} дансанд хийснээр аялал баталгаажна."
    )
    data["balanceParagraph"] = (
        f"5.3.2. Аяллын үлдэгдэл төлбөр болох {data['balanceAmount']} төгрөгийг {format_balance_due_date(data['balanceDueDate'])} өдөр “Дэлхий Трэвел Икс” ХХК-ний {bank_phrase} дансанд хийнэ."
    )
    return data


def validate_contract_data(data):
    # contractSerial is intentionally NOT required: managers can opt
    # out via the "Skip serial" checkbox for off-cycle contracts.
    # The auto-numbering loop ignores empty serials so the next
    # numbered contract still picks up where the sequence left off.
    required = [
        "contractDate",
        "managerLastName",
        "managerFirstName",
        "touristLastName",
        "touristFirstName",
        "touristRegister",
        "destination",
        "tripStartDate",
        "tripEndDate",
        "tripDuration",
        "totalPrice",
        "depositAmount",
        "depositDueDate",
        "balanceDueDate",
    ]
    missing = [field for field in required if not str(data.get(field, "")).strip()]
    if missing:
        return f"Missing required fields: {', '.join(missing)}"
    return None


def qname(tag):
    return f"{{{WORD_NS}}}{tag}"


def paragraph_text(paragraph):
    return "".join(node.text or "" for node in paragraph.findall(f".//{qname('t')}")).strip()


def set_paragraph_text(paragraph, text):
    paragraph_props = None
    first_run_props = None

    for child in list(paragraph):
        if child.tag == qname("pPr"):
            paragraph_props = copy.deepcopy(child)
        if first_run_props is None and child.tag == qname("r"):
            run_props = child.find(qname("rPr"))
            if run_props is not None:
                first_run_props = copy.deepcopy(run_props)

    paragraph.clear()

    if paragraph_props is not None:
        paragraph.append(paragraph_props)

    if not text:
        return

    run = ET.Element(qname("r"))
    if first_run_props is not None:
        run.append(first_run_props)

    text_node = ET.Element(qname("t"))
    if text[:1].isspace() or text[-1:].isspace():
        text_node.set(f"{{{XML_NS}}}space", "preserve")
    text_node.text = text
    run.append(text_node)
    paragraph.append(run)


def ensure_child(parent, tag):
    child = parent.find(qname(tag))
    if child is None:
        child = ET.Element(qname(tag))
        parent.append(child)
    return child


def set_or_create(parent, tag, attributes):
    node = parent.find(qname(tag))
    if node is None:
        node = ET.Element(qname(tag))
        parent.append(node)
    for key, value in attributes.items():
        node.set(qname(key), value)
    return node


def update_docx_styles(styles_xml):
    root = ET.fromstring(styles_xml)

    doc_defaults = root.find(qname("docDefaults"))
    if doc_defaults is None:
        doc_defaults = ET.Element(qname("docDefaults"))
        root.insert(0, doc_defaults)

    rpr_default = ensure_child(doc_defaults, "rPrDefault")
    rpr = ensure_child(rpr_default, "rPr")
    set_or_create(rpr, "rFonts", {"ascii": "Times New Roman", "hAnsi": "Times New Roman", "cs": "Times New Roman"})
    set_or_create(rpr, "sz", {"val": "24"})
    set_or_create(rpr, "szCs", {"val": "24"})

    for style in root.findall(qname("style")):
        style_type = style.get(qname("type"))
        style_id = style.get(qname("styleId"))
        if style_type == "paragraph" and style_id in {"Normal", "BodyText", "DefaultParagraphFont"}:
            style_rpr = ensure_child(style, "rPr")
            set_or_create(style_rpr, "rFonts", {"ascii": "Times New Roman", "hAnsi": "Times New Roman", "cs": "Times New Roman"})
            set_or_create(style_rpr, "sz", {"val": "24"})
            set_or_create(style_rpr, "szCs", {"val": "24"})

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def replace_template_paragraphs(root, data):
    manager_formal_name = get_manager_contract_formal_name(data)
    manager_signature_name = get_manager_signature_name(data)
    trip_range_phrase = format_trip_range_phrase(data["tripStartDate"], data["tripEndDate"])
    traveler_count_sentence = build_traveler_count_sentence(data)
    traveler_representation_phrase = build_traveler_representation_phrase(data)
    replacements = {
        "Дугаар: DTX-09А-26-_____": f"Дугаар: {data['contractSerial']}",
        "2026 оны 01 сарын 26 өдөр                                                 Улаанбаатар хот":
            f"{format_contract_header_date(data['contractDate'])}                                                 Улаанбаатар хот",
        "Монгол Улсын Аялал Жуулчлалын тухай хуулийн 13.1 дүгээр зүйл, Иргэний хуулийн 370-379 дүгээр зүйлийг үндэслэн нэг талаас “Дэлхий Трэвел Икс” ХХК (РД:6925073) цаашид Дэлхий Трэвел Икс гэхийг төлөөлөн аяллын менежер албан тушаалтай Чулуунбаатар овогтой Нямбаяр, нөгөө талаас 2 жуулчин төлөөлөн Батмөнх овогтой Уранчимэг (РД:ШД84011762) нар харилцан тохиролцож энэхүү аялал жуулчлалын гэрээг байгуулав.":
            "Монгол Улсын Аялал Жуулчлалын тухай хуулийн 13.1 дүгээр зүйл, Иргэний хуулийн 370-379 дүгээр зүйлийг үндэслэн нэг талаас “Дэлхий Трэвел Икс” ХХК (РД:6925073) цаашид Дэлхий Трэвел Икс гэхийг төлөөлөн аяллын менежер албан тушаалтай "
            f"{manager_formal_name}, нөгөө талаас цаашид Жуулчин гэхийг {traveler_representation_phrase} төлөөлөн, {data['touristLastName']} овогтой {data['touristFirstName']} нар харилцан тохиролцож энэхүү аялал жуулчлалын гэрээг байгуулав.",
        "Энэхүү гэрээгээр Дэлхий Трэвел Икс нь 2026/02/17-2026/02/25 хооронд Египет аяллын хөтөлбөртэй үйлчилгээг үзүүлэх, аялал зохион байгуулах, Жуулчин нь гэрээний нөхцөлийн дагуу төлбөрийг төлөх, аяллын үйлчилгээ авахтай холбоотой талуудын эдлэх эрх, үүрэг, хариуцлага, төлбөр тооцоотой холбогдон үүссэн харилцааг зохицуулна.":
            f"Энэхүү гэрээгээр Дэлхий Трэвел Икс нь {trip_range_phrase} {data['destination']} чиглэлийн аяллын хөтөлбөртэй үйлчилгээг үзүүлэх, аялал зохион байгуулах, Жуулчин нь гэрээний нөхцөлийн дагуу төлбөрийг төлөх, аяллын үйлчилгээ авахтай холбоотой талуудын эдлэх эрх, үүрэг, хариуцлага, төлбөр тооцоотой холбогдон үүссэн харилцааг зохицуулна.",
        "Энэхүү гэрээгээр аялагчийн төлбөр нь том хүний 7,340,000 төгрөг буюу нийт 2 хүний 14,680,000 төгрөг байхаар харилцан тохиролцож гэрээ байгуулав. Аялал зохион байгуулагч нь НӨАТ төлөгч биш болно.":
            data["paymentParagraph"],
        "Аяллын төлбөр дараах байдлаар хийгдэнэ. 5.3.1.Аяллын урьдчилгаа төлбөр болох 4,404,000 төгрөгийг 2026 оны 01-р сарын 26 өдөр “Дэлхий Трэвел Икс”  ХХК-ний  Төрийн Банкны MN03 0034 3432":
            data["depositParagraph"],
        "7777 9999 дугаартай дансанд хийснээр аялал баталгаажна.":
            "",
        "5.3.2 Аяллын үлдэгдэл төлбөр болох 10,276,000 төгрөгийг 2026 оны 02 сарын 06 өдөр “Дэлхий Трэвел Икс”  ХХК-ний  Төрийн Банкны MN03 0034 3432":
            data["balanceParagraph"],
        "7777 9999 дугаартай дансанд хийнэ.":
            "",
        "Б. Уранчимэг":
            f"{data['touristLastName']} {data['touristFirstName']}",
        "Утас:":
            f"Утас: {data.get('clientPhone') or ''}".strip(),
        "VIBER:":
            f"Яаралтай үед холбоо барих: {data.get('emergencyContactName') or ''}".strip(),
        "FACEBOOK:":
            f"Холбоо барих утас: {data.get('emergencyContactPhone') or ''}".strip(),
        "Хаяг: _____________ хот, ___________":
            "",
        "дүүрэг, ______________________ гудамж,":
            "",
        "______ хороо, _________________хотхон":
            "",
        "______ байр ___тоот":
            "",
        "Яаралтай үед холбоо барих утасны дугаар:":
            f"Яаралтай үед холбоо барих утас: {data.get('emergencyContactPhone') or ''}".strip(),
        "Таны хэн болох:":
            f"Таны хэн болох: {data.get('emergencyContactRelation') or ''}".strip(),
        "2026 оны 03 сарын 13 өдөр                                                                                Улаанбаатар хот":
            f"{format_contract_header_date(data['contractDate'])}                                                                                Улаанбаатар хот",
        "Монгол Улсын Аялал Жуулчлалын тухай хуулийн 13.1 дүгээр зүйл, Иргэний хуулийн 370-379 дүгээр зүйлийг үндэслэн нэг талаас “Дэлхий Трэвел Икс” ХХК (РД:6925073) цаашид Дэлхий Трэвел Икс гэхийг төлөөлөн аяллын менежер албан тушаалтай Чулуунбаатар овогтой Нямбаяр, нөгөө талаас Жуулчин цаашид “Жуулчин” гэхийг төлөөлөн Цэдэн-Иш овогтой Чинзориг (РД: ШЕ77111832) нар харилцан тохиролцож энэхүү аялал жуулчлалын гэрээг байгуулав.":
            "Монгол Улсын Аялал Жуулчлалын тухай хуулийн 13.1 дүгээр зүйл, Иргэний хуулийн 370-379 дүгээр зүйлийг үндэслэн нэг талаас “Дэлхий Трэвел Икс” ХХК (РД:6925073) цаашид Дэлхий Трэвел Икс гэхийг төлөөлөн аяллын менежер албан тушаалтай "
            f"{manager_formal_name}, нөгөө талаас цаашид Жуулчин гэхийг {traveler_representation_phrase} төлөөлөн, "
            f"{data['touristLastName']} овогтой {data['touristFirstName']} нар харилцан тохиролцож энэхүү аялал жуулчлалын гэрээг байгуулав.",
        "Энэхүү гэрээгээр Дэлхий Трэвел Икс нь 2026/03/28-2026/04/03 хооронд Турк аялал, 7 өдөр 6 шөнө, хөтөлбөртэй аяллын дагуу 5 аялагчдад үйлчилгээг үзүүлэх, аялал зохион байгуулах, Жуулчин нь гэрээний нөхцөлийн дагуу төлбөрийг төлөх, аяллын үйлчилгээ авахтай холбоотой талуудын эдлэх эрх, үүрэг, хариуцлага, төлбөр тооцоотой холбогдон үүссэн харилцааг зохицуулна.":
            f"Энэхүү гэрээгээр Дэлхий Трэвел Икс нь {trip_range_phrase} {data['destination']} чиглэлийн аяллын хөтөлбөртэй, {data['tripDuration']}, үйлчилгээг үзүүлэх, аялал зохион байгуулах, Жуулчин нь гэрээний нөхцөлийн дагуу төлбөрийг төлөх, аяллын үйлчилгээ авахтай холбоотой талуудын эдлэх эрх, үүрэг, хариуцлага, төлбөр тооцоотой холбогдон үүссэн харилцааг зохицуулна.",
        "Энэхүү гэрээгээр аялагчийн төлбөр нь 3 том хүний 3,990,000 төгрөг, 1 хүүхдийн 3,390,000  төгрөг, 1 хүний онгоц ороогүй дүн 2,590,000 төгрөг,  нийт 5 аялагчийн аяллын төлбөр 17,450,000 төгрөг байхаар харилцан тохиролцож гэрээ байгуулав. Аялал зохион байгуулагч нь НӨАТ төлөгч биш болно.":
            data["paymentParagraph"],
        "Аяллын төлбөр дараах байдлаар хийгдэнэ. 5.3.1.Аяллын урьдчилгаа төлбөр болох 8,725,000 төгрөгийг 2026 оны 03-р сарын":
            f"Аяллын төлбөр дараах байдлаар хийгдэнэ.\n5.3.1. Аяллын урьдчилгаа төлбөр болох {data['depositAmount']} төгрөгийг {format_due_date_ordinal(data['depositDueDate'])} дотор “Дэлхий Трэвел Икс” ХХК-ний Төрийн Банкны MN030034343277779999 дугаартай дансанд хийснээр аялал баталгаажна.",
        "13 -ий дотор “Дэлхий Трэвел Икс” ХХК-ний Төрийн Банкны":
            "",
        "MN030034343277779999 дугаартай дансанд хийснээр аялал баталгаажна.":
            "",
        "5.3.2. Аяллын үлдэгдэлийг 2026 оны 03 сарын 20-ий дотор “Дэлхий Трэвел Икс”  ХХК-ний  Төрийн Банкны MN030034343277779999 дугаартай дансанд хийхээр тохиролцов.":
            f"5.3.2. Аяллын үлдэгдэлийг {format_balance_due_date(data['balanceDueDate'])} дотор “Дэлхий Трэвел Икс” ХХК-ний Төрийн Банкны MN030034343277779999 дугаартай дансанд хийхээр тохиролцов.",
        "Ц.Чинзориг":
            data["touristSignature"],
        "2025 оны 01 сарын  21 өдөр                     № 25 / ПУКЕТ                      Улаанбаатар хот":
            f"{format_contract_header_date(data['contractDate'])}                     № {data['contractSerial']} / {data['destination']}                      Улаанбаатар хот",
        "Нэг талаас Дэлхий Трэвел Икс ХХК (6925073 )/цаашид “Аялал зохион байгуулагч” гэх/ түүнийг төлөөлөн менежер Ч.Нямбаяр,":
            f"Нэг талаас Дэлхий Трэвел Икс ХХК (6925073 )/цаашид “Аялал зохион байгуулагч” гэх/ түүнийг төлөөлөн менежер {manager_signature_name},",
        "Нөгөө талаас 21 аялагчийг төлөөлөн, ХХХХХХХХ овогтой XXXXXXXX (РД: ДЙ91101311) /цаашид “Захиалагч” гэх/ нар дор дурдсан нөхцөлөөр харилцан тохиролцож  байгуулав.":
            f"Нөгөө талаас {data['travelerCount']} аялагчийг төлөөлөн, {data['touristLastName']} овогтой {data['touristFirstName']} (РД: {data['touristRegister']}) /цаашид “Захиалагч” гэх/ нар дор дурдсан нөхцөлөөр харилцан тохиролцож  байгуулав.",
        "Энэхүү гэрээгээр Аялал зохион байгуулагч нь захиалагчийн хүсэлтээр Тайланд улсын Пукет арлаар аялах хөтөлбөртэй аяллыг 2025/02/16 – 2025/02/23-ны хооронд 8 өдөр 7 шөнөөр тооцож энэ гэрээнд заагдсан аяллыг зохион байгуулах,":
            f"Энэхүү гэрээгээр Аялал зохион байгуулагч нь захиалагчийн хүсэлтээр {data['destination']} чиглэлд аялах хөтөлбөртэй аяллыг {data['tripStartDate']} – {data['tripEndDate']}-ны хооронд {data['tripDuration']} тооцож энэ гэрээнд заагдсан аяллыг зохион байгуулах,",
        "1 том хүний 4’590’000 төгрөг, нийт 21 хүний 96,390,000 төгрөг байхаар харилцан тохиров.":
            data["paymentParagraph"],
        "Аяллын урьдчилгаа төлбөр болох 10,980,000 төгрөгийг  2026 оны 01 сарын 24 –ны өдрийн дотор Төрийн банкны MN030034 3432 7777 9999 тоот төгрөгийн дансанд шилжүүлнэ.":
            f"Аяллын урьдчилгаа төлбөр болох {data['depositAmount']} төгрөгийг  {format_balance_due_date(data['depositDueDate'])} өдрийн дотор Төрийн банкны MN030034 3432 7777 9999 тоот төгрөгийн дансанд шилжүүлнэ.",
        "Аяллын үлдэгдэл болох 10,980,000 төгрөгийг  2026 оны 01 сарын 24 –ны өдрийн дотор Төрийн банкны MN030034 3432 7777 9999 тоот төгрөгийн дансанд шилжүүлнэ.":
            f"Аяллын үлдэгдэл болох {data['balanceAmount']} төгрөгийг  {format_balance_due_date(data['balanceDueDate'])} өдрийн дотор Төрийн банкны MN030034 3432 7777 9999 тоот төгрөгийн дансанд шилжүүлнэ.",
        "XXXXX Овогтой XXXXXXXX":
            f"{data['touristLastName']} Овогтой {data['touristFirstName']}",
    }

    for paragraph in root.findall(f".//{qname('p')}"):
        current_text = paragraph_text(paragraph)
        if "хараар оршин сууж байсан удаатай" in current_text:
            set_paragraph_text(paragraph, current_text.replace("хараар оршин сууж байсан удаатай", "хараар оршин сууж байсан"))
            continue
        if current_text.strip().startswith("Цоо шинэ паспорттой эсэх."):
            set_paragraph_text(paragraph, "")
            continue
        if current_text.strip().startswith("Гадаад паспортоо өмнө нь хаяж гээгдүүлж байсан эсэх"):
            set_paragraph_text(paragraph, "")
            continue
        if current_text in replacements:
            set_paragraph_text(paragraph, replacements[current_text])
            continue
        if current_text.startswith("Хөтөлбөрт аяллаар"):
            set_paragraph_text(paragraph, traveler_count_sentence)
            continue
        if current_text.startswith("Зорчих чиглэл нь визтэй бол"):
            set_paragraph_text(paragraph, "Зорчих чиглэл нь визтэй эсвэл визний тусгай нөхцөлтэй чиглэл бол энэхүү гэрээний Хавсралт 2 - Харилцан ойлголцлын санамж бичиг нь гэрээний салшгүй хэсэг байна.")


def get_contract_template_path():
    template_path = TEMPLATE_FILE
    if not template_path.exists() and ALT_TEMPLATE.exists():
        template_path = ALT_TEMPLATE
    if not template_path.exists() and TEMPLATE_FALLBACK.exists():
        template_path = TEMPLATE_FALLBACK
    return template_path


def load_contract_template_root(data):
    template_path = get_contract_template_path()
    if not template_path.exists():
        return None

    with zipfile.ZipFile(template_path, "r") as source_zip:
        document_xml = source_zip.read("word/document.xml")

    root = ET.fromstring(document_xml)
    replace_template_paragraphs(root, data)
    return root


def normalize_contract_heading(text):
    normalized = " ".join((text or "").strip().upper().split())
    normalized = normalized.replace("YY", "ҮҮ")
    normalized = normalized.replace(" YY", " ҮҮ")
    normalized = normalized.replace("ЖУУЛЧИН Ы", "ЖУУЛЧИНЫ")
    normalized = normalized.replace("ЖУУЛЧНЫ", "ЖУУЛЧИНЫ")
    normalized = normalized.replace(", Ү Ү Р Э Г", ", ҮҮРЭГ")
    normalized = normalized.replace(", Ү ҮРЭГ", ", ҮҮРЭГ")
    normalized = normalized.replace(", YYРЭГ", ", ҮҮРЭГ")
    normalized = normalized.replace("  ", " ")
    return normalized.strip()


def extract_contract_blocks(data):
    root = load_contract_template_root(data)
    if root is None:
        return []

    body = root.find(qname("body"))
    if body is None:
        return []

    section_heading_map = {
        "ЕРӨНХИЙ ЗҮЙЛ": "1. ЕРӨНХИЙ ЗҮЙЛ",
        "ГЭРЭЭНИЙ ХУГАЦАА": "2. ГЭРЭЭНИЙ ХУГАЦАА",
        "АЯЛАЛ ЗОХИОН БАЙГУУЛАГЧИЙН ЭРХ, ҮҮРЭГ": "3. АЯЛАЛ ЗОХИОН БАЙГУУЛАГЧИЙН ЭРХ, ҮҮРЭГ",
        "ЖУУЛЧИНЫ ЭРХ, ҮҮРЭГ": "4. ЖУУЛЧИНЫ ЭРХ, ҮҮРЭГ",
        "АЯЛЛЫН ЗАРДАЛ, ТӨЛБӨР ТООЦОО": "5. АЯЛЛЫН ЗАРДАЛ, ТӨЛБӨР ТООЦОО",
        "ТАЛУУДЫН ХАРИУЦЛАГА": "6. ТАЛУУДЫН ХАРИУЦЛАГА",
        "ХИЛИЙН ШАЛГАН НЭВТРҮҮЛЭХ ХЭСЭГ": "7. ХИЛИЙН ШАЛГАН НЭВТРҮҮЛЭХ ХЭСЭГ",
        "БУСАД ЗҮЙЛ": "8. БУСАД ЗҮЙЛ",
    }

    blocks = []
    current_section = None
    subsection_index = 0

    for element in body:
        tag = element.tag.split("}")[-1]
        if tag == "p":
            text = paragraph_text(element).strip()
            if not text:
                continue
            text = text.replace("хараар оршин сууж байсан удаатай", "хараар оршин сууж байсан")
            if text.startswith("Цоо шинэ паспорттой эсэх."):
                continue
            if text.startswith("Гадаад паспортоо өмнө нь хаяж гээгдүүлж байсан эсэх"):
                continue
            normalized = normalize_contract_heading(text)
            if normalized in {"ГЭРЭЭГ БАЙГУУЛСАН:", "ГЭРЭЭГ БАЙГУУЛСАН"}:
                break
            if normalized in section_heading_map:
                numbered_heading = section_heading_map[normalized]
                current_section = numbered_heading.split(".", 1)[0]
                subsection_index = 0
                blocks.append({"type": "heading", "text": numbered_heading})
            elif current_section is not None:
                nested_subsection_match = re.match(
                    rf"^(?P<prefix>.*?)(?:\s+)?{re.escape(current_section)}\.(?P<outer>\d+)\.(?P<inner>\d+)\.\s*(?P<rest>.*)$",
                    text,
                )
                if nested_subsection_match:
                    prefix = nested_subsection_match.group("prefix").strip()
                    outer = nested_subsection_match.group("outer")
                    inner = nested_subsection_match.group("inner")
                    rest = nested_subsection_match.group("rest").strip()
                    subsection_index = max(subsection_index, int(outer))
                    if prefix:
                        blocks.append(
                            {
                                "type": "numbered-paragraph",
                                "number": f"{current_section}.{outer}.",
                                "text": prefix,
                            }
                        )
                    blocks.append(
                        {
                            "type": "numbered-paragraph",
                            "number": f"{current_section}.{outer}.{inner}.",
                            "text": rest,
                        }
                    )
                    continue
                inline_subsections = list(
                    re.finditer(
                        rf"({re.escape(current_section)}\.(\d+)\.\s*)(.*?)(?=(?:\s+{re.escape(current_section)}\.\d+\.)|$)",
                        text,
                    )
                )
                if inline_subsections:
                    lead_text = text[:inline_subsections[0].start()].strip()
                    if lead_text:
                        subsection_index += 1
                        blocks.append(
                            {
                                "type": "numbered-paragraph",
                                "number": f"{current_section}.{subsection_index}.",
                                "text": lead_text,
                            }
                        )
                    for match in inline_subsections:
                        subsection_index = max(subsection_index, int(match.group(2)))
                        blocks.append(
                            {
                                "type": "numbered-paragraph",
                                "number": f"{current_section}.{match.group(2)}.",
                                "text": match.group(3).strip(),
                            }
                        )
                    continue
                split_lines = [line.strip() for line in re.split(r"\n+", text) if line.strip()]
                explicit_subsections = [
                    line for line in split_lines[1:]
                    if re.match(rf"^{re.escape(current_section)}\.\d+\.\s*", line)
                ]
                if split_lines and explicit_subsections:
                    subsection_index += 1
                    blocks.append(
                        {
                            "type": "numbered-paragraph",
                            "number": f"{current_section}.{subsection_index}.",
                            "text": split_lines[0],
                        }
                    )
                    for line in split_lines[1:]:
                        match = re.match(rf"^{re.escape(current_section)}\.(\d+)\.\s*(.*)$", line)
                        if match:
                            subsection_index = max(subsection_index, int(match.group(1)))
                            blocks.append(
                                {
                                    "type": "numbered-paragraph",
                                    "number": f"{current_section}.{match.group(1)}.",
                                    "text": match.group(2).strip(),
                                }
                            )
                        else:
                            blocks.append({"type": "paragraph", "text": line})
                    continue
                subsection_index += 1
                blocks.append(
                    {
                        "type": "numbered-paragraph",
                        "number": f"{current_section}.{subsection_index}.",
                        "text": text,
                    }
                )
            else:
                blocks.append({"type": "paragraph", "text": text})
        elif tag == "tbl":
            rows = []
            for row in element.findall(f".//{qname('tr')}"):
                cells = []
                for cell in row.findall(f".//{qname('tc')}"):
                    cell_text = " ".join(
                        paragraph_text(p).strip()
                        for p in cell.findall(f".//{qname('p')}")
                    ).strip()
                    cells.append(cell_text)
                if cells:
                    rows.append(cells)
            if rows:
                blocks.append({"type": "table", "rows": rows})
    return blocks


def generate_docx(data, output_path):
    ensure_data_store()
    template_path = get_contract_template_path()
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found at {template_path}")

    with zipfile.ZipFile(template_path, "r") as source_zip:
        document_xml = source_zip.read("word/document.xml")
        root = ET.fromstring(document_xml)
        replace_template_paragraphs(root, data)
        new_document_xml = ET.tostring(root, encoding="utf-8", xml_declaration=True)
        styles_xml = None
        if "word/styles.xml" in source_zip.namelist():
            styles_xml = update_docx_styles(source_zip.read("word/styles.xml"))

        with zipfile.ZipFile(output_path, "w") as target_zip:
            for item in source_zip.infolist():
                content = source_zip.read(item.filename)
                if item.filename == "word/document.xml":
                    content = new_document_xml
                elif item.filename == "word/styles.xml" and styles_xml is not None:
                    content = styles_xml
                target_zip.writestr(item, content)


def render_docx_to_html(data):
    if not get_contract_template_path().exists():
        return "<p>Template not found.</p>"
    blocks = extract_contract_blocks(data)
    if not blocks:
        return "<p>Template is empty.</p>"

    parts = []
    for block in blocks:
        if block["type"] == "heading":
            parts.append(f"<h2>{html.escape(block['text'])}</h2>")
        elif block["type"] == "numbered-paragraph":
            parts.append(
                "<p class=\"contract-numbered\">"
                f"<span class=\"contract-number\">{html.escape(block['number'])}</span>"
                f"<span class=\"contract-text\">{html.escape(block['text'])}</span>"
                "</p>"
            )
        elif block["type"] == "paragraph":
            parts.append(f"<p>{html.escape(block['text'])}</p>")
        elif block["type"] == "table":
            rows = []
            for row in block["rows"]:
                cells = "".join(f"<td>{html.escape(cell)}</td>" for cell in row)
                rows.append(f"<tr>{cells}</tr>")
            parts.append(f"<table>{''.join(rows)}</table>")
    return "\n".join(parts)


def get_contract_display_blocks(data):
    blocks = []
    for block in extract_contract_blocks(data):
        if block["type"] == "paragraph":
            text = block["text"].strip()
            if text == "АЯЛАЛ ЖУУЛЧЛАЛЫН ГЭРЭЭ":
                continue
            if text.startswith("Дугаар:"):
                continue
            if "Улаанбаатар хот" in text and any(ch.isdigit() for ch in text):
                continue
        blocks.append(block)
    return blocks


# Template tokens — used both to seed the editor (sample values
# in the DOCX get rewritten to {{tokens}} so the manager sees them
# highlighted in red) and to expand at contract-render time.
TEMPLATE_TOKEN_SUBSTITUTIONS = [
    # Date ranges + individual dates (longest first to avoid
    # partially matching the range).
    ("2026/02/17-2026/02/25", "{{tripStartDate}}-{{tripEndDate}}"),
    ("2026/03/28-2026/04/03", "{{tripStartDate}}-{{tripEndDate}}"),
    ("2026/02/17", "{{tripStartDate}}"),
    ("2026/02/25", "{{tripEndDate}}"),
    ("2026/03/28", "{{tripStartDate}}"),
    ("2026/04/03", "{{tripEndDate}}"),
    ("2026 оны 01 сарын 26 өдөр", "{{contractDate}}"),
    ("2026 оны 03 сарын 13 өдөр", "{{contractDate}}"),
    ("2026 оны 02 сарын 06 өдөр", "{{balanceDueDate}}"),
    # Money — longest first so 14,680,000 doesn't get cut by 4,680.
    ("14,680,000 төгрөг", "{{totalPrice}} төгрөг"),
    ("17,450,000 төгрөг", "{{totalPrice}} төгрөг"),
    ("10,276,000 төгрөг", "{{balanceAmount}} төгрөг"),
    ("8,725,000 төгрөг", "{{depositAmount}} төгрөг"),
    ("4,404,000 төгрөг", "{{depositAmount}} төгрөг"),
    ("7,340,000 төгрөг", "{{adultPrice}} төгрөг"),
    ("3,990,000 төгрөг", "{{adultPrice}} төгрөг"),
    ("3,390,000 төгрөг", "{{childPrice}} төгрөг"),
    ("2,590,000 төгрөг", "{{landOnlyPrice}} төгрөг"),
    # Destinations
    ("Египет", "{{destination}}"),
    ("Турк", "{{destination}}"),
    # Compound name patterns first (so we don't double-tokenise).
    ("Чулуунбаатар овогтой Нямбаяр", "{{managerLastName}} овогтой {{managerFirstName}}"),
    ("Батмөнх овогтой Уранчимэг", "{{touristLastName}} овогтой {{touristFirstName}}"),
    ("Цэдэн-Иш овогтой Чинзориг", "{{touristLastName}} овогтой {{touristFirstName}}"),
    ("Б. Уранчимэг", "{{touristLastName}} {{touristFirstName}}"),
    ("(РД: ШД84011762)", "(РД: {{touristRegister}})"),
    ("(РД: ШЕ77111832)", "(РД: {{touristRegister}})"),
    ("ШД84011762", "{{touristRegister}}"),
    ("ШЕ77111832", "{{touristRegister}}"),
    # Counts
    ("2 жуулчин", "{{travelerCount}} жуулчин"),
    ("5 аялагч", "{{travelerCount}} аялагч"),
    ("3 том хүн", "{{adultCount}} том хүн"),
    ("1 хүүхэд", "{{childCount}} хүүхэд"),
    ("2 хүн", "{{travelerCount}} хүн"),
    # Trip duration phrases
    ("8 өдөр 7 шөнө", "{{tripDuration}}"),
    ("7 өдөр 6 шөнө", "{{tripDuration}}"),
]


def _apply_template_tokens(text):
    s = str(text or "")
    for src, dst in TEMPLATE_TOKEN_SUBSTITUTIONS:
        if src and src in s:
            s = s.replace(src, dst)
    return s


def _template_token_values(data):
    """Token name → value map used by render_template_body_html when
    a contract is generated from a saved template. Covers every
    field the contract form captures plus a few computed phrases."""
    contract_date = data.get("contractDate") or ""
    return {
        "contractSerial":           data.get("contractSerial", "") or "",
        "contractDate":             format_contract_header_date(contract_date) if contract_date else "",
        "tripStartDate":            data.get("tripStartDate", "") or "",
        "tripEndDate":              data.get("tripEndDate", "") or "",
        "tripDuration":             data.get("tripDuration", "") or "",
        "destination":              data.get("destination", "") or "",
        "managerLastName":          data.get("managerLastName", "") or "",
        "managerFirstName":         data.get("managerFirstName", "") or "",
        "managerFormalName":        get_manager_contract_formal_name(data),
        "managerSignatureName":     get_manager_signature_name(data),
        "managerEmail":             data.get("managerEmail", "") or "",
        "managerPhone":             data.get("managerPhone", "") or "",
        "touristLastName":          data.get("touristLastName", "") or "",
        "touristFirstName":         data.get("touristFirstName", "") or "",
        "touristRegister":          data.get("touristRegister", "") or "",
        "travelerCount":            str(data.get("travelerCount") or ""),
        "adultCount":               str(data.get("adultCount") or ""),
        "childCount":               str(data.get("childCount") or ""),
        "infantCount":              str(data.get("infantCount") or ""),
        "ticketOnlyCount":          str(data.get("ticketOnlyCount") or ""),
        "landOnlyCount":            str(data.get("landOnlyCount") or ""),
        "totalPrice":               data.get("totalPrice", "") or "",
        "adultPrice":               data.get("adultPrice", "") or "",
        "childPrice":               data.get("childPrice", "") or "",
        "infantPrice":              data.get("infantPrice", "") or "",
        "ticketOnlyPrice":          data.get("ticketOnlyPrice", "") or "",
        "landOnlyPrice":            data.get("landOnlyPrice", "") or "",
        "depositAmount":            data.get("depositAmount", "") or "",
        "balanceAmount":            data.get("balanceAmount", "") or "",
        "depositDueDate":           data.get("depositDueDate", "") or "",
        "balanceDueDate":           data.get("balanceDueDate", "") or "",
        "paymentParagraph":         data.get("paymentParagraph", "") or "",
        "depositParagraph":         data.get("depositParagraph", "") or "",
        "balanceParagraph":         data.get("balanceParagraph", "") or "",
        "clientPhone":              data.get("clientPhone", "") or "",
        "emergencyContactName":     data.get("emergencyContactName", "") or "",
        "emergencyContactPhone":    data.get("emergencyContactPhone", "") or "",
        "emergencyContactRelation": data.get("emergencyContactRelation", "") or "",
        "tripRangePhrase":          format_trip_range_phrase(data.get("tripStartDate"), data.get("tripEndDate")),
        "bankPhrase":               data.get("bankPhrase", "") or "",
        "travelerRepresentationPhrase": build_traveler_representation_phrase(data),
    }


def _extract_contract_blocks_raw():
    """Same shape as extract_contract_blocks but skips the
    replace_template_paragraphs() pass, so the returned text is the
    DOCX's raw source sentences (with the original sample dates /
    names / prices). Used to seed the contract-template editor.
    Saved templates therefore contain the same source sentences
    that render_template_body_html knows how to replace at contract
    render time, so manager-edited templates still substitute live
    data without any extra work.
    """
    template_path = get_contract_template_path()
    if not template_path.exists():
        return []
    with zipfile.ZipFile(template_path, "r") as source_zip:
        document_xml = source_zip.read("word/document.xml")
    root = ET.fromstring(document_xml)
    # Re-walk the same heading/numbered-paragraph extraction logic as
    # extract_contract_blocks but inline so we don't double-call
    # replace_template_paragraphs on the tree.
    body = root.find(qname("body"))
    if body is None:
        return []
    section_heading_map = {
        "ЕРӨНХИЙ ЗҮЙЛ": "1. ЕРӨНХИЙ ЗҮЙЛ",
        "ГЭРЭЭНИЙ ХУГАЦАА": "2. ГЭРЭЭНИЙ ХУГАЦАА",
        "АЯЛАЛ ЗОХИОН БАЙГУУЛАГЧИЙН ЭРХ, ҮҮРЭГ": "3. АЯЛАЛ ЗОХИОН БАЙГУУЛАГЧИЙН ЭРХ, ҮҮРЭГ",
        "ЖУУЛЧИНЫ ЭРХ, ҮҮРЭГ": "4. ЖУУЛЧИНЫ ЭРХ, ҮҮРЭГ",
        "АЯЛЛЫН ЗАРДАЛ, ТӨЛБӨР ТООЦОО": "5. АЯЛЛЫН ЗАРДАЛ, ТӨЛБӨР ТООЦОО",
        "ТАЛУУДЫН ХАРИУЦЛАГА": "6. ТАЛУУДЫН ХАРИУЦЛАГА",
        "ХИЛИЙН ШАЛГАН НЭВТРҮҮЛЭХ ХЭСЭГ": "7. ХИЛИЙН ШАЛГАН НЭВТРҮҮЛЭХ ХЭСЭГ",
        "БУСАД ЗҮЙЛ": "8. БУСАД ЗҮЙЛ",
    }
    blocks = []
    current_section = None
    for element in body:
        tag = element.tag.split("}")[-1]
        if tag != "p":
            continue
        text = paragraph_text(element).strip()
        if not text:
            continue
        normalized = normalize_contract_heading(text)
        if normalized in {"ГЭРЭЭГ БАЙГУУЛСАН:", "ГЭРЭЭГ БАЙГУУЛСАН"}:
            break
        # Locked header lines (page title, contract number, date/city
        # row) — the renderer always emits these so they don't belong
        # in the editable body.
        if text == "АЯЛАЛ ЖУУЛЧЛАЛЫН ГЭРЭЭ":
            continue
        if text.startswith("Дугаар:"):
            continue
        if "Улаанбаатар хот" in text and any(ch.isdigit() for ch in text):
            continue
        if normalized in section_heading_map:
            current_section = section_heading_map[normalized].split(".", 1)[0]
            blocks.append({"type": "heading", "text": section_heading_map[normalized]})
        elif current_section is not None:
            # Strip a leading "N.M." or "N.M.K." numeric prefix if
            # the DOCX paragraph started with one — section position
            # gives us numbering on render.
            cleaned = re.sub(rf"^\s*{re.escape(current_section)}\.\d+(?:\.\d+)?\.\s*", "", text)
            # Replace the DOCX's hard-coded sample values (dates,
            # destinations, names, prices) with {{token}} markers so
            # the editor can render them in red and the renderer can
            # substitute the live contract data later.
            cleaned = _apply_template_tokens(cleaned)
            blocks.append({"type": "numbered-paragraph", "text": cleaned})
        else:
            # Free-form paragraph that lives BEFORE the first section
            # heading — the legal preamble that introduces both
            # parties. Tokenised the same way so the editor can red-
            # highlight names / counts.
            blocks.append({"type": "paragraph", "text": _apply_template_tokens(text)})
    return blocks


def get_default_template_payload():
    """Read the built-in DOCX and return both the legal preamble
    paragraphs (intro, before Section 1) and the structural
    section list. Re-parses on each call so any future edit to the
    canonical DOCX flows through.
    """
    blocks = _extract_contract_blocks_raw()
    intro = []
    sections = []
    current = None
    for block in blocks:
        if block["type"] == "paragraph":
            intro.append(block["text"])
        elif block["type"] == "heading":
            current = {"title": block["text"], "paragraphs": []}
            sections.append(current)
        elif block["type"] == "numbered-paragraph":
            if current is not None:
                current["paragraphs"].append(block["text"])
    return {"bigTitle": "АЯЛАЛ ЖУУЛЧЛАЛЫН ГЭРЭЭ", "intro": intro, "sections": sections}


_TEMPLATE_ALLOWED_TAGS = {"strong", "b", "em", "i", "u", "ul", "ol", "li", "br", "p"}
_TEMPLATE_TAG_RE = re.compile(r"</?\s*([a-zA-Z][a-zA-Z0-9]*)\b[^>]*>")


def _sanitize_template_html(s):
    """Tag-and-attribute strip to the allowlist used by the editor.
    Anything outside the list is removed (text content survives).
    """
    if not s:
        return ""
    def repl(m):
        full = m.group(0)
        tag = m.group(1).lower()
        if tag not in _TEMPLATE_ALLOWED_TAGS:
            return ""
        is_close = full.startswith("</")
        is_self = full.rstrip(">").endswith("/")
        if is_close:
            return f"</{tag}>"
        if tag == "br" or is_self:
            return f"<{tag}/>"
        return f"<{tag}>"
    return _TEMPLATE_TAG_RE.sub(repl, str(s))


def render_template_body_html(template, data):
    """Render a saved template's sections into the contract body
    HTML, applying the same sentence-substitution + token expansion
    the DOCX flow uses so dates / names / prices substitute live.
    Templates may include light HTML (lists, bold, italic, etc); we
    sanitise to a small allowlist before emitting and substitute
    {{tokens}} regardless of whether they live in plain text or
    inside a tag's text content.
    """
    sections = (template or {}).get("sections") or []

    tokens = _template_token_values(data)

    def format_text(text):
        s = str(text or "")
        for src, dst in _contract_text_replacements(data).items():
            if src and src in s:
                s = s.replace(src, dst)
        s = re.sub(r"\{\{\s*([a-zA-Z][a-zA-Z0-9_]*)\s*\}\}", lambda m: str(tokens.get(m.group(1), "") or ""), s)
        return s

    def html_text(text):
        # Templates store sanitised HTML; expand tokens inside the
        # raw string (since they live in text segments anyway), then
        # sanitise again to be safe and bold any money amounts.
        s = format_text(text)
        s = _sanitize_template_html(s)
        # Bold money amounts only inside plain text segments — apply
        # to the whole string is fine because the regex doesn't match
        # inside tag names.
        s = re.sub(
            r"(\d{1,3}(?:,\d{3})+(?:\s*төгрөг(?:ийг|ийн|өөр|өөс|төгрөг)?)?)",
            r"<strong>\1</strong>",
            s,
        )
        return s

    parts = []
    intro = (template or {}).get("intro") or []
    for para in intro:
        if not (para or "").strip():
            continue
        # Use the same opening-paragraph class the DOCX flow uses so
        # spacing / indentation matches Section 1's neighbour.
        parts.append(f"<p class=\"contract-opening-paragraph\">{html_text(para)}</p>")
    for section_index, section in enumerate(sections, start=1):
        title = format_text(section.get("title") or "").strip()
        if title:
            parts.append(f"<h2>{html.escape(title)}</h2>")
        for para_index, para in enumerate(section.get("paragraphs") or [], start=1):
            number = f"{section_index}.{para_index}."
            parts.append(
                "<p class=\"contract-numbered\">"
                f"<span class=\"contract-number\">{html.escape(number)}</span>"
                f"<span class=\"contract-text\">{html_text(para)}</span>"
                "</p>"
            )
    return "\n".join(parts) or "<p>Template is empty.</p>"


def _contract_text_replacements(data):
    """Same mapping replace_template_paragraphs uses inline, exposed
    as a function so render_template_body_html can apply it without
    re-walking the XML tree."""
    manager_formal_name = get_manager_contract_formal_name(data)
    trip_range_phrase = format_trip_range_phrase(data.get("tripStartDate"), data.get("tripEndDate"))
    traveler_representation_phrase = build_traveler_representation_phrase(data)
    return {
        "Дугаар: DTX-09А-26-_____": f"Дугаар: {data.get('contractSerial', '')}",
        "Энэхүү гэрээгээр Дэлхий Трэвел Икс нь 2026/02/17-2026/02/25 хооронд Египет аяллын хөтөлбөртэй үйлчилгээг үзүүлэх, аялал зохион байгуулах, Жуулчин нь гэрээний нөхцөлийн дагуу төлбөрийг төлөх, аяллын үйлчилгээ авахтай холбоотой талуудын эдлэх эрх, үүрэг, хариуцлага, төлбөр тооцоотой холбогдон үүссэн харилцааг зохицуулна.":
            f"Энэхүү гэрээгээр Дэлхий Трэвел Икс нь {trip_range_phrase} {data.get('destination', '')} чиглэлийн аяллын хөтөлбөртэй үйлчилгээг үзүүлэх, аялал зохион байгуулах, Жуулчин нь гэрээний нөхцөлийн дагуу төлбөрийг төлөх, аяллын үйлчилгээ авахтай холбоотой талуудын эдлэх эрх, үүрэг, хариуцлага, төлбөр тооцоотой холбогдон үүссэн харилцааг зохицуулна.",
        "Энэхүү гэрээгээр аялагчийн төлбөр нь том хүний 7,340,000 төгрөг буюу нийт 2 хүний 14,680,000 төгрөг байхаар харилцан тохиролцож гэрээ байгуулав. Аялал зохион байгуулагч нь НӨАТ төлөгч биш болно.":
            data.get("paymentParagraph", ""),
        "Аяллын төлбөр дараах байдлаар хийгдэнэ. 5.3.1.Аяллын урьдчилгаа төлбөр болох 4,404,000 төгрөгийг 2026 оны 01-р сарын 26 өдөр “Дэлхий Трэвел Икс”  ХХК-ний  Төрийн Банкны MN03 0034 3432":
            data.get("depositParagraph", ""),
        "5.3.2 Аяллын үлдэгдэл төлбөр болох 10,276,000 төгрөгийг 2026 оны 02 сарын 06 өдөр “Дэлхий Трэвел Икс”  ХХК-ний  Төрийн Банкны MN03 0034 3432":
            data.get("balanceParagraph", ""),
    }


def build_contract_body_html(data):
    template_id = (data or {}).get("templateId") or ""
    if template_id:
        for tpl in read_contract_templates():
            if tpl.get("id") == template_id:
                return render_template_body_html(tpl, data)
    blocks = get_contract_display_blocks(data)

    if not blocks:
        return "<p>Template is empty.</p>"

    def format_contract_text_html(text):
        escaped = html.escape(text)
        highlighted = re.sub(
            r"(\d{1,3}(?:,\d{3})+(?:\s*төгрөг(?:ийг|ийн|өөр|өөс|төгрөг)?)?)",
            r"<strong>\1</strong>",
            escaped,
        )
        return highlighted.replace("\n", "<br />")

    parts = []
    for block in blocks:
        if block["type"] == "heading":
            parts.append(f"<h2>{html.escape(block['text'])}</h2>")
        elif block["type"] == "numbered-paragraph":
            parts.append(
                "<p class=\"contract-numbered\">"
                f"<span class=\"contract-number\">{html.escape(block['number'])}</span>"
                f"<span class=\"contract-text\">{format_contract_text_html(block['text'])}</span>"
                "</p>"
            )
        elif block["type"] == "paragraph":
            paragraph_class = "contract-opening-paragraph" if block["text"].startswith("Монгол Улсын Аялал Жуулчлалын тухай хуулийн 13.1 дүгээр зүйл") else ""
            class_attr = f' class="{paragraph_class}"' if paragraph_class else ""
            parts.append(f"<p{class_attr}>{format_contract_text_html(block['text'])}</p>")
        elif block["type"] == "table":
            rows = []
            for row in block["rows"]:
                cells = "".join(f"<td>{format_contract_text_html(cell)}</td>" for cell in row)
                rows.append(f"<tr>{cells}</tr>")
            parts.append(f"<table>{''.join(rows)}</table>")
    return "\n".join(parts)


def build_contract_html(data, signature_path=None, asset_mode="web", contract_id=None, paper_mode=False):
    content = build_contract_body_html(data)
    # Big page title — manager-built templates can override the
    # default "АЯЛАЛ ЖУУЛЧЛАЛЫН ГЭРЭЭ" wording (e.g. memos use
    # "ХАРИЛЦАН ОЙЛГОЛЦЛЫН САНАМЖ БИЧИГ"). Falls back to default
    # when the chosen template has no bigTitle or no template
    # is selected.
    big_title_raw = "АЯЛАЛ ЖУУЛЧЛАЛЫН ГЭРЭЭ"
    template_id = (data or {}).get("templateId") or ""
    if template_id:
        for tpl in read_contract_templates():
            if tpl.get("id") == template_id:
                tpl_title = (tpl.get("bigTitle") or "").strip()
                if tpl_title:
                    big_title_raw = tpl_title
                break
    big_title = html.escape(big_title_raw)
    organizer_name = html.escape(get_manager_display_name(data))
    customer_name = html.escape(data.get("touristSignature") or "")
    manager_display_name = html.escape(get_manager_signature_name(data))
    manager_email = html.escape(get_manager_contract_email(data))
    manager_phone = html.escape(get_manager_contract_phone(data))
    contract_date = html.escape(format_contract_header_date(data["contractDate"]))
    contract_serial = html.escape(data.get("contractSerial", "") or "")
    contract_serial_line = (
        f'<p class="contract-number-line"><span class="contract-number-label">Дугаар:</span> {contract_serial}</p>'
        if contract_serial else ""
    )
    signature_markup = ""
    if signature_path:
        signature_src = signature_path
        if asset_mode == "file":
            sig_file = (GENERATED_DIR / signature_path.replace("/generated/", "", 1)).resolve()
            if sig_file.exists():
                signature_src = sig_file.as_uri()
        signature_markup = f'<img class="tourist-signature-image" src="{html.escape(signature_src)}" alt="Tourist signature" />'

    def asset_src(filename):
        if asset_mode == "file":
            return (PUBLIC_DIR / "assets" / filename).resolve().as_uri()
        return f"/assets/{filename}"

    manager_signature_src = asset_src("nyambayar-signature-cropped.png")
    manager_signature_path = normalize_text(data.get("managerSignaturePath"))
    if manager_signature_path:
        manager_signature_src = manager_signature_path
        if asset_mode == "file" and manager_signature_path.startswith("/generated/"):
            manager_file = (GENERATED_DIR / manager_signature_path.replace("/generated/", "", 1)).resolve()
            if manager_file.exists():
                manager_signature_src = manager_file.as_uri()

    # Paper mode strips the stamp + manager signature so the office
    # can print a clean copy and the client / manager hand-sign it.
    if paper_mode:
        manager_signature_block = ""
    else:
        manager_signature_block = (
            '<div class="signature-stack">'
            f'<img class="stamp-image" src="{asset_src("dtx-stamp-cropped.png")}" alt="DTX stamp" />'
            f'<img class="company-signature-image" src="{html.escape(manager_signature_src)}" alt="Manager signature" />'
            '</div>'
        )

    download_href = f"/api/contracts/{contract_id}/document?mode=download" if contract_id else ""
    _dl_title = quote(data.get("contractSerial", "") or (contract_id or ""), safe="")
    download_button = (
        f'<a href="/pdf-viewer?src={quote(download_href, safe="")}&title={_dl_title}" target="_blank" rel="noreferrer">PDF Татах</a>'
        if download_href
        else '<button onclick="window.print()">Print</button>'
    )

    return f"""<!DOCTYPE html>
<html lang="mn">
  <head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Аялал жуулчлалын гэрээ</title>
    <link rel="icon" type="image/png" href="{asset_src('favicon-dtx-x.png')}" />
    <style>
      @font-face {{
        font-family: "TravelXTimes";
        src:
          local("Times New Roman"),
          local("Times New Roman Regular"),
          url("{asset_src('fonts/times-new-roman.ttf')}") format("truetype");
        font-weight: 400;
        font-style: normal;
      }}
      @font-face {{
        font-family: "TravelXTimes";
        src:
          local("Times New Roman Bold"),
          local("Times New Roman"),
          url("{asset_src('fonts/times-new-roman-bold.ttf')}") format("truetype");
        font-weight: 700;
        font-style: normal;
      }}
      :root {{
        color-scheme: light;
        --ink: #1d1d1b;
        --paper: #ffffff;
        --accent: #1d2b4f;
        --muted: #6e645f;
        --page-width: 210mm;
        --page-height: 297mm;
        --page-top: 2cm;
        --page-right: 3cm;
        --page-bottom: 2cm;
        --page-left: 3cm;
      }}
      @page {{
        size: A4;
        margin: 2cm 3cm 2cm 3cm;
      }}
      * {{ box-sizing: border-box; }}
      body {{
        margin: 0;
        background: #f4f2ee;
        color: var(--ink);
        font-family: "TravelXTimes", "Times New Roman", "Liberation Serif", serif;
        line-height: 1.5;
      }}
      .toolbar {{
        position: sticky;
        top: 0;
        display: flex;
        gap: 12px;
        justify-content: center;
        padding: 16px;
        background: rgba(255, 255, 255, 0.96);
        border-bottom: 1px solid rgba(0, 0, 0, 0.08);
        backdrop-filter: blur(10px);
      }}
      .toolbar button, .toolbar a {{
        padding: 12px 18px;
        border-radius: 999px;
        border: none;
        background: var(--accent);
        color: white;
        text-decoration: none;
        font: 600 14px/1.2 sans-serif;
        cursor: pointer;
      }}
      .toolbar a.secondary {{
        background: #2f4858;
      }}
      .page {{
        width: var(--page-width);
        min-height: var(--page-height);
        margin: 24px auto 56px;
        padding: var(--page-top) var(--page-right) var(--page-bottom) var(--page-left);
        background: var(--paper);
        box-shadow: 0 20px 60px rgba(71, 53, 43, 0.12);
      }}
      .doc-logo {{
        display: flex;
        justify-content: center;
        margin-bottom: 14px;
      }}
      .doc-logo-image {{
        width: 220px;
        max-width: 100%;
        height: auto;
        display: block;
      }}
      h1 {{
        margin: 0 0 10px;
        text-align: center;
        font-size: 12pt;
      }}
      .contract-number-line {{
        margin: 0 0 18px;
        text-align: center;
        font-size: 12pt;
      }}
      .contract-number-label {{
        text-decoration: none;
      }}
      .contract-date-row {{
        display: flex;
        justify-content: space-between;
        gap: 24px;
        margin-bottom: 24px;
        font-size: 12pt;
      }}
      h2 {{
        margin: 22px 0 12px;
        font-size: 12pt;
        text-align: center;
        text-transform: uppercase;
      }}
      p {{
        margin: 0 0 12px;
        font-size: 12pt;
        line-height: 1.5;
        text-align: justify;
      }}
      .contract-opening-paragraph {{
        text-indent: 1.25cm;
        margin-top: 8px;
      }}
      .contract-numbered {{
        display: flex;
        align-items: flex-start;
        gap: 8px;
      }}
      .contract-number {{
        flex: 0 0 auto;
        font-weight: 700;
      }}
      .contract-text {{
        flex: 1 1 auto;
      }}
      table {{
        width: 100%;
        border-collapse: collapse;
        margin: 16px 0 24px;
      }}
      td, th {{
        border: 1px solid #1d1d1b;
        padding: 8px 10px;
        font-size: 12pt;
        line-height: 1.5;
        text-align: center;
      }}
      .signature-section {{
        margin-top: 42px;
        padding-top: 10px;
        page-break-inside: auto;
      }}
      .signature-heading {{
        margin: 0 0 22px;
        text-align: center;
        font-size: 12pt;
        font-weight: 700;
      }}
      .signature-grid {{
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 40px;
        align-items: start;
      }}
      .signature-column {{
        display: flex;
        flex-direction: column;
      }}
      .signature-prelude {{
        min-height: 88px;
      }}
      .signature-title {{
        margin-bottom: 12px;
        font-size: 12pt;
        font-weight: 700;
      }}
      .signature-org-name {{
        margin-bottom: 12px;
        font-size: 12pt;
        font-weight: 700;
      }}
      .signature-sign-label {{
        margin: 0 0 8px;
        font-size: 12pt;
        font-weight: 400;
      }}
      .signature-sign-area {{
        position: relative;
        width: 100%;
        max-width: 360px;
        height: 144px;
        margin: 0 0 10px;
        display: flex;
        align-items: flex-end;
        border-bottom: 1px solid rgba(0, 0, 0, 0.28);
        overflow: visible;
      }}
      .signature-stack {{
        position: relative;
        width: 100%;
        height: 100%;
      }}
      .signature-stack img {{
        position: absolute;
        object-fit: contain;
      }}
      .tourist-signature-image {{
        max-width: 300px;
        max-height: 120px;
        object-fit: contain;
      }}
      .stamp-image {{
        left: -8px;
        top: -52px;
        width: 300px;
        height: 300px;
        z-index: 3;
      }}
      .company-signature-image {{
        left: 22px;
        top: 12px;
        width: 260px;
        height: 107px;
        z-index: 2;
      }}
      .signature-contact {{
        margin-top: 10px;
        position: relative;
        z-index: 1;
      }}
      .signer-contact {{
        margin-top: 10px;
        position: relative;
        z-index: 1;
      }}
      .signature-contact p,
      .signer-contact p {{
        margin-bottom: 8px;
        line-height: 1.5;
      }}
      .signature-label {{
        display: inline-block;
        min-width: 68px;
        font-weight: 400;
      }}
      .signature-name {{
        font-size: 12pt;
        font-weight: 400;
        margin: 0 0 6px;
        min-height: 1.75em;
        display: flex;
        align-items: flex-start;
      }}
      .signature-subtitle {{
        margin: 0 0 4px;
        font-size: 12pt;
        font-weight: 400;
        min-height: 1.75em;
        display: flex;
        align-items: flex-start;
      }}
      .signer-signature-space {{
        width: 100%;
        height: 100%;
        display: flex;
        align-items: flex-end;
        justify-content: center;
        padding-bottom: 20px;
      }}
      .signature-role {{
        margin-top: 6px;
        font-size: 12pt;
        font-weight: 400;
      }}
      @media print {{
        body {{ background: white; }}
        .toolbar {{ display: none; }}
        .page {{
          width: auto;
          min-height: auto;
          margin: 0;
          padding: 0;
          box-shadow: none;
        }}
      }}
      @media (max-width: 720px) {{
        :root {{
          --page-top: 16px;
          --page-right: 12px;
          --page-bottom: 20px;
          --page-left: 12px;
        }}
        body {{
          background: #ffffff;
        }}
        .toolbar {{
          position: static;
          padding: 12px;
          border-bottom: 0;
          background: #ffffff;
          justify-content: center;
        }}
        .page {{
          width: 100%;
          min-height: auto;
          margin: 0;
          padding: var(--page-top) var(--page-right) var(--page-bottom) var(--page-left);
          box-shadow: none;
        }}
        .doc-logo-image {{
          width: min(220px, 72vw);
        }}
        .contract-date-row {{
          gap: 12px;
          flex-wrap: wrap;
        }}
        .signature-grid {{
          grid-template-columns: 1fr;
          gap: 28px;
        }}
        .signature-prelude {{
          min-height: auto;
        }}
        .signature-sign-area {{
          max-width: 100%;
        }}
      }}
    </style>
  </head>
  <body>
    <div class="toolbar">
      {download_button}
    </div>
    <main class="page">
      <div class="doc-logo">
        <img src="{asset_src('logo.png')}" alt="Дэлхий Трэвел Икс" class="doc-logo-image" />
      </div>
      <h1>{big_title}</h1>
      {contract_serial_line}
      <div class="contract-date-row">
        <span>{contract_date}</span>
        <span>Улаанбаатар хот</span>
      </div>
      {content}
      <section class="signature-section">
        <h2 class="signature-heading">ГЭРЭЭГ БАЙГУУЛСАН:</h2>
        <div class="signature-grid">
          <div class="signature-column">
            <div class="signature-prelude">
              <div class="signature-title">Аялал зохион байгуулагчийг төлөөлж:</div>
              <div class="signature-org-name">“Дэлхий Трэвел Икс” ХХК -ийн</div>
            </div>
            <p class="signature-sign-label">Гарын үсэг:</p>
            <div class="signature-sign-area">
              {manager_signature_block}
            </div>
            <div class="signature-contact">
              <p class="signature-name">{manager_display_name}</p>
              <p class="signature-subtitle">Аяллын менежер</p>
              <p><span class="signature-label">Гар утас:</span> {manager_phone}</p>
              <p><span class="signature-label">Утас:</span> 72007722</p>
              <p><span class="signature-label">И-мэйл:</span> {manager_email}</p>
              <p><span class="signature-label"></span> info@travelx.mn</p>
              <p><span class="signature-label">Вэбсайт:</span> www.travelx.mn</p>
              <p><span class="signature-label">Хаяг:</span> Улаанбаатар хот, Хан-Уул дүүрэг, Их Монгол Улс гудамж, 17 хороо, Кинг Тауэр 121-102 тоот</p>
            </div>
          </div>
          <div class="signature-column">
            <div class="signature-prelude">
              <div class="signature-title">Жуулчныг төлөөлж:</div>
            </div>
            <p class="signature-sign-label">Гарын үсэг:</p>
            <div class="signature-sign-area">
              <div class="signer-signature-space">{signature_markup}</div>
            </div>
            <div class="signer-contact">
              <p class="signature-name">{customer_name}</p>
              <p class="signature-subtitle">Аялагч</p>
              <p><span class="signature-label">Утас:</span> {html.escape(data.get("clientPhone") or "")}</p>
              <p>Яаралтай үед холбогдох дугаар: {html.escape(data.get("emergencyContactPhone") or "")}</p>
              <p>Таны хэн болох: {html.escape(data.get("emergencyContactRelation") or "")}</p>
            </div>
          </div>
        </div>
      </section>
    </main>
  </body>
</html>
"""


def register_contract_font():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_name = "Times-Roman"
    bold_font_name = "Times-Bold"

    for candidate in [
        "/usr/share/fonts/truetype/noto/NotoSerif-Regular.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSerif.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSerif-Regular.ttf",
        "/Library/Fonts/Times New Roman.ttf",
        "/System/Library/Fonts/Supplemental/Times New Roman.ttf",
        "/Library/Fonts/Arial Unicode.ttf",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
    ]:
        if Path(candidate).exists():
            pdfmetrics.registerFont(TTFont("TravelXContract", candidate))
            font_name = "TravelXContract"
            bold_font_name = "TravelXContract"
            break

    return font_name, bold_font_name


def draw_wrapped_text(pdf, text, x, y, max_width, font_name, font_size, leading=16):
    from reportlab.lib.utils import simpleSplit

    lines = simpleSplit(text, font_name, font_size, max_width)
    pdf.setFont(font_name, font_size)
    for line in lines:
        pdf.drawString(x, y, line)
        y -= leading
    return y


def draw_wrapped_text_with_indent(pdf, text, x, y, max_width, font_name, font_size, leading=16, first_line_indent=0):
    from reportlab.lib.utils import simpleSplit

    if first_line_indent <= 0:
        return draw_wrapped_text(pdf, text, x, y, max_width, font_name, font_size, leading)

    first_width = max_width - first_line_indent
    lines = simpleSplit(text, font_name, font_size, first_width)
    pdf.setFont(font_name, font_size)
    for index, line in enumerate(lines):
        line_x = x + first_line_indent if index == 0 else x
        pdf.drawString(line_x, y, line)
        y -= leading
    return y


def build_invoice_line_items(data):
    destination = normalize_text(data.get("destination")) or "Аяллын үйлчилгээ"
    custom_label = normalize_text(data.get("customPriceLabel")) or "Нэмэлт үйлчилгээ"
    item_specs = [
        ("adultCount", "adultPrice", "Том хүн"),
        ("childCount", "childPrice", "Хүүхэд"),
        ("infantCount", "infantPrice", "Нярай"),
        ("ticketOnlyCount", "ticketOnlyPrice", "Зөвхөн билет"),
        ("landOnlyCount", "landOnlyPrice", "Газрын үйлчилгээ"),
        ("customCount", "customPrice", custom_label),
    ]
    populated = []
    for count_key, price_key, label in item_specs:
        count = parse_int(data.get(count_key))
        unit_price = parse_int(data.get(price_key))
        if count <= 0 or unit_price <= 0:
            continue
        populated.append((count, unit_price, label))

    if not populated:
        total_price = parse_int(data.get("totalPrice"))
        return [{"description": destination, "quantity": 1, "unitPrice": total_price, "totalPrice": total_price}]

    rows = []
    for count, unit_price, label in populated:
        rows.append(
            {
                "description": label or destination,
                "quantity": count,
                "unitPrice": unit_price,
                "totalPrice": count * unit_price,
            }
        )
    return rows

def strip_invoice_destination_prefix(text, destination):
    value = normalize_text(text)
    base = normalize_text(destination)
    if not value or not base:
        return value

    lowered = value.lower()
    base_lowered = base.lower()
    if lowered == base_lowered:
        return value

    removable_tail_keywords = (
        "урьдчилгаа төлбөр",
        "үлдэгдэл төлбөр",
        "том хүн",
        "хүүхэд",
        "нярай",
        "зөвхөн билет",
        "газрын үйлчилгээ",
    )
    for separator in (" / ", "/", " - ", "-", " "):
        prefix = f"{base_lowered}{separator}"
        if not lowered.startswith(prefix):
            continue
        stripped = value[len(base) + len(separator) :].strip(" /-")
        if stripped and any(keyword in stripped.lower() for keyword in removable_tail_keywords):
            return stripped
    return value


def normalize_invoice_line_items(record):
    destination = normalize_text((record.get("data") or {}).get("destination"))
    invoice_meta = record.get("invoiceMeta") if isinstance(record.get("invoiceMeta"), dict) else {}
    raw_items = invoice_meta.get("lineItems")
    normalized = []
    if isinstance(raw_items, list):
        for item in raw_items:
            if not isinstance(item, dict):
                continue
            description = strip_invoice_destination_prefix(item.get("description"), destination)
            quantity = parse_int(item.get("quantity"))
            unit_price = parse_int(item.get("unitPrice"))
            total_price = parse_int(item.get("totalPrice")) or quantity * unit_price
            if not description or quantity <= 0:
                continue
            normalized.append(
                {
                    "key": normalize_text(item.get("key")) or f"item-{len(normalized) + 1}",
                    "description": description,
                    "quantity": quantity,
                    "unitPrice": unit_price,
                    "totalPrice": total_price,
                }
            )
    if normalized:
        return normalized
    return [
        {
            "key": f"item-{index}",
            "description": item["description"],
            "quantity": item["quantity"],
            "unitPrice": item["unitPrice"],
            "totalPrice": item["totalPrice"],
        }
        for index, item in enumerate(build_invoice_line_items(record.get("data") or {}), start=1)
    ]


INVOICE_STATUS_META = {
    "paid": {"label": "Төлөгдсөн", "className": "paid"},
    "waiting": {"label": "Хүлээгдэж буй", "className": "waiting"},
    "overdue": {"label": "Хугацаа хэтэрсэн", "className": "overdue"},
}

INVOICE_BANK_ACCOUNTS = {
    "state": {
        "bankName": "Төрийн Банк",
        "prefix": "MN030034",
        "accountNumber": "3432 7777 9999",
    },
    "golomt": {
        "bankName": "Голомт Банк",
        "prefix": "MN80001500",
        "accountNumber": "3675114666",
    },
}


def normalize_invoice_status(value, fallback="waiting"):
    normalized = normalize_text(value).strip().lower()
    if normalized in INVOICE_STATUS_META:
        return normalized
    return fallback


def _invoice_banks_from_settings():
    """Map every Settings → bankAccount into the same {bankName, prefix,
    accountNumber} shape the invoice template has always rendered. Splits
    accountNumber on its first space so existing data ("MN030034 3432 7777
    9999") becomes prefix="MN030034" + accountNumber="3432 7777 9999",
    matching the legacy hardcoded pair pixel-for-pixel."""
    out = {}
    for entry in (read_settings().get("bankAccounts") or []):
        key = (entry.get("id") or "").strip().lower()
        if not key:
            continue
        bank_name = entry.get("bankName") or entry.get("label") or ""
        full = (entry.get("accountNumber") or "").strip()
        prefix, _, rest = full.partition(" ")
        if not rest:
            # No space → keep prefix empty so the slash/slash formatting
            # still works without printing the number twice.
            prefix, rest = "", full
        out[key] = {
            "bankName": bank_name,
            "prefix": prefix,
            "accountNumber": rest,
        }
    if not out:
        # Defensive: fall back to the legacy hardcoded pair.
        return dict(INVOICE_BANK_ACCOUNTS)
    return out


def normalize_invoice_bank_account(value, fallback="state"):
    normalized = normalize_text(value).strip().lower()
    banks = _invoice_banks_from_settings()
    if normalized in banks:
        return normalized
    if fallback in banks:
        return fallback
    return next(iter(banks.keys())) if banks else "state"


def _stub_for_legacy_passthrough(value, fallback="state"):
    normalized = normalize_text(value).strip().lower()
    if normalized in INVOICE_BANK_ACCOUNTS:
        return normalized
    return fallback


def format_invoice_input_date(value):
    normalized = normalize_text(value).replace(".", "-")
    parsed = parse_date_safe(normalized)
    if parsed:
        return parsed.isoformat()
    match = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", normalized)
    if match:
        return normalized
    return ""


def build_invoice_payment_rows(record):
    data = record.get("data") or {}
    created_at = record.get("createdAt")
    signed_at = record.get("signedAt")
    today = now_mongolia().date()
    issue_date = format_iso_date_display(data.get("contractDate") or created_at)
    signed_date = format_iso_date_display(signed_at or created_at or data.get("contractDate"))
    balance_due = parse_date_safe(data.get("balanceDueDate"))
    balance_amount = parse_int(data.get("balanceAmount"))
    invoice_meta = record.get("invoiceMeta") if isinstance(record.get("invoiceMeta"), dict) else {}
    payment_meta = invoice_meta.get("payments") if isinstance(invoice_meta.get("payments"), dict) else {}
    payment_meta = payment_meta if isinstance(payment_meta, dict) else {}

    custom_rows = []
    if isinstance(payment_meta.get("rows"), list):
        for row in payment_meta.get("rows"):
            if not isinstance(row, dict):
                continue
            amount = parse_int(row.get("amount"))
            title = normalize_text(row.get("title"))
            if not title or amount <= 0:
                continue
            status_key = normalize_invoice_status(row.get("status"), fallback="waiting")
            custom_rows.append(
                {
                    "key": normalize_text(row.get("key")) or f"payment-{len(custom_rows) + 1}",
                    "title": title,
                    "created": format_iso_date_display(row.get("created")),
                    "secondaryLabel": normalize_text(row.get("secondaryLabel")) or "Эцсийн хугацаа",
                    "secondaryValue": format_iso_date_display(row.get("secondaryValue")),
                    "statusKey": status_key,
                    "status": INVOICE_STATUS_META[status_key]["label"],
                    "statusClass": INVOICE_STATUS_META[status_key]["className"],
                    "amount": amount,
                }
            )
    if custom_rows:
        return custom_rows

    rows = []
    deposit_amount = parse_int(data.get("depositAmount"))
    if deposit_amount > 0:
        deposit_status = normalize_invoice_status(
            (payment_meta.get("deposit") or {}).get("status"),
            fallback="paid",
        )
        rows.append(
            {
                "key": "deposit",
                "title": "Урьдчилгаа",
                "created": issue_date,
                "secondaryLabel": "Төлсөн огноо" if deposit_status == "paid" else "Эцсийн хугацаа",
                "secondaryValue": signed_date if deposit_status == "paid" else format_iso_date_display(data.get("depositDueDate")),
                "statusKey": deposit_status,
                "status": INVOICE_STATUS_META[deposit_status]["label"],
                "statusClass": INVOICE_STATUS_META[deposit_status]["className"],
                "amount": deposit_amount,
            }
        )

    if balance_amount > 0:
        overdue = bool(balance_due and balance_due < today)
        default_balance_status = "overdue" if overdue else "waiting"
        balance_status = normalize_invoice_status(
            (payment_meta.get("balance") or {}).get("status"),
            fallback=default_balance_status,
        )
        rows.append(
            {
                "key": "balance",
                "title": "Аяллын үлдэгдэл",
                "created": issue_date,
                "secondaryLabel": "Эцсийн хугацаа",
                "secondaryValue": format_iso_date_display(data.get("balanceDueDate")),
                "statusKey": balance_status,
                "status": INVOICE_STATUS_META[balance_status]["label"],
                "statusClass": INVOICE_STATUS_META[balance_status]["className"],
                "amount": balance_amount,
            }
        )
    return rows


def build_invoice_html(record, asset_mode="web"):
    data = record.get("data") or {}
    invoice_meta = record.get("invoiceMeta") if isinstance(record.get("invoiceMeta"), dict) else {}
    # Banks now come from Settings (any number, any name). Fall back to the
    # legacy hardcoded pair if Settings is unreadable.
    invoice_banks = _invoice_banks_from_settings()
    bank_account_key = normalize_invoice_bank_account(invoice_meta.get("bankAccountKey"))
    if bank_account_key not in invoice_banks:
        bank_account_key = next(iter(invoice_banks.keys()))
    bank_account = invoice_banks[bank_account_key]
    invoice_number = html.escape(f"{normalize_text(data.get('contractSerial')) or record.get('id', '')}-1")
    tourist_name = normalize_person_name(
        f"{normalize_text(data.get('touristLastName'))} {normalize_text(data.get('touristFirstName'))}"
    ).strip()
    customer_name = html.escape(tourist_name.upper() or "CLIENT")
    items = normalize_invoice_line_items(record)
    payment_rows = build_invoice_payment_rows(record)
    total_amount = sum(item["totalPrice"] for item in items)
    bank_options_markup = "".join(
        f'<option value="{account_key}"{" selected" if account_key == bank_account_key else ""}>{html.escape(account["bankName"])} / {html.escape(account["prefix"])} / {html.escape(account["accountNumber"])}</option>'
        for account_key, account in invoice_banks.items()
    )

    def status_options_for(current_key):
        current = normalize_invoice_status(current_key)
        return "".join(
            f'<option value="{status_key}"{" selected" if status_key == current else ""}>{html.escape(status_meta["label"])}</option>'
            for status_key, status_meta in INVOICE_STATUS_META.items()
        )

    def asset_src(filename):
        if asset_mode == "file":
            return (PUBLIC_DIR / "assets" / filename).resolve().as_uri()
        return f"/assets/{filename}"

    download_href = "" if asset_mode == "file" else f"/api/contracts/{record.get('id')}/invoice?mode=download"
    _inv_raw_number = f"{normalize_text(data.get('contractSerial')) or record.get('id', '')}-1"
    _inv_viewer_href = f"/pdf-viewer?src={quote(download_href, safe='')}&title={quote(_inv_raw_number, safe='')}" if download_href else ""
    toolbar_markup = ""
    script_markup = ""
    notice_markup = ""
    if asset_mode != "file":
        # The auto-generated contract invoice should be read-only
        # from this preview — managers edit invoice content via the
        # Trip-finance "+ Add invoice" flow, not by tweaking the
        # contract's automatic one. So we skip the View / Edit /
        # Save toolbar buttons here and leave only "PDF Татах".
        toolbar_markup = f"""
    <div class="toolbar">
      <a href="{html.escape(_inv_viewer_href)}" target="_blank" rel="noreferrer">PDF Татах</a>
    </div>"""
        notice_markup = '<div class="save-notice" data-save-notice hidden>Saved successfully</div>'
        script_markup = f"""
    <script>
      (() => {{
        const statusMeta = {json.dumps(INVOICE_STATUS_META, ensure_ascii=False)};
        const bankAccountMeta = {json.dumps(invoice_banks, ensure_ascii=False)};
        const modeButtons = Array.from(document.querySelectorAll("[data-invoice-mode]"));
        const saveButton = document.querySelector("[data-save-invoice]");
        const itemRowsBody = document.querySelector("[data-invoice-items-body]");
        const paymentStack = document.querySelector("[data-payment-stack]");
        const bankSelect = document.querySelector("[data-bank-account-select]");
        const bankName = document.querySelector("[data-bank-name]");
        const bankPrefix = document.querySelector("[data-bank-prefix]");
        const bankNumber = document.querySelector("[data-bank-number]");
        const notice = document.querySelector("[data-save-notice]");
        const setMode = (mode) => {{
          document.body.classList.toggle("is-editing", mode === "edit");
          modeButtons.forEach((button) => button.classList.toggle("is-active", button.dataset.invoiceMode === mode));
          if (saveButton) saveButton.hidden = mode !== "edit";
        }};
        const formatMoney = (value) => new Intl.NumberFormat("en-US").format(Number(value || 0)) + " ₮";
        const syncItems = () => {{
          let total = 0;
          Array.from(itemRowsBody?.querySelectorAll("tr[data-item-key]") || []).forEach((row, index) => {{
            const descriptionInput = row.querySelector('[data-item-field="description"]');
            const quantityInput = row.querySelector('[data-item-field="quantity"]');
            const unitInput = row.querySelector('[data-item-field="unitPrice"]');
            const totalInput = row.querySelector('[data-item-field="totalPrice"]');
            const quantity = Number(quantityInput?.value || 0);
            const unitPrice = Number(unitInput?.value || 0);
            const lineTotal = quantity * unitPrice;
            if (totalInput) totalInput.value = String(lineTotal);
            total += lineTotal;
            row.querySelector("[data-item-index]").textContent = index + 1;
            row.querySelector('[data-view-field="description"]').textContent = descriptionInput?.value || "";
            row.querySelector('[data-view-field="quantity"]').textContent = quantity || 0;
            row.querySelector('[data-view-field="unitPrice"]').textContent = formatMoney(unitPrice);
            row.querySelector('[data-view-field="totalPrice"]').textContent = formatMoney(lineTotal);
          }});
          const totalCell = document.querySelector("[data-invoice-total]");
          if (totalCell) totalCell.textContent = formatMoney(total);
        }};
        const syncPayments = () => {{
          Array.from(paymentStack?.querySelectorAll("[data-payment-key]") || []).forEach((card) => {{
            const titleInput = card.querySelector('[data-payment-field="title"]');
            const createdInput = card.querySelector('[data-payment-field="created"]');
            const secondaryLabelInput = card.querySelector('[data-payment-field="secondaryLabel"]');
            const secondaryValueInput = card.querySelector('[data-payment-field="secondaryValue"]');
            const amountInput = card.querySelector('[data-payment-field="amount"]');
            const select = card.querySelector("[data-status-select]");
            const badge = card.querySelector("[data-status-badge]");
            const meta = statusMeta[select?.value || "waiting"] || statusMeta.waiting;
            card.querySelector('[data-view-field="payment-title"]').textContent = titleInput?.value || "";
            card.querySelector('[data-view-field="created"]').textContent = createdInput?.value || "-";
            card.querySelector('[data-view-field="secondary-label"]').textContent = secondaryLabelInput?.value || "Эцсийн хугацаа";
            card.querySelector('[data-view-field="secondary-value"]').textContent = secondaryValueInput?.value || "-";
            card.querySelector('[data-view-field="amount"]').textContent = formatMoney(amountInput?.value || 0);
            if (badge) {{
              badge.textContent = meta.label;
              badge.className = "payment-status payment-status-view " + meta.className;
            }}
          }});
        }};
        const syncBank = () => {{
          if (!bankSelect) return;
          const meta = bankAccountMeta[bankSelect.value] || bankAccountMeta.state;
          if (bankName) bankName.textContent = meta.bankName;
          if (bankPrefix) bankPrefix.textContent = meta.prefix;
          if (bankNumber) bankNumber.textContent = meta.accountNumber;
        }};
        itemRowsBody?.addEventListener("input", syncItems);
        paymentStack?.addEventListener("input", syncPayments);
        paymentStack?.addEventListener("change", syncPayments);
        bankSelect?.addEventListener("change", syncBank);
        modeButtons.forEach((button) => button.addEventListener("click", () => setMode(button.dataset.invoiceMode || "view")));
        saveButton?.addEventListener("click", async () => {{
          saveButton.disabled = true;
          saveButton.textContent = "Saving...";
          try {{
            const items = Array.from(itemRowsBody?.querySelectorAll("tr[data-item-key]") || []).map((row) => ({{
              key: row.dataset.itemKey || "",
              description: row.querySelector('[data-item-field="description"]')?.value || "",
              quantity: Number(row.querySelector('[data-item-field="quantity"]')?.value || 0),
              unitPrice: Number(row.querySelector('[data-item-field="unitPrice"]')?.value || 0),
              totalPrice: Number(row.querySelector('[data-item-field="totalPrice"]')?.value || 0),
            }}));
            const payments = Array.from(paymentStack?.querySelectorAll("[data-payment-key]") || []).map((card) => ({{
              key: card.dataset.paymentKey || "",
              title: card.querySelector('[data-payment-field="title"]')?.value || "",
              created: card.querySelector('[data-payment-field="created"]')?.value || "",
              secondaryLabel: card.querySelector('[data-payment-field="secondaryLabel"]')?.value || "",
              secondaryValue: card.querySelector('[data-payment-field="secondaryValue"]')?.value || "",
              status: card.querySelector('[data-status-select]')?.value || "waiting",
              amount: Number(card.querySelector('[data-payment-field="amount"]')?.value || 0),
            }}));
            const response = await fetch("/api/contracts/{record.get('id')}/invoice", {{
              method: "POST",
              headers: {{ "Content-Type": "application/json" }},
              credentials: "same-origin",
              body: JSON.stringify({{
                items,
                payments,
                bankAccountKey: bankSelect?.value || "state",
              }}),
            }});
            const payload = await response.json().catch(() => ({{}}));
            if (!response.ok) throw new Error(payload.error || "Could not save invoice.");
            syncItems();
            syncPayments();
            syncBank();
            setMode("view");
            if (notice) {{
              notice.hidden = false;
              setTimeout(() => notice.hidden = true, 2200);
            }}
          }} catch (error) {{
            window.alert(error.message || "Could not save invoice.");
          }} finally {{
            saveButton.disabled = false;
            saveButton.textContent = "Save";
          }}
        }});
        syncItems();
        syncPayments();
        syncBank();
        setMode("view");
      }})();
    </script>"""

    items_markup = "".join(
        f"""
          <tr data-item-key="{html.escape(item['key'])}">
            <td data-item-index>{index}</td>
            <td>
              <span class="invoice-view-text" data-view-field="description">{html.escape(item['description'])}</span>
              <input class="invoice-edit-input" data-item-field="description" value="{html.escape(item['description'])}" />
            </td>
            <td>
              <span class="invoice-view-text" data-view-field="quantity">{item['quantity']}</span>
              <input class="invoice-edit-input" data-item-field="quantity" type="number" min="1" value="{item['quantity']}" />
            </td>
            <td>
              <span class="invoice-view-text" data-view-field="unitPrice">{format_money(item['unitPrice'])} ₮</span>
              <input class="invoice-edit-input" data-item-field="unitPrice" type="number" min="0" value="{item['unitPrice']}" />
            </td>
            <td>
              <span class="invoice-view-text" data-view-field="totalPrice">{format_money(item['totalPrice'])} ₮</span>
              <input class="invoice-edit-input" data-item-field="totalPrice" type="number" min="0" value="{item['totalPrice']}" readonly />
            </td>
          </tr>
        """
        for index, item in enumerate(items, start=1)
    )

    payment_markup = "".join(
        f"""
          <div class="payment-card" data-payment-key="{html.escape(row['key'])}">
            <div class="payment-main">
              <span class="invoice-view-text" data-view-field="payment-title">{html.escape(row['title'])}</span>
              <input class="invoice-edit-input" data-payment-field="title" value="{html.escape(row['title'])}" />
            </div>
            <div class="payment-meta">
              <span class="meta-label">Нэхэмжилсэн огноо</span>
              <span class="meta-value invoice-view-text" data-view-field="created">{html.escape(row['created'])}</span>
              <input class="invoice-edit-input" data-payment-field="created" type="date" value="{html.escape(format_invoice_input_date(row['created']))}" />
            </div>
            <div class="payment-meta">
              <span class="meta-label invoice-view-text" data-view-field="secondary-label">{html.escape(row['secondaryLabel'])}</span>
              <input class="invoice-edit-input" data-payment-field="secondaryLabel" value="{html.escape(row['secondaryLabel'])}" />
              <span class="meta-value invoice-view-text" data-view-field="secondary-value">{html.escape(row['secondaryValue'])}</span>
              <input class="invoice-edit-input" data-payment-field="secondaryValue" type="date" value="{html.escape(format_invoice_input_date(row['secondaryValue']))}" />
            </div>
            <div class="payment-meta">
              <span class="meta-label">Төлөв</span>
              <span class="payment-status payment-status-view {row['statusClass']}" data-status-badge>{html.escape(row['status'])}</span>
              <select class="payment-status-select" data-status-select>{status_options_for(row['statusKey'])}</select>
            </div>
            <div class="payment-amount">
              <span class="invoice-view-text" data-view-field="amount">{format_money(row['amount'])} ₮</span>
              <input class="invoice-edit-input" data-payment-field="amount" type="number" min="0" value="{row['amount']}" />
            </div>
          </div>
        """
        for row in payment_rows
    )
    invoice_font_stack = '"Nunito", Arial, sans-serif'
    invoice_font_faces = f"""
      @font-face {{
        font-family: "Nunito";
        src: url("{asset_src('fonts/nunito-variable.ttf')}") format("truetype");
        font-weight: 200 1000;
        font-style: normal;
      }}"""

    return f"""<!DOCTYPE html>
<html lang="mn">
  <head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Нэхэмжлэх</title>
    <link rel="icon" type="image/png" href="{asset_src('favicon-dtx-x.png')}" />
    <style>
{invoice_font_faces}
      @page {{
        size: 768px 1180px;
        margin: 0;
      }}
      * {{ box-sizing: border-box; }}
      body {{
        margin: 0;
        background: #ffffff;
        color: #27272a;
        font-family: {invoice_font_stack};
        font-size: 13px;
      }}
      .toolbar {{
        position: sticky;
        top: 0;
        z-index: 10;
        display: flex;
        justify-content: center;
        gap: 14px;
        padding: 18px 16px;
        background: rgba(255, 255, 255, 0.96);
        border-bottom: 1px solid #eef1f7;
        backdrop-filter: blur(10px);
      }}
      .toolbar a,
      .toolbar-button {{
        min-width: 86px;
        min-height: 46px;
        padding: 11px 18px;
        border: none;
        border-radius: 999px;
        background: #253776;
        color: #fff;
        text-decoration: none;
        font: 800 16px/1.2 {invoice_font_stack};
        cursor: pointer;
      }}
      .toolbar-button {{
        background: #e9eef8;
        color: #2a3c78;
      }}
      .toolbar-button.is-active {{
        background: #253776;
        color: #fff;
      }}
      .toolbar-save {{
        background: #157347;
        color: #fff;
      }}
      .save-notice {{
        position: fixed;
        right: 18px;
        top: 18px;
        padding: 10px 14px;
        border-radius: 10px;
        background: #1f8550;
        color: #fff;
        font: 600 13px/1.2 {invoice_font_stack};
        z-index: 20;
      }}
      .page {{
        width: min(720px, calc(100vw - 48px));
        margin: 20px auto 40px;
        padding: 20px;
        background: #fff;
        border: 1px solid #dce3ef;
        border-radius: 14px;
        box-shadow: none;
      }}
      .invoice-number {{
        margin: 0 0 18px;
        color: #27272a;
        font-size: 18px;
        line-height: 1.15;
        font-weight: 500;
      }}
      .header-grid {{
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 20px;
        align-items: start;
        margin-bottom: 28px;
      }}
      .invoice-logo {{
        width: 154px;
        max-width: 100%;
        display: block;
        margin-bottom: 10px;
      }}
      .company-name {{
        margin: 0 0 12px;
        font-size: 14px;
        line-height: 1.25;
        font-weight: 700;
        color: #27272a;
      }}
      .company-block p,
      .customer-block p,
      .meta-note {{
        margin: 0;
        font-size: 13px;
        line-height: 1.38;
      }}
      .meta-note {{
        text-align: right;
        color: #27272a;
      }}
      .customer-block {{
        padding-top: 80px;
      }}
      .customer-block .label {{
        display: block;
        margin-bottom: 4px;
        color: #64748b;
        font-size: 13px;
        font-weight: 600;
      }}
      .section-title {{
        margin: 0 0 10px;
        color: #64748b;
        font-size: 13px;
        font-weight: 600;
      }}
      .invoice-items-table {{
        width: 100%;
        border-collapse: separate;
        border-spacing: 0;
        border-radius: 12px;
        border: 1px solid #cfd8e6;
        overflow: hidden;
        margin-bottom: 28px;
      }}
      th, td {{
        padding: 10px 12px;
        border-bottom: 1px solid #cfd8e6;
        text-align: left;
        font-size: 13px;
        line-height: 1.25;
      }}
      th {{
        background: #fbfcfe;
        color: #27272a;
        font-weight: 700;
      }}
      th:first-child,
      td:first-child {{
        width: 46px;
      }}
      td:last-child,
      th:last-child,
      td:nth-last-child(2),
      th:nth-last-child(2) {{
        text-align: right;
      }}
      .total-row td {{
        font-weight: 700;
        background: #fff;
        border-bottom: 0;
      }}
      .payment-stack {{
        display: grid;
        gap: 16px;
        margin-bottom: 28px;
      }}
      .payment-card {{
        display: grid;
        grid-template-columns: 1.35fr 1.08fr 1.08fr 0.9fr 0.92fr;
        gap: 10px;
        align-items: center;
        min-height: 74px;
        padding: 16px 18px;
        border: 1px solid #cfd8e6;
        border-radius: 12px;
        background: #fff;
      }}
      .payment-main,
      .payment-amount {{
        min-width: 0;
        font-size: 13px;
        font-weight: 600;
        color: #27272a;
      }}
      .payment-amount {{
        text-align: right;
        white-space: nowrap;
      }}
      .payment-meta {{
        display: grid;
        gap: 4px;
        min-width: 0;
      }}
      .meta-value {{
        font-size: 13px;
        font-weight: 600;
        color: #27272a;
      }}
      .meta-label {{
        color: #64748b;
        font-size: 13px;
        font-weight: 600;
        white-space: nowrap;
      }}
      .payment-status {{
        display: inline-flex;
        align-items: center;
        justify-content: center;
        width: fit-content;
        min-width: 82px;
        min-height: 30px;
        padding: 5px 11px;
        border-radius: 999px;
        font-size: 12px;
        font-weight: 700;
      }}
      .payment-status.paid {{
        background: #dcf4e3;
        color: #1f8550;
      }}
      .payment-status.overdue {{
        background: #f8dede;
        color: #c44747;
      }}
      .payment-status.waiting {{
        background: #fff5bd;
        color: #8a4b12;
      }}
      .invoice-edit-input,
      .payment-status-select,
      .bank-account-select {{
        display: none;
        width: 100%;
        min-width: 0;
        min-height: 38px;
        padding: 8px 10px;
        border: 1px solid #cbd6ee;
        border-radius: 10px;
        background: #fff;
        color: #1f2937;
        font: 600 13px/1.2 {invoice_font_stack};
      }}
      .invoice-edit-input[type="date"] {{
        min-width: 0;
      }}
      body.is-editing .invoice-edit-input,
      body.is-editing .payment-status-select,
      body.is-editing .bank-account-select {{
        display: block;
      }}
      body.is-editing .invoice-view-text,
      body.is-editing .payment-status-view,
      body.is-editing .bank-view {{
        display: none;
      }}
      body.is-editing .page {{
        width: min(1120px, calc(100vw - 48px));
      }}
      body.is-editing .payment-card {{
        grid-template-columns: minmax(210px, 1.25fr) minmax(150px, 0.9fr) minmax(150px, 0.9fr) minmax(150px, 0.9fr) minmax(170px, 1fr);
        align-items: end;
        gap: 12px;
      }}
      body.is-editing .payment-main {{
        grid-column: auto;
      }}
      body.is-editing .payment-meta,
      body.is-editing .payment-amount {{
        display: grid;
        gap: 6px;
        min-width: 0;
      }}
      body.is-editing .payment-amount {{
        text-align: right;
      }}
      .bank-section {{
        margin-top: 0;
        padding-bottom: 28px;
        border-bottom: 1px solid #d9e0ea;
      }}
      .bank-grid {{
        display: grid;
        gap: 4px;
        font-size: 13px;
        line-height: 1.35;
        color: #27272a;
      }}
      .bank-line {{
        display: flex;
        flex-wrap: wrap;
        gap: 18px;
        align-items: baseline;
      }}
      .bank-prefix {{
        color: #64748b;
      }}
      .bank-account-number {{
        font-weight: 800;
      }}
      .signature-grid {{
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 28px;
        margin-top: 24px;
        align-items: start;
      }}
      .signature-card {{
        position: relative;
        min-height: 218px;
        padding-top: 0;
      }}
      .signature-label {{
        position: relative;
        z-index: 3;
        min-height: 18px;
        margin-bottom: 112px;
        color: #64748b;
        font-size: 13px;
        font-weight: 600;
        background: #fff;
      }}
      .signature-line {{
        border-bottom: 1px dashed #d5ddec;
      }}
      .accountant-stamp {{
        position: absolute;
        left: 2px;
        bottom: 36px;
        width: 218px;
        z-index: 1;
        opacity: 0.98;
      }}
      .accountant-signature {{
        position: absolute;
        left: 64px;
        bottom: -34px;
        width: 290px;
        z-index: 2;
      }}
      .signature-name {{
        position: relative;
        z-index: 4;
        margin-top: 14px;
        font-size: 13px;
        font-weight: 700;
        color: #27272a;
      }}
      .signature-role {{
        position: relative;
        z-index: 4;
        color: #27272a;
        font-size: 13px;
      }}
      @media print {{
        .toolbar, .save-notice {{
          display: none;
        }}
      }}
      @media (max-width: 980px) {{
        body {{
          font-size: 16px;
        }}
        .toolbar {{
          gap: 10px;
          padding: 14px 10px;
          flex-wrap: wrap;
        }}
        .toolbar a,
        .toolbar-button {{
          min-width: 82px;
          min-height: 46px;
          padding: 10px 16px;
          font-size: 16px;
        }}
        .page {{
          width: calc(100vw - 24px);
          margin: 12px auto 24px;
          padding: 22px;
        }}
        .header-grid,
        .signature-grid {{
          grid-template-columns: 1fr;
        }}
        .invoice-number {{
          font-size: 24px;
        }}
        .invoice-logo {{
          width: 220px;
        }}
        .customer-block {{
          padding-top: 0;
        }}
        .company-block p,
        .customer-block p,
        .meta-note,
        .section-title,
        th,
        td,
        .payment-main,
        .payment-amount,
        .meta-value,
        .meta-label,
        .bank-grid,
        .signature-label,
        .signature-name,
        .signature-role {{
          font-size: 16px;
        }}
        .meta-note {{
          text-align: left;
        }}
        .invoice-items-table {{
          display: block;
          overflow-x: auto;
        }}
        .payment-card {{
          grid-template-columns: 1fr;
          min-height: auto;
          padding: 18px;
        }}
        body.is-editing .payment-card {{
          grid-template-columns: 1fr;
        }}
        .payment-amount {{
          text-align: left;
        }}
        body.is-editing .payment-amount {{
          text-align: left;
        }}
      }}
    </style>
  </head>
  <body>
    {toolbar_markup}
    {notice_markup}
    <div class="page">
      <p class="invoice-number">Нэхэмжлэх #{invoice_number}</p>
      <div class="header-grid">
        <div class="company-block">
          <img class="invoice-logo" src="{asset_src('dtx-logo-blue-yellow.png')}" alt="Дэлхий Трэвел" />
          <p class="company-name">Дэлхий Трэвел Икс ХХК (6925073)</p>
          <p>Улаанбаатар хот, ХУД, 17-р хороо</p>
          <p>Их Монгол Улс гудамж, Кинг Тауэр, 121 байр, 102 тоот</p>
          <p>info@travelx.mn</p>
          <p>+976 72007722</p>
        </div>
        <div>
          <p class="meta-note">Сангийн сайдын 2017 оны 12 дугаар сарын 05</p>
          <p class="meta-note">өдрийн 347 тоот тушаалын хавсралт</p>
          <div class="customer-block">
            <span class="label">Төлөгч</span>
            <p><strong>{customer_name}</strong></p>
          </div>
        </div>
      </div>
      <p class="section-title">Үнийн мэдээлэл</p>
      <table class="invoice-items-table">
        <thead>
          <tr>
            <th>№</th>
            <th>Утга</th>
            <th>Аялагч</th>
            <th>Нэгжийн үнэ</th>
            <th>Нийт үнэ</th>
          </tr>
        </thead>
        <tbody data-invoice-items-body>
          {items_markup}
          <tr class="total-row">
            <td colspan="4">Нийт үнэ</td>
            <td data-invoice-total>{format_money(total_amount)} ₮</td>
          </tr>
        </tbody>
      </table>
      <p class="section-title">Төлбөрийн хуваарь</p>
      <div class="payment-stack" data-payment-stack>
        {payment_markup}
      </div>
      <div class="bank-section">
        <p class="section-title">Дансны мэдээлэл</p>
        <select class="bank-account-select" data-bank-account-select>{bank_options_markup}</select>
        <div class="bank-grid bank-view">
          <div>Дэлхий Трэвел Икс</div>
          <div class="bank-line">
            <span data-bank-name>{html.escape(bank_account['bankName'])}</span>
            <span class="bank-prefix" data-bank-prefix>{html.escape(bank_account['prefix'])}</span>
            <strong class="bank-account-number" data-bank-number>{html.escape(bank_account['accountNumber'])}</strong>
          </div>
        </div>
      </div>
      <div class="signature-grid">
        <div class="signature-card">
          <div class="signature-label">Дэлхий Трэвел Икс ХХК</div>
          <div class="signature-line"></div>
          <img class="accountant-stamp" src="{asset_src('invoice-finance-stamp.png')}" alt="Finance stamp" />
          <img class="accountant-signature" src="{asset_src('invoice-finance-signature.png')}" alt="Accountant signature" />
          <div class="signature-name">Нягтлан</div>
          <div class="signature-role">Г.Басгалаан</div>
        </div>
        <div class="signature-card">
          <div class="signature-label">Төлөгч</div>
          <div class="signature-line"></div>
          <div class="signature-name">{customer_name}</div>
        </div>
      </div>
    </div>
    {script_markup}
  </body>
</html>"""


def save_invoice_pdf(record):
    ensure_data_store()
    pdf_filename = f"invoice-{record['id']}.pdf"
    pdf_path = GENERATED_DIR / pdf_filename
    try:
        from weasyprint import HTML
    except Exception as exc:
        raise RuntimeError(f"WeasyPrint not available: {exc}") from exc

    html_string = build_invoice_html(record, asset_mode="file")
    try:
        HTML(string=html_string, base_url=str(BASE_DIR), media_type="screen").write_pdf(str(pdf_path))
    except Exception as exc:
        raise RuntimeError(f"HTML PDF generation failed: {exc}") from exc
    return f"/generated/{pdf_filename}"


# ── Standalone invoice (new model) — Mongolian invoice template ──
INVOICE_STATUS_LABELS = {
    "pending": ("Хүлээгдэж буй", "waiting"),
    "waiting": ("Хүлээгдэж буй", "waiting"),
    "paid":    ("Төлөгдсөн", "paid"),
    "overdue": ("Хугацаа хэтэрсэн", "overdue"),
}


def _fmt_money(value):
    try:
        n = float(value or 0)
    except Exception:
        n = 0
    return f"{int(round(n)):,} ₮".replace(",", ",")


# Currency symbols + formatting for the USM invoice template (matches the
# user's reference PDFs: "2,723 $", "12,000 €" — symbol AFTER the number).
_CURRENCY_SYMBOLS = {
    "USD": "$",
    "EUR": "€",
    "GBP": "£",
    "JPY": "¥",
    "CNY": "¥",
    "KRW": "₩",
    "RUB": "₽",
    "AUD": "A$",
    "MNT": "₮",
}


def _fmt_money_ccy(value, currency):
    try:
        n = float(value or 0)
    except Exception:
        n = 0
    sym = _CURRENCY_SYMBOLS.get((currency or "MNT").upper(), (currency or "").upper() or "")
    return f"{int(round(n)):,} {sym}".rstrip()


# ── USM (Unlock Steppe Mongolia) invoice template ──
# Per the legal-separation rule (DTX vs USM are two licensed companies),
# USM invoices use a distinct layout that mirrors the printable PDFs the
# user supplied. Currency-aware: switches the bank account number,
# IBAN, and intermediary banks based on invoice.currency.

USM_COMPANY = {
    "name": "Unlock Steppe Mongolia LLC (8415315)",
    "address_lines": [
        "Khan-Uul district, 17 Khoroo, Ikh Mongol Uls Street,",
        "King Tower, Apt 121, Door 102, Ulaanbaatar, 17012,",
        "Mongolia",
    ],
    "email": "info@steppe-mongolia.com",
    "phone": "+976 77044040",
    "logo": "usm-logo-horizontal.png",
    "stamp": "usm-stamp.png",
    "signature": "invoice-finance-signature.png",
    "accountant": "G.Bayasgalan",
}

USM_BENEFICIARY_NAME = "UNLOCK STEPPE MONGOLIA"
USM_BENEFICIARY_ADDRESS = (
    "Khan-Uul district, 17 Khoroo, Ikh Mongol Uls Street, "
    "King Tower, Apt 121, Door 102, Ulaanbaatar, 17012, Mongolia"
)
USM_BENEFICIARY_BANK = {
    "swift": "TDBMMNUB",
    "name": "TRADE AND DEVELOPMENT BANK JSC,",
    "address": "14210 Peace avenue 19, Sukhbaatar district, 1st khoroo, Ulaanbaatar, Mongolia, Tel: 1800-1977",
}

USM_BANK_BY_CURRENCY = {
    "USD": {
        "account": "436044172",
        "iban": "MN53 0004000 436044172",
        "intermediaries": [
            {"swift": "IRVTUS3N", "name": "THE BANK OF NEW YORK MELLON",
             "address": "240 GREENWICH STREET NEWYORK NY 10286, US"},
            {"swift": "OCBCSGSG", "chips": "010275",
             "name": "OVERSEA-CHINESE BANKING CORPORATION LIMITED",
             "address": "OCBC CENTRE FLOOR 10 63 CHULIA STREET SINGAPORE"},
            {"swift": "KASITHBK", "chips": "008942",
             "name": "KASIKORNBANK PUBLIC COMPANY LIMITED",
             "address": "27/1 RATBURANA RD. 1 SOI RATBURANA BANGKOK, THAILAND"},
        ],
    },
    "EUR": {
        "account": "436044173",
        "iban": "MN26 0004000 436044173",
        "intermediaries": [
            {"swift": "COBADEFF", "name": "COMMERZBANK AG",
             "address": "FRANKFURT AM MAIN, GERMANY"},
            {"swift": "BYLADEMM", "name": "BAYERISCHE LANDESBANK, MUENCHEN",
             "address": "MUENCHEN DE"},
            {"swift": "GIBAATWG", "name": "ERSTE GROUP BANK AG",
             "address": "VIENNA AUSTRIA"},
            {"swift": "SOLADEST", "name": "LANDESBANK BADEN-WUERTTEMBERG",
             "address": "STUTTGART, GERMANY"},
        ],
    },
}


def _company_from_invoice(invoice):
    """Infer the issuing company from the invoice serial prefix (T- = DTX,
    S- = USM). Falls back to the trip's company if the serial is not
    prefixed (shouldn't happen for new invoices)."""
    serial = str(invoice.get("serial") or "").upper()
    if serial.startswith("T-"):
        return "DTX"
    if serial.startswith("S-"):
        return "USM"
    trip_id = invoice.get("tripId")
    if trip_id:
        trip = next((t for t in read_camp_trips() if t.get("id") == trip_id), None)
        if trip:
            return normalize_company(trip.get("company") or "") or "DTX"
    return "DTX"


_USM_STATUS_LABELS = {
    "pending":  ("Pending",  "waiting"),
    "waiting":  ("Pending",  "waiting"),
    "paid":     ("Paid",     "paid"),
    "confirmed":("Paid",     "paid"),
    "overdue":  ("Overdue",  "overdue"),
    "cancelled":("Cancelled","cancelled"),
}

_USM_STATUS_LABELS_FR = {
    "pending":  ("En attente", "waiting"),
    "waiting":  ("En attente", "waiting"),
    "paid":     ("Payé",       "paid"),
    "confirmed":("Payé",       "paid"),
    "overdue":  ("En retard",  "overdue"),
    "cancelled":("Annulé",     "cancelled"),
}

# Full label set for the USM template, indexed by language. EUR invoices
# go out to French-speaking customers (the user's reference PDF was the
# Origins Voyages/French version), so EUR → fr, everything else → en.
_USM_LABELS = {
    "en": {
        "title":              "Invoice",
        "bill_to":            "Bill to",
        "billing_address":    "Billing address",
        "price_detail":       "Price detail",
        "col_n":              "№",
        "col_description":    "Description",
        "col_amount":         "Amount",
        "col_unit_price":     "Unit price",
        "col_total":          "Total",
        "row_total":          "Total",
        "installments":       "Installments",
        "issue_date":         "Issue date",
        "due_date":           "Due date",
        "status":             "Status",
        "ben_name_addr":      "BENEFICIARY'S ACCOUNT NAME, ADDRESS:",
        "ben_account_no":     "BENEFICIARY'S ACCOUNT NUMBER:",
        "ben_bank":           "BENEFICIARY BANK:",
        "ben_intermediary":   "INTERMEDIARY BANK:",
        "name":               "Name",
        "address":            "Address",
        "account_for":        "Account number / {ccy}",
        "iban_number":        "IBAN number",
        "suggestion":         "SUGGESTION",
        "suggestion_intro":   'Please pay attention to the following suggestions when filling out "Beneficiary\'s information":',
        "suggestion_li_name": "Please write full beneficiary's account name in Roman alphabet.",
        "suggestion_li_acct": "Please write 9 digit TDB account number.",
        "suggestion_outro":   "Please keep in mind to give a correct, completed and clear payment purpose as in case of unclear payment purpose the payment will be stopped or rejected by our bank.",
        "company_label":      "Unlock Steppe Mongolia LLC",
        "accountant":         "Accountant",
    },
    "fr": {
        "title":              "Facture",
        "bill_to":            "Facturé à",
        "billing_address":    "Adresse de facturation",
        "price_detail":       "Détail du tarif",
        "col_n":              "№",
        "col_description":    "Description",
        "col_amount":         "Montant",
        "col_unit_price":     "Prix unitaire",
        "col_total":          "Total",
        "row_total":          "Total",
        "installments":       "Versements",
        "issue_date":         "En date du",
        "due_date":           "À régler avant le",
        "status":             "Statut",
        "ben_name_addr":      "BÉNÉFICIAIRE — NOM ET ADRESSE :",
        "ben_account_no":     "NUMÉRO DE COMPTE BÉNÉFICIAIRE :",
        "ben_bank":           "BANQUE BÉNÉFICIAIRE :",
        "ben_intermediary":   "BANQUE INTERMÉDIAIRE :",
        "name":               "Nom",
        "address":            "Adresse",
        "account_for":        "Numéro de compte / {ccy}",
        "iban_number":        "Numéro IBAN",
        "suggestion":         "REMARQUE",
        "suggestion_intro":   "Veuillez prêter attention aux suggestions suivantes lors de la saisie des « informations du bénéficiaire » :",
        "suggestion_li_name": "Veuillez indiquer le nom complet du bénéficiaire en alphabet latin.",
        "suggestion_li_acct": "Veuillez indiquer le numéro de compte TDB à 9 chiffres.",
        "suggestion_outro":   "Veuillez fournir un motif de paiement correct, complet et clair, faute de quoi notre banque pourra suspendre ou refuser le paiement.",
        "company_label":      "Unlock Steppe Mongolia LLC",
        "accountant":         "Comptable",
    },
    "mn": {
        "title":              "Нэхэмжлэх",
        "bill_to":            "Төлөгч",
        "billing_address":    "Хаяг",
        "price_detail":       "Үнийн мэдээлэл",
        "col_n":              "№",
        "col_description":    "Утга",
        "col_amount":         "Тоо ширхэг",
        "col_unit_price":     "Нэгжийн үнэ",
        "col_total":          "Нийт",
        "row_total":          "Нийт",
        "installments":       "Төлбөрийн хуваарь",
        "issue_date":         "Нэхэмжилсэн огноо",
        "due_date":           "Эцсийн хугацаа",
        "status":             "Төлөв",
        "ben_name_addr":      "ХҮЛЭЭН АВАГЧИЙН НЭР, ХАЯГ:",
        "ben_account_no":     "ХҮЛЭЭН АВАГЧИЙН ДАНСНЫ ДУГААР:",
        "ben_bank":           "ХҮЛЭЭН АВАГЧ БАНК:",
        "ben_intermediary":   "ЗУУЧЛАГЧ БАНК:",
        "name":               "Нэр",
        "address":            "Хаяг",
        "account_for":        "Дансны дугаар / {ccy}",
        "iban_number":        "IBAN дугаар",
        "suggestion":         "АНХААРУУЛГА",
        "suggestion_intro":   "«Хүлээн авагчийн мэдээлэл» бөглөхдөө дараах зүйлд анхаарна уу:",
        "suggestion_li_name": "Хүлээн авагчийн нэрийг латин үсгээр бүтнээр бичнэ үү.",
        "suggestion_li_acct": "9 оронтой TDB-ийн дансны дугаарыг бичнэ үү.",
        "suggestion_outro":   "Төлбөрийн зорилгыг тодорхой, бүрэн бичээгүй тохиолдолд банк гүйлгээг түр зогсоох эсвэл буцаах магадлалтай.",
        "company_label":      "Unlock Steppe Mongolia LLC",
        "accountant":         "Нягтлан",
    },
}


def _usm_lang_for_currency(currency):
    return "fr" if (currency or "").upper() == "EUR" else "en"


# Trip language → invoice language. DTX is always Mongolian regardless,
# so this only matters for USM (S-) invoices.
def _usm_lang_for_trip(trip_language, currency):
    name = (trip_language or "").strip().lower()
    if name == "french" or name == "français":
        return "fr"
    if name == "english":
        return "en"
    if name == "mongolian" or name == "монгол":
        return "mn"
    # Unknown / "Other" — fall back to currency-based heuristic so the
    # invoice still renders sensibly.
    return _usm_lang_for_currency(currency)


def build_standalone_invoice_html_usm(invoice):
    """USM-branded invoice. Language follows currency: EUR → French (the
    Origins Voyages reference); everything else → English. Bank account
    section also swaps by currency."""
    serial = html.escape(str(invoice.get("serial") or invoice.get("id") or ""))
    customer = html.escape(str((invoice.get("payerName") or "CLIENT")).strip() or "CLIENT")
    billing_address = html.escape(str(invoice.get("payerAddress") or "").strip())
    currency = (invoice.get("currency") or "USD").upper()
    # Use the parent trip's language when available; otherwise fall back to
    # the currency-based heuristic.
    trip_id = invoice.get("tripId")
    trip_lang_field = ""
    if trip_id:
        trip_obj = next((t for t in read_camp_trips() if t.get("id") == trip_id), None)
        if trip_obj:
            trip_lang_field = trip_obj.get("language") or ""
    lang = _usm_lang_for_trip(trip_lang_field, currency)
    L = _USM_LABELS[lang]
    if lang == "fr":
        status_map = _USM_STATUS_LABELS_FR
    elif lang == "mn":
        status_map = INVOICE_STATUS_LABELS  # the Mongolian dict already in scope
    else:
        status_map = _USM_STATUS_LABELS
    items = invoice.get("items") or []
    grand = sum((float(it.get("qty") or 0) * float(it.get("price") or 0)) for it in items)

    def _fmt_qty(q):
        try:
            f = float(q or 0)
        except (TypeError, ValueError):
            return html.escape(str(q or 0))
        return str(int(f)) if f.is_integer() else str(f)

    items_rows = "".join(
        f"<tr><td>{i+1}</td>"
        f"<td>{html.escape(str(it.get('description') or ''))}</td>"
        f"<td>{_fmt_qty(it.get('qty'))}</td>"
        f"<td>{_fmt_money_ccy(it.get('price'), currency)}</td>"
        f"<td>{_fmt_money_ccy(float(it.get('qty') or 0) * float(it.get('price') or 0), currency)}</td></tr>"
        for i, it in enumerate(items)
    )
    payments_html = ""
    for inst in (invoice.get("installments") or []):
        status_key = (inst.get("status") or "pending").lower()
        label, klass = status_map.get(status_key, status_map["pending"])
        payments_html += f"""
        <div class="payment-card">
          <div class="payment-main">{html.escape(str(inst.get('description') or ''))}</div>
          <div class="payment-meta"><span class="meta-label">{html.escape(L['issue_date'])}</span><span class="meta-value">{html.escape(str(inst.get('issueDate') or '-'))}</span></div>
          <div class="payment-meta"><span class="meta-label">{html.escape(L['due_date'])}</span><span class="meta-value">{html.escape(str(inst.get('dueDate') or '-'))}</span></div>
          <div class="payment-meta"><span class="meta-label">{html.escape(L['status'])}</span><span class="payment-status {klass}">{label}</span></div>
          <div class="payment-amount">{_fmt_money_ccy(inst.get('amount'), currency)}</div>
        </div>
        """

    def _asset(name):
        p = (BASE_DIR / "public" / "assets" / name)
        return p.resolve().as_uri() if p.exists() else ""
    logo_src = _asset(USM_COMPANY["logo"])
    stamp_src = _asset(USM_COMPANY["stamp"])
    sig_src = _asset(USM_COMPANY["signature"])

    bank_meta = USM_BANK_BY_CURRENCY.get(currency)
    if bank_meta:
        account_label = L["account_for"].format(ccy=currency)
        account_html = f"""<p>{html.escape(account_label)}: <strong>{html.escape(bank_meta['account'])}</strong></p>
            <p>{html.escape(L['iban_number'])}: <strong>{html.escape(bank_meta['iban'])}</strong></p>"""
        intermediaries_html = "".join(
            f"""<div class="bank-block">
                <p><strong>*SWIFT/BIC: {html.escape(b['swift'])}</strong>{(" CHIPS UID: " + html.escape(b['chips'])) if b.get('chips') else ''}</p>
                <p>{html.escape(b['name'])}</p>
                <p>{html.escape(b['address'])}</p>
            </div>"""
            for b in bank_meta["intermediaries"]
        )
    else:
        snap = invoice.get("bankAccount") or {}
        account_html = (
            f"""<p>{html.escape(L['name'])}: <strong>{html.escape(snap.get('accountName') or USM_BENEFICIARY_NAME)}</strong></p>
            <p>{html.escape(L['col_n'] if False else 'Account number')}: <strong>{html.escape(snap.get('accountNumber') or '-')}</strong></p>"""
            if snap else "<p><em>—</em></p>"
        )
        intermediaries_html = "<p><em>—</em></p>"

    company_address = "".join(f"<p>{html.escape(line)}</p>" for line in USM_COMPANY["address_lines"])
    css = """
      @page { size: A4; margin: 16mm 14mm; }
      * { box-sizing: border-box; }
      body { margin: 0; background: #fff; color: #27272a;
        font-family: 'Nunito', Arial, sans-serif; font-size: 12px; }
      .page { padding: 0; }
      .invoice-number { margin: 0 0 14px; font-size: 18px; font-weight: 500; color: #27272a; }
      .header-grid { display: grid; grid-template-columns: 1.05fr 0.95fr;
        gap: 28px; align-items: start; margin-bottom: 24px; }
      .invoice-logo { width: 175px; max-width: 100%; display: block; margin-bottom: 12px; }
      .company-name { margin: 0 0 10px; font-size: 13px; font-weight: 700; }
      .company-block p, .customer-block p { margin: 0; font-size: 12px; line-height: 1.45; }
      /* Pad-top so "Bill to" lines up with the company name in the left
         column rather than the top of the logo. */
      .customer-block { padding-top: 96px; }
      .customer-block .label { display: block; margin-bottom: 4px; color: #64748b;
        font-size: 12px; font-weight: 600; }
      .customer-block strong { font-size: 13px; }
      .section-title { margin: 18px 0 8px; color: #5d6b87; font-size: 12px; font-weight: 600; }
      .invoice-items-table { width: 100%; border-collapse: separate; border-spacing: 0;
        border-radius: 10px; border: 1px solid #cfd8e6; overflow: hidden; margin-bottom: 16px; }
      th, td { padding: 9px 12px; border-bottom: 1px solid #cfd8e6;
        text-align: left; font-size: 12px; line-height: 1.3; }
      th { background: #fbfcfe; font-weight: 700; }
      th:first-child, td:first-child { width: 36px; }
      td:nth-child(3), th:nth-child(3) { text-align: center; }
      td:last-child, th:last-child, td:nth-last-child(2), th:nth-last-child(2) { text-align: right; }
      .total-row td { font-weight: 700; background: #fff; border-bottom: 0; }
      .payment-stack { display: grid; gap: 12px; margin-bottom: 16px; }
      .payment-card { display: grid;
        grid-template-columns: 1.3fr 1fr 1fr 0.85fr 0.95fr;
        gap: 10px; align-items: center; min-height: 64px; padding: 12px 14px;
        border: 1px solid #cfd8e6; border-radius: 10px; }
      .payment-main, .payment-amount { font-size: 12px; font-weight: 600; }
      .payment-amount { text-align: right; white-space: nowrap; }
      .payment-meta { display: grid; gap: 3px; }
      .meta-value { font-size: 12px; font-weight: 600; }
      .meta-label { color: #5d6b87; font-size: 11px; font-weight: 600; }
      .payment-status { display: inline-flex; align-items: center; justify-content: center;
        min-width: 70px; padding: 4px 10px; border-radius: 999px;
        font-size: 11px; font-weight: 700; }
      .payment-status.paid { background: #dcf4e3; color: #1f8550; }
      .payment-status.overdue { background: #f8dede; color: #c44747; }
      .payment-status.waiting { background: #fff5bd; color: #8a4b12; }
      .payment-status.cancelled { background: #f1f5f9; color: #64748b; }
      .bank-grid { display: grid; grid-template-columns: 1fr 2fr; gap: 14px 18px;
        margin-bottom: 12px; }
      .bank-grid > .label { color: #5d6b87; font-size: 11px; font-weight: 700; padding-top: 2px; }
      .bank-grid > .value { font-size: 12px; line-height: 1.45; }
      .bank-grid > .value p { margin: 0 0 4px; }
      .bank-grid > .value strong { color: #1f2937; }
      .bank-block { margin-bottom: 8px; }
      .suggestion-block { margin: 18px 0 8px; }
      .suggestion-title { color: #5d6b87; font-size: 12px; font-weight: 700; margin-bottom: 6px; }
      .suggestion-block p, .suggestion-block li {
        margin: 0 0 4px; font-size: 11px; line-height: 1.45; color: #27272a; }
      .suggestion-block ul { margin: 0 0 6px 18px; padding: 0; }
      .suggestion-block .accent { color: #c44747; font-weight: 700; }
      /* Signature block — same proven layout the DTX template uses, just
         with usm-stamp.png. The label gets pushed up by margin-bottom,
         the line sits below it, and the stamp + handwritten signature
         are positioned absolute over both. Extra margin-top so the
         stamp doesn't bleed into the SUGGESTION block above. */
      .signature-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 28px;
        margin-top: 60px; align-items: start; }
      .signature-card { position: relative; min-height: 218px; padding-top: 0; }
      .signature-label { position: relative; z-index: 3; min-height: 18px;
        margin-bottom: 112px; color: #64748b; font-size: 13px;
        font-weight: 600; background: #fff; }
      .signature-line { border-bottom: 1px dashed #d5ddec; }
      .accountant-stamp { position: absolute; left: 2px; bottom: 36px;
        width: 218px; z-index: 1; opacity: 0.98; }
      .accountant-signature { position: absolute; left: 64px; bottom: -34px;
        width: 290px; z-index: 2; }
      .signature-name { position: relative; z-index: 4; margin-top: 14px;
        font-size: 13px; font-weight: 700; color: #27272a; }
      .signature-role { position: relative; z-index: 4; color: #27272a;
        font-size: 13px; }
    """
    return f"""<!DOCTYPE html>
<html lang="{lang}"><head><meta charset="UTF-8"><title>{html.escape(L['title'])} #{serial}</title>
<style>{css}</style></head><body><div class="page">
  <p class="invoice-number">{html.escape(L['title'])} #{serial}</p>
  <div class="header-grid">
    <div class="company-block">
      {f'<img class="invoice-logo" src="{logo_src}" alt="">' if logo_src else ''}
      <p class="company-name">{html.escape(USM_COMPANY['name'])}</p>
      {company_address}
      <p>{html.escape(USM_COMPANY['email'])}</p>
      <p>{html.escape(USM_COMPANY['phone'])}</p>
    </div>
    <div class="customer-block">
      <span class="label">{html.escape(L['bill_to'])}</span>
      <p><strong>{customer}</strong></p>
      {f'<span class="label" style="margin-top:8px;">' + html.escape(L['billing_address']) + '</span><p>' + billing_address + '</p>' if billing_address else ''}
    </div>
  </div>
  <p class="section-title">{html.escape(L['price_detail'])}</p>
  <table class="invoice-items-table">
    <thead><tr><th>{html.escape(L['col_n'])}</th><th>{html.escape(L['col_description'])}</th><th>{html.escape(L['col_amount'])}</th><th>{html.escape(L['col_unit_price'])}</th><th>{html.escape(L['col_total'])}</th></tr></thead>
    <tbody>{items_rows}<tr class="total-row"><td colspan="4">{html.escape(L['row_total'])}</td><td>{_fmt_money_ccy(grand, currency)}</td></tr></tbody>
  </table>
  <p class="section-title">{html.escape(L['installments'])}</p>
  <div class="payment-stack">{payments_html}</div>
  <div class="bank-grid">
    <div class="label">{html.escape(L['ben_name_addr'])}</div>
    <div class="value">
      <p>{html.escape(L['name'])}: <strong>{html.escape(USM_BENEFICIARY_NAME)}</strong></p>
      <p>{html.escape(L['address'])}: <strong>{html.escape(USM_BENEFICIARY_ADDRESS)}</strong></p>
    </div>
    <div class="label">{html.escape(L['ben_account_no'])}</div>
    <div class="value">{account_html}</div>
    <div class="label">{html.escape(L['ben_bank'])}</div>
    <div class="value">
      <p>SWIFT/BIC: <strong>{html.escape(USM_BENEFICIARY_BANK['swift'])}</strong></p>
      <p>{html.escape(USM_BENEFICIARY_BANK['name'])}</p>
      <p>{html.escape(L['address'])}: <strong>{html.escape(USM_BENEFICIARY_BANK['address'])}</strong></p>
    </div>
    <div class="label">{html.escape(L['ben_intermediary'])}</div>
    <div class="value">{intermediaries_html}</div>
  </div>
  <div class="suggestion-block">
    <p class="suggestion-title">{html.escape(L['suggestion'])}</p>
    <p><span class="accent">*</span> {html.escape(L['suggestion_intro'])}</p>
    <ul>
      <li>{html.escape(L['suggestion_li_name'])}</li>
      <li>{html.escape(L['suggestion_li_acct'])}</li>
    </ul>
    <p><span class="accent">**</span> {html.escape(L['suggestion_outro'])}</p>
  </div>
  <div class="signature-grid">
    <div class="signature-card">
      <div class="signature-label">{html.escape(L['company_label'])}</div>
      <div class="signature-line"></div>
      {f'<img class="accountant-stamp" src="{stamp_src}" alt="">' if stamp_src else ''}
      {f'<img class="accountant-signature" src="{sig_src}" alt="">' if sig_src else ''}
      <div class="signature-name">{html.escape(L['accountant'])}</div>
      <div class="signature-role">{html.escape(USM_COMPANY['accountant'])}</div>
    </div>
    <div class="signature-card">
      <div class="signature-label">{html.escape(L['bill_to'])}</div>
      <div class="signature-line"></div>
      <div class="signature-name">{customer}</div>
    </div>
  </div>
</div></body></html>"""


def build_standalone_invoice_html(invoice):
    """Render the new-model invoice as Mongolian printable HTML for WeasyPrint.
    CSS mirrors the working contract-invoice template (build_invoice_html).
    """
    serial = html.escape(str(invoice.get("serial") or invoice.get("id") or ""))
    customer = html.escape(str((invoice.get("payerName") or "CLIENT")).upper())
    items = invoice.get("items") or []
    grand = sum((float(it.get("qty") or 0) * float(it.get("price") or 0)) for it in items)
    def _fmt_qty(q):
        try:
            f = float(q or 0)
        except (TypeError, ValueError):
            return html.escape(str(q or 0))
        return str(int(f)) if f.is_integer() else str(f)
    items_rows = "".join(
        f"<tr><td>{i+1}</td>"
        f"<td>{html.escape(str(it.get('description') or ''))}</td>"
        f"<td>{_fmt_qty(it.get('qty'))}</td>"
        f"<td>{_fmt_money(it.get('price'))}</td>"
        f"<td>{_fmt_money(float(it.get('qty') or 0) * float(it.get('price') or 0))}</td></tr>"
        for i, it in enumerate(items)
    )
    payments_html = ""
    for inst in (invoice.get("installments") or []):
        status_key = (inst.get("status") or "pending").lower()
        label, klass = INVOICE_STATUS_LABELS.get(status_key, INVOICE_STATUS_LABELS["pending"])
        payments_html += f"""
        <div class="payment-card">
          <div class="payment-main">{html.escape(str(inst.get('description') or ''))}</div>
          <div class="payment-meta"><span class="meta-label">Нэхэмжилсэн огноо</span><span class="meta-value">{html.escape(str(inst.get('issueDate') or '-'))}</span></div>
          <div class="payment-meta"><span class="meta-label">Эцсийн хугацаа</span><span class="meta-value">{html.escape(str(inst.get('dueDate') or '-'))}</span></div>
          <div class="payment-meta"><span class="meta-label">Төлөв</span><span class="payment-status {klass}">{label}</span></div>
          <div class="payment-amount">{_fmt_money(inst.get('amount'))}</div>
        </div>
        """
    # Use Path.as_uri() so WeasyPrint loads the local files reliably (handles spaces).
    def _asset(name):
        p = (BASE_DIR / "public" / "assets" / name)
        return p.resolve().as_uri() if p.exists() else ""
    logo_src = _asset("dtx-logo-blue-yellow.png")
    stamp_src = _asset("invoice-finance-stamp.png")
    sig_src = _asset("invoice-finance-signature.png")
    css = """
      @page { size: A4; margin: 16mm 14mm; }
      * { box-sizing: border-box; }
      body { margin: 0; background: #fff; color: #27272a;
        font-family: 'Nunito', Arial, sans-serif; font-size: 13px; }
      .page { padding: 0; }
      .invoice-number { margin: 0 0 18px; color: #27272a; font-size: 18px;
        line-height: 1.15; font-weight: 500; }
      .header-grid { display: grid; grid-template-columns: 1.05fr 0.95fr;
        gap: 20px; align-items: start; margin-bottom: 28px; }
      .invoice-logo { width: 154px; max-width: 100%; display: block; margin-bottom: 10px; }
      .company-name { margin: 0 0 12px; font-size: 14px; line-height: 1.25;
        font-weight: 700; color: #27272a; }
      .company-block p, .customer-block p, .meta-note {
        margin: 0; font-size: 13px; line-height: 1.38; }
      .meta-note { text-align: right; color: #27272a; white-space: nowrap; }
      .customer-block { padding-top: 80px; }
      .customer-block .label { display: block; margin-bottom: 4px; color: #64748b;
        font-size: 13px; font-weight: 600; }
      .section-title { margin: 0 0 10px; color: #64748b; font-size: 13px; font-weight: 600; }
      .invoice-items-table { width: 100%; border-collapse: separate; border-spacing: 0;
        border-radius: 12px; border: 1px solid #cfd8e6; overflow: hidden; margin-bottom: 28px; }
      th, td { padding: 10px 12px; border-bottom: 1px solid #cfd8e6;
        text-align: left; font-size: 13px; line-height: 1.25; }
      th { background: #fbfcfe; color: #27272a; font-weight: 700; }
      th:first-child, td:first-child { width: 46px; }
      td:last-child, th:last-child, td:nth-last-child(2), th:nth-last-child(2) { text-align: right; }
      .total-row td { font-weight: 700; background: #fff; border-bottom: 0; }
      .payment-stack { display: grid; gap: 16px; margin-bottom: 28px; }
      .payment-card { display: grid;
        grid-template-columns: 1.35fr 1.08fr 1.08fr 0.9fr 0.92fr;
        gap: 10px; align-items: center; min-height: 74px; padding: 16px 18px;
        border: 1px solid #cfd8e6; border-radius: 12px; background: #fff; }
      .payment-main, .payment-amount { min-width: 0; font-size: 13px; font-weight: 600; color: #27272a; }
      .payment-amount { text-align: right; white-space: nowrap; }
      .payment-meta { display: grid; gap: 4px; min-width: 0; }
      .meta-value { font-size: 13px; font-weight: 600; color: #27272a; }
      .meta-label { color: #64748b; font-size: 13px; font-weight: 600; white-space: nowrap; }
      .payment-status { display: inline-flex; align-items: center; justify-content: center;
        min-width: 82px; min-height: 30px; padding: 5px 11px; border-radius: 999px;
        font-size: 12px; font-weight: 700; }
      .payment-status.paid { background: #dcf4e3; color: #1f8550; }
      .payment-status.overdue { background: #f8dede; color: #c44747; }
      .payment-status.waiting { background: #fff5bd; color: #8a4b12; }
      .bank-section { margin-top: 0; padding-bottom: 28px; border-bottom: 1px solid #d9e0ea; }
      .bank-grid { display: grid; gap: 4px; font-size: 13px; line-height: 1.35; color: #27272a; }
      .bank-line { display: flex; flex-wrap: wrap; gap: 18px; align-items: baseline; }
      .bank-prefix { color: #64748b; }
      .bank-account-number { font-weight: 800; }
      .signature-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 28px;
        margin-top: 24px; align-items: start; }
      .signature-card { position: relative; min-height: 218px; padding-top: 0; }
      .signature-label { position: relative; z-index: 3; min-height: 18px; margin-bottom: 112px;
        color: #64748b; font-size: 13px; font-weight: 600; background: #fff; }
      .signature-line { border-bottom: 1px dashed #d5ddec; }
      .accountant-stamp { position: absolute; left: 2px; bottom: 36px; width: 218px;
        z-index: 1; opacity: 0.98; }
      .accountant-signature { position: absolute; left: 64px; bottom: -34px; width: 290px; z-index: 2; }
      .signature-name { position: relative; z-index: 4; margin-top: 14px; font-size: 13px;
        font-weight: 700; color: #27272a; }
      .signature-role { position: relative; z-index: 4; color: #27272a; font-size: 13px; }
    """
    return f"""<!DOCTYPE html>
<html lang="mn"><head><meta charset="UTF-8"><title>Нэхэмжлэх #{serial}</title>
<style>{css}</style></head><body><div class="page">
  <p class="invoice-number">Нэхэмжлэх #{serial}</p>
  <div class="header-grid">
    <div class="company-block">
      {f'<img class="invoice-logo" src="{logo_src}" alt="">' if logo_src else ''}
      <p class="company-name">Дэлхий Трэвел Икс ХХК (6925073)</p>
      <p>Улаанбаатар хот, ХУД, 17-р хороо</p>
      <p>Их Монгол Улс гудамж, Кинг Тауэр, 121 байр, 102 тоот</p>
      <p>info@travelx.mn</p><p>+976 72007722</p>
    </div>
    <div>
      <p class="meta-note">Сангийн сайдын 2017 оны 12 дугаар сарын 05</p>
      <p class="meta-note">өдрийн 347 тоот тушаалын хавсралт</p>
      <div class="customer-block"><span class="label">Төлөгч</span><p><strong>{customer}</strong></p></div>
    </div>
  </div>
  <p class="section-title">Үнийн мэдээлэл</p>
  <table class="invoice-items-table">
    <thead><tr><th>№</th><th>Утга</th><th>Аялагч</th><th>Нэгжийн үнэ</th><th>Нийт үнэ</th></tr></thead>
    <tbody>{items_rows}<tr class="total-row"><td colspan="4">Нийт үнэ</td><td>{_fmt_money(grand)}</td></tr></tbody>
  </table>
  <p class="section-title">Төлбөрийн хуваарь</p>
  <div class="payment-stack">{payments_html}</div>
  <div class="bank-section"><p class="section-title">Дансны мэдээлэл</p>
    <div class="bank-grid"><div>Дэлхий Трэвел Икс</div>
      <div class="bank-line"><span>Төрийн Банк</span><span class="bank-prefix">MN030034</span><strong class="bank-account-number">3432 7777 9999</strong></div>
    </div>
  </div>
  <div class="signature-grid">
    <div class="signature-card"><div class="signature-label">Дэлхий Трэвел Икс ХХК</div><div class="signature-line"></div>
      {f'<img class="accountant-stamp" src="{stamp_src}" alt="">' if stamp_src else ''}
      {f'<img class="accountant-signature" src="{sig_src}" alt="">' if sig_src else ''}
      <div class="signature-name">Нягтлан</div><div class="signature-role">Г.Басгалаан</div>
    </div>
    <div class="signature-card"><div class="signature-label">Төлөгч</div><div class="signature-line"></div>
      <div class="signature-name">{customer}</div>
    </div>
  </div>
</div></body></html>"""


def save_standalone_invoice_pdf(invoice):
    ensure_data_store()
    pdf_filename = f"std-invoice-{invoice['id']}.pdf"
    pdf_path = GENERATED_DIR / pdf_filename
    try:
        from weasyprint import HTML
    except Exception as exc:
        raise RuntimeError(f"WeasyPrint not available: {exc}") from exc
    # Pick the template by issuing company. USM (S- serial) gets the
    # English steppe-mongolia template with currency-aware bank info.
    # DTX (T- serial) keeps the existing Mongolian template.
    company = _company_from_invoice(invoice)
    if company == "USM":
        html_string = build_standalone_invoice_html_usm(invoice)
    else:
        html_string = build_standalone_invoice_html(invoice)
    try:
        HTML(string=html_string, base_url=str(BASE_DIR), media_type="screen").write_pdf(str(pdf_path))
    except Exception as exc:
        raise RuntimeError(f"PDF generation failed: {exc}") from exc
    return f"/generated/{pdf_filename}"


def handle_standalone_invoice_pdf(environ, start_response, invoice_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    invoices = read_invoices()
    invoice = next((i for i in invoices if i.get("id") == invoice_id), None)
    if not invoice:
        return json_response(start_response, "404 Not Found", {"error": "Invoice not found"})
    try:
        pdf_url = save_standalone_invoice_pdf(invoice)
    except Exception as exc:
        return json_response(start_response, "500 Internal Server Error", {"error": f"Could not generate invoice PDF: {exc}"})
    safe_path = (GENERATED_DIR / unquote(pdf_url.replace("/generated/", "", 1))).resolve()
    if not str(safe_path).startswith(str(GENERATED_DIR.resolve())) or not safe_path.exists():
        return json_response(start_response, "404 Not Found", {"error": "Invoice PDF not found"})
    return file_response(start_response, safe_path, extra_headers=generated_download_headers(safe_path))


def save_contract_pdf(record):
    ensure_data_store()
    pdf_filename = f"contract-{record['id']}.pdf"
    pdf_path = GENERATED_DIR / pdf_filename
    try:
        from weasyprint import HTML
    except Exception as exc:
        raise RuntimeError(f"WeasyPrint not available: {exc}") from exc

    html_string = build_contract_html(
        record["data"],
        signature_path=record.get("signaturePath"),
        asset_mode="file",
        contract_id=record.get("id"),
    )
    try:
        HTML(string=html_string, base_url=str(BASE_DIR)).write_pdf(str(pdf_path))
    except Exception as exc:
        raise RuntimeError(f"HTML PDF generation failed: {exc}") from exc
    return f"/generated/{pdf_filename}"


def save_contract_files(data):
    contract_id = str(uuid4())
    timestamp = datetime.now(timezone.utc).astimezone(MONGOLIA_TZ).strftime("%Y%m%d%H%M%S")
    filename_stem = slugify(f"contract-{data['contractSerial']}-{timestamp}-{contract_id[:8]}")
    docx_filename = f"{filename_stem}.docx"
    html_filename = f"{filename_stem}.html"

    docx_path = GENERATED_DIR / docx_filename
    html_path = GENERATED_DIR / html_filename

    generate_docx(data, docx_path)
    html_path.write_text(build_contract_html(data), encoding="utf-8")

    record = {
        "id": contract_id,
        "createdAt": datetime.now(timezone.utc).isoformat(),
        "status": "pending",
        "signedAt": None,
        "signaturePath": None,
        "signerName": None,
        "data": data,
        "docxPath": f"/generated/{docx_filename}",
        "pdfViewPath": f"/generated/{html_filename}",
        "pdfPath": None,
        "invoiceViewPath": None,
        "invoicePath": None,
        "invoiceMeta": {"payments": {}, "bankAccountKey": "state"},
    }
    return record


def save_signature_image(data_url, contract_id):
    if not data_url or "base64," not in data_url:
        return None
    header, encoded = data_url.split("base64,", 1)
    if "image/png" not in header:
        return None
    try:
        raw = base64.b64decode(encoded)
    except Exception:
        return None
    filename = f"contract-signature-{contract_id}.png"
    path = GENERATED_DIR / filename
    path.write_bytes(raw)
    return f"/generated/{filename}"


def save_manager_signature_image(data_url, user_id):
    if not data_url or "base64," not in data_url:
        return None
    header, encoded = data_url.split("base64,", 1)
    if "image/png" not in header:
        return None
    try:
        raw = base64.b64decode(encoded)
    except Exception:
        return None
    filename = f"manager-signature-{user_id}-{uuid4().hex[:8]}.png"
    path = GENERATED_DIR / filename
    path.write_bytes(raw)
    return f"/generated/{filename}"


def save_user_avatar_image(data_url, user_id):
    if not data_url or "base64," not in data_url:
        return None
    header, encoded = data_url.split("base64,", 1)
    ext = "png"
    if "image/jpeg" in header or "image/jpg" in header:
        ext = "jpg"
    elif "image/webp" in header:
        ext = "webp"
    elif "image/png" not in header:
        return None
    try:
        raw = base64.b64decode(encoded)
    except Exception:
        return None
    if len(raw) > 4 * 1024 * 1024:
        return None
    filename = f"avatar-{user_id}-{uuid4().hex[:8]}.{ext}"
    path = GENERATED_DIR / filename
    path.write_bytes(raw)
    return f"/generated/{filename}"


def build_document_html(title, subtitle, sections):
    section_markup = []
    for heading, content in sections:
        section_markup.append(
            f"<section><h2>{html.escape(heading)}</h2><p>{html.escape(content)}</p></section>"
        )

    return f"""<!DOCTYPE html>
<html lang="en">
  <head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>{html.escape(title)}</title>
    <style>
      body {{
        margin: 0;
        padding: 48px 24px;
        background: #f5efe8;
        color: #1f1713;
        font-family: "Times New Roman", Tinos, "Liberation Serif", serif;
      }}
      main {{
        width: min(860px, 100%);
        margin: 0 auto;
        padding: 40px;
        background: white;
        box-shadow: 0 20px 60px rgba(0, 0, 0, 0.08);
      }}
      h1 {{
        margin-bottom: 8px;
      }}
      .subtitle {{
        margin: 0 0 28px;
        color: #70584a;
      }}
      section {{
        margin-top: 24px;
      }}
      h2 {{
        margin: 0 0 10px;
        font-size: 1.05rem;
        letter-spacing: 0.04em;
        text-transform: uppercase;
      }}
      p {{
        margin: 0;
        line-height: 1.65;
        white-space: pre-wrap;
      }}
    </style>
  </head>
  <body>
    <main>
      <h1>{html.escape(title)}</h1>
      <p class="subtitle">{html.escape(subtitle)}</p>
      {''.join(section_markup)}
    </main>
  </body>
</html>
"""


def save_simple_document(prefix, title, subtitle, sections):
    doc_id = str(uuid4())
    filename_stem = slugify(f"{prefix}-{title}-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}")
    html_filename = f"{filename_stem}.html"
    doc_filename = f"{filename_stem}.doc"
    html_path = GENERATED_DIR / html_filename
    doc_path = GENERATED_DIR / doc_filename
    html = build_document_html(title, subtitle, sections)
    html_path.write_text(html, encoding="utf-8")
    doc_path.write_text(html, encoding="utf-8")
    return {
        "id": doc_id,
        "wordPath": f"/generated/{doc_filename}",
        "pdfViewPath": f"/generated/{html_filename}",
    }


def format_pdf_date(value):
    parts = split_date_parts(value)
    if not parts["year"]:
        return "-"
    return f"{parts['year']}.{parts['month']}.{parts['day']}"


def format_iso_date(value):
    parts = split_date_parts(value)
    if not parts["year"]:
        return "-"
    return f"{parts['year']}-{parts['month']}-{parts['day']}"


def camp_reservation_title(record):
    reservation_type = normalize_text(record.get("reservationType")).lower() or "camp"
    return RESERVATION_TYPE_LABELS.get(reservation_type, RESERVATION_TYPE_LABELS["camp"])


def is_hotel_reservation(record):
    reservation_type = normalize_text(record.get("reservationType")).lower()
    camp_name = normalize_text(record.get("campName")).lower()
    title = camp_reservation_title(record)
    return reservation_type == "hotel" or "hotel" in camp_name or title == RESERVATION_TYPE_LABELS["hotel"]


def camp_reservation_meals(record):
    return " / ".join(
        [
            meal
            for meal in [
                record.get("breakfast") == "Yes" and "Breakfast",
                record.get("lunch") == "Yes" and "Lunch",
                record.get("dinner") == "Yes" and "Dinner",
            ]
            if meal
        ]
    ) or "No meal"


def camp_reservation_meals_mn(record):
    return " / ".join(
        [
            meal
            for meal in [
                record.get("breakfast") == "Yes" and "Өглөөний цай",
                record.get("lunch") == "Yes" and "Өдрийн хоол",
                record.get("dinner") == "Yes" and "Оройн хоол",
            ]
            if meal
        ]
    ) or "Хоолгүй"


def build_camp_document_html(record, pdf_href):
    meals = camp_reservation_meals_mn(record)
    reservation_title = camp_reservation_title(record)
    manager_name = record.get("staffAssignment") or STEPPE_MANAGER
    is_hotel = is_hotel_reservation(record)
    unit_label = "Өрөөний тоо" if is_hotel else "Гэрийн тоо"
    room_type_label = "Өрөөний төрөл" if is_hotel else "Гэрийн төрөл"
    return f"""<!DOCTYPE html>
<html lang="mn">
  <head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Camp reservation</title>
    <style>
      body {{
        margin: 0;
        background: #f1f0ec;
        color: #161616;
        font-family: "Times New Roman", Tinos, "Liberation Serif", serif;
      }}
      .toolbar {{
        position: sticky;
        top: 0;
        display: flex;
        gap: 12px;
        justify-content: center;
        padding: 16px;
        background: rgba(255,255,255,0.94);
        border-bottom: 1px solid rgba(0,0,0,0.08);
      }}
      .toolbar button, .toolbar a {{
        padding: 12px 18px;
        border: none;
        border-radius: 999px;
        background: #17365f;
        color: #fff;
        text-decoration: none;
        font: 600 14px/1 sans-serif;
        cursor: pointer;
      }}
      .page {{
        width: 920px;
        margin: 22px auto 40px;
        padding: 56px 52px 72px;
        background: white;
        box-shadow: 0 20px 60px rgba(0,0,0,0.1);
      }}
      .logo {{
        color: #1d2f86;
        font: 800 30px/1 "Poppins", Arial, sans-serif;
        letter-spacing: 0.04em;
        text-transform: uppercase;
      }}
      .top {{
        display: grid;
        grid-template-columns: 1.1fr 1.2fr;
        gap: 24px;
        align-items: start;
      }}
      .brand-sub {{
        margin-top: 14px;
        color: #4e4e4e;
        line-height: 1.55;
        font-size: 15px;
      }}
      .doc-title {{
        text-align: right;
      }}
      .doc-title h1 {{
        margin: 0;
        font-size: 24px;
        color: #1d2f86;
        white-space: nowrap;
      }}
      .doc-meta {{
        display: flex;
        justify-content: space-between;
        margin-top: 8px;
        font-size: 15px;
      }}
      .center-title {{
        margin: 42px 0 28px;
        text-align: center;
      }}
      .center-title h2 {{
        margin: 0 0 22px;
        font-size: 22px;
      }}
      .center-title p {{
        margin: 0;
        font-size: 18px;
        font-weight: 700;
      }}
      table {{
        width: 100%;
        border-collapse: collapse;
        margin-top: 20px;
      }}
      th, td {{
        border: 1px solid #2a2a2a;
        padding: 12px 10px;
        text-align: center;
        vertical-align: top;
        font-size: 16px;
      }}
      th {{
        background: #d8e4f2;
      }}
      .footer {{
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 24px;
        margin-top: 54px;
      }}
      .status-box {{
        padding: 18px;
        border-radius: 18px;
        background: #f7f8fb;
      }}
      .status-box p {{
        margin: 0 0 8px;
      }}
      @media print {{
        @page {{
          size: A4 landscape;
          margin: 0;
        }}
        body {{
          background: #fff;
        }}
        .toolbar {{
          display: none;
        }}
        .page {{
          width: auto;
          min-height: 100vh;
          margin: 0;
          padding: 34px 40px 40px;
          box-shadow: none;
          box-sizing: border-box;
        }}
        .logo {{
          font-size: 24px;
        }}
        .brand-sub {{
          font-size: 13px;
        }}
        .center-title {{
          margin: 28px 0 22px;
        }}
        th, td {{
          padding: 8px 7px;
          font-size: 13px;
        }}
        .footer {{
          margin-top: 34px;
        }}
      }}
    </style>
  </head>
  <body>
    <div class="toolbar">
      <button onclick="window.print()">Print / Save PDF</button>
      <a href="{html.escape(pdf_href)}" download>Download PDF</a>
    </div>
    <div class="page">
      <div class="top">
        <div>
          <div class="logo">Unlock Steppe Mongolia</div>
          <div class="brand-sub">
            {"<br/>".join(html.escape(line) for line in STEPPE_ADDRESS_LINES)}<br/>
            Утас: {html.escape(STEPPE_PHONES)}<br/>
            И-мэйл: {html.escape(STEPPE_EMAIL)}
          </div>
        </div>
        <div class="doc-title">
          <h1 style="font-size:24px; white-space:nowrap;">{html.escape(STEPPE_COMPANY_NAME)}</h1>
          <div class="doc-meta">
            <span></span>
            <span>{html.escape(STEPPE_CITY)}</span>
          </div>
          <div class="doc-meta" style="justify-content:flex-end; margin-top: 2px;">
            <span>{format_pdf_date(record['createdDate'])}</span>
          </div>
        </div>
      </div>

      <div class="center-title">
        <h2>Аяллын нэр: {html.escape(record.get('reservationName') or record['tripName'])}</h2>
        <p>{reservation_title} - “{html.escape(record['campName'])}”</p>
      </div>

      <table>
        <thead>
          <tr>
            <th>Жуулчны тоо</th>
            <th>Ажилчдын тоо</th>
            <th>Ирэх өдөр</th>
            <th>Явах өдөр</th>
            <th>Хоногийн тоо</th>
            <th>{unit_label}</th>
            <th>{room_type_label}</th>
            <th>Хоол</th>
          </tr>
        </thead>
        <tbody>
          <tr>
            <td>{record['clientCount']}</td>
            <td>{record['staffCount']}</td>
            <td>{format_pdf_date(record['checkIn'])}</td>
            <td>{format_pdf_date(record['checkOut'])}</td>
            <td>{record['nights']}</td>
            <td>{record['gerCount']}</td>
            <td>{html.escape(record['roomType'])}</td>
            <td>{html.escape(meals)}</td>
          </tr>
        </tbody>
      </table>

      <div style="margin-top: 18px; font-size: 18px; line-height: 1.7;">
        <strong>Нэмэлт тэмдэглэл:</strong> {html.escape(record['notes'] or '-')}
      </div>

      <div class="footer">
        <div></div>
        <div class="status-box">
          <p><strong>Захиалгын менежер:</strong> {html.escape(manager_name)}</p>
          <p><strong>Харилцах утас:</strong> {html.escape(STEPPE_CONTACT_PHONES)}</p>
          <p><strong>Цахим шуудан:</strong> {html.escape(STEPPE_EMAIL)}</p>
        </div>
      </div>
    </div>
  </body>
</html>
"""


def build_camp_bundle_document_html(records, pdf_href):
    if not records:
        return build_document_html("Camp reservations", "No records", [])

    first = records[0]
    reservation_title = camp_reservation_title(first)
    manager_name = first.get("staffAssignment") or STEPPE_MANAGER
    is_hotel_bundle = is_hotel_reservation(first)
    place_label = "Буудал" if is_hotel_bundle else "Бааз"
    location_label = "Байршил"
    reservations_label = "Захиалга"
    manager_label = "Менежер"
    unit_label = "Өрөө" if is_hotel_bundle else "Гэр"
    room_type_label = "Өрөөний төрөл" if is_hotel_bundle else "Гэрийн төрөл"
    common_labels = {
        "reservation": "Захиалга",
        "pax": "Жуулчин",
        "staff": "Ажилчид",
        "check_in": "Ирэх өдөр",
        "check_out": "Явах өдөр",
        "nights": "Хоног",
        "meals": "Хоол",
    }
    row_markup = "".join(
        f"""
          <tr>
            <td>{index + 1}</td>
            <td>{html.escape(record.get('reservationName') or record['tripName'])}</td>
            <td>{record['clientCount']}</td>
            <td>{record['staffCount']}</td>
            <td>{format_pdf_date(record['checkIn'])}</td>
            <td>{format_pdf_date(record['checkOut'])}</td>
            <td>{record['nights']}</td>
            <td>{record['gerCount']}</td>
            <td>{html.escape(record['roomType'])}</td>
            <td>{html.escape(camp_reservation_meals_mn(record))}</td>
          </tr>
        """
        for index, record in enumerate(records)
    )
    return f"""<!DOCTYPE html>
<html lang="mn">
  <head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Camp reservation export</title>
    <style>
      body {{
        margin: 0;
        background: #f1f0ec;
        color: #161616;
        font-family: "Times New Roman", Tinos, "Liberation Serif", serif;
      }}
      .toolbar {{
        position: sticky;
        top: 0;
        display: flex;
        gap: 12px;
        justify-content: center;
        padding: 16px;
        background: rgba(255,255,255,0.94);
        border-bottom: 1px solid rgba(0,0,0,0.08);
      }}
      .toolbar button, .toolbar a {{
        padding: 12px 18px;
        border: none;
        border-radius: 999px;
        background: #17365f;
        color: #fff;
        text-decoration: none;
        font: 600 14px/1 sans-serif;
        cursor: pointer;
      }}
      .page {{
        width: 920px;
        margin: 22px auto 40px;
        padding: 56px 52px 72px;
        background: white;
        box-shadow: 0 20px 60px rgba(0,0,0,0.1);
      }}
      .logo {{
        color: #1d2f86;
        font: 800 30px/1 "Poppins", Arial, sans-serif;
        letter-spacing: 0.04em;
        text-transform: uppercase;
      }}
      .top {{
        display: grid;
        grid-template-columns: 1.1fr 1.2fr;
        gap: 24px;
        align-items: start;
      }}
      .brand-sub {{
        margin-top: 14px;
        color: #4e4e4e;
        line-height: 1.55;
        font-size: 15px;
      }}
      .doc-title {{
        text-align: right;
      }}
      .doc-title h1 {{
        margin: 0;
        font-size: 24px;
        white-space: nowrap;
        color: #1d2f86;
      }}
      .doc-meta {{
        display: flex;
        justify-content: space-between;
        margin-top: 8px;
        font-size: 15px;
      }}
      .center-title {{
        margin: 42px 0 28px;
        text-align: center;
      }}
      .center-title h2 {{
        margin: 0 0 22px;
        font-size: 22px;
      }}
      .center-title p {{
        margin: 0;
        font-size: 18px;
        font-weight: 700;
      }}
      table {{
        width: 100%;
        border-collapse: collapse;
        margin-top: 20px;
      }}
      th, td {{
        border: 1px solid #2a2a2a;
        padding: 10px 8px;
        text-align: center;
        vertical-align: top;
        font-size: 15px;
      }}
      th {{
        background: #d8e4f2;
      }}
      .footer {{
        display: grid;
        grid-template-columns: 1fr minmax(360px, 48%);
        gap: 24px;
        margin-top: 54px;
      }}
      .status-box {{
        padding: 18px;
        border-radius: 18px;
        background: #f7f8fb;
      }}
      .status-box p {{
        margin: 0 0 8px;
      }}
    </style>
  </head>
  <body>
    <div class="toolbar">
      <button onclick="window.print()">Print / Save PDF</button>
      <a href="{html.escape(pdf_href)}" download>Download PDF</a>
    </div>
    <div class="page">
      <div class="top">
        <div>
          <div class="logo">Unlock Steppe Mongolia</div>
          <div class="brand-sub">
            {"<br/>".join(html.escape(line) for line in STEPPE_ADDRESS_LINES)}<br/>
            Утас: {html.escape(STEPPE_PHONES)}<br/>
            И-мэйл: {html.escape(STEPPE_EMAIL)}
          </div>
        </div>
        <div class="doc-title">
          <h1>{html.escape(STEPPE_COMPANY_NAME)}</h1>
          <div class="doc-meta">
            <span></span>
            <span>{html.escape(STEPPE_CITY)}</span>
          </div>
          <div class="doc-meta" style="justify-content:flex-end; margin-top: 2px;">
            <span>{format_pdf_date(first['createdDate'])}</span>
          </div>
        </div>
      </div>
      <div class="center-title">
        <p>{reservation_title} - “{html.escape(first['campName'])}”</p>
      </div>
      <table>
        <thead>
          <tr>
            <th>#</th>
            <th>{common_labels["reservation"]}</th>
            <th>{common_labels["pax"]}</th>
            <th>{common_labels["staff"]}</th>
            <th>{common_labels["check_in"]}</th>
            <th>{common_labels["check_out"]}</th>
            <th>{common_labels["nights"]}</th>
            <th>{unit_label}</th>
            <th>{room_type_label}</th>
            <th>{common_labels["meals"]}</th>
          </tr>
        </thead>
        <tbody>{row_markup}</tbody>
      </table>
      <div class="footer">
        <div></div>
        <div class="status-box">
          <p><strong>Захиалгын менежер:</strong> {html.escape(manager_name)}</p>
          <p><strong>Харилцах утас:</strong> {html.escape(STEPPE_CONTACT_PHONES)}</p>
          <p><strong>Цахим шуудан:</strong> {html.escape(STEPPE_EMAIL)}</p>
        </div>
      </div>
    </div>
  </body>
</html>
"""


def save_camp_reservation_document(record):
    ensure_data_store()
    filename_stem = slugify(
        f"camp-{record.get('reservationName') or record['tripName']}-{record['campName']}-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}"
    )
    html_filename = f"{filename_stem}.html"
    pdf_filename = f"{filename_stem}.pdf"
    html_path = GENERATED_DIR / html_filename
    pdf_path = GENERATED_DIR / pdf_filename

    pdf_href = f"/generated/{pdf_filename}"
    html_path.write_text(build_camp_document_html(record, pdf_href), encoding="utf-8")

    pdf_ready = False

    try:
        from weasyprint import HTML

        html_string = build_camp_document_html(record, pdf_href)
        HTML(string=html_string, base_url=str(BASE_DIR)).write_pdf(str(pdf_path))
        pdf_ready = True
    except Exception as exc:
        print(f"Camp reservation PDF generation failed: {exc}", flush=True)
        pdf_href = f"/generated/{html_filename}"

    return {
        "pdfViewPath": f"/generated/{html_filename}",
        "pdfPath": pdf_href if pdf_ready else f"/generated/{html_filename}",
    }


def save_camp_reservations_bundle(records):
    ensure_data_store()
    first = records[0]
    filename_stem = slugify(
        f"camp-bundle-{first['tripName']}-{first['campName']}-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}"
    )
    html_filename = f"{filename_stem}.html"
    pdf_filename = f"{filename_stem}.pdf"
    html_path = GENERATED_DIR / html_filename
    pdf_path = GENERATED_DIR / pdf_filename

    pdf_href = f"/generated/{pdf_filename}"
    html_path.write_text(build_camp_bundle_document_html(records, pdf_href), encoding="utf-8")

    pdf_ready = False
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        font_name = "Helvetica"
        bold_font_name = "Helvetica-Bold"
        for candidate in [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
            "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf",
            "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
            "/Library/Fonts/Arial Unicode.ttf",
            "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        ]:
            if Path(candidate).exists():
                pdfmetrics.registerFont(TTFont("TravelXUnicode", candidate))
                font_name = "TravelXUnicode"
                bold_font_name = "TravelXUnicode"
                break

        reservation_title = camp_reservation_title(first)
        manager_name = first.get("staffAssignment") or STEPPE_MANAGER
        is_hotel_bundle = is_hotel_reservation(first)
        place_label = "Буудал" if is_hotel_bundle else "Бааз"
        location_label = "Байршил"
        reservations_label = "Захиалга"
        manager_label = "Менежер"
        unit_label = "Өрөө" if is_hotel_bundle else "Гэр"
        room_type_label = "Өрөөний төрөл" if is_hotel_bundle else "Гэрийн төрөл"
        common_labels = {
            "reservation": "Захиалга",
            "pax": "Жуулчин",
            "staff": "Ажилчид",
            "check_in": "Ирэх өдөр",
            "check_out": "Явах өдөр",
            "nights": "Хоног",
            "meals": "Хоол",
        }

        styles = {
            "brand": ParagraphStyle(
                "CampBrand",
                fontName=bold_font_name,
                fontSize=18,
                leading=22,
                textColor=colors.HexColor("#1d2f86"),
                spaceAfter=2,
            ),
            "company": ParagraphStyle(
                "CampCompany",
                fontName=bold_font_name,
                fontSize=13,
                leading=16,
                alignment=2,
                textColor=colors.HexColor("#1d2f86"),
            ),
            "meta": ParagraphStyle("CampMeta", fontName=font_name, fontSize=9, leading=12),
            "meta_right": ParagraphStyle("CampMetaRight", fontName=font_name, fontSize=9, leading=12, alignment=2),
            "title": ParagraphStyle("CampTitle", fontName=bold_font_name, fontSize=16, leading=20, alignment=1, spaceAfter=4),
            "subtitle": ParagraphStyle("CampSubtitle", fontName=bold_font_name, fontSize=13, leading=17, alignment=1, spaceAfter=10),
            "summary": ParagraphStyle("CampSummary", fontName=font_name, fontSize=9, leading=12),
            "th": ParagraphStyle("CampHeaderCell", fontName=bold_font_name, fontSize=7.5, leading=9, alignment=1),
            "td": ParagraphStyle("CampBodyCell", fontName=font_name, fontSize=7.5, leading=9, alignment=1),
            "td_left": ParagraphStyle("CampBodyCellLeft", fontName=font_name, fontSize=7.5, leading=9, alignment=0),
            "footer": ParagraphStyle("CampFooter", fontName=font_name, fontSize=9, leading=12),
        }

        def p(value, style="td"):
            return Paragraph(html.escape(str(value or "-")).replace("\n", "<br/>"), styles[style])

        page_width, _ = landscape(A4)
        margin = 10 * mm
        usable_width = page_width - (margin * 2)
        doc = SimpleDocTemplate(
            str(pdf_path),
            pagesize=landscape(A4),
            rightMargin=margin,
            leftMargin=margin,
            topMargin=10 * mm,
            bottomMargin=9 * mm,
            title=f"{first['campName']} reservations",
        )

        story = []
        header_table = Table(
            [[
                Paragraph("UNLOCK STEPPE MONGOLIA", styles["brand"]),
                Paragraph(f"{html.escape(STEPPE_COMPANY_NAME)}<br/>{html.escape(STEPPE_CITY)}<br/>{format_pdf_date(now_mongolia().date().isoformat())}", styles["company"]),
            ], [
                Paragraph(
                    "<br/>".join(html.escape(line) for line in STEPPE_ADDRESS_LINES)
                    + f"<br/>Утас: {html.escape(STEPPE_PHONES)}<br/>И-мэйл: {html.escape(STEPPE_EMAIL)}",
                    styles["meta"],
                ),
                Paragraph("", styles["meta_right"]),
            ]],
            colWidths=[usable_width * 0.48, usable_width * 0.52],
        )
        header_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.extend([
            header_table,
            Spacer(1, 10),
            Paragraph(f"{html.escape(reservation_title)} - “{html.escape(first['campName'])}”", styles["subtitle"]),
        ])

        summary = Table(
            [[
                p(f"{place_label}: {first['campName']}", "summary"),
                p(f"{location_label}: {first.get('locationName') or '-'}", "summary"),
                p(f"{reservations_label}: {len(records)}", "summary"),
                p(f"{manager_label}: {manager_name}", "summary"),
            ]],
            colWidths=[usable_width * 0.26, usable_width * 0.28, usable_width * 0.18, usable_width * 0.28],
        )
        summary.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f6fb")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#c8d4e6")),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#d7e0ec")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.extend([summary, Spacer(1, 9)])

        table_rows = [[
            p("#", "th"),
            p(common_labels["reservation"], "th"),
            p(common_labels["pax"], "th"),
            p(common_labels["staff"], "th"),
            p(common_labels["check_in"], "th"),
            p(common_labels["check_out"], "th"),
            p(common_labels["nights"], "th"),
            p(unit_label, "th"),
            p(room_type_label, "th"),
            p(common_labels["meals"], "th"),
        ]]
        for index, record in enumerate(records, start=1):
            table_rows.append([
                p(index),
                p(record.get("reservationName") or record["tripName"], "td_left"),
                p(record["clientCount"]),
                p(record["staffCount"]),
                p(format_iso_date(record["checkIn"])),
                p(format_iso_date(record["checkOut"])),
                p(record["nights"]),
                p(record["gerCount"]),
                p(record["roomType"], "td_left"),
                p(camp_reservation_meals_mn(record), "td_left"),
            ])

        col_widths = [
            usable_width * 0.035,
            usable_width * 0.18,
            usable_width * 0.045,
            usable_width * 0.05,
            usable_width * 0.095,
            usable_width * 0.095,
            usable_width * 0.05,
            usable_width * 0.05,
            usable_width * 0.22,
            usable_width * 0.18,
        ]
        table = Table(
            table_rows,
            colWidths=col_widths,
            repeatRows=1,
        )
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d8e4f2")),
            ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#1f2937")),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ]))
        story.append(table)
        contact_box = Table(
            [[
                Paragraph(
                    f"<b>Захиалгын менежер:</b> {html.escape(manager_name)}<br/>"
                    f"<b>Харилцах утас:</b> {html.escape(STEPPE_CONTACT_PHONES)}<br/>"
                    f"<b>Цахим шуудан:</b> {html.escape(STEPPE_EMAIL)}",
                    styles["footer"],
                ),
            ]],
            colWidths=[usable_width * 0.48],
            style=TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f7f8fb")),
                ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#d7e0ec")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]),
        )
        footer_row = Table(
            [["", contact_box]],
            colWidths=[usable_width * 0.52, usable_width * 0.48],
            style=TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]),
        )
        story.extend([Spacer(1, 10), footer_row])

        doc.build(story)
        pdf_ready = True
    except Exception as exc:
        print(f"Camp bundle PDF generation failed: {exc}", flush=True)
        pdf_href = f"/generated/{html_filename}"

    return {
        "pdfViewPath": f"/generated/{html_filename}",
        "pdfPath": pdf_href if pdf_ready else f"/generated/{html_filename}",
    }


def clean_ds160_payload(payload):
    cleaned = {}
    for key, value in (payload or {}).items():
        if isinstance(value, list):
            cleaned_list = []
            for item in value:
                if isinstance(item, dict):
                    cleaned_list.append({k: normalize_text(v) for k, v in item.items()})
                else:
                    cleaned_list.append(normalize_text(item))
            cleaned[key] = cleaned_list
        elif isinstance(value, dict):
            cleaned[key] = {k: normalize_text(v) for k, v in value.items()}
        elif key in ("photo", "passportScan") and isinstance(value, str) and value.startswith("data:"):
            cleaned[key] = value
        else:
            cleaned[key] = normalize_text(value)
    return cleaned


def flatten_ds160_answers(payload):
    cleaned = clean_ds160_payload(payload)
    surname = cleaned.get("surname", "")
    given_name = cleaned.get("givenName", "")
    applicant_name = normalize_text(f"{surname} {given_name}")
    return {
        **cleaned,
        "applicantName": applicant_name,
    }


def normalize_ds160_status(value):
    status = normalize_text(value).lower()
    if status in {"draft", "sent", "submitted", "reviewed"}:
        return status
    return "submitted" if status == "complete" else "sent"


def normalize_ds160_record(record):
    if not isinstance(record, dict):
        return {}
    payload = clean_ds160_payload(record.get("payload"))
    legacy_payload = {
        key: value
        for key, value in record.items()
        if key
        not in {
            "id",
            "clientToken",
            "status",
            "clientName",
            "clientEmail",
            "clientPhone",
            "managerName",
            "managerEmail",
            "managerPhone",
            "shareUrl",
            "internalNotes",
            "appId",
            "createdAt",
            "updatedAt",
            "submittedAt",
            "appointmentDate",
            "appointmentTime",
            "createdBy",
            "updatedBy",
            "payload",
            "applicantName",
            "officialFlowNote",
        }
    }
    if not payload:
        payload = clean_ds160_payload(legacy_payload)
    flattened = flatten_ds160_answers(payload)
    created_at = normalize_text(record.get("createdAt")) or datetime.now(timezone.utc).isoformat()
    updated_at = normalize_text(record.get("updatedAt")) or normalize_text(record.get("submittedAt")) or created_at
    submitted_at = normalize_text(record.get("submittedAt"))
    status = normalize_ds160_status(record.get("status") or ("submitted" if payload else "sent"))
    if status == "submitted" and not submitted_at:
        submitted_at = updated_at

    return {
        "id": normalize_text(record.get("id")) or str(uuid4()),
        "clientToken": normalize_text(record.get("clientToken")) or secrets.token_urlsafe(18),
        "status": status,
        "clientName": normalize_text(record.get("clientName")) or flattened.get("applicantName", ""),
        "clientEmail": normalize_text(record.get("clientEmail") or flattened.get("email")).lower(),
        "clientPhone": normalize_text(record.get("clientPhone") or flattened.get("primaryPhone")),
        "managerName": normalize_text(record.get("managerName") or (record.get("createdBy") or {}).get("name")),
        "managerEmail": normalize_text(record.get("managerEmail")),
        "managerPhone": normalize_text(record.get("managerPhone")),
        "shareUrl": normalize_text(record.get("shareUrl")),
        "internalNotes": normalize_text(record.get("internalNotes") or record.get("notes")),
        "appId": normalize_text(record.get("appId")),
        "createdAt": created_at,
        "updatedAt": updated_at,
        "submittedAt": submitted_at,
        "appointmentDate": normalize_text(record.get("appointmentDate")),
        "appointmentTime": normalize_text(record.get("appointmentTime")),
        "createdBy": record.get("createdBy") if isinstance(record.get("createdBy"), dict) else {"id": "", "email": "", "name": ""},
        "updatedBy": record.get("updatedBy") if isinstance(record.get("updatedBy"), dict) else {"id": "", "email": "", "name": ""},
        "payload": payload,
        **flattened,
        "officialFlowNote": "Энэ нь дотоод мэдээлэл авах маягт бөгөөд албан ёсны DS-160 маягтыг орлохгүй.",
    }


def build_ds160_invitation(payload, actor):
    now_iso = datetime.now(timezone.utc).isoformat()
    record = normalize_ds160_record(
        {
            "id": str(uuid4()),
            "clientToken": secrets.token_urlsafe(18),
            "status": "sent",
            "clientName": payload.get("clientName"),
            "clientEmail": normalize_text(payload.get("clientEmail")).lower(),
            "clientPhone": payload.get("clientPhone"),
            "managerName": payload.get("managerName") or actor.get("fullName") or actor.get("email"),
            "managerEmail": normalize_text(payload.get("managerEmail") or actor.get("email")).lower(),
            "managerPhone": payload.get("managerPhone"),
            "internalNotes": payload.get("internalNotes"),
            "appId": payload.get("appId"),
            "createdAt": now_iso,
            "updatedAt": now_iso,
            "appointmentDate": payload.get("appointmentDate"),
            "appointmentTime": payload.get("appointmentTime"),
            "createdBy": actor_snapshot(actor),
            "updatedBy": actor_snapshot(actor),
            "payload": {},
        }
    )
    return record


def build_ds160_application(payload):
    now_iso = datetime.now(timezone.utc).isoformat()
    flattened = flatten_ds160_answers(payload)
    return normalize_ds160_record(
        {
            "id": str(uuid4()),
            "clientToken": secrets.token_urlsafe(18),
            "status": "submitted",
            "clientName": flattened.get("applicantName", ""),
            "clientEmail": flattened.get("email", "").lower(),
            "clientPhone": flattened.get("primaryPhone", ""),
            "createdAt": now_iso,
            "updatedAt": now_iso,
            "submittedAt": now_iso,
            "appointmentDate": payload.get("appointmentDate"),
            "appointmentTime": payload.get("appointmentTime"),
            "payload": cleaned if (cleaned := clean_ds160_payload(payload)) else {},
        }
    )


def validate_ds160_application(data):
    required = [
        "surname",
        "givenName",
        "nativeFullName",
        "dateOfBirth",
        "birthCountry",
        "nationality",
        "registerNumber",
        "email",
        "primaryPhone",
        "passportNumber",
        "passportIssueDate",
        "passportExpiryDate",
        "tripPurposeCategory",
    ]
    missing = [field for field in required if not data.get(field)]
    if normalize_text(data.get("hasSpecificTravelPlans")).upper() == "ТИЙМ" and not data.get("intendedArrivalDate"):
        missing.append("intendedArrivalDate")
    if missing:
        return f"Missing required fields: {', '.join(missing)}"
    return None


def validate_ds160_invitation(data):
    if len(normalize_text(data.get("clientName"))) < 2:
        return "Client name is required"
    if "@" not in normalize_text(data.get("clientEmail")):
        return "Client email is required"
    return None


def find_ds160_record(record_id=None, client_token=None):
    for record in read_ds160_applications():
        if record_id and record.get("id") == record_id:
            return record
        if client_token and record.get("clientToken") == client_token:
            return record
    return None


def build_finance_entry(payload):
    amount = parse_int(payload.get("amount"))
    bonus_rate = parse_int(payload.get("bonusRate"))
    bonus_amount = round(amount * bonus_rate / 100)
    entry_type = normalize_text(payload.get("type")).lower() or "income"
    return {
        "id": str(uuid4()),
        "createdAt": datetime.now(timezone.utc).isoformat(),
        "employeeName": normalize_text(payload.get("employeeName")),
        "clientName": normalize_text(payload.get("clientName")),
        "category": normalize_text(payload.get("category")),
        "type": entry_type,
        "amount": amount,
        "bonusRate": bonus_rate,
        "bonusAmount": bonus_amount if entry_type == "income" else 0,
        "status": normalize_text(payload.get("status")) or "open",
        "notes": normalize_text(payload.get("notes")),
    }


def validate_finance_entry(data):
    required = ["employeeName", "category", "type"]
    missing = [field for field in required if not data.get(field)]
    if missing or data.get("amount", 0) <= 0:
        return "Employee, category, type, and a positive amount are required"
    return None


def build_booking(payload):
    return {
        "id": str(uuid4()),
        "createdAt": datetime.now(timezone.utc).isoformat(),
        "guestName": normalize_text(payload.get("guestName")),
        "hotelName": normalize_text(payload.get("hotelName")),
        "city": normalize_text(payload.get("city")),
        "checkIn": normalize_text(payload.get("checkIn")),
        "checkOut": normalize_text(payload.get("checkOut")),
        "status": normalize_text(payload.get("status")) or "pending",
        "confirmationCode": normalize_text(payload.get("confirmationCode")),
        "notes": normalize_text(payload.get("notes")),
    }


def validate_booking(data):
    required = ["guestName", "hotelName", "checkIn", "checkOut", "status"]
    missing = [field for field in required if not data.get(field)]
    if missing:
        return f"Missing required fields: {', '.join(missing)}"
    return None


def build_reservation(payload):
    return {
        "id": str(uuid4()),
        "createdAt": datetime.now(timezone.utc).isoformat(),
        "customerName": normalize_text(payload.get("customerName")),
        "serviceType": normalize_text(payload.get("serviceType")),
        "reservationDate": normalize_text(payload.get("reservationDate")),
        "status": normalize_text(payload.get("status")) or "draft",
        "amount": parse_int(payload.get("amount")),
        "notes": normalize_text(payload.get("notes")),
    }


def validate_reservation(data):
    required = ["customerName", "serviceType", "reservationDate", "status"]
    missing = [field for field in required if not data.get(field)]
    if missing:
        return f"Missing required fields: {', '.join(missing)}"
    return None


def normalize_tag_list(value):
    if isinstance(value, str):
        items = [t.strip() for t in value.split(",")]
    elif isinstance(value, list):
        items = [normalize_text(t) for t in value]
    else:
        items = []
    seen = []
    for item in items:
        if item and item not in seen:
            seen.append(item)
    return seen


def build_camp_trip(payload, actor=None):
    company = normalize_company(payload.get("company"))
    trip_type = normalize_text(payload.get("tripType")).lower() or "git"
    if trip_type not in {"fit", "git"}:
        trip_type = "git"
    raw_rates = payload.get("exchangeRates") or {}
    fx = {}
    for k in ("USD", "EUR", "CNY", "JPY", "KRW", "RUB"):
        try:
            v = float(raw_rates.get(k) or 0)
        except (TypeError, ValueError):
            v = 0
        if v > 0:
            fx[k] = v
    try:
        margin_pct = float(payload.get("marginPct") or 0)
    except (TypeError, ValueError):
        margin_pct = 0
    tourists = _normalize_tourist_breakdown(payload.get("tourists"))
    staff = _normalize_staff_breakdown(payload.get("staff"))
    pax_total = parse_int(payload.get("participantCount"))
    if not pax_total:
        pax_total = sum(tourists.values())
    staff_total = parse_int(payload.get("staffCount"))
    if not staff_total:
        staff_total = sum(staff.values())
    currency = (normalize_text(payload.get("currency")) or "MNT").upper()
    if currency not in {"MNT", "USD", "EUR", "CNY", "JPY", "KRW", "RUB"}:
        currency = "MNT"
    return {
        "id": str(uuid4()),
        "serial": next_trip_serial(company),
        "createdAt": now_mongolia().isoformat(),
        "tripName": normalize_text(payload.get("tripName")),
        "tripType": trip_type,
        "groupName": normalize_text(payload.get("groupName")),
        "reservationName": normalize_text(payload.get("reservationName")) or normalize_text(payload.get("tripName")),
        "startDate": normalize_text(payload.get("startDate")),
        "endDate": normalize_text(payload.get("endDate")),
        "totalDays": parse_int(payload.get("totalDays")) or 1,
        "participantCount": pax_total,
        "staffCount": staff_total,
        "tourists": tourists,
        "staff": staff,
        "currency": currency,
        "guideName": normalize_text(payload.get("guideName")),
        "driverName": normalize_text(payload.get("driverName")),
        "cookName": normalize_text(payload.get("cookName")),
        "language": normalize_text(payload.get("language")) or "Other",
        "status": normalize_text(payload.get("status")).lower() or "offer",
        "tags": normalize_tag_list(payload.get("tags")),
        "expenseLines": _normalize_trip_template_lines(payload.get("expenseLines")),
        "marginPct": margin_pct,
        "exchangeRates": fx,
        "inboundCompany": "Unlock Steppe Mongolia",
        "company": company,
        "createdBy": actor_snapshot(actor),
        "updatedAt": "",
        "updatedBy": actor_snapshot(actor),
    }


_TOURIST_KEYS = ("adults", "teen", "child", "infant", "foc")
_STAFF_KEYS = ("guide", "driver", "cook", "tourLeader", "shaman", "shamanAssistant")


def _normalize_tourist_breakdown(value):
    out = {k: 0 for k in _TOURIST_KEYS}
    if isinstance(value, dict):
        for k in _TOURIST_KEYS:
            out[k] = max(0, parse_int(value.get(k)) or 0)
    return out


def _normalize_staff_breakdown(value):
    out = {k: 0 for k in _STAFF_KEYS}
    if isinstance(value, dict):
        for k in _STAFF_KEYS:
            out[k] = max(0, parse_int(value.get(k)) or 0)
    return out


def validate_camp_trip(data):
    required = ["tripName", "startDate"]
    missing = [field for field in required if not data.get(field)]
    if missing:
        return f"Missing required fields: {', '.join(missing)}"
    if data.get("participantCount", 0) <= 0:
        return "Number of participants must be greater than 0"
    if data.get("totalDays", 0) <= 0:
        return "Total days must be greater than 0"
    trip_type = (data.get("tripType") or "").lower()
    if trip_type and trip_type not in {"fit", "git"}:
        return "Trip type must be FIT or GIT"
    return None


def find_camp_trip(trip_id):
    for trip in read_camp_trips():
        if trip["id"] == trip_id:
            return trip
    return None


def find_tourist_group(group_id):
    for record in read_tourist_groups():
        if record["id"] == group_id:
            return record
    return None


def find_tourist(tourist_id):
    for record in read_tourists():
        if record["id"] == tourist_id:
            return record
    return None


def build_tourist_group(payload, actor=None):
    trip_id = normalize_text(payload.get("tripId"))
    trip = find_camp_trip(trip_id)
    # Self-heal any company / serial-prefix mismatch (e.g. trip was once
    # USM, got an S- serial, then switched to DTX). This rewrites the
    # trip serial to the company-correct prefix so the new group
    # inherits the right one.
    trip_serial = reconcile_trip_serial_with_company(trip) if trip else ""
    return {
        "id": str(uuid4()),
        "serial": next_group_serial(trip_serial, trip_id),
        "tripId": trip_id,
        "tripSerial": trip_serial,
        "name": normalize_text(payload.get("name")),
        "leaderName": normalize_text(payload.get("leaderName")),
        "leaderEmail": normalize_text(payload.get("leaderEmail")).lower(),
        "leaderPhone": normalize_text(payload.get("leaderPhone")),
        "leaderNationality": normalize_text(payload.get("leaderNationality")),
        "headcount": parse_int(payload.get("headcount")) or 0,
        "notes": normalize_text(payload.get("notes")),
        "status": (normalize_text(payload.get("status")) or "pending").lower(),
        "createdAt": now_mongolia().isoformat(),
        "createdBy": actor_snapshot(actor),
        "updatedAt": "",
        "updatedBy": actor_snapshot(actor),
    }


def validate_tourist_group(data):
    if not data.get("tripId"):
        return "Trip is required"
    if not data.get("name"):
        return "Group name is required"
    return None


VALID_ROOM_TYPES = {"single", "double", "twin", "triple", "family", "other"}


def normalize_room_type(value):
    v = normalize_text(value).lower()
    return v if v in VALID_ROOM_TYPES else ""


def upper_text(value):
    return normalize_text(value).upper()


def _calc_age_from_dob(dob):
    """Return integer age (full years) from a yyyy-mm-dd dob string, or None."""
    s = (dob or "").strip()
    if not s or len(s) < 10:
        return None
    try:
        y, m, d = int(s[0:4]), int(s[5:7]), int(s[8:10])
    except Exception:
        return None
    today = now_mongolia().date()
    age = today.year - y
    if (today.month, today.day) < (m, d):
        age -= 1
    return age if age >= 0 else None


def _resolve_marketing_status(payload_status, dob, prior_status=None):
    """Children (under 18) are forced to 'child' regardless of input.
    Adults: use payload status, falling back to prior status, then 'standard'.
    If a record was previously 'child' and grew up, default back to 'standard'."""
    raw = (payload_status or "").strip().lower()
    age = _calc_age_from_dob(dob)
    if age is not None and age < 18:
        return "child"
    if raw and raw != "child":
        return raw
    prior = (prior_status or "").strip().lower()
    if prior and prior != "child":
        return prior
    return "standard"


def build_tourist(payload, actor=None, environ=None):
    group_id = normalize_text(payload.get("groupId"))
    group = find_tourist_group(group_id) if group_id else None
    trip_id = (group or {}).get("tripId") or normalize_text(payload.get("tripId"))
    trip = find_camp_trip(trip_id) if trip_id else None
    # Self-heal a stale T-/S- prefix on the trip's serial before deriving
    # the tourist's. Cascades to the group serial too — see
    # reconcile_trip_serial_with_company for details.
    trip_serial = reconcile_trip_serial_with_company(trip) if trip else ""
    # Re-fetch the group so a cascaded serial rewrite is reflected here.
    if group_id:
        group = find_tourist_group(group_id) or group
    group_serial = (group or {}).get("serial") or ""
    # Promo-only contacts (no trip + no group) skip the per-group serial scheme
    # since there's no parent group counter to use; we fall back to a workspace-
    # scoped "PR-####" sequence so they still sort and look identifiable.
    if group_id:
        serial = next_tourist_serial(group_serial, group_id)
    else:
        serial = next_promo_tourist_serial()
    # Trip-less tourists need a company tag of their own so the workspace
    # filter can still scope them; default to the active workspace if known.
    explicit_company = normalize_company(payload.get("company")) if payload.get("company") else ""
    if trip:
        company = normalize_company(trip.get("company"))
    elif explicit_company:
        company = explicit_company
    elif environ is not None:
        company = active_workspace(environ) or DEFAULT_COMPANY
    else:
        company = DEFAULT_COMPANY
    return {
        "id": str(uuid4()),
        "serial": serial,
        "tripId": trip_id,
        "tripSerial": trip_serial,
        "groupId": group_id,
        "groupSerial": group_serial,
        "groupName": (group or {}).get("name") or "",
        "company": company,
        "firstName": upper_text(payload.get("firstName")),
        "lastName": upper_text(payload.get("lastName")),
        "gender": normalize_text(payload.get("gender")).lower(),
        "dob": normalize_text(payload.get("dob")),
        "nationality": upper_text(payload.get("nationality")),
        "passportNumber": upper_text(payload.get("passportNumber")),
        "passportIssueDate": normalize_text(payload.get("passportIssueDate")),
        "passportExpiry": normalize_text(payload.get("passportExpiry")),
        "passportIssuePlace": upper_text(payload.get("passportIssuePlace")),
        "registrationNumber": upper_text(payload.get("registrationNumber")),
        "phone": normalize_text(payload.get("phone")),
        "email": normalize_text(payload.get("email")).lower(),
        "marketingStatus": _resolve_marketing_status(payload.get("marketingStatus"), normalize_text(payload.get("dob"))),
        "tags": normalize_tag_list(payload.get("tags")),
        "roomType": normalize_room_type(payload.get("roomType")),
        "roomCode": normalize_text(payload.get("roomCode")),
        "passportScanPath": normalize_text(payload.get("passportScanPath")),
        "photoPath": normalize_text(payload.get("photoPath")),
        "notes": normalize_text(payload.get("notes")),
        "createdAt": now_mongolia().isoformat(),
        "createdBy": actor_snapshot(actor),
        "updatedAt": "",
        "updatedBy": actor_snapshot(actor),
    }


def validate_tourist(data):
    # Promo-only contacts (no trip, no group) are allowed — they live in the
    # Tourist Directory for marketing reach. Trip-bound tourists still need
    # both ids so the trip dossier stays consistent: if one is set, both must
    # be set. Either way, a name is required.
    has_trip = bool(data.get("tripId"))
    has_group = bool(data.get("groupId"))
    if has_trip and not has_group:
        return "Group is required when a trip is selected"
    if has_group and not has_trip:
        return "Trip is required when a group is selected"
    if not data.get("firstName") and not data.get("lastName"):
        return "Tourist name is required"
    return None


def save_tourist_image(data_url, prefix, tourist_id):
    if not data_url or "base64," not in data_url:
        return ""
    header, encoded = data_url.split("base64,", 1)
    ext = "png"
    if "image/jpeg" in header or "image/jpg" in header:
        ext = "jpg"
    elif "image/webp" in header:
        ext = "webp"
    elif "image/png" not in header:
        return ""
    try:
        raw = base64.b64decode(encoded)
    except Exception:
        return ""
    if len(raw) > 6 * 1024 * 1024:
        return ""
    filename = f"{prefix}-{tourist_id}-{uuid4().hex[:8]}.{ext}"
    path = GENERATED_DIR / filename
    path.write_bytes(raw)
    return f"/generated/{filename}"


def handle_list_tourist_groups(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    params = parse_qs(environ.get("QUERY_STRING", ""))
    trip_id = (params.get("tripId", [""])[0] or "").strip()
    workspace = active_workspace(environ)
    groups = read_tourist_groups()
    if trip_id:
        groups = [g for g in groups if g.get("tripId") == trip_id]
    elif workspace:
        trip_ids = {t["id"] for t in read_camp_trips() if normalize_company(t.get("company")) == workspace}
        groups = [g for g in groups if g.get("tripId") in trip_ids]
    groups.sort(key=lambda g: (g.get("tripSerial") or "", g.get("serial") or ""))
    return json_response(start_response, "200 OK", {"entries": groups})


def handle_create_tourist_group(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    record = build_tourist_group(payload, actor)
    error = validate_tourist_group(record)
    if error:
        return json_response(start_response, "400 Bad Request", {"error": error})
    records = read_tourist_groups()
    records.insert(0, record)
    write_tourist_groups(records)
    try:
        log_notification(
            "group.created",
            actor,
            "New group added",
            detail=f"{record.get('serial', '')} {record.get('name', '')}".strip(),
            meta={"id": record["id"], "tripId": record.get("tripId")},
        )
    except Exception:
        pass
    return json_response(start_response, "201 Created", {"ok": True, "entry": record})


def handle_update_tourist_group(environ, start_response, group_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    records = read_tourist_groups()
    for index, record in enumerate(records):
        if record.get("id") != group_id:
            continue
        merged = {**record}
        for key in ["name", "leaderName", "leaderEmail", "leaderPhone", "leaderNationality", "notes"]:
            if key in payload:
                merged[key] = normalize_text(payload.get(key))
        if "headcount" in payload:
            merged["headcount"] = parse_int(payload.get("headcount")) or 0
        if "status" in payload:
            merged["status"] = (normalize_text(payload.get("status")) or "pending").lower()
        error = validate_tourist_group(merged)
        if error:
            return json_response(start_response, "400 Bad Request", {"error": error})
        merged["updatedAt"] = now_mongolia().isoformat()
        merged["updatedBy"] = actor_snapshot(actor)
        records[index] = merged
        write_tourist_groups(records)
        # Cascade groupName update to tourists
        if "name" in payload:
            tourists = read_tourists()
            changed = False
            for t in tourists:
                if t.get("groupId") == group_id and t.get("groupName") != merged["name"]:
                    t["groupName"] = merged["name"]
                    changed = True
            if changed:
                write_tourists(tourists)
        try:
            log_notification(
                "group.updated",
                actor,
                "Group updated",
                detail=f"{merged.get('serial', '')} {merged.get('name', '')}".strip(),
                meta={"id": merged.get("id"), "tripId": merged.get("tripId")},
            )
        except Exception:
            pass
        return json_response(start_response, "200 OK", {"ok": True, "entry": merged})
    return json_response(start_response, "404 Not Found", {"error": "Group not found"})


def handle_delete_tourist_group(environ, start_response, group_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    records = read_tourist_groups()
    remaining = [r for r in records if r.get("id") != group_id]
    if len(remaining) == len(records):
        return json_response(start_response, "404 Not Found", {"error": "Group not found"})
    write_tourist_groups(remaining)
    # Also delete tourists in that group
    tourists = read_tourists()
    new_tourists = [t for t in tourists if t.get("groupId") != group_id]
    if len(new_tourists) != len(tourists):
        write_tourists(new_tourists)
    return json_response(start_response, "200 OK", {"ok": True, "deletedId": group_id})


def handle_list_tourists(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    params = parse_qs(environ.get("QUERY_STRING", ""))
    trip_id = (params.get("tripId", [""])[0] or "").strip()
    group_id = (params.get("groupId", [""])[0] or "").strip()
    workspace = active_workspace(environ)
    tourists = read_tourists()
    if group_id:
        tourists = [t for t in tourists if t.get("groupId") == group_id]
    elif trip_id:
        tourists = [t for t in tourists if t.get("tripId") == trip_id]
    elif workspace:
        trip_ids = {t["id"] for t in read_camp_trips() if normalize_company(t.get("company")) == workspace}
        # Include trip-bound tourists whose trip is in the active workspace,
        # plus promo-only contacts (no tripId) tagged with the same workspace
        # company directly.
        tourists = [
            t for t in tourists
            if (t.get("tripId") in trip_ids)
            or (not t.get("tripId") and normalize_company(t.get("company")) == workspace)
        ]
    tourists.sort(key=lambda t: (t.get("tripSerial") or "", t.get("serial") or ""))
    return json_response(start_response, "200 OK", {"entries": tourists})


def handle_promo_email(environ, start_response):
    """POST /api/tourists/promo-email — admin-driven marketing send. Body:
    {touristIds: [...], subject, body, workspace?}. Sends an individual email
    per tourist (so they can't see each other), filtering out anyone with
    marketingStatus="do_not_contact" or no email on file."""
    actor = require_login(environ, start_response)
    if not actor:
        return []
    try:
        body_raw = environ["wsgi.input"].read(int(environ.get("CONTENT_LENGTH") or "0"))
        data = json.loads(body_raw)
    except Exception:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid JSON"})
    ids = data.get("touristIds") or []
    subject = (data.get("subject") or "").strip()
    body_text = (data.get("body") or "").strip()
    body_html_in = (data.get("bodyHtml") or "").strip()
    workspace = (data.get("workspace") or "").strip().upper()
    if not isinstance(ids, list) or not ids:
        return json_response(start_response, "400 Bad Request", {"error": "touristIds (non-empty list) is required"})
    if not subject or not body_text:
        return json_response(start_response, "400 Bad Request", {"error": "subject and body are required"})

    tourists_by_id = {t.get("id"): t for t in read_tourists()}
    targets = []
    skipped_no_email = 0
    skipped_optout = 0
    skipped_child = 0
    for tid in ids:
        t = tourists_by_id.get(tid)
        if not t:
            continue
        status = (t.get("marketingStatus") or "").lower()
        if status == "do_not_contact":
            skipped_optout += 1
            continue
        # Defense in depth: never email children even if frontend sent the id.
        age = _calc_age_from_dob(t.get("dob"))
        if status == "child" or (age is not None and age < 18):
            skipped_child += 1
            continue
        email = (t.get("email") or "").strip()
        if not email or "@" not in email:
            skipped_no_email += 1
            continue
        targets.append((t, email))

    company_name = "Unlock Steppe Mongolia" if workspace == "USM" else "Дэлхий Трэвел Икс"
    sent = 0
    failures = []
    for t, email in targets:
        first = t.get("firstName") or ""
        greet = f"Сайн байна уу{(', ' + first.title()) if first else ''},"
        personalized = f"{greet}\n\n{body_text}"
        send_args = {
            "to": email,
            "subject": subject,
            "body": personalized,
            "_company_name": company_name,
        }
        if body_html_in:
            # Prepend the personalized greeting as a paragraph above the
            # rich-editor HTML, then let _tool_send_email use the override
            # verbatim (instead of escaping the plain body).
            greet_html = "<p>" + html.escape(greet) + "</p>"
            send_args["_body_html_override"] = greet_html + body_html_in
        result = _tool_send_email(send_args, actor)
        if result.get("error"):
            failures.append(f"{email}: {result['error'][:80]}")
        else:
            sent += 1

    return json_response(start_response, "200 OK", {
        "ok": True,
        "sent": sent,
        "skippedNoEmail": skipped_no_email,
        "skippedOptOut": skipped_optout,
        "skippedChild": skipped_child,
        "failures": failures[:20],
    })


def handle_create_tourist(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    record = build_tourist(payload, actor, environ=environ)
    error = validate_tourist(record)
    if error:
        return json_response(start_response, "400 Bad Request", {"error": error})
    if payload.get("passportScanData"):
        record["passportScanPath"] = save_tourist_image(payload["passportScanData"], "tourist-passport", record["id"])
    if payload.get("photoData"):
        record["photoPath"] = save_tourist_image(payload["photoData"], "tourist-photo", record["id"])
    passport_token = (payload.get("passportFileToken") or "").strip()
    if passport_token:
        doc_name = (record.get("firstName") or record.get("lastName") or "passport").strip()
        try:
            doc = consume_passport_token(passport_token, record.get("tripId"), record["id"], doc_name, actor)
            if doc:
                record["passportDocumentId"] = doc["id"]
        except Exception:
            pass
    # When the participant is being copied from an existing tourist (the
    # "Existing tourist" picker in the add-participant flow), pull their
    # passport scan over so the manager doesn't have to re-upload it.
    copy_from_id = normalize_text(payload.get("copyFromTouristId"))
    if copy_from_id and not record.get("passportDocumentId"):
        try:
            doc = clone_passport_from_tourist(
                copy_from_id,
                record.get("tripId"),
                record["id"],
                f"{record.get('lastName', '')} {record.get('firstName', '')}".strip() or "passport",
                actor,
            )
            if doc:
                record["passportDocumentId"] = doc["id"]
                # Mirror the source's passportScanPath onto the new record
                # so /api/tourists rows reflect that a scan is on file.
                source = next((t for t in read_tourists() if t.get("id") == copy_from_id), None)
                if source and source.get("passportScanPath"):
                    record["passportScanPath"] = source.get("passportScanPath")
        except Exception as exc:
            print(f"[clone-passport] failed: {exc}", flush=True)
    records = read_tourists()
    records.insert(0, record)
    write_tourists(records)
    try:
        log_notification(
            "tourist.created",
            actor,
            "New tourist added",
            detail=f"{record.get('serial', '')} {record.get('lastName', '')} {record.get('firstName', '')}".strip(),
            meta={"id": record["id"], "tripId": record.get("tripId"), "groupId": record.get("groupId")},
        )
    except Exception:
        pass
    return json_response(start_response, "201 Created", {"ok": True, "entry": record})


def handle_update_tourist(environ, start_response, tourist_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    records = read_tourists()
    for index, record in enumerate(records):
        if record.get("id") != tourist_id:
            continue
        merged = {**record}
        upper_keys = {"firstName", "lastName", "nationality", "passportNumber", "passportIssuePlace", "registrationNumber"}
        for key in ["firstName", "lastName", "gender", "dob", "nationality", "passportNumber", "passportIssueDate", "passportExpiry", "passportIssuePlace", "registrationNumber", "phone", "email", "notes", "roomCode", "marketingStatus"]:
            if key in payload:
                value = normalize_text(payload.get(key))
                if key == "email":
                    value = value.lower()
                if key == "gender":
                    value = value.lower()
                if key in upper_keys:
                    value = value.upper()
                merged[key] = value
        if "roomType" in payload:
            merged["roomType"] = normalize_room_type(payload.get("roomType"))
        if "tags" in payload:
            merged["tags"] = normalize_tag_list(payload.get("tags"))
        if "orderIndex" in payload:
            merged["orderIndex"] = parse_int(payload.get("orderIndex")) or 0
        merged["marketingStatus"] = _resolve_marketing_status(
            merged.get("marketingStatus"),
            merged.get("dob"),
            prior_status=record.get("marketingStatus"),
        )
        if payload.get("passportScanData"):
            merged["passportScanPath"] = save_tourist_image(payload["passportScanData"], "tourist-passport", tourist_id)
        if payload.get("photoData"):
            merged["photoPath"] = save_tourist_image(payload["photoData"], "tourist-photo", tourist_id)
        if payload.get("removePassportScan"):
            merged["passportScanPath"] = ""
        if payload.get("removePhoto"):
            merged["photoPath"] = ""
        error = validate_tourist(merged)
        if error:
            return json_response(start_response, "400 Bad Request", {"error": error})
        merged["updatedAt"] = now_mongolia().isoformat()
        merged["updatedBy"] = actor_snapshot(actor)
        records[index] = merged
        write_tourists(records)
        return json_response(start_response, "200 OK", {"ok": True, "entry": merged})
    return json_response(start_response, "404 Not Found", {"error": "Tourist not found"})


def handle_delete_tourist(environ, start_response, tourist_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    records = read_tourists()
    deleted = next((r for r in records if r.get("id") == tourist_id), None)
    if not deleted:
        return json_response(start_response, "404 Not Found", {"error": "Tourist not found"})
    remaining = [r for r in records if r.get("id") != tourist_id]
    write_tourists(remaining)
    # Mark — but don't delete — any passport/document the tourist had attached
    # to a trip. The Documents page surfaces these as "Removed from <trip>"
    # so the file isn't lost when a participant gets removed.
    try:
        trips = read_camp_trips()
        deleted_at = datetime.now(timezone.utc).isoformat()
        deleted_name = (
            f"{normalize_text(deleted.get('lastName'))} {normalize_text(deleted.get('firstName'))}".strip()
            or normalize_text(deleted.get("touristName"))
        )
        changed = False
        for trip in trips:
            for doc in (trip.get("documents") or []):
                if doc.get("touristId") != tourist_id:
                    continue
                if doc.get("touristRemovedAt"):
                    continue
                doc["touristRemovedAt"] = deleted_at
                doc["touristRemovedTripName"] = (
                    f"{normalize_text(trip.get('serial'))} · {normalize_text(trip.get('tripName'))}".strip(" ·")
                )
                doc["touristRemovedName"] = deleted_name
                changed = True
        if changed:
            write_camp_trips(trips)
    except Exception as exc:
        print(f"[tourist-delete] document mark failed: {exc}", flush=True)
    return json_response(start_response, "200 OK", {"ok": True, "deletedId": tourist_id})


def handle_list_documents(environ, start_response):
    """GET /api/documents — flatten every document attached to every trip in
    the active workspace, returning enriched rows ({ tripId, tripSerial,
    tripName, ...doc }) so the global Documents page can show context and
    filter without N+1 fetches."""
    if not require_login(environ, start_response):
        return []
    workspace = active_workspace(environ)
    # Index tourists by id so we can split touristName into first + last for
    # the Documents page (Bataa wants both columns separately, with passport-
    # case spelling preserved).
    tourist_by_id = {}
    for t in read_tourists():
        tid = t.get("id")
        if tid:
            tourist_by_id[tid] = t
    rows = []
    for trip in read_camp_trips():
        if workspace and normalize_company(trip.get("company")) != workspace:
            continue
        for doc in (trip.get("documents") or []):
            tourist = tourist_by_id.get(doc.get("touristId") or "")
            rows.append({
                **doc,
                "tripId": trip.get("id"),
                "tripSerial": trip.get("serial") or "",
                "tripName": trip.get("tripName") or "",
                "destinations": list(trip.get("tags") or []),
                "touristFirstName": (tourist.get("firstName") if tourist else "") or "",
                "touristLastName": (tourist.get("lastName") if tourist else "") or "",
            })
    rows.sort(key=lambda d: d.get("uploadedAt") or "", reverse=True)
    return json_response(start_response, "200 OK", {"entries": rows})


def handle_list_invoices(environ, start_response):
    if not require_login(environ, start_response):
        return []
    qs = parse_qs(environ.get("QUERY_STRING", ""))
    trip_id = (qs.get("tripId") or [""])[0].strip()
    group_id = (qs.get("groupId") or [""])[0].strip()
    records = read_invoices()
    if trip_id:
        records = [r for r in records if r.get("tripId") == trip_id]
    if group_id:
        records = [r for r in records if r.get("groupId") == group_id]
    # Scope by active workspace so DTX admins don't see USM invoices and vice
    # versa. Invoices don't carry their own `company` field — they're tied to
    # a trip, and the trip carries the workspace. So filter via the trip set.
    workspace = active_workspace(environ)
    if workspace and not trip_id:
        ws_trip_ids = {t["id"] for t in read_camp_trips() if normalize_company(t.get("company")) == workspace}
        records = [r for r in records if r.get("tripId") in ws_trip_ids]
    records.sort(key=lambda r: r.get("createdAt") or "", reverse=True)
    # Enrich each invoice with the parent trip's language so the
    # invoice-view page can render labels in the right language without
    # an extra round-trip.
    trip_lang_by_id = {t.get("id"): normalize_text(t.get("language")) for t in read_camp_trips()}
    enriched = []
    for r in records:
        copy = dict(r)
        copy["tripLanguage"] = trip_lang_by_id.get(r.get("tripId"), "") or ""
        enriched.append(copy)
    return json_response(start_response, "200 OK", {"entries": enriched})


def _invoice_notification_detail(record):
    serial = normalize_text(record.get("invoiceSerial")) or normalize_text(record.get("serial"))
    name = normalize_text(record.get("clientName")) or normalize_text(record.get("touristName"))
    parts = [p for p in [serial, name] if p]
    return " · ".join(parts) if parts else "Invoice"


def _log_invoice(kind, title, record, actor):
    try:
        log_notification(
            kind,
            actor,
            title,
            detail=_invoice_notification_detail(record),
            meta={
                "id": record.get("id"),
                "tripId": record.get("tripId"),
                "groupId": record.get("groupId"),
            },
        )
    except Exception:
        pass


def handle_create_invoice(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    record = build_invoice(payload, actor)
    error = validate_invoice(record)
    if error:
        return json_response(start_response, "400 Bad Request", {"error": error})
    records = read_invoices()
    records.append(record)
    write_invoices(records)
    _log_invoice("invoice.created", "New invoice added", record, actor)
    return json_response(start_response, "201 Created", {"ok": True, "entry": record})


def handle_update_invoice(environ, start_response, invoice_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    records = read_invoices()
    for index, record in enumerate(records):
        if record.get("id") != invoice_id:
            continue
        rebuilt = build_invoice({**record, **payload}, actor)
        rebuilt["id"] = record["id"]
        rebuilt["serial"] = record["serial"]
        rebuilt["createdAt"] = record.get("createdAt", now_mongolia().isoformat())
        rebuilt["createdBy"] = record.get("createdBy", actor_snapshot(actor))
        rebuilt["status"] = record.get("status", "draft")
        rebuilt["updatedAt"] = now_mongolia().isoformat()
        rebuilt["updatedBy"] = actor_snapshot(actor)
        error = validate_invoice(rebuilt)
        if error:
            return json_response(start_response, "400 Bad Request", {"error": error})
        records[index] = rebuilt
        write_invoices(records)
        _log_invoice("invoice.updated", "Invoice updated", rebuilt, actor)
        return json_response(start_response, "200 OK", {"ok": True, "entry": rebuilt})
    return json_response(start_response, "404 Not Found", {"error": "Invoice not found"})


def handle_delete_invoice(environ, start_response, invoice_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    records = read_invoices()
    deleted = next((r for r in records if r.get("id") == invoice_id), None)
    remaining = [r for r in records if r.get("id") != invoice_id]
    if len(remaining) == len(records):
        return json_response(start_response, "404 Not Found", {"error": "Invoice not found"})
    write_invoices(remaining)
    if deleted:
        _log_invoice("invoice.deleted", "Invoice deleted", deleted, actor)
    return json_response(start_response, "200 OK", {"ok": True, "deletedId": invoice_id})


def handle_publish_invoice(environ, start_response, invoice_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    records = read_invoices()
    for index, record in enumerate(records):
        if record.get("id") != invoice_id:
            continue
        record["status"] = "published"
        record["publishedAt"] = now_mongolia().isoformat()
        record["updatedAt"] = now_mongolia().isoformat()
        record["updatedBy"] = actor_snapshot(actor)
        records[index] = record
        write_invoices(records)
        _log_invoice("invoice.published", "Invoice published", record, actor)
        return json_response(start_response, "200 OK", {"ok": True, "entry": record})
    return json_response(start_response, "404 Not Found", {"error": "Invoice not found"})


def handle_invoice_payment(environ, start_response, invoice_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ) or {}
    inst_index = payload.get("installmentIndex")
    if inst_index is None:
        return json_response(start_response, "400 Bad Request", {"error": "installmentIndex required"})
    try:
        inst_index = int(inst_index)
    except Exception:
        return json_response(start_response, "400 Bad Request", {"error": "installmentIndex must be integer"})
    new_status = normalize_text(payload.get("status")) or "paid"
    paid_date = normalize_text(payload.get("paidDate"))
    note = normalize_text(payload.get("note"))
    bank_account_id = normalize_text(payload.get("bankAccountId"))
    is_admin = (actor or {}).get("role") == "admin"
    records = read_invoices()
    for index, record in enumerate(records):
        if record.get("id") != invoice_id:
            continue
        installments = record.get("installments") or []
        if inst_index < 0 or inst_index >= len(installments):
            return json_response(start_response, "400 Bad Request", {"error": "installment out of range"})
        target = installments[inst_index]
        # Lock once paid: only admins can edit a registered payment.
        # Non-admins clicking "Register payment" on something already paid
        # is a UI bug (the button hides when paid), so reject loudly.
        was_paid = (target.get("status") or "").lower() in ("paid", "confirmed")
        if was_paid and not is_admin:
            return json_response(
                start_response,
                "403 Forbidden",
                {"error": "This payment is already registered. Ask an admin to edit it."},
            )
        target["status"] = new_status
        if paid_date:
            target["paidDate"] = paid_date
        if "note" in payload:
            target["note"] = note
        # paidAmount is what the client actually transferred. Defaults to
        # the expected installment amount; gets recorded as-typed if the
        # customer sent a different number (over- or under-paid).
        if "paidAmount" in payload:
            try:
                target["paidAmount"] = float(payload["paidAmount"])
            except (TypeError, ValueError):
                target["paidAmount"] = float(target.get("amount") or 0)
        if "bankAccountId" in payload:
            target["bankAccountId"] = bank_account_id
            # Snapshot the bank info at time of payment so the record stays
            # correct even if the bank account is renamed/deleted later.
            bank_snapshot = None
            if bank_account_id:
                for b in read_settings().get("bankAccounts") or []:
                    if b.get("id") == bank_account_id:
                        bank_snapshot = b
                        break
            target["bankAccount"] = bank_snapshot or payload.get("bankAccount") or {}
        # Track who registered/edited and when, so the audit trail makes the
        # admin-edit policy meaningful.
        target["paidBy"] = actor_snapshot(actor)
        target["paidAt"] = now_mongolia().isoformat()
        installments[inst_index] = target
        record["installments"] = installments
        record["updatedAt"] = now_mongolia().isoformat()
        record["updatedBy"] = actor_snapshot(actor)
        records[index] = record
        write_invoices(records)
        return json_response(start_response, "200 OK", {"ok": True, "entry": record})
    return json_response(start_response, "404 Not Found", {"error": "Invoice not found"})


# Payment-request flow:
#   1) any logged-in user opens an invoice and submits a "Register payment"
#      form. Instead of mutating the invoice immediately, this endpoint
#      stores a pending payment-request record.
#   2) accountants/admins see the request in the ₮ queue.
#   3) accountant approves → register the payment on the invoice + attach
#      the proof document to the trip's "Paid documents" category.
def _payment_request_workspace(invoice):
    """Workspace assigned to a payment_request derived from an invoice.

    Looks up the invoice's trip and returns its company (DTX/USM).
    Falls back to the serial prefix only if the trip can't be found
    (e.g. legacy invoice without a tripId, or trip deleted).
    """
    trip_id = (invoice or {}).get("tripId") or ""
    if trip_id:
        trip = next((t for t in read_camp_trips() if t.get("id") == trip_id), None)
        if trip:
            return normalize_company(trip.get("company")) or "DTX"
    serial = (invoice.get("serial") or "").upper()
    if serial.startswith("S-"):
        return "USM"
    return "DTX"


def _can_approve_payment_request(user):
    role = (user or {}).get("role", "").lower()
    return role in ("admin", "accountant")


# "Paid" variants used across camp / flight / transfer reservations.
# Only admin / accountant can flip to one of these — managers raise
# a request and the accountant confirms (mirrors the invoice flow).
PAID_PAYMENT_STATUSES = {"paid", "paid_100", "paid_deposit"}


def _check_payment_status_guard(actor, payload, existing=None):
    """Return an error-response tuple (status, body) if the payload
    is asking to set paymentStatus to a paid variant and the actor is
    not an admin/accountant. Returns None when the change is allowed
    (either the actor is admin/accountant, or the status isn't a paid
    variant, or the status didn't actually change)."""
    incoming = (payload.get("paymentStatus") or "").strip().lower() if isinstance(payload, dict) else ""
    if not incoming or incoming not in PAID_PAYMENT_STATUSES:
        return None
    prior = ((existing or {}).get("paymentStatus") or "").strip().lower()
    if incoming == prior:
        return None  # no change — leave it alone
    if _can_approve_payment_request(actor):
        return None
    return (
        "403 Forbidden",
        {"error": "Only admin or accountant can mark a payment as paid. Ask your accountant to confirm."},
    )


def _attach_vendor_invoice(upload, actor):
    """Store the manager-uploaded vendor invoice (or quote/contract)
    under data/trip-uploads/_vendor/. Returned metadata is stashed on
    the payment_request so the Accountant + trip P&L pages can link
    to it. Returns None when no file or invalid extension."""
    if not upload:
        return None
    original_name = upload.get("filename") or "vendor-invoice"
    ext = Path(original_name).suffix.lower()
    if ext not in ALLOWED_UPLOAD_EXTENSIONS:
        return None
    data = upload.get("data") or b""
    if not data:
        return None
    if len(data) > MAX_UPLOAD_BYTES:
        return None
    ensure_data_store()
    doc_id = str(uuid4())
    out_dir = TRIP_UPLOADS_DIR / "_vendor"
    out_dir.mkdir(parents=True, exist_ok=True)
    stored_name = doc_id + ext
    (out_dir / stored_name).write_bytes(data)
    return {
        "id": doc_id,
        "storedName": stored_name,
        "originalName": original_name,
        "mimeType": upload.get("content_type") or "application/octet-stream",
        "size": len(data),
        "uploadedAt": now_mongolia().isoformat(),
        "uploadedBy": actor_snapshot(actor),
    }


def handle_create_payment_request(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    # The expense modal sends multipart when a vendor invoice is
    # attached; the legacy invoice "Request payment" flow stays JSON.
    content_type = environ.get("CONTENT_TYPE", "")
    if "multipart/form-data" in content_type:
        fields, files = parse_multipart(environ)
        payload = {k: v for k, v in fields.items()}
        vendor_upload = files.get("vendorInvoice")
    else:
        payload = collect_json(environ) or {}
        vendor_upload = None
    direction = (normalize_text(payload.get("direction")) or "incoming").lower()
    if direction not in ("incoming", "outgoing"):
        direction = "incoming"

    if direction == "incoming":
        # Existing flow: client paid us → record against an invoice +
        # installment.
        invoice_id = normalize_text(payload.get("invoiceId"))
        if not invoice_id:
            return json_response(start_response, "400 Bad Request", {"error": "invoiceId required"})
        inst_index_raw = payload.get("installmentIndex")
        try:
            inst_index = int(inst_index_raw)
        except Exception:
            return json_response(start_response, "400 Bad Request", {"error": "installmentIndex required"})
        invoices = read_invoices()
        invoice = next((r for r in invoices if r.get("id") == invoice_id), None)
        if not invoice:
            return json_response(start_response, "404 Not Found", {"error": "Invoice not found"})
        installments = invoice.get("installments") or []
        if inst_index < 0 or inst_index >= len(installments):
            return json_response(start_response, "400 Bad Request", {"error": "installment out of range"})
        # Reject if the installment is already paid — the manager should use
        # the receipt-upload flow instead, not register a fresh payment.
        inst_status = (installments[inst_index].get("status") or "").lower()
        if inst_status in ("paid", "confirmed"):
            return json_response(start_response, "409 Conflict", {"error": "Installment is already marked paid."})
        # Reject duplicate pending requests on the same (invoice, installment).
        # Per-installment lock — different installments are still independent.
        existing_requests = read_payment_requests()
        dup = next(
            (
                r for r in existing_requests
                if r.get("invoiceId") == invoice_id
                and r.get("installmentIndex") == inst_index
                and (r.get("status") or "") == "pending"
            ),
            None,
        )
        if dup:
            return json_response(start_response, "409 Conflict", {"error": "A payment request is already pending for this installment."})
        paid_amount_raw = payload.get("paidAmount")
        try:
            paid_amount = float(paid_amount_raw)
        except (TypeError, ValueError):
            paid_amount = float(installments[inst_index].get("amount") or 0)
        if paid_amount <= 0:
            return json_response(start_response, "400 Bad Request", {"error": "paidAmount must be greater than zero."})
        # Look up the group name so the accountant's approve modal can
        # show "Trip · Group · Date" without an extra round-trip. Falls
        # back to "" if the invoice has no group attached.
        invoice_group_id = invoice.get("groupId") or ""
        invoice_group_name = ""
        if invoice_group_id:
            try:
                groups_data = json.loads(GROUPS_FILE.read_text("utf-8")) if GROUPS_FILE.exists() else []
                grp = next((g for g in groups_data if g.get("id") == invoice_group_id), None)
                if grp:
                    invoice_group_name = grp.get("name") or grp.get("groupName") or ""
            except Exception:
                pass
        request = {
            "id": uuid4().hex,
            "direction": "incoming",
            "scope": "trip",
            "category": "Invoice",
            "invoiceId": invoice_id,
            "invoiceSerial": invoice.get("serial") or "",
            "tripId": invoice.get("tripId") or "",
            "groupId": invoice_group_id,
            "groupName": invoice_group_name,
            "installmentIndex": inst_index,
            "installmentDescription": installments[inst_index].get("description") or "",
            "paidDate": normalize_text(payload.get("paidDate")),
            "paidAmount": paid_amount,
            "currency": (invoice.get("currency") or "MNT").upper(),
            "bankAccountId": normalize_text(payload.get("bankAccountId")),
            "note": normalize_text(payload.get("note")),
            "payerName": invoice.get("payerName") or "",
            "payeeName": "",
            "dueDate": "",
            "referenceDocId": "",
            "workspace": _payment_request_workspace(invoice),
            "status": "pending",
            "requestedBy": actor_snapshot(actor),
            "requestedAt": now_mongolia().isoformat(),
            "approvedBy": None,
            "approvedAt": "",
            "rejectedBy": None,
            "rejectedAt": "",
            "rejectReason": "",
            "paidDocumentId": "",
        }
    else:
        # New outgoing flow: manager asks the company to pay a vendor /
        # employee / overhead bill. Trip is optional (office expenses).
        category = normalize_text(payload.get("category")) or "Other"
        payee = normalize_text(payload.get("payeeName"))
        if not payee:
            return json_response(start_response, "400 Bad Request", {"error": "Payee name is required."})
        try:
            paid_amount = float(payload.get("paidAmount") or 0)
        except (TypeError, ValueError):
            paid_amount = 0.0
        if paid_amount <= 0:
            return json_response(start_response, "400 Bad Request", {"error": "Amount must be greater than zero."})
        scope = (normalize_text(payload.get("scope")) or "office").lower()
        if scope not in ("trip", "office", "other"):
            scope = "other"
        trip_id = normalize_text(payload.get("tripId"))
        trip_name = ""
        if trip_id:
            trip = next((t for t in read_camp_trips() if t.get("id") == trip_id), None)
            if trip:
                trip_name = trip.get("tripName") or ""
        workspace_field = (active_workspace(environ) or "").upper() or "DTX"
        # Persist the category on the workspace settings so it shows up
        # next time someone opens the dropdown.
        try:
            settings = read_settings()
            cats = list(settings.get("expenseCategories") or [])
            if category and category.lower() not in {c.lower() for c in cats}:
                cats.append(category)
                payees = list(settings.get("expensePayees") or [])
                if payee and payee.lower() not in {p.lower() for p in payees}:
                    payees.append(payee)
                write_settings({**settings, "expenseCategories": cats, "expensePayees": payees})
            elif payee:
                payees = list(settings.get("expensePayees") or [])
                if payee.lower() not in {p.lower() for p in payees}:
                    payees.append(payee)
                    write_settings({**settings, "expensePayees": payees})
        except Exception:
            pass
        vendor_meta = _attach_vendor_invoice(vendor_upload, actor) if vendor_upload else None
        # Optional link to a flight / transfer / camp reservation. When
        # set, approving this payment_request flips the underlying
        # reservation's paymentStatus to "paid" (mirrors how invoice
        # approval flips the installment status).
        record_type = (normalize_text(payload.get("recordType")) or "").lower()
        # camp_group is a fan-out type: recordId is "<tripId>::<campName>"
        # and approval flips every camp_reservation whose tripId + campName
        # match. Mirrors the trip-detail "Camp payments" card flow.
        valid_record_types = {"flight_reservation", "transfer_reservation", "camp_reservation", "camp_group"}
        record_id = normalize_text(payload.get("recordId"))
        if record_type and record_type not in valid_record_types:
            return json_response(start_response, "400 Bad Request", {"error": f"Unknown recordType: {record_type}"})
        if record_type and not record_id:
            return json_response(start_response, "400 Bad Request", {"error": "recordId is required when recordType is set."})
        # Reject duplicate pending requests on the same reservation —
        # spam-click guard, mirrors the per-installment lock used for
        # invoices.
        # Optional stage qualifier for camp_group: "deposit" or "balance"
        # so each payment stage has its own independent request lock and
        # the approval flips only that stage.
        stage = (normalize_text(payload.get("stage")) or "").lower()
        if stage and stage not in ("deposit", "balance", "third", "fourth"):
            return json_response(start_response, "400 Bad Request", {"error": "Invalid stage (deposit | balance | third | fourth)."})
        if record_type and record_id:
            existing = next(
                (
                    r for r in read_payment_requests()
                    if r.get("recordType") == record_type
                    and r.get("recordId") == record_id
                    and (r.get("stage") or "") == stage
                    and (r.get("status") or "") == "pending"
                ),
                None,
            )
            if existing:
                return json_response(start_response, "409 Conflict", {"error": "A payment request is already pending for this reservation."})
        request = {
            "id": uuid4().hex,
            "direction": "outgoing",
            "scope": scope,
            "category": category,
            "invoiceId": "",
            "invoiceSerial": "",
            "tripId": trip_id,
            "tripName": trip_name,
            "recordType": record_type,
            "recordId": record_id,
            "stage": stage,
            "installmentIndex": -1,
            "installmentDescription": category,
            "paidDate": "",
            "dueDate": normalize_text(payload.get("dueDate")),
            "paidAmount": paid_amount,
            "currency": (normalize_text(payload.get("currency")) or "MNT").upper(),
            "bankAccountId": normalize_text(payload.get("bankAccountId")),
            "note": normalize_text(payload.get("note")),
            "payerName": "",
            "payeeName": payee,
            "referenceDocId": (vendor_meta or {}).get("id", "") or normalize_text(payload.get("referenceDocId")),
            "vendorInvoiceMeta": vendor_meta,
            "workspace": workspace_field,
            "status": "pending",
            "requestedBy": actor_snapshot(actor),
            "requestedAt": now_mongolia().isoformat(),
            "approvedBy": None,
            "approvedAt": "",
            "rejectedBy": None,
            "rejectedAt": "",
            "rejectReason": "",
            "paidDocumentId": "",
        }
    records = read_payment_requests()
    records.insert(0, request)
    write_payment_requests(records)

    if direction == "incoming":
        notif_title = f"Payment request · {request.get('invoiceSerial') or ''}"
        notif_detail = (
            f"{actor.get('fullName') or actor.get('email')} requested payment for "
            f"{request['installmentDescription'] or 'installment'} "
            f"({_fmt_money_ccy(paid_amount, request['currency'])})"
        )
    else:
        notif_title = f"Expense request · {request.get('category') or 'Other'}"
        notif_detail = (
            f"{actor.get('fullName') or actor.get('email')} asked to pay "
            f"{request.get('payeeName') or ''} "
            f"{_fmt_money_ccy(paid_amount, request['currency'])} for "
            f"{request.get('category') or 'an expense'}"
        )
    log_notification(
        kind="payment_request.created",
        actor=actor,
        title=notif_title,
        detail=notif_detail,
        meta={
            "paymentRequestId": request["id"],
            "invoiceId": request.get("invoiceId") or "",
            "tripId": request.get("tripId") or "",
            "workspace": request.get("workspace") or "",
            "direction": request.get("direction"),
        },
    )
    return json_response(start_response, "200 OK", {"ok": True, "entry": request})


def handle_list_payment_requests(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    qs = parse_qs(environ.get("QUERY_STRING", ""))
    status_filter = (qs.get("status", [""])[0] or "").lower()
    workspace_filter = (qs.get("workspace", [""])[0] or "").upper()
    invoice_filter = (qs.get("invoiceId", [""])[0] or "").strip()
    mine_only = (qs.get("mine", [""])[0] or "").strip() in ("1", "true", "yes")
    if not workspace_filter:
        workspace_filter = active_workspace(environ) or ""
    records = read_payment_requests()
    filtered = []
    actor_id = (actor or {}).get("id") or ""
    # Build a tripId → documents lookup once so each row's
    # paidDocumentUrl can be resolved without reading every trip's
    # documents inline. Skipped when there are no trip-bound requests.
    trips_by_id = None
    # Some legacy payment_requests have a stale `workspace` field
    # (an older heuristic stamped every "S-" serial as USM, even for
    # DTX trips). When the request points at a trip, trust the
    # trip's `company` field as the source of truth and ignore the
    # stored workspace for filtering purposes.
    def _workspace_of(record):
        trip_id = record.get("tripId") or ""
        if trip_id:
            nonlocal trips_by_id
            if trips_by_id is None:
                trips_by_id = {t.get("id"): t for t in read_camp_trips()}
            trip = trips_by_id.get(trip_id)
            if trip:
                return (normalize_company(trip.get("company")) or "DTX").upper()
        return (record.get("workspace") or "").upper()
    for r in records:
        if status_filter and (r.get("status") or "").lower() != status_filter:
            continue
        if workspace_filter and _workspace_of(r) != workspace_filter:
            continue
        if invoice_filter and r.get("invoiceId") != invoice_filter:
            continue
        if mine_only and ((r.get("requestedBy") or {}).get("id") or "") != actor_id:
            continue
        # Denormalise paidDocumentUrl + paidDocumentName so every UI
        # surface (invoice side panel, accountant ledger, ₮ popover)
        # reads receipt links from one field instead of building them
        # ad-hoc with broken fallbacks.
        out = dict(r)
        url = ""
        name = ""
        meta = r.get("paidDocumentMeta") or {}
        doc_id = r.get("paidDocumentId") or ""
        if meta.get("storedName"):
            url = f"/trip-uploads/_office/{meta['storedName']}"
            name = meta.get("originalName") or ""
        elif doc_id and r.get("tripId"):
            if trips_by_id is None:
                trips_by_id = {t.get("id"): t for t in read_camp_trips()}
            trip = trips_by_id.get(r.get("tripId")) or {}
            for d in (trip.get("documents") or []):
                if d.get("id") == doc_id:
                    url = f"/trip-uploads/{r['tripId']}/{d.get('storedName')}"
                    name = d.get("originalName") or ""
                    break
        out["paidDocumentUrl"] = url
        out["paidDocumentName"] = name
        filtered.append(out)
    return json_response(start_response, "200 OK", {"entries": filtered})


def handle_payment_request_count(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    if not _can_approve_payment_request(actor):
        return json_response(start_response, "200 OK", {"count": 0})
    workspace_filter = active_workspace(environ) or ""
    records = read_payment_requests()
    count = 0
    for r in records:
        if (r.get("status") or "").lower() != "pending":
            continue
        if workspace_filter and (r.get("workspace") or "").upper() != workspace_filter:
            continue
        count += 1
    return json_response(start_response, "200 OK", {"count": count})


def _apply_payment_to_invoice(invoice, request, actor, paid_amount, paid_date, bank_account_id, note):
    """Mirrors handle_invoice_payment's mutation but without the HTTP
    response — used when an accountant approves a request and we need to
    flip the underlying installment to paid in one transaction."""
    installments = invoice.get("installments") or []
    inst_index = int(request.get("installmentIndex") or 0)
    if inst_index < 0 or inst_index >= len(installments):
        return False, "installment out of range"
    target = installments[inst_index]
    target["status"] = "paid"
    if paid_date:
        target["paidDate"] = paid_date
    target["paidAmount"] = float(paid_amount or 0)
    target["bankAccountId"] = bank_account_id or ""
    target["note"] = note or target.get("note") or ""
    bank_snapshot = None
    if bank_account_id:
        for b in read_settings().get("bankAccounts") or []:
            if b.get("id") == bank_account_id:
                bank_snapshot = b
                break
    target["bankAccount"] = bank_snapshot or {}
    target["paidBy"] = actor_snapshot(actor)
    target["paidAt"] = now_mongolia().isoformat()
    installments[inst_index] = target
    invoice["installments"] = installments
    invoice["updatedAt"] = now_mongolia().isoformat()
    invoice["updatedBy"] = actor_snapshot(actor)
    return True, ""


def _attach_orphan_paid_document(upload, actor):
    """For outgoing office/overhead expenses where there's no trip to
    attach the receipt to. Stored under data/trip-uploads/_office/ so
    the existing trip-uploads route serves it like any other file.
    Returns a dict with the metadata the request record needs to keep
    so the Accountant page can link to it later."""
    if not upload:
        return None
    original_name = upload.get("filename") or "paid-document"
    ext = Path(original_name).suffix.lower()
    if ext not in ALLOWED_UPLOAD_EXTENSIONS:
        return None
    data = upload.get("data") or b""
    if len(data) > MAX_UPLOAD_BYTES:
        return None
    ensure_data_store()
    doc_id = str(uuid4())
    out_dir = TRIP_UPLOADS_DIR / "_office"
    out_dir.mkdir(parents=True, exist_ok=True)
    stored_name = doc_id + ext
    (out_dir / stored_name).write_bytes(data)
    return {
        "id": doc_id,
        "storedName": stored_name,
        "originalName": original_name,
        "mimeType": upload.get("content_type") or "application/octet-stream",
        "size": len(data),
        "uploadedAt": now_mongolia().isoformat(),
        "uploadedBy": actor_snapshot(actor),
    }


def _attach_paid_document_to_trip(trip_id, upload, actor):
    """When the accountant uploads the proof file, attach it to the
    trip's documents under category "Paid documents" and return the
    new document id. Same shape handle_upload_trip_document writes —
    so the file shows up in the trip's Documents list and the future
    Accountant page (which queries paid-category docs across trips)."""
    if not trip_id or not upload:
        return None
    trips = read_camp_trips()
    trip_index = next((i for i, t in enumerate(trips) if t.get("id") == trip_id), None)
    if trip_index is None:
        return None
    original_name = upload.get("filename") or "paid-document"
    ext = Path(original_name).suffix.lower()
    if ext not in ALLOWED_UPLOAD_EXTENSIONS:
        return None
    data = upload.get("data") or b""
    if len(data) > MAX_UPLOAD_BYTES:
        return None
    ensure_data_store()
    doc_id = str(uuid4())
    trip_upload_dir = TRIP_UPLOADS_DIR / trip_id
    trip_upload_dir.mkdir(parents=True, exist_ok=True)
    stored_name = doc_id + ext
    (trip_upload_dir / stored_name).write_bytes(data)
    doc = {
        "id": doc_id,
        "originalName": original_name,
        "storedName": stored_name,
        "mimeType": upload.get("content_type") or "application/octet-stream",
        "size": len(data),
        "category": "Paid documents",
        "touristId": "",
        "touristName": "",
        "uploadedAt": now_mongolia().isoformat(),
        "uploadedBy": actor_snapshot(actor),
    }
    trip = trips[trip_index]
    documents = list(trip.get("documents") or [])
    documents.append(doc)
    trips[trip_index] = {**trip, "documents": documents}
    write_camp_trips(trips)
    return doc_id


def _attach_invoice_document_to_trip(trip_id, upload, actor):
    """Save the camp's invoice PDF/image to the trip's documents
    bucket under category "Invoice". Mirrors _attach_paid_document_-
    to_trip so the file shows up in the trip's Documents list."""
    if not trip_id or not upload:
        return None, None
    trips = read_camp_trips()
    trip_index = next((i for i, t in enumerate(trips) if t.get("id") == trip_id), None)
    if trip_index is None:
        return None, None
    original_name = upload.get("filename") or "invoice"
    ext = Path(original_name).suffix.lower()
    if ext not in ALLOWED_UPLOAD_EXTENSIONS:
        return None, None
    data = upload.get("data") or b""
    if len(data) > MAX_UPLOAD_BYTES:
        return None, None
    ensure_data_store()
    doc_id = str(uuid4())
    trip_upload_dir = TRIP_UPLOADS_DIR / trip_id
    trip_upload_dir.mkdir(parents=True, exist_ok=True)
    stored_name = doc_id + ext
    (trip_upload_dir / stored_name).write_bytes(data)
    doc = {
        "id": doc_id,
        "originalName": original_name,
        "storedName": stored_name,
        "mimeType": upload.get("content_type") or "application/octet-stream",
        "size": len(data),
        "category": "Invoice",
        "touristId": "",
        "touristName": "",
        "uploadedAt": now_mongolia().isoformat(),
        "uploadedBy": actor_snapshot(actor),
    }
    trip = trips[trip_index]
    documents = list(trip.get("documents") or [])
    documents.append(doc)
    trips[trip_index] = {**trip, "documents": documents}
    write_camp_trips(trips)
    return doc_id, original_name, stored_name


def handle_upload_camp_group_invoice(environ, start_response, group_key):
    """Upload a camp invoice PDF for a tripId+campName group. The
    file is saved to the trip's Documents under "Invoice" and every
    camp_reservation in the group is stamped with invoiceDocumentId
    so the row's "View invoice" link survives reloads.

    group_key format: "<tripId>::<campName>" (URL-encoded by the
    client).
    """
    actor = require_login(environ, start_response)
    if not actor:
        return []
    raw_key = unquote(group_key or "")
    if "::" not in raw_key:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid group key."})
    trip_id, _, camp_name = raw_key.partition("::")
    if not trip_id or not camp_name:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid group key."})
    content_type = environ.get("CONTENT_TYPE", "")
    if "multipart/form-data" not in content_type:
        return json_response(start_response, "400 Bad Request", {"error": "Multipart upload required."})
    fields, files = parse_multipart(environ)
    upload = files.get("file")
    if not upload:
        return json_response(start_response, "400 Bad Request", {"error": "No file uploaded."})
    stage = (normalize_text(fields.get("stage")) or "deposit").lower()
    if stage not in ("deposit", "balance", "third", "fourth"):
        return json_response(start_response, "400 Bad Request", {"error": "Invalid stage (deposit | balance | third | fourth)."})
    result = _attach_invoice_document_to_trip(trip_id, upload, actor)
    if not result or not result[0]:
        return json_response(start_response, "400 Bad Request", {"error": "Could not attach invoice (file too large or unsupported type)."})
    doc_id, original_name, stored_name = result
    rows = read_camp_reservations()
    now_iso = now_mongolia().isoformat()
    actor_meta = actor_snapshot(actor)
    prefix = {
        "deposit": "depositInvoice",
        "balance": "balanceInvoice",
        "third":   "thirdInvoice",
        "fourth":  "fourthInvoice",
    }[stage]
    changed = False
    for i, r in enumerate(rows):
        if r.get("tripId") == trip_id and (r.get("campName") or "") == camp_name:
            rows[i] = {
                **r,
                f"{prefix}DocumentId": doc_id,
                f"{prefix}DocumentName": original_name,
                f"{prefix}StoredName": stored_name,
                f"{prefix}UploadedAt": now_iso,
                f"{prefix}UploadedBy": actor_meta,
                "updatedAt": now_iso,
                "updatedBy": actor_meta,
            }
            changed = True
    if changed:
        write_camp_reservations(rows)
    return json_response(start_response, "200 OK", {
        "ok": True,
        "stage": stage,
        "documentId": doc_id,
        "documentName": original_name,
        "storedName": stored_name,
        "uploadedAt": now_iso,
    })


def handle_approve_payment_request(environ, start_response, request_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    if not _can_approve_payment_request(actor):
        return json_response(start_response, "403 Forbidden", {"error": "Only admins / accountants can approve payment requests."})

    # Accept either JSON (no file) or multipart (with the proof file).
    content_type = environ.get("CONTENT_TYPE", "")
    if "multipart/form-data" in content_type:
        fields, files = parse_multipart(environ)
        payload = {k: v for k, v in fields.items()}
        upload = files.get("file")
    else:
        payload = collect_json(environ) or {}
        upload = None

    records = read_payment_requests()
    target = next((r for r in records if r.get("id") == request_id), None)
    if not target:
        return json_response(start_response, "404 Not Found", {"error": "Request not found"})
    if (target.get("status") or "").lower() != "pending":
        return json_response(start_response, "409 Conflict", {"error": "Request already resolved."})

    is_outgoing = (target.get("direction") or "incoming") == "outgoing"
    invoices = read_invoices()
    invoice = None if is_outgoing else next((r for r in invoices if r.get("id") == target.get("invoiceId")), None)
    if not is_outgoing and not invoice:
        return json_response(start_response, "404 Not Found", {"error": "Invoice has been deleted."})

    paid_amount = payload.get("paidAmount")
    if paid_amount is None:
        paid_amount = target.get("paidAmount")
    paid_date = normalize_text(payload.get("paidDate")) or target.get("paidDate") or ""
    # Outgoing requests don't carry a paid date from the manager — the
    # accountant pays it later. Default to today so the ledger column
    # isn't blank when the file gets uploaded.
    if not paid_date:
        paid_date = now_mongolia().date().isoformat()
    bank_account_id = normalize_text(payload.get("bankAccountId")) or target.get("bankAccountId") or ""
    note = normalize_text(payload.get("note")) or target.get("note") or ""
    paid_document_id = normalize_text(payload.get("paidDocumentId"))

    # Auto-attach the uploaded proof to the trip's Documents under
    # "Paid documents" when there is a trip — outgoing office expenses
    # don't have one, so the file is stored on the request only.
    if upload and target.get("tripId"):
        attached_id = _attach_paid_document_to_trip(target["tripId"], upload, actor)
        if attached_id:
            paid_document_id = attached_id
    elif upload and not target.get("tripId"):
        attached = _attach_orphan_paid_document(upload, actor)
        if attached:
            paid_document_id = attached["id"]
            # Store the office-doc metadata on the request so the
            # Accountant page can render a link without a trip lookup.
            target["paidDocumentMeta"] = attached
    # Skip-document path: accountant in a hurry, can attach the
    # receipt later via the Accountant page. Approval still flips the
    # invoice to paid and writes the audit trail.
    skip_document = bool(payload.get("skipDocument") in (True, "true", "1", 1))
    if not paid_document_id and not skip_document:
        return json_response(start_response, "400 Bad Request", {"error": "Upload the paid receipt before approving, or use Register without document."})

    if not is_outgoing:
        ok, err = _apply_payment_to_invoice(invoice, target, actor, paid_amount, paid_date, bank_account_id, note)
        if not ok:
            return json_response(start_response, "400 Bad Request", {"error": err})
        for i, r in enumerate(invoices):
            if r.get("id") == invoice.get("id"):
                invoices[i] = invoice
                break
        write_invoices(invoices)
    else:
        # Outgoing payment_request linked to a flight / transfer / camp
        # reservation — flip the reservation's paymentStatus to "paid"
        # so the trip dashboards reflect the confirmation. Skipped when
        # recordType is empty (legacy office-overhead expenses don't
        # link to a reservation).
        rec_type = (target.get("recordType") or "").lower()
        rec_id = target.get("recordId") or ""
        readers_writers = {
            "flight_reservation":   (read_flight_reservations,   write_flight_reservations),
            "transfer_reservation": (read_transfer_reservations, write_transfer_reservations),
            "camp_reservation":     (read_camp_reservations,     write_camp_reservations),
        }
        if rec_type in readers_writers and rec_id:
            reader, writer = readers_writers[rec_type]
            rows = reader()
            for i, r in enumerate(rows):
                if r.get("id") == rec_id:
                    rows[i] = {
                        **r,
                        "paymentStatus": "paid",
                        "paidDate": paid_date or r.get("paidDate") or "",
                        "updatedAt": now_mongolia().isoformat(),
                        "updatedBy": actor_snapshot(actor),
                    }
                    writer(rows)
                    break
        elif rec_type == "camp_group" and rec_id and "::" in rec_id:
            # recordId format: "<tripId>::<campName>". When the request
            # carries a stage ("deposit" or "balance") only that stage's
            # paid date is set; if every positive stage is now paid, the
            # whole camp's paymentStatus also flips to paid. With no
            # stage (legacy), every reservation in the pair flips to
            # paid in one shot.
            group_trip_id, _, group_camp_name = rec_id.partition("::")
            stage = (target.get("stage") or "").lower()
            rows = read_camp_reservations()
            changed = False
            now_iso = now_mongolia().isoformat()
            actor_meta = actor_snapshot(actor)
            for i, r in enumerate(rows):
                if r.get("tripId") != group_trip_id or (r.get("campName") or "") != group_camp_name:
                    continue
                if stage in ("deposit", "balance", "third", "fourth"):
                    date_field_for_stage = {
                        "deposit": "depositPaidDate",
                        "balance": "secondPaidDate",
                        "third":   "thirdPaidDate",
                        "fourth":  "fourthPaidDate",
                    }[stage]
                    updated = {**r, date_field_for_stage: paid_date or r.get(date_field_for_stage) or now_iso[:10],
                               "updatedAt": now_iso, "updatedBy": actor_meta}
                    # Whole camp flips to paid only when every stage
                    # that has a positive amount is now paid. Empty
                    # stages (amount = 0) are treated as already done
                    # so a 2-stage camp doesn't need stage 3/4 set.
                    stage_amount_field = {
                        "deposit": "deposit",
                        "balance": "secondPayment",
                        "third":   "thirdPayment",
                        "fourth":  "fourthPayment",
                    }
                    stage_date_field = {
                        "deposit": "depositPaidDate",
                        "balance": "secondPaidDate",
                        "third":   "thirdPaidDate",
                        "fourth":  "fourthPaidDate",
                    }
                    any_positive = False
                    all_done = True
                    for s in ("deposit", "balance", "third", "fourth"):
                        amt = float(updated.get(stage_amount_field[s]) or 0)
                        if amt > 0:
                            any_positive = True
                            if not updated.get(stage_date_field[s]):
                                all_done = False
                    if any_positive and all_done:
                        updated["paymentStatus"] = "paid"
                        updated["paidDate"] = paid_date or updated.get("paidDate") or now_iso[:10]
                    rows[i] = updated
                else:
                    rows[i] = {
                        **r,
                        "paymentStatus": "paid",
                        "paidDate": paid_date or r.get("paidDate") or "",
                        "updatedAt": now_iso,
                        "updatedBy": actor_meta,
                    }
                changed = True
            if changed:
                write_camp_reservations(rows)

    target["status"] = "approved"
    target["approvedBy"] = actor_snapshot(actor)
    target["approvedAt"] = now_mongolia().isoformat()
    target["paidAmount"] = float(paid_amount or 0)
    target["paidDate"] = paid_date
    target["bankAccountId"] = bank_account_id
    target["note"] = note
    if paid_document_id:
        target["paidDocumentId"] = paid_document_id
    for i, r in enumerate(records):
        if r.get("id") == request_id:
            records[i] = target
            break
    write_payment_requests(records)

    # Notify the manager who originally requested the payment so they
    # see in their bell + ₮ that the accountant registered it.
    requester = target.get("requestedBy") or {}
    log_notification(
        kind="payment_request.approved",
        actor=actor,
        title=f"Payment registered · {target.get('invoiceSerial') or ''}",
        detail=f"{actor.get('fullName') or actor.get('email')} registered "
               f"{_fmt_money_ccy(target.get('paidAmount'), target.get('currency'))} "
               f"for {target.get('installmentDescription') or 'this installment'}.",
        meta={
            "paymentRequestId": target.get("id"),
            "invoiceId": target.get("invoiceId"),
            "tripId": target.get("tripId"),
            "workspace": target.get("workspace"),
            "recipientUserId": requester.get("id") or "",
        },
    )
    return json_response(start_response, "200 OK", {"ok": True, "entry": target, "invoice": invoice})


def handle_attach_payment_document(environ, start_response, request_id):
    """Late-attach a receipt to an already-approved payment-request.
    Used by the Accountant page when an entry was either approved
    via "Register without document" or backfilled from a legacy
    paid invoice. Same storage paths the regular approve flow uses,
    so the file shows up under the trip's Paid documents (or the
    _office bucket when there's no trip)."""
    actor = require_login(environ, start_response)
    if not actor:
        return []
    if not _can_approve_payment_request(actor):
        return json_response(start_response, "403 Forbidden", {"error": "Only admin / accountant can attach receipts."})
    content_type = environ.get("CONTENT_TYPE", "")
    if "multipart/form-data" not in content_type:
        return json_response(start_response, "400 Bad Request", {"error": "Multipart upload required."})
    fields, files = parse_multipart(environ)
    upload = files.get("file")
    if not upload:
        return json_response(start_response, "400 Bad Request", {"error": "No file uploaded."})

    records = read_payment_requests()
    target = next((r for r in records if r.get("id") == request_id), None)
    if not target:
        return json_response(start_response, "404 Not Found", {"error": "Request not found"})

    paid_doc_id = ""
    if target.get("tripId"):
        paid_doc_id = _attach_paid_document_to_trip(target["tripId"], upload, actor) or ""
    if not paid_doc_id:
        attached = _attach_orphan_paid_document(upload, actor)
        if attached:
            paid_doc_id = attached["id"]
            target["paidDocumentMeta"] = attached
    if not paid_doc_id:
        return json_response(start_response, "400 Bad Request", {"error": "Could not save the file (unsupported type or too large)."})

    target["paidDocumentId"] = paid_doc_id
    for i, r in enumerate(records):
        if r.get("id") == request_id:
            records[i] = target
            break
    write_payment_requests(records)
    return json_response(start_response, "200 OK", {"ok": True, "entry": target})


def handle_delete_payment_document(environ, start_response, request_id):
    """Remove the receipt attached to a payment_request. Office-bucket
    receipts (no tripId) live under data/trip-uploads/_office/ and are
    deleted directly here. Trip-bound receipts go through
    handle_delete_trip_document, which already cascades the
    paidDocumentId clear back onto the request."""
    actor = require_login(environ, start_response)
    if not actor:
        return []
    if not _can_approve_payment_request(actor):
        return json_response(start_response, "403 Forbidden", {"error": "Only admin / accountant can delete receipts."})
    records = read_payment_requests()
    target = next((r for r in records if r.get("id") == request_id), None)
    if not target:
        return json_response(start_response, "404 Not Found", {"error": "Request not found"})
    meta = target.get("paidDocumentMeta") or {}
    stored_name = meta.get("storedName") or ""
    if stored_name:
        office_dir = (TRIP_UPLOADS_DIR / "_office").resolve()
        file_path = (office_dir / stored_name).resolve()
        if str(file_path).startswith(str(office_dir)) and file_path.exists():
            try:
                file_path.unlink()
            except Exception:
                pass
    target["paidDocumentMeta"] = None
    target["paidDocumentId"] = ""
    for i, r in enumerate(records):
        if r.get("id") == request_id:
            records[i] = target
            break
    write_payment_requests(records)
    return json_response(start_response, "200 OK", {"ok": True})


def handle_attach_invoice_receipt(environ, start_response, invoice_id, inst_index):
    """Backfill a bank-transfer receipt onto a paid installment from
    the invoice side panel. Saves the file under the trip's "Paid
    documents", and either:
      • links it to the existing approved payment_request (if one
        was created via the new flow), OR
      • creates a synthetic approved payment_request that mirrors the
        installment's snapshot, so the file shows up on the
        Accountant ledger like every other paid row."""
    actor = require_login(environ, start_response)
    if not actor:
        return []
    if not _can_approve_payment_request(actor):
        return json_response(start_response, "403 Forbidden", {"error": "Only admin / accountant can attach receipts."})
    content_type = environ.get("CONTENT_TYPE", "")
    if "multipart/form-data" not in content_type:
        return json_response(start_response, "400 Bad Request", {"error": "Multipart upload required."})
    fields, files = parse_multipart(environ)
    upload = files.get("file")
    if not upload:
        return json_response(start_response, "400 Bad Request", {"error": "No file uploaded."})

    invoices = read_invoices()
    invoice = next((r for r in invoices if r.get("id") == invoice_id), None)
    if not invoice:
        return json_response(start_response, "404 Not Found", {"error": "Invoice not found"})
    installments = invoice.get("installments") or []
    if inst_index < 0 or inst_index >= len(installments):
        return json_response(start_response, "400 Bad Request", {"error": "Installment out of range"})
    inst = installments[inst_index]

    # Save the file. Trip-attached when the invoice has a tripId,
    # _office bucket otherwise.
    paid_doc_id = ""
    paid_doc_meta = None
    trip_id = invoice.get("tripId") or ""
    if trip_id:
        paid_doc_id = _attach_paid_document_to_trip(trip_id, upload, actor) or ""
    if not paid_doc_id:
        attached = _attach_orphan_paid_document(upload, actor)
        if attached:
            paid_doc_id = attached["id"]
            paid_doc_meta = attached
    if not paid_doc_id:
        return json_response(start_response, "400 Bad Request", {"error": "Could not save the file."})

    # Link the receipt to a payment_request — either an existing
    # approved one for this installment, or a freshly synthesized
    # record so the Accountant page reflects the new file.
    records = read_payment_requests()
    target = next(
        (r for r in records
         if r.get("invoiceId") == invoice_id and r.get("installmentIndex") == inst_index
         and (r.get("status") or "").lower() == "approved"),
        None,
    )
    if target:
        target["paidDocumentId"] = paid_doc_id
        if paid_doc_meta:
            target["paidDocumentMeta"] = paid_doc_meta
        for i, r in enumerate(records):
            if r.get("id") == target.get("id"):
                records[i] = target
                break
    else:
        synth = {
            "id": uuid4().hex,
            "direction": "incoming",
            "scope": "trip" if trip_id else "other",
            "category": "Invoice",
            "invoiceId": invoice_id,
            "invoiceSerial": invoice.get("serial") or "",
            "tripId": trip_id,
            "installmentIndex": inst_index,
            "installmentDescription": inst.get("description") or "",
            "paidDate": inst.get("paidDate") or "",
            "paidAmount": float(inst.get("paidAmount") or inst.get("amount") or 0),
            "currency": (invoice.get("currency") or "MNT").upper(),
            "bankAccountId": inst.get("bankAccountId") or "",
            "note": inst.get("note") or "",
            "payerName": invoice.get("payerName") or "",
            "payeeName": "",
            "workspace": _payment_request_workspace(invoice),
            "status": "approved",
            "requestedBy": inst.get("paidBy") or actor_snapshot(actor),
            "requestedAt": inst.get("paidAt") or now_mongolia().isoformat(),
            "approvedBy": actor_snapshot(actor),
            "approvedAt": now_mongolia().isoformat(),
            "rejectedBy": None,
            "rejectedAt": "",
            "rejectReason": "",
            "paidDocumentId": paid_doc_id,
            "paidDocumentMeta": paid_doc_meta,
        }
        records.insert(0, synth)
    write_payment_requests(records)
    return json_response(start_response, "200 OK", {"ok": True, "paidDocumentId": paid_doc_id})


def handle_reject_payment_request(environ, start_response, request_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    if not _can_approve_payment_request(actor):
        return json_response(start_response, "403 Forbidden", {"error": "Only admins / accountants can reject payment requests."})
    payload = collect_json(environ) or {}
    reason = normalize_text(payload.get("reason"))
    if not reason:
        return json_response(start_response, "400 Bad Request", {"error": "Rejection reason is required so the manager understands what to fix."})
    records = read_payment_requests()
    target = next((r for r in records if r.get("id") == request_id), None)
    if not target:
        return json_response(start_response, "404 Not Found", {"error": "Request not found"})
    if (target.get("status") or "").lower() != "pending":
        return json_response(start_response, "409 Conflict", {"error": "Request already resolved."})
    target["status"] = "rejected"
    target["rejectedBy"] = actor_snapshot(actor)
    target["rejectedAt"] = now_mongolia().isoformat()
    target["rejectReason"] = reason
    for i, r in enumerate(records):
        if r.get("id") == request_id:
            records[i] = target
            break
    write_payment_requests(records)

    # Notify the manager who originally requested the payment so they
    # see the rejection (with reason) in their ₮ popover and bell.
    requester = target.get("requestedBy") or {}
    log_notification(
        kind="payment_request.rejected",
        actor=actor,
        title=f"Payment request rejected · {target.get('invoiceSerial') or ''}",
        detail=f"{actor.get('fullName') or actor.get('email')} rejected your "
               f"{target.get('installmentDescription') or 'installment'} "
               f"payment request — {reason}",
        meta={
            "paymentRequestId": target.get("id"),
            "invoiceId": target.get("invoiceId"),
            "tripId": target.get("tripId"),
            "workspace": target.get("workspace"),
            "recipientUserId": requester.get("id") or "",
            "rejectReason": reason,
        },
    )
    return json_response(start_response, "200 OK", {"ok": True, "entry": target})


def handle_delete_payment_request(environ, start_response, request_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []

    # Legacy ledger rows are synthetic — they're paid-invoice
    # installments surfaced into the Accountant page that have no real
    # payment_request record. Their id is "legacy-<invoiceId>-<idx>".
    # Deleting one means: unmark that invoice installment as paid.
    if request_id.startswith("legacy-"):
        if not _can_approve_payment_request(actor):
            return json_response(start_response, "403 Forbidden", {"error": "Only admin / accountant can unmark paid invoices."})
        # Strip the "legacy-" prefix and split off the trailing -<idx>.
        rest = request_id[len("legacy-"):]
        try:
            invoice_id, idx_str = rest.rsplit("-", 1)
            inst_index = int(idx_str)
        except (ValueError, AttributeError):
            return json_response(start_response, "404 Not Found", {"error": "Bad legacy id"})
        invoices = read_invoices()
        for i, inv in enumerate(invoices):
            if inv.get("id") != invoice_id:
                continue
            installments = list(inv.get("installments") or [])
            if inst_index < 0 or inst_index >= len(installments):
                return json_response(start_response, "400 Bad Request", {"error": "Installment out of range"})
            inst = installments[inst_index]
            inst["status"] = "pending"
            for k in ("paidDate", "paidAmount", "bankAccountId", "bankAccount", "paidBy", "paidAt"):
                if k in inst:
                    inst[k] = "" if k != "paidAmount" else 0
            installments[inst_index] = inst
            invoices[i] = {**inv, "installments": installments,
                           "updatedAt": now_mongolia().isoformat(),
                           "updatedBy": actor_snapshot(actor)}
            write_invoices(invoices)
            return json_response(start_response, "200 OK", {"ok": True, "legacy": True})
        return json_response(start_response, "404 Not Found", {"error": "Invoice not found"})

    records = read_payment_requests()
    target = next((r for r in records if r.get("id") == request_id), None)
    if not target:
        return json_response(start_response, "404 Not Found", {"error": "Request not found"})
    # Requesters can cancel their own pending request; admins/accountants
    # can delete anything (including approved, which keeps the invoice
    # paid status — the row just disappears from the ledger).
    can_delete = (
        _can_approve_payment_request(actor)
        or ((target.get("requestedBy") or {}).get("id") == actor.get("id")
            and (target.get("status") or "").lower() == "pending")
    )
    if not can_delete:
        return json_response(start_response, "403 Forbidden", {"error": "Only the requester or an admin can cancel."})
    records = [r for r in records if r.get("id") != request_id]
    write_payment_requests(records)
    return json_response(start_response, "200 OK", {"ok": True})


def handle_list_accountant_paid(environ, start_response):
    """Joins payment requests with the latest invoice / trip snapshots
    so the Accountant page can render one row per request — both
    pending (still waiting on the accountant) and approved (paid) —
    with all the columns the user asked for. Workspace-scoped so DTX
    and USM lists never bleed."""
    actor = require_login(environ, start_response)
    if not actor:
        return []
    workspace_filter = (active_workspace(environ) or "").upper()
    requests_all = read_payment_requests()
    invoices_all = read_invoices()
    invoices_by_id = {r.get("id"): r for r in invoices_all}
    trips_by_id = {t.get("id"): t for t in read_camp_trips()}
    banks_by_id = {b.get("id"): b for b in (read_settings().get("bankAccounts") or [])}
    # Resolve groupId → groupName once so legacy payment_requests that
    # never had groupName denormalised still render the group column.
    try:
        _groups_raw = json.loads(GROUPS_FILE.read_text("utf-8")) if GROUPS_FILE.exists() else []
    except Exception:
        _groups_raw = []
    groups_by_id = {g.get("id"): g for g in _groups_raw}
    # Track which (invoiceId, installmentIndex) pairs are already
    # represented by a payment_request so we don't double-count them
    # when we also surface paid installments below.
    request_index_pairs = {
        (r.get("invoiceId"), r.get("installmentIndex"))
        for r in requests_all
        if r.get("invoiceId") and r.get("installmentIndex") is not None
    }
    # Same trust-the-trip rule used in handle_list_payment_requests:
    # legacy rows have a stale `workspace` field stamped by the old
    # serial-prefix heuristic. Derive workspace from the linked
    # invoice's trip whenever possible.
    def _ws_of(record):
        invoice = invoices_by_id.get(record.get("invoiceId")) or {}
        if invoice.get("tripId"):
            trip = trips_by_id.get(invoice.get("tripId"))
            if trip:
                return (normalize_company(trip.get("company")) or "DTX").upper()
        if record.get("tripId"):
            trip = trips_by_id.get(record.get("tripId"))
            if trip:
                return (normalize_company(trip.get("company")) or "DTX").upper()
        return (record.get("workspace") or "").upper()
    rows = []
    for r in requests_all:
        status = (r.get("status") or "").lower()
        # Skip rejected / cancelled — only pending and approved are useful
        # in the Accountant ledger view.
        if status not in ("pending", "approved"):
            continue
        if workspace_filter and _ws_of(r) != workspace_filter:
            continue
        invoice = invoices_by_id.get(r.get("invoiceId")) or {}
        trip = trips_by_id.get(r.get("tripId")) or {}
        bank = banks_by_id.get(r.get("bankAccountId")) or {}
        # Find the trip document this approval saved. Office expenses
        # don't have a trip, so fall back to the metadata stashed on
        # the request itself.
        paid_doc = None
        paid_doc_url = ""
        for d in (trip.get("documents") or []):
            if d.get("id") == r.get("paidDocumentId"):
                paid_doc = d
                if d.get("storedName"):
                    paid_doc_url = f"/trip-uploads/{r.get('tripId')}/{d.get('storedName')}"
                break
        if not paid_doc and r.get("paidDocumentMeta"):
            meta = r.get("paidDocumentMeta") or {}
            if meta.get("storedName"):
                paid_doc = meta
                paid_doc_url = f"/trip-uploads/_office/{meta.get('storedName')}"
        manager_name = ""
        for source in (invoice.get("createdBy"), invoice.get("updatedBy"), r.get("requestedBy")):
            if source and source.get("name"):
                manager_name = source.get("name")
                break
            if source and source.get("email"):
                manager_name = source.get("email")
                break
        rows.append({
            "id": r.get("id"),
            "status": status,
            "direction": r.get("direction") or "incoming",
            "scope": r.get("scope") or "trip",
            "category": r.get("category") or "Invoice",
            "paidDate": r.get("paidDate") or "",
            "dueDate": r.get("dueDate") or "",
            "requestedAt": r.get("requestedAt") or "",
            "tripId": r.get("tripId") or "",
            "tripName": trip.get("tripName") or r.get("tripName") or "",
            "tripSerial": trip.get("serial") or "",
            "tripStartDate": trip.get("startDate") or "",
            "groupId": r.get("groupId") or invoice.get("groupId") or "",
            "groupName": (r.get("groupName")
                          or (groups_by_id.get(r.get("groupId") or invoice.get("groupId") or "") or {}).get("name")
                          or (groups_by_id.get(r.get("groupId") or invoice.get("groupId") or "") or {}).get("groupName")
                          or ""),
            "invoiceId": r.get("invoiceId") or "",
            "invoiceSerial": r.get("invoiceSerial") or invoice.get("serial") or "",
            "installmentDescription": r.get("installmentDescription") or "",
            "installmentIndex": r.get("installmentIndex") if r.get("installmentIndex") is not None else -1,
            "payerName": r.get("payerName") or invoice.get("payerName") or "",
            "payeeName": r.get("payeeName") or "",
            "amount": r.get("paidAmount") or 0,
            "currency": r.get("currency") or invoice.get("currency") or "MNT",
            "bankAccountId": r.get("bankAccountId") or "",
            "bankLabel": (bank.get("label") or bank.get("bankName") or ""),
            "bankAccountNumber": bank.get("accountNumber") or "",
            "manager": manager_name,
            "approvedBy": (r.get("approvedBy") or {}).get("name") or "",
            "approvedAt": r.get("approvedAt") or "",
            "requestedBy": r.get("requestedBy") or {},
            "note": r.get("note") or "",
            "paidDocumentId": r.get("paidDocumentId") or "",
            "paidDocumentName": (paid_doc or {}).get("originalName") or "",
            "paidDocumentMime": (paid_doc or {}).get("mimeType") or "",
            "paidDocumentUrl": paid_doc_url,
            "vendorInvoiceMeta": r.get("vendorInvoiceMeta") or None,
            "workspace": r.get("workspace") or "",
        })
    # Backfill: any paid installment on an existing invoice that has no
    # matching payment_request shows up as a synthetic incoming row so
    # the ledger reflects the company's actual paid history (records
    # that were marked paid before the request flow existed).
    for invoice in invoices_all:
        ws = _payment_request_workspace(invoice)
        if workspace_filter and ws != workspace_filter:
            continue
        installments = invoice.get("installments") or []
        for idx, inst in enumerate(installments):
            inst_status = (inst.get("status") or "").lower()
            if inst_status not in ("paid", "confirmed"):
                continue
            if (invoice.get("id"), idx) in request_index_pairs:
                continue
            trip = trips_by_id.get(invoice.get("tripId")) or {}
            bank_id = inst.get("bankAccountId") or ""
            bank = banks_by_id.get(bank_id) or inst.get("bankAccount") or {}
            paid_amount = inst.get("paidAmount")
            if paid_amount is None:
                paid_amount = inst.get("amount") or 0
            paid_by = inst.get("paidBy") or {}
            rows.append({
                "id": f"legacy-{invoice.get('id')}-{idx}",
                "status": "approved",
                "direction": "incoming",
                "scope": "trip",
                "category": "Invoice",
                "paidDate": inst.get("paidDate") or "",
                "dueDate": inst.get("dueDate") or "",
                "requestedAt": inst.get("paidAt") or invoice.get("createdAt") or "",
                "tripId": invoice.get("tripId") or "",
                "tripName": trip.get("tripName") or "",
                "tripSerial": trip.get("serial") or "",
                "tripStartDate": trip.get("startDate") or "",
                "groupId": invoice.get("groupId") or "",
                "groupName": (groups_by_id.get(invoice.get("groupId") or "") or {}).get("name")
                              or (groups_by_id.get(invoice.get("groupId") or "") or {}).get("groupName")
                              or "",
                "invoiceId": invoice.get("id") or "",
                "invoiceSerial": invoice.get("serial") or "",
                "installmentDescription": inst.get("description") or "",
                "installmentIndex": idx,
                "payerName": invoice.get("payerName") or "",
                "payeeName": "",
                "amount": paid_amount,
                "currency": invoice.get("currency") or "MNT",
                "bankAccountId": bank_id,
                "bankLabel": (bank.get("label") or bank.get("bankName") or ""),
                "bankAccountNumber": bank.get("accountNumber") or "",
                "manager": (paid_by.get("name") or paid_by.get("email") or ""),
                "approvedBy": (paid_by.get("name") or ""),
                "approvedAt": inst.get("paidAt") or "",
                "requestedBy": paid_by,
                "note": inst.get("note") or "",
                "paidDocumentId": "",
                "paidDocumentName": "",
                "paidDocumentMime": "",
                "paidDocumentUrl": "",
                "workspace": ws,
                "isLegacy": True,
            })

    # Pending rows float to top (so the accountant sees what to act on),
    # then approved rows ordered by paid date desc. Sort by the
    # 10-char ISO prefix of paidDate so a legacy record with garbage
    # in the field (e.g. "2026-04-30 manager-name 12345") doesn't
    # crash the whole endpoint.
    def sort_key(x):
        is_pending = (x.get("status") or "") == "pending"
        date_str = str(x.get("paidDate") or "")[:10]
        return (0 if is_pending else 1, "" if is_pending else date_str)
    # First sort approved rows by date desc among themselves, then
    # pending rows pile on top in their original order.
    rows.sort(key=lambda x: sort_key(x)[1], reverse=True)
    rows.sort(key=lambda x: sort_key(x)[0])
    return json_response(start_response, "200 OK", {"entries": rows})


def handle_list_trip_templates(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    workspace_filter = (active_workspace(environ) or "").upper()
    records = read_trip_templates()
    if workspace_filter:
        records = [
            r for r in records
            if (r.get("workspace") or "").upper() in ("", "BOTH", workspace_filter)
        ]
    return json_response(start_response, "200 OK", {"entries": records})


def handle_create_trip_template(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ) or {}
    name = normalize_text(payload.get("name"))
    if not name:
        return json_response(start_response, "400 Bad Request", {"error": "Name is required"})
    workspace = (normalize_text(payload.get("workspace")) or active_workspace(environ) or "DTX").upper()
    if workspace not in ("DTX", "USM", "BOTH"):
        workspace = "DTX"
    try:
        days = max(1, int(payload.get("days") or 1))
    except (TypeError, ValueError):
        days = 1
    record = {
        "id": uuid4().hex,
        "name": name,
        "workspace": workspace,
        "days": days,
        "marginPct": float(payload.get("marginPct") or 0),
        "notes": normalize_text(payload.get("notes")),
        "expenseLines": _normalize_trip_template_lines(payload.get("expenseLines")),
        "createdAt": now_mongolia().isoformat(),
        "createdBy": actor_snapshot(actor),
        "updatedAt": now_mongolia().isoformat(),
    }
    records = read_trip_templates()
    records.insert(0, record)
    write_trip_templates(records)
    return json_response(start_response, "200 OK", {"ok": True, "entry": record})


def handle_update_trip_template(environ, start_response, template_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ) or {}
    records = read_trip_templates()
    target = next((r for r in records if r.get("id") == template_id), None)
    if not target:
        return json_response(start_response, "404 Not Found", {"error": "Template not found"})
    if "name" in payload:
        target["name"] = normalize_text(payload.get("name")) or target.get("name")
    if "workspace" in payload:
        ws = (normalize_text(payload.get("workspace")) or "").upper()
        if ws in ("DTX", "USM", "BOTH"):
            target["workspace"] = ws
    if "days" in payload:
        try:
            target["days"] = max(1, int(payload.get("days") or 1))
        except (TypeError, ValueError):
            pass
    if "marginPct" in payload:
        try:
            target["marginPct"] = float(payload.get("marginPct") or 0)
        except (TypeError, ValueError):
            pass
    if "notes" in payload:
        target["notes"] = normalize_text(payload.get("notes"))
    if "expenseLines" in payload:
        target["expenseLines"] = _normalize_trip_template_lines(payload.get("expenseLines"))
    target["updatedAt"] = now_mongolia().isoformat()
    target["updatedBy"] = actor_snapshot(actor)
    for i, r in enumerate(records):
        if r.get("id") == template_id:
            records[i] = target
            break
    write_trip_templates(records)
    return json_response(start_response, "200 OK", {"ok": True, "entry": target})


def handle_delete_trip_template(environ, start_response, template_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    records = read_trip_templates()
    if not any(r.get("id") == template_id for r in records):
        return json_response(start_response, "404 Not Found", {"error": "Template not found"})
    records = [r for r in records if r.get("id") != template_id]
    write_trip_templates(records)
    return json_response(start_response, "200 OK", {"ok": True})


# ── Service templates ────────────────────────────────────────────
# A service is a costed activity (museum visit, pony riding, …) with:
#   - prices per pax/staff type (empty/0 = excluded from the calc)
#   - a multilingual prose description that may embed [[content:slug|Title]]
#     references resolved at trip-render time.

_SERVICE_PAX_KEYS = (
    "adults", "teen", "child", "infant", "foc",
    "guide", "driver", "cook", "tourLeader", "shaman", "shamanAssistant",
)
_SERVICE_LANG_CODES = ("mn", "en", "fr", "it", "es", "ko", "zh", "ja", "ru")
_SERVICE_CURRENCIES = ("MNT", "USD", "EUR", "CNY", "JPY", "KRW", "RUB")


def _normalize_service_prices(raw):
    out = {k: 0 for k in _SERVICE_PAX_KEYS}
    if isinstance(raw, dict):
        for k in _SERVICE_PAX_KEYS:
            try:
                v = float(raw.get(k) or 0)
            except (TypeError, ValueError):
                v = 0
            out[k] = v if v > 0 else 0
    return out


def _normalize_service_descriptions(raw):
    out = {k: "" for k in _SERVICE_LANG_CODES}
    if isinstance(raw, dict):
        for k in _SERVICE_LANG_CODES:
            v = raw.get(k)
            if isinstance(v, str):
                out[k] = v.strip()
    return out


def _build_service_template(payload, actor, workspace):
    name = normalize_text(payload.get("name"))
    workspace = (workspace or "DTX").upper()
    if workspace not in ("DTX", "USM"):
        workspace = "DTX"
    currency = (normalize_text(payload.get("currency")) or "MNT").upper()
    if currency not in _SERVICE_CURRENCIES:
        currency = "MNT"
    return {
        "id": uuid4().hex,
        "name": name,
        "workspace": workspace,
        "currency": currency,
        "prices": _normalize_service_prices(payload.get("prices")),
        "descriptions": _normalize_service_descriptions(payload.get("descriptions")),
        "createdAt": now_mongolia().isoformat(),
        "createdBy": actor_snapshot(actor),
        "updatedAt": now_mongolia().isoformat(),
        "updatedBy": actor_snapshot(actor),
    }


def handle_list_service_templates(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    workspace_filter = (active_workspace(environ) or "").upper()
    records = read_service_templates()
    if workspace_filter:
        records = [
            r for r in records
            if (r.get("workspace") or "").upper() in ("", "BOTH", workspace_filter)
        ]
    return json_response(start_response, "200 OK", {"entries": records})


def handle_create_service_template(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ) or {}
    if not normalize_text(payload.get("name")):
        return json_response(start_response, "400 Bad Request", {"error": "Name is required"})
    # Workspace is locked to the active session — a service created on USM
    # only ever shows up on USM. Never accept a client-supplied value here.
    workspace = (active_workspace(environ) or "DTX").upper()
    record = _build_service_template(payload, actor, workspace)
    records = read_service_templates()
    records.insert(0, record)
    write_service_templates(records)
    return json_response(start_response, "200 OK", {"ok": True, "entry": record})


def handle_update_service_template(environ, start_response, template_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ) or {}
    records = read_service_templates()
    target = next((r for r in records if r.get("id") == template_id), None)
    if not target:
        return json_response(start_response, "404 Not Found", {"error": "Template not found"})
    if "name" in payload:
        target["name"] = normalize_text(payload.get("name")) or target.get("name")
    # Workspace is locked at create time and never changes via update — keep it.
    if "currency" in payload:
        ccy = (normalize_text(payload.get("currency")) or "").upper()
        if ccy in _SERVICE_CURRENCIES:
            target["currency"] = ccy
    if "prices" in payload:
        target["prices"] = _normalize_service_prices(payload.get("prices"))
    if "descriptions" in payload:
        target["descriptions"] = _normalize_service_descriptions(payload.get("descriptions"))
    target["updatedAt"] = now_mongolia().isoformat()
    target["updatedBy"] = actor_snapshot(actor)
    for i, r in enumerate(records):
        if r.get("id") == template_id:
            records[i] = target
            break
    write_service_templates(records)
    return json_response(start_response, "200 OK", {"ok": True, "entry": target})


def handle_delete_service_template(environ, start_response, template_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    records = read_service_templates()
    if not any(r.get("id") == template_id for r in records):
        return json_response(start_response, "404 Not Found", {"error": "Template not found"})
    records = [r for r in records if r.get("id") != template_id]
    write_service_templates(records)
    return json_response(start_response, "200 OK", {"ok": True})


# ── Claude-backed prose translation ────────────────────────────
# Used by Service templates (and reusable for other multilingual prose).
# Preserves [[content:slug|label]] markers verbatim — slug stays identical,
# label is translated to match the target language.

_LANG_LABELS = {
    "mn": "Mongolian",
    "en": "English",
    "fr": "French",
    "it": "Italian",
    "es": "Spanish",
    "ko": "Korean",
    "zh": "Chinese (Simplified)",
    "ja": "Japanese",
    "ru": "Russian",
}


def _call_claude_for_translation(text, source_code, target_codes):
    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        return {"error": "ANTHROPIC_API_KEY is not set on the server."}
    source_label = _LANG_LABELS.get(source_code, source_code)
    target_lines = "\n".join(
        f"- {code}: {_LANG_LABELS.get(code, code)}"
        for code in target_codes
    )
    system = (
        "You are a professional travel-industry translator. "
        "You translate short tour-itinerary prose between languages while preserving "
        "every [[content:slug|label]] marker EXACTLY. The slug (after 'content:') and the "
        "brackets must stay identical; the label after the pipe must be translated to "
        "natural wording in the target language. Output strict JSON only — no markdown, "
        "no commentary, no code fences."
    )
    user = (
        f"Translate this prose from {source_label} ({source_code}) into the listed target "
        f"languages. Return a single JSON object whose keys are the target language codes "
        f"and whose values are the translated strings. Do not include the source language "
        f"in the output. Keep punctuation, line breaks, and [[content:...]] markers intact.\n\n"
        f"Source ({source_label}):\n{text}\n\n"
        f"Target languages:\n{target_lines}\n"
    )
    body = {
        "model": os.environ.get("AGENT_MODEL", "claude-sonnet-4-5-20250929"),
        "max_tokens": 2000,
        "system": system,
        "messages": [{"role": "user", "content": user}],
    }
    try:
        data = json.dumps(body).encode("utf-8")
    except Exception as e:
        return {"error": f"Could not serialize request: {e}"}
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=data, method="POST",
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=90) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        try: msg = e.read().decode("utf-8")[:500]
        except Exception: msg = str(e)
        return {"error": f"Anthropic API error {e.code}: {msg}"}
    except Exception as e:
        return {"error": f"{type(e).__name__}: {e}"}
    # Extract the text content block.
    out_text = ""
    for block in payload.get("content", []):
        if isinstance(block, dict) and block.get("type") == "text":
            out_text += block.get("text", "")
    out_text = out_text.strip()
    # Strip occasional ```json fences just in case.
    if out_text.startswith("```"):
        out_text = re.sub(r"^```(?:json)?\s*", "", out_text)
        out_text = re.sub(r"\s*```$", "", out_text)
    try:
        translations = json.loads(out_text)
    except Exception as e:
        return {"error": f"Claude response was not valid JSON: {e}", "raw": out_text[:500]}
    if not isinstance(translations, dict):
        return {"error": "Claude response was not a JSON object."}
    # Keep only requested target codes.
    cleaned = {code: str(translations.get(code) or "") for code in target_codes}
    return {"translations": cleaned}


def handle_translate_prose(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ) or {}
    text = (payload.get("text") or "").strip()
    source_code = (payload.get("sourceLang") or "en").strip().lower()
    raw_targets = payload.get("targetLangs") or []
    if not isinstance(raw_targets, list):
        return json_response(start_response, "400 Bad Request", {"error": "targetLangs must be a list."})
    target_codes = [c for c in (str(t).strip().lower() for t in raw_targets) if c and c != source_code and c in _LANG_LABELS]
    if not text:
        return json_response(start_response, "400 Bad Request", {"error": "Source text is empty."})
    if source_code not in _LANG_LABELS:
        return json_response(start_response, "400 Bad Request", {"error": f"Unsupported source language: {source_code}"})
    if not target_codes:
        return json_response(start_response, "200 OK", {"translations": {}})
    result = _call_claude_for_translation(text, source_code, target_codes)
    if "error" in result:
        return json_response(start_response, "502 Bad Gateway", result)
    return json_response(start_response, "200 OK", result)


def handle_accountant_download_zip(environ, start_response):
    """POST { ids: ["paymentRequestId", ...] } → streams a ZIP of every
    paid receipt for the given approved requests. Used by the
    Accountant page's bulk-download — sequential <a download> clicks
    were unreliable across browsers."""
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ) or {}
    raw_ids = payload.get("ids") or []
    if not isinstance(raw_ids, list) or not raw_ids:
        return json_response(start_response, "400 Bad Request", {"error": "Pick at least one row."})
    ids = {str(x) for x in raw_ids if x}
    workspace_filter = (active_workspace(environ) or "").upper()
    requests_all = read_payment_requests()
    trips_by_id = {t.get("id"): t for t in read_camp_trips()}

    import io
    import zipfile
    buf = io.BytesIO()
    written = 0
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        used_names = set()
        for r in requests_all:
            if r.get("id") not in ids:
                continue
            if workspace_filter and (r.get("workspace") or "").upper() != workspace_filter:
                continue
            if not r.get("paidDocumentId"):
                continue
            trip = trips_by_id.get(r.get("tripId")) or {}
            doc = next((d for d in (trip.get("documents") or []) if d.get("id") == r.get("paidDocumentId")), None)
            if not doc:
                continue
            file_path = (TRIP_UPLOADS_DIR / r.get("tripId", "") / (doc.get("storedName") or "")).resolve()
            if not str(file_path).startswith(str(TRIP_UPLOADS_DIR.resolve())) or not file_path.exists():
                continue
            # Avoid name collisions inside the zip (two managers might
            # have uploaded "receipt.png").
            base = doc.get("originalName") or file_path.name
            stem = Path(base).stem
            ext = Path(base).suffix
            arcname = base
            n = 1
            while arcname in used_names:
                arcname = f"{stem} ({n}){ext}"
                n += 1
            used_names.add(arcname)
            zf.write(file_path, arcname=arcname)
            written += 1
    if not written:
        return json_response(start_response, "404 Not Found", {"error": "None of the selected rows have a downloadable receipt."})
    data = buf.getvalue()
    workspace_label = workspace_filter or "ALL"
    filename = f"travelx-{workspace_label.lower()}-receipts-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.zip"
    headers = [
        ("Content-Type", "application/zip"),
        ("Content-Length", str(len(data))),
        ("Content-Disposition", f'attachment; filename="{filename}"'),
    ]
    start_response("200 OK", headers)
    return [data]


def handle_export_tourists(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ) or {}
    ids = payload.get("ids") or []
    trip_id = normalize_text(payload.get("tripId"))
    raw_trip_ids = payload.get("tripIds") or []
    if not isinstance(raw_trip_ids, list):
        raw_trip_ids = []
    trip_ids = [normalize_text(t) for t in raw_trip_ids if normalize_text(t)]
    if not trip_ids and trip_id:
        trip_ids = [trip_id]

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
    except Exception as exc:
        return json_response(start_response, "500 Internal Server Error", {"error": f"Excel export unavailable: {exc}"})

    all_tourists = read_tourists()
    trips_by_id = {t["id"]: t for t in read_camp_trips()}
    groups_by_id = {g["id"]: g for g in read_tourist_groups()}

    # Tourists carry an orderIndex set by the manual Move up / Move down
    # buttons on the trip page. The Excel export must follow that ordering
    # so the rooming list comes out in the same sequence the manager sees
    # on screen, with roommates already paired up.
    def tourist_sort_key(t):
        idx = t.get("orderIndex")
        if not isinstance(idx, (int, float)):
            try:
                idx = int(idx)
            except (TypeError, ValueError):
                idx = 9999
        return (idx, str(t.get("serial") or ""))

    # Build sections (trip, [tourists]). For "selected ids" we treat the
    # whole selection as a single section under that trip's banner.
    sections = []
    if ids:
        id_set = set(ids)
        records = sorted(
            [r for r in all_tourists if r.get("id") in id_set],
            key=tourist_sort_key,
        )
        if records:
            tid = records[0].get("tripId", "")
            sections.append((trips_by_id.get(tid), records))
    else:
        for tid in trip_ids:
            records = sorted(
                [r for r in all_tourists if r.get("tripId") == tid],
                key=tourist_sort_key,
            )
            if not records:
                continue
            sections.append((trips_by_id.get(tid), records))

    # DTX gets the streamlined rooming layout (no Group / Phone / Email,
    # extra Note column). Use the workspace the user is currently signed
    # into rather than the per-trip company so the layout is consistent
    # within a single download.
    workspace = active_workspace(environ) or ""
    is_dtx = workspace == "DTX"

    if is_dtx:
        columns = [
            ("#", "_index"),
            ("Serial", "serial"),
            ("Last name", "lastName"),
            ("First name", "firstName"),
            ("Gender", "gender"),
            ("Date of birth", "dob"),
            ("Nationality", "nationality"),
            ("Passport #", "passportNumber"),
            ("Passport issue date", "passportIssueDate"),
            ("Passport expiry", "passportExpiry"),
            ("Passport issued at", "passportIssuePlace"),
            ("Registration #", "registrationNumber"),
            ("Room", "_room"),
            ("Note", "notes"),
            ("Starting date", "_tripStartDate"),
        ]
    else:
        columns = [
            ("#", "_index"),
            ("Serial", "serial"),
            ("Last name", "lastName"),
            ("First name", "firstName"),
            ("Group", "_group"),
            ("Gender", "gender"),
            ("Date of birth", "dob"),
            ("Nationality", "nationality"),
            ("Passport #", "passportNumber"),
            ("Passport issue date", "passportIssueDate"),
            ("Passport expiry", "passportExpiry"),
            ("Passport issued at", "passportIssuePlace"),
            ("Registration #", "registrationNumber"),
            ("Phone", "phone"),
            ("Email", "email"),
            ("Room", "_room"),
            ("Starting date", "_tripStartDate"),
        ]

    column_widths = {
        "#": 5, "Serial": 14, "Last name": 16, "First name": 16, "Group": 18,
        "Gender": 8, "Date of birth": 12, "Nationality": 14,
        "Passport #": 14, "Passport issue date": 14, "Passport expiry": 14,
        "Passport issued at": 18, "Registration #": 14, "Phone": 14,
        "Email": 22, "Room": 14, "Note": 26, "Starting date": 12,
    }

    # Same palette as the front-end (trip-extras.js ROOM_PALETTE) so the
    # colour groupings on the printed Excel match what the manager sees
    # on screen.
    PALETTE = [
        ("FDE2E4", "7A1D2A"),
        ("DBEAFE", "1E3A8A"),
        ("DCFCE7", "14532D"),
        ("FEF3C7", "78350F"),
        ("E9D5FF", "5B21B6"),
        ("FED7AA", "7C2D12"),
        ("CFFAFE", "155E75"),
        ("FBCFE8", "831843"),
        ("D9F99D", "365314"),
        ("FDE68A", "713F12"),
    ]

    def room_key(t):
        if not t.get("roomType"):
            return None
        code = str(t.get("roomCode") or "").strip().lower()
        return (t.get("groupId") or "", t.get("roomType"), code)

    ROOM_TYPE_SHORT = {
        "single": "SGL", "double": "DBL", "twin": "TWIN",
        "triple": "TPL", "family": "FAM", "other": "OTH",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Rooming"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="20356F")
    centre = Alignment(horizontal="center", vertical="center")
    title_font = Font(bold=True, size=13, color="20356F")
    title_align = Alignment(horizontal="left", vertical="center")

    n_cols = len(columns)
    room_col_idx = next((i + 1 for i, (lbl, _) in enumerate(columns) if lbl == "Room"), None)

    # ── Letterhead ──────────────────────────────────────────────
    # Text-only banner. Embedding the PNG logo had it rendering at
    # native size and covering the agency name; a clean two-line text
    # block is safer and still reads like a proper agency document.
    if is_dtx:
        agency_name = "Delkhii Travel X"
    else:
        agency_name = "Steppe Mongolia"

    name_cell = ws.cell(row=1, column=1, value=agency_name)
    name_cell.font = Font(bold=True, size=20, color="20356F")
    name_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.row_dimensions[1].height = 32

    sub_cell = ws.cell(row=2, column=1, value=f"{agency_name} agency · Rooming list")
    sub_cell.font = Font(bold=True, size=12, color="475569")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.row_dimensions[2].height = 22

    current_row = 4  # blank row 3 separates letterhead from first section

    for section_idx, (trip, tourists) in enumerate(sections):
        # Section title (merged across all columns).
        if trip:
            tname = trip.get("tripName", "") or ""
            tserial = trip.get("serial", "") or ""
            tstart = trip.get("startDate", "") or ""
            tend = trip.get("endDate", "") or ""
            title_bits = []
            if tserial:
                title_bits.append(tserial)
            if tname:
                title_bits.append(tname)
            date_bit = ""
            if tstart and tend:
                date_bit = f"{tstart} → {tend}"
            elif tstart:
                date_bit = tstart
            title_text = " · ".join(title_bits)
            if date_bit:
                title_text = f"{title_text} · {date_bit}" if title_text else date_bit
        else:
            title_text = "Tourists"
        title_cell = ws.cell(row=current_row, column=1, value=title_text or "Tourists")
        title_cell.font = title_font
        title_cell.alignment = title_align
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=n_cols)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Column headers.
        for col_idx, (label, _key) in enumerate(columns, start=1):
            c = ws.cell(row=current_row, column=col_idx, value=label)
            c.font = header_font
            c.fill = header_fill
            c.alignment = centre
        current_row += 1

        # Assign a colour per (groupId, roomType, roomCode) bucket in the
        # order they first appear in the section, so the same room maps
        # to the same colour across rows even if the list isn't sorted.
        room_color_map = {}
        for t in tourists:
            k = room_key(t)
            if k is None:
                continue
            if k not in room_color_map:
                room_color_map[k] = PALETTE[len(room_color_map) % len(PALETTE)]

        section_data_start = current_row
        for row_index, t in enumerate(tourists, start=1):
            group = groups_by_id.get(t.get("groupId")) or {}
            tstart = (trip or {}).get("startDate", "")
            for col_idx, (label, key) in enumerate(columns, start=1):
                if key == "_index":
                    val = row_index
                elif key == "_group":
                    val = group.get("name") or t.get("groupSerial", "")
                elif key == "_room":
                    if t.get("roomType"):
                        rt_short = ROOM_TYPE_SHORT.get(
                            str(t.get("roomType", "")).lower(),
                            str(t.get("roomType", "")).upper()[:4],
                        )
                        val = f"{t.get('roomCode') or '—'} {rt_short}"
                    else:
                        val = ""
                elif key == "_tripStartDate":
                    val = tstart
                else:
                    val = t.get(key, "")
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                if label == "Room":
                    k = room_key(t)
                    if k and k in room_color_map:
                        bg, fg = room_color_map[k]
                        cell.fill = PatternFill("solid", fgColor=bg)
                        cell.font = Font(color=fg, bold=True)
                    cell.alignment = centre
            current_row += 1

        # Merge contiguous same-room cells in the Room column so two
        # roommates show up as one shared block — what the user calls
        # "rooming merged with relative person".
        if room_col_idx and tourists:
            run_start = section_data_start
            run_key = room_key(tourists[0])
            for i in range(1, len(tourists)):
                t_key = room_key(tourists[i])
                if t_key == run_key and t_key is not None:
                    continue
                run_end = section_data_start + i - 1
                if run_key is not None and run_end > run_start:
                    ws.merge_cells(start_row=run_start, start_column=room_col_idx,
                                   end_row=run_end, end_column=room_col_idx)
                run_start = section_data_start + i
                run_key = t_key
            last_row = section_data_start + len(tourists) - 1
            if run_key is not None and last_row > run_start:
                ws.merge_cells(start_row=run_start, start_column=room_col_idx,
                               end_row=last_row, end_column=room_col_idx)

        # Blank separator between trip sections.
        current_row += 1

    # Column widths.
    for col_idx, (label, _) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = column_widths.get(label, 14)

    # If no sections were resolved (e.g. selected ids that no longer
    # exist) at least leave a single header row so the file isn't empty.
    if not sections:
        for col_idx, (label, _key) in enumerate(columns, start=1):
            c = ws.cell(row=1, column=col_idx, value=label)
            c.font = header_font
            c.fill = header_fill
            c.alignment = centre

    buf = BytesIO()
    wb.save(buf)
    body = buf.getvalue()

    if len(sections) > 1:
        filename = f"rooming-list-{len(sections)}-trips.xlsx"
    elif sections and sections[0][0]:
        serial = (sections[0][0] or {}).get("serial", "") or ""
        filename = f"rooming-{serial or 'trip'}.xlsx"
    else:
        filename = "tourists.xlsx"
    headers_out = [
        ("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        ("Content-Length", str(len(body))),
        ("Content-Disposition", f'attachment; filename="{filename}"'),
    ]
    start_response("200 OK", headers_out)
    return [body]


def handle_email_rooming(environ, start_response):
    """POST /api/tourists/email-rooming — same xlsx the Download
    rooming button generates, sent to a recipient via Resend with a
    friendly English body. Reuses handle_export_tourists internally
    via a captured-response shim so the rooming layout stays in sync
    with the download path."""
    import io
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ) or {}
    recipient = (payload.get("to") or payload.get("recipientEmail") or "").strip()
    if not recipient or "@" not in recipient:
        return json_response(start_response, "400 Bad Request", {"error": "Recipient email is required."})
    recipient_name = (payload.get("recipientName") or "").strip()
    extra_message = (payload.get("message") or "").strip()

    # Capture the xlsx body from handle_export_tourists by feeding it a
    # shim start_response and reusing the same input payload.
    captured = {"status": "", "headers": [], "body": b""}
    def shim_start_response(status, headers):
        captured["status"] = status
        captured["headers"] = headers
        return lambda data: None
    # Re-pack the trip ids into a fresh environ-like input. Easier to
    # just pass the JSON body forward by re-injecting via a local wsgi
    # input. handle_export_tourists reads collect_json(environ), so we
    # duplicate the environ with a fresh stream.
    body_json = json.dumps({
        "tripIds": payload.get("tripIds") or ([] if not payload.get("tripId") else [payload.get("tripId")]),
        "ids": payload.get("ids") or [],
        "tripId": payload.get("tripId") or "",
    }).encode("utf-8")
    fake_environ = dict(environ)
    fake_environ["wsgi.input"] = io.BytesIO(body_json)
    fake_environ["CONTENT_LENGTH"] = str(len(body_json))
    fake_environ["REQUEST_METHOD"] = "POST"
    chunks = handle_export_tourists(fake_environ, shim_start_response)
    if not captured["status"].startswith("200"):
        # Surface the export error verbatim.
        try:
            err_text = b"".join(chunks).decode("utf-8", "replace")
        except Exception:
            err_text = "Export failed."
        return json_response(start_response, "400 Bad Request", {"error": err_text})
    body = b"".join(chunks)
    filename = "rooming.xlsx"
    for k, v in captured["headers"]:
        if k.lower() == "content-disposition":
            m = re.search(r'filename="([^"]+)"', v)
            if m:
                filename = m.group(1)
            break

    # Friendly English body. Trip serials referenced so the recipient
    # can match the file to the booking immediately.
    trip_ids = payload.get("tripIds") or ([] if not payload.get("tripId") else [payload.get("tripId")])
    trips_by_id = {t["id"]: t for t in read_camp_trips()}
    trip_lines = []
    for tid in trip_ids:
        t = trips_by_id.get(tid)
        if not t:
            continue
        serial = t.get("serial") or ""
        name = t.get("tripName") or ""
        start = (t.get("startDate") or "")[:10]
        end = (t.get("endDate") or "")[:10]
        date_part = f" — {start}" + (f" → {end}" if end else "") if start else ""
        trip_lines.append(f"  • {serial} · {name}{date_part}")

    greeting = f"Dear {recipient_name},\n\n" if recipient_name else "Hello,\n\n"
    body_text_parts = [greeting + "Please find the rooming list attached for the trip" + ("s" if len(trip_lines) > 1 else "") + " below:\n"]
    if trip_lines:
        body_text_parts.append("\n".join(trip_lines) + "\n")
    if extra_message:
        body_text_parts.append("\n" + extra_message + "\n")
    body_text_parts.append(
        "\nThe attached spreadsheet contains the full participant list with passport details, "
        "room assignments and contact information. Let us know if any change is needed.\n\n"
        "Best regards,"
    )
    email_body = "".join(body_text_parts)

    workspace = (active_workspace(environ) or "DTX").upper()
    company_name = "Unlock Steppe Mongolia" if workspace == "USM" else "Дэлхий Трэвел Икс"

    subject = "Rooming list — " + (
        f"{len(trip_lines)} trip{'s' if len(trip_lines) != 1 else ''}"
        if trip_lines else "TravelX"
    )

    result = _tool_send_email({
        "to": recipient,
        "subject": subject,
        "body": email_body,
        "attachments": [{
            "filename": filename,
            "content": base64.b64encode(body).decode("ascii"),
            "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }],
        "_company_name": company_name,
        "_skip_footer": True,
    }, actor)
    if result.get("error"):
        return json_response(start_response, "500 Internal Server Error", {"error": result["error"]})
    return json_response(start_response, "200 OK", {"ok": True, "to": recipient, "filename": filename})


def ensure_checkout_from_nights(check_in, nights, check_out):
    return normalize_stay_fields(check_in, nights, check_out)[2]


def build_camp_reservation(payload, actor=None):
    check_in, nights, check_out = normalize_stay_fields(payload.get("checkIn"), payload.get("nights"), payload.get("checkOut"))
    created_date = normalize_text(payload.get("createdDate")) or today_mongolia()
    return {
        "id": str(uuid4()),
        "createdAt": now_mongolia().isoformat(),
        "createdDate": created_date,
        "tripId": normalize_text(payload.get("tripId")),
        "tripName": normalize_text(payload.get("tripName")),
        "reservationName": normalize_text(payload.get("reservationName")) or normalize_text(payload.get("tripName")),
        "language": normalize_text(payload.get("language")) or "Other",
        "campName": normalize_text(payload.get("campName")),
        "locationName": normalize_text(payload.get("locationName")),
        "newCampName": normalize_text(payload.get("newCampName")),
        "reservationType": normalize_text(payload.get("reservationType")).lower() or "camp",
        "checkIn": check_in,
        "checkOut": check_out,
        "clientCount": parse_int(payload.get("clientCount")),
        "staffCount": parse_int(payload.get("staffCount")),
        "staffAssignment": normalize_text(payload.get("staffAssignment")),
        "gerCount": parse_int(payload.get("gerCount")),
        "nights": nights,
        "roomType": normalize_text(payload.get("roomType")),
        "breakfast": normalize_text(payload.get("breakfast")) or "No",
        "lunch": normalize_text(payload.get("lunch")) or "No",
        "dinner": normalize_text(payload.get("dinner")) or "No",
        "status": normalize_text(payload.get("status")).lower() or "pending",
        "deposit": parse_int(payload.get("deposit")),
        "depositPaidDate": normalize_text(payload.get("depositPaidDate")),
        "secondPayment": parse_int(payload.get("secondPayment")),
        "secondPaidDate": normalize_text(payload.get("secondPaidDate")),
        # Optional 3rd / 4th payment stages — camps occasionally
        # invoice in 3 or 4 installments. Empty by default; the
        # Edit camp payment modal exposes them via "+ Add payment".
        "thirdPayment": parse_int(payload.get("thirdPayment")),
        "thirdPaidDate": normalize_text(payload.get("thirdPaidDate")),
        "fourthPayment": parse_int(payload.get("fourthPayment")),
        "fourthPaidDate": normalize_text(payload.get("fourthPaidDate")),
        "totalPayment": parse_int(payload.get("totalPayment")),
        "balancePayment": parse_int(payload.get("balancePayment")),
        "paidAmount": parse_int(payload.get("paidAmount")),
        "paymentStatus": normalize_text(payload.get("paymentStatus")) or "in_progress",
        "notes": normalize_text(payload.get("notes")),
        "company": normalize_company(payload.get("company")),
        "createdBy": actor_snapshot(actor),
        "updatedAt": "",
        "updatedBy": actor_snapshot(actor),
    }


def validate_camp_reservation(data):
    required = ["tripId", "tripName", "reservationName", "campName", "reservationType", "checkIn", "checkOut", "status", "roomType"]
    missing = [field for field in required if not data.get(field)]
    if missing:
        return f"Missing required fields: {', '.join(missing)}"
    if data.get("clientCount", 0) <= 0:
        return "Number of clients must be greater than 0"
    if data.get("gerCount", 0) <= 0:
        return "Number of gers must be greater than 0"
    if data.get("nights", 0) <= 0:
        return "Number of nights must be greater than 0"
    check_in_date = parse_date_input(data.get("checkIn"))
    check_out_date = parse_date_input(data.get("checkOut"))
    if check_in_date and check_out_date and check_out_date <= check_in_date:
        return "Check-out must be after check-in"
    return None


def camp_summary(records):
    return {
        "total": len(records),
        "confirmed": len([record for record in records if record["status"] == "confirmed"]),
        "pending": len([record for record in records if record["status"] == "pending"]),
        "cancelled": len([record for record in records if record["status"] == "cancelled"]),
        "rejected": len([record for record in records if record["status"] == "rejected"]),
        "trips": len({record["tripId"] for record in records if record.get("tripId")}),
    }


def finance_summary(entries):
    income = sum(entry["amount"] for entry in entries if entry["type"] == "income")
    expense = sum(entry["amount"] for entry in entries if entry["type"] == "expense")
    bonus = sum(entry["bonusAmount"] for entry in entries)
    return {
        "income": income,
        "expense": expense,
        "profit": income - expense,
        "bonus": bonus,
    }


def build_manager_summary(store):
    today = today_mongolia()
    tasks = store.get("tasks", [])
    reminders = store.get("reminders", [])
    contacts = store.get("contacts", [])
    return {
        "tasks": {
            "total": len(tasks),
            "open": len([item for item in tasks if item.get("status") != "done"]),
            "done": len([item for item in tasks if item.get("status") == "done"]),
            "highPriority": len([item for item in tasks if item.get("priority") == "high"]),
        },
        "reminders": {
            "total": len(reminders),
            "active": len([item for item in reminders if item.get("status") != "done"]),
            "today": len([item for item in reminders if str(item.get("reminderDate") or "")[:10] == today and item.get("status") != "done"]),
            "overdue": len(
                [
                    item
                    for item in reminders
                    if item.get("status") != "done"
                    and normalize_text(item.get("reminderDate"))[:10]
                    and normalize_text(item.get("reminderDate"))[:10] < today
                ]
            ),
        },
        "contacts": {
            "total": len(contacts),
            "priority": len([item for item in contacts if item.get("status") == "priority"]),
            "clients": len([item for item in contacts if item.get("type") == "client"]),
            "partners": len([item for item in contacts if item.get("type") == "partner"]),
        },
    }


def _task_due_datetime(task):
    """Compose dueDate + dueTime (default 09:00) into a Mongolia-aware datetime, or None."""
    raw_date = normalize_text(task.get("dueDate"))[:10]
    if not raw_date or not re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw_date):
        return None
    due_time = normalize_text(task.get("dueTime")) or "09:00"
    if not re.fullmatch(r"\d{2}:\d{2}", due_time):
        due_time = "09:00"
    try:
        return datetime.fromisoformat(f"{raw_date}T{due_time}:00").replace(tzinfo=MONGOLIA_TZ)
    except Exception:
        return None


def _find_user_by_name(name):
    target = normalize_text(name).lower()
    if not target:
        return None
    for user in read_users():
        if user.get("status") != "approved":
            continue
        if normalize_text(user.get("fullName")).lower() == target:
            return user
    return None


def _task_owner_names(task):
    raw = task.get("owners")
    if isinstance(raw, list) and raw:
        return [normalize_text(name) for name in raw if normalize_text(name)]
    single = normalize_text(task.get("owner"))
    return [single] if single else []


def send_task_assignment_email(task, actor, owner_filter=None):
    """Email each task assignee that a manager assigned this task to them.
    Best-effort: any failure is logged and swallowed so the task save still
    succeeds. Skips silently if no assignee has a resolvable email.

    If `owner_filter` is provided (a set of lowercased names), only those
    owners get an email — used on update so we don't re-spam owners who
    were already assigned before this edit."""
    owner_names = _task_owner_names(task)
    if not owner_names:
        return
    title = normalize_text(task.get("title")) or "Untitled task"
    assigner = normalize_text((actor or {}).get("fullName")) or normalize_text((actor or {}).get("email")) or "Менежер"
    due_dt = _task_due_datetime(task)
    due_label = due_dt.strftime("%Y-%m-%d %H:%M") if due_dt else "—"
    # Don't run the note through normalize_text — that collapses line
    # breaks into single spaces and the multi-paragraph note the user
    # wrote ends up as a one-liner. Trim, escape, then convert real
    # newlines to <br> so the email preserves the original formatting.
    note_raw = str(task.get("note") or "").strip()
    note_html = (
        f"<div style=\"margin-top:12px;color:#475569;white-space:pre-wrap;\">"
        f"<strong>Note:</strong><br>{html.escape(note_raw).replace(chr(10), '<br>')}"
        f"</div>"
    ) if note_raw else ""
    for owner_name in owner_names:
        if owner_filter is not None and owner_name.lower() not in owner_filter:
            continue
        user = _find_user_by_name(owner_name)
        owner_email = normalize_text(user.get("email")) if user else ""
        if not owner_email:
            continue
        body_html = (
            f"<p>Сайн байна уу, {html.escape(owner_name)}.</p>"
            f"<p><strong>{html.escape(assigner)}</strong> танд <strong>{html.escape(title)}</strong> "
            f"гэсэн шинэ ажил оноолоо.</p>"
            f"<p>Дуусах хугацаа: <strong>{html.escape(due_label)}</strong></p>"
            f"{note_html}"
        )
        try:
            _tool_send_email(
                {
                    "to": owner_email,
                    "subject": f"NEW TASK · {title}",
                    "body": "",
                    "_body_html_override": body_html,
                    "_skip_footer": True,
                },
                None,
            )
        except Exception as exc:
            print(f"[task-assignment] email failed for {owner_email}: {exc}", flush=True)


_TASK_REMINDER_SCAN_AT = [0.0]
_TASK_REMINDER_MIN_INTERVAL_SEC = 60


def scan_task_reminders():
    """Send 6h-before-due email + bell notification for tasks. Idempotent via reminderSentAt.

    Triggered inline on dashboard/notification reads. Throttled to once per minute per worker
    to keep request latency low when many polls overlap.
    """
    nowmono = time.monotonic()
    if nowmono - _TASK_REMINDER_SCAN_AT[0] < _TASK_REMINDER_MIN_INTERVAL_SEC:
        return
    _TASK_REMINDER_SCAN_AT[0] = nowmono

    try:
        store = read_manager_dashboard()
    except Exception:
        return
    tasks = store.get("tasks") or []
    if not tasks:
        return

    now_local = now_mongolia()
    horizon = timedelta(hours=6)
    changed = False

    for task in tasks:
        if task.get("status") in {"done", "cancelled"}:
            continue
        due_dt = _task_due_datetime(task)
        if not due_dt:
            continue
        delta = due_dt - now_local

        owner_names = _task_owner_names(task)
        title = normalize_text(task.get("title")) or "Untitled task"
        due_label = due_dt.strftime("%Y-%m-%d %H:%M")
        owner_summary = ", ".join(owner_names)
        # Preserve newlines in the user's note so multi-line content keeps
        # its formatting in both the reminder and overdue emails.
        note_raw = str(task.get("note") or "").strip()
        note_html = (
            f"<div style=\"margin-top:12px;color:#475569;white-space:pre-wrap;\">"
            f"<strong>Note:</strong><br>{html.escape(note_raw).replace(chr(10), '<br>')}"
            f"</div>"
        ) if note_raw else ""
        # Reference images: attach all as real email attachments + reference
        # each inline via cid so the recipient sees them in-body in clients
        # that support content-id (Gmail, Apple Mail, Outlook). If a client
        # strips inline images, the attachments are still visible.
        image_attachments = []
        image_html = ""
        images = _task_images_list(task)
        if images:
            tile_html = []
            for idx, img in enumerate(images, start=1):
                p = _task_image_path(task.get("id"), img["id"], img["ext"])
                if not p or not p.exists():
                    continue
                try:
                    image_bytes = p.read_bytes()
                except OSError:
                    continue
                ext = (img["ext"] or ".jpg").lstrip(".").lower() or "jpg"
                mime = {
                    "jpg": "image/jpeg", "jpeg": "image/jpeg",
                    "png": "image/png", "gif": "image/gif",
                    "webp": "image/webp", "heic": "image/heic", "heif": "image/heif",
                }.get(ext, "application/octet-stream")
                fn = f"task-reference-{idx}.{ext}"
                content_id = f"task-image-{task.get('id')}-{img['id']}"
                image_attachments.append({
                    "filename": fn,
                    "content": base64.b64encode(image_bytes).decode("ascii"),
                    "content_id": content_id,
                    "type": mime,
                })
                tile_html.append(
                    f"<img src=\"cid:{content_id}\" alt=\"task reference {idx}\" "
                    f"style=\"max-width:280px;width:100%;height:auto;border-radius:8px;border:1px solid #e5e7eb;margin:0 8px 8px 0;\" />"
                )
            if tile_html:
                image_html = (
                    f"<div style=\"margin-top:14px;\">"
                    f"<p style=\"color:#475569;\">📎 Хавсаргасан зураг ({len(tile_html)}):</p>"
                    + "".join(tile_html) +
                    f"</div>"
                )

        # Stage 1: 6 hours before due — "approaching" reminder.
        if (
            not normalize_text(task.get("reminderSentAt"))
            and delta > timedelta(seconds=0)
            and delta <= horizon
        ):
            # Real hours remaining at send time, rounded to whole hours
            # (1-6). The user's reference text says "6 цаг үлдлээ" but the
            # number should match reality so the manager isn't confused
            # if the cron fires slightly later.
            hours_left = max(1, int(round(delta.total_seconds() / 3600)))
            for owner_name in owner_names or [""]:
                user = _find_user_by_name(owner_name) if owner_name else None
                owner_email = normalize_text(user.get("email")) if user else ""
                body_html = (
                    f"<p>Сайн байна уу{', ' + html.escape(owner_name) if owner_name else ''}.</p>"
                    f"<p>Таны <strong>{html.escape(title)}</strong> ажлын хугацаа дуусахад <strong>{hours_left} цаг</strong> үлдлээ.</p>"
                    f"<p>Дуусах хугацаа: <strong>{html.escape(due_label)}</strong></p>"
                    "<p>Та цагтаа ажлаа дуусгахаа бүү мартаарай.</p>"
                    f"{note_html}"
                    f"{image_html}"
                )
                if owner_email:
                    try:
                        _tool_send_email(
                            {
                                "to": owner_email,
                                "subject": f"REMINDER TASK · {title}",
                                "body": "",
                                "_body_html_override": body_html,
                                "_skip_footer": True,
                                "attachments": image_attachments,
                            },
                            None,
                        )
                    except Exception as exc:
                        print(f"[task-reminder] approaching email failed for {owner_email}: {exc}", flush=True)
                try:
                    log_notification(
                        "task.reminder",
                        {"id": "system", "email": "noreply", "name": "TravelX"},
                        f"REMINDER TASK · {title}",
                        detail=f"Due {due_label}" + (f" · {owner_summary}" if owner_summary else ""),
                        meta={"id": task.get("id"), "key": "tasks", "userId": (user or {}).get("id") or ""},
                    )
                except Exception:
                    pass
            task["reminderSentAt"] = now_local.isoformat()
            changed = True

        # Stage 2: due time has passed — "overdue" notification, sent once.
        if (
            not normalize_text(task.get("overdueSentAt"))
            and delta <= timedelta(seconds=0)
        ):
            for owner_name in owner_names or [""]:
                user = _find_user_by_name(owner_name) if owner_name else None
                owner_email = normalize_text(user.get("email")) if user else ""
                body_html = (
                    f"<p>Сайн байна уу{', ' + html.escape(owner_name) if owner_name else ''}.</p>"
                    f"<p>Таны <strong>{html.escape(title)}</strong> ажлын дуусах хугацаа хэтэрсэн байна. "
                    "Аль болох түргэн гүйцэтгэж дуусгана уу.</p>"
                    f"{note_html}"
                    f"{image_html}"
                )
                if owner_email:
                    try:
                        _tool_send_email(
                            {
                                "to": owner_email,
                                "subject": f"OVERDUE TASK · {title}",
                                "body": "",
                                "_body_html_override": body_html,
                                "_skip_footer": True,
                                "attachments": image_attachments,
                            },
                            None,
                        )
                    except Exception as exc:
                        print(f"[task-reminder] overdue email failed for {owner_email}: {exc}", flush=True)
                try:
                    log_notification(
                        "task.overdue",
                        {"id": "system", "email": "noreply", "name": "TravelX"},
                        f"OVERDUE TASK · {title}",
                        detail=f"Due {due_label}" + (f" · {owner_summary}" if owner_summary else ""),
                        meta={"id": task.get("id"), "key": "tasks", "userId": (user or {}).get("id") or ""},
                    )
                except Exception:
                    pass
            task["overdueSentAt"] = now_local.isoformat()
            changed = True

    if changed:
        try:
            write_manager_dashboard(store)
        except Exception as exc:
            print(f"[task-reminder] write failed: {exc}", flush=True)


def handle_get_manager_dashboard(start_response):
    try:
        scan_task_reminders()
    except Exception as exc:
        print(f"[task-reminder] scan error: {exc}", flush=True)
    store = read_manager_dashboard()
    return json_response(
        start_response,
        "200 OK",
        {
            **store,
            "summary": build_manager_summary(store),
        },
    )


def _normalize_owner_list(payload):
    """Accept either `owners` (list) or legacy `owner` (string).
    Returns (owners_list, primary_owner_string) — primary is owners[0] or "".
    Used to keep email + reminder lookup working while moving to multi-assign."""
    raw = payload.get("owners")
    owners: list[str] = []
    if isinstance(raw, list):
        owners = [normalize_text(item) for item in raw if normalize_text(item)]
    elif isinstance(raw, str):
        owners = [normalize_text(part) for part in raw.split(",") if normalize_text(part)]
    if not owners:
        single = normalize_text(payload.get("owner"))
        if single:
            owners = [single]
    # de-dupe while preserving order
    seen = set()
    deduped: list[str] = []
    for name in owners:
        key = name.lower()
        if key in seen:
            continue
        seen.add(key)
        deduped.append(name)
    return deduped, deduped[0] if deduped else ""


def build_manager_task(payload):
    owners, primary = _normalize_owner_list(payload)
    record = {
        "id": str(uuid4()),
        "title": normalize_text(payload.get("title")),
        "owners": owners,
        # Keep `owner` populated for older callers / agents that still read it.
        "owner": primary,
        "priority": normalize_text(payload.get("priority")).lower() or "medium",
        "status": normalize_text(payload.get("status")).lower() or "todo",
        "dueDate": normalize_text(payload.get("dueDate")),
        "dueTime": normalize_text(payload.get("dueTime")),
        "destinations": normalize_tag_list(payload.get("destinations")),
        "note": normalize_text(payload.get("note")),
        "reminderSentAt": normalize_text(payload.get("reminderSentAt")),
        "createdAt": now_mongolia().isoformat(),
        "updatedAt": now_mongolia().isoformat(),
    }
    # Preserve image attachments across updates — the actual files are
    # uploaded via a separate endpoint, not part of this JSON payload.
    images = payload.get("images")
    if isinstance(images, list) and images:
        clean = []
        for entry in images:
            if isinstance(entry, dict) and entry.get("id") and entry.get("ext"):
                clean.append({"id": str(entry["id"]), "ext": str(entry["ext"])})
        if clean:
            record["images"] = clean
    elif normalize_text(payload.get("imageExt")):
        # Legacy single-image field — preserve as-is until the next upload
        # converts it into the array shape.
        record["imageExt"] = normalize_text(payload.get("imageExt"))
    return record


def validate_manager_task(task):
    if len(task.get("title", "")) < 2:
        return "Task title must be at least 2 characters"
    if task.get("priority") not in {"low", "medium", "high"}:
        return "Task priority is invalid"
    if task.get("status") not in {"todo", "in-progress", "done", "cancelled"}:
        return "Task status is invalid"
    if task.get("dueDate") and not parse_date_input(task.get("dueDate")):
        return "Task due date must be in YYYY-MM-DD format"
    due_time = task.get("dueTime")
    if due_time and not re.fullmatch(r"\d{2}:\d{2}", due_time):
        return "Task due time must be HH:MM"
    return None


# ── Task reference images (multiple attachments per task, throwaway) ──
# Stored at TASK_ATTACHMENTS_DIR/<task_id>__<img_id>.<ext>. Not in the
# gallery — these are quick context images (screenshots, photos) that
# get deleted with the task.

def _task_images_list(task):
    """Return the task's images as a list of {id, ext}, migrating the
    legacy single-image `imageExt` field on the fly."""
    images = task.get("images")
    if isinstance(images, list) and images:
        out = []
        for entry in images:
            if isinstance(entry, dict) and entry.get("id") and entry.get("ext"):
                out.append({"id": str(entry["id"]), "ext": str(entry["ext"])})
        return out
    legacy_ext = task.get("imageExt")
    if legacy_ext:
        return [{"id": "legacy", "ext": str(legacy_ext)}]
    return []


def _task_image_path(task_id, image_id, ext):
    """Resolve the on-disk path for a specific task image."""
    if not task_id or not image_id or not ext:
        return None
    if image_id == "legacy":
        return TASK_ATTACHMENTS_DIR / f"{task_id}{ext}"
    return TASK_ATTACHMENTS_DIR / f"{task_id}__{image_id}{ext}"


def _delete_task_image_file(task_id, image_id, ext):
    p = _task_image_path(task_id, image_id, ext)
    if p and p.exists():
        try:
            p.unlink()
        except OSError:
            pass


def _delete_all_task_images(task_id):
    """Cascade-delete every file belonging to this task — covers both
    the legacy single-file naming and the new __<imgId> naming."""
    if not task_id:
        return
    for candidate in TASK_ATTACHMENTS_DIR.glob(f"{task_id}.*"):
        if candidate.is_file():
            try:
                candidate.unlink()
            except OSError:
                pass
    for candidate in TASK_ATTACHMENTS_DIR.glob(f"{task_id}__*"):
        if candidate.is_file():
            try:
                candidate.unlink()
            except OSError:
                pass


def handle_task_image_upload(environ, start_response, task_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    store = read_manager_dashboard()
    task = next((t for t in store.get("tasks", []) if t.get("id") == task_id), None)
    if not task:
        return json_response(start_response, "404 Not Found", {"error": "Task not found"})
    fields, files = parse_multipart(environ)
    upload = files.get("file")
    if not upload or not upload.get("data"):
        return json_response(start_response, "400 Bad Request", {"error": "No file provided"})
    original_name = upload.get("filename") or "image"
    ext = Path(original_name).suffix.lower()
    if ext not in TASK_IMAGE_EXTS:
        return json_response(start_response, "400 Bad Request", {"error": f"Image type {ext} not allowed"})
    data = upload["data"]
    if len(data) > MAX_UPLOAD_BYTES:
        return json_response(start_response, "400 Bad Request", {"error": "Image too large (max 10 MB)"})
    ensure_data_store()
    image_id = uuid4().hex[:12]
    target = TASK_ATTACHMENTS_DIR / f"{task_id}__{image_id}{ext}"
    target.write_bytes(data)
    images = _task_images_list(task)
    images.append({"id": image_id, "ext": ext})
    task["images"] = images
    if "imageExt" in task:
        # Promote the legacy single image into the list once a second is
        # added, so we don't keep the dual representation around.
        del task["imageExt"]
    task["updatedAt"] = now_mongolia().isoformat()
    write_manager_dashboard(store)
    return json_response(start_response, "200 OK", {"ok": True, "image": {"id": image_id, "ext": ext}, "images": images})


def handle_task_image_serve(environ, start_response, task_id, image_id=None):
    if not require_login(environ, start_response):
        return []
    store = read_manager_dashboard()
    task = next((t for t in store.get("tasks", []) if t.get("id") == task_id), None)
    if not task:
        return json_response(start_response, "404 Not Found", {"error": "Task not found"})
    images = _task_images_list(task)
    if not images:
        return json_response(start_response, "404 Not Found", {"error": "No image"})
    target = None
    if image_id:
        match = next((img for img in images if img["id"] == image_id), None)
        if match:
            target = _task_image_path(task_id, match["id"], match["ext"])
    else:
        # Back-compat: serve the first image when no id is given.
        target = _task_image_path(task_id, images[0]["id"], images[0]["ext"])
    if not target or not target.exists():
        return json_response(start_response, "404 Not Found", {"error": "Image file missing"})
    return file_response(start_response, target)


def handle_task_image_delete(environ, start_response, task_id, image_id=None):
    if not require_login(environ, start_response):
        return []
    store = read_manager_dashboard()
    task = next((t for t in store.get("tasks", []) if t.get("id") == task_id), None)
    if not task:
        return json_response(start_response, "404 Not Found", {"error": "Task not found"})
    images = _task_images_list(task)
    if not images:
        return json_response(start_response, "200 OK", {"ok": True})
    if image_id:
        keep = []
        removed = False
        for img in images:
            if img["id"] == image_id and not removed:
                _delete_task_image_file(task_id, img["id"], img["ext"])
                removed = True
                continue
            keep.append(img)
        task["images"] = keep
    else:
        # No id — delete every image (legacy behaviour).
        for img in images:
            _delete_task_image_file(task_id, img["id"], img["ext"])
        task["images"] = []
    if "imageExt" in task:
        del task["imageExt"]
    task["updatedAt"] = now_mongolia().isoformat()
    write_manager_dashboard(store)
    return json_response(start_response, "200 OK", {"ok": True, "images": task["images"]})


def build_manager_reminder(payload):
    return {
        "id": str(uuid4()),
        "title": normalize_text(payload.get("title")),
        "reminderDate": normalize_text(payload.get("reminderDate")),
        "status": normalize_text(payload.get("status")).lower() or "active",
        "audience": normalize_text(payload.get("audience")).lower() or "team",
        "note": normalize_text(payload.get("note")),
        "createdAt": now_mongolia().isoformat(),
        "updatedAt": now_mongolia().isoformat(),
    }


def validate_manager_reminder(reminder):
    if len(reminder.get("title", "")) < 2:
        return "Reminder title must be at least 2 characters"
    if reminder.get("status") not in {"active", "done"}:
        return "Reminder status is invalid"
    if reminder.get("audience") not in {"team", "client", "internal"}:
        return "Reminder audience is invalid"
    reminder_date = normalize_text(reminder.get("reminderDate"))
    if reminder_date and not re.fullmatch(r"\d{4}-\d{2}-\d{2}(T\d{2}:\d{2})?", reminder_date):
        return "Reminder date must be YYYY-MM-DD or YYYY-MM-DDTHH:MM"
    return None


def build_manager_contact(payload):
    return {
        "id": str(uuid4()),
        "name": normalize_text(payload.get("name")),
        "phone": normalize_text(payload.get("phone")),
        "company": normalize_text(payload.get("company")),
        "type": normalize_text(payload.get("type")).lower() or "client",
        "status": normalize_text(payload.get("status")).lower() or "new",
        "lastContacted": normalize_text(payload.get("lastContacted")),
        "destinations": normalize_tag_list(payload.get("destinations")),
        "note": normalize_text(payload.get("note")),
        "createdAt": now_mongolia().isoformat(),
        "updatedAt": now_mongolia().isoformat(),
    }


def validate_manager_contact(contact):
    if len(contact.get("name", "")) < 2:
        return "Contact name must be at least 2 characters"
    if len(contact.get("phone", "")) < 5:
        return "Phone number must be at least 5 characters"
    if contact.get("type") not in {"client", "staff", "partner", "vendor"}:
        return "Contact type is invalid"
    if contact.get("status") not in {"new", "warm", "priority", "inactive"}:
        return "Contact status is invalid"
    if contact.get("lastContacted") and not parse_date_input(contact.get("lastContacted")):
        return "Last contacted date must be in YYYY-MM-DD format"
    return None


def fifa_active_sales(sales):
    return [sale for sale in sales if normalize_text(sale.get("saleStatus")).lower() != "cancelled"]


def fifa_committed_sales(sales):
    return [sale for sale in sales if normalize_text(sale.get("saleStatus")).lower() == "confirmed"]


def fifa_ticket_sales(sales, ticket_id, excluded_sale_id=None):
    return [
        sale
        for sale in fifa_committed_sales(sales)
        if sale.get("id") != excluded_sale_id
        and (
            sale.get("ticketId") == ticket_id
            or ticket_id in (sale.get("ticketIds") or [])
        )
    ]


def fifa_sale_quantity_for_ticket(sale, ticket_id):
    ticket_ids = [normalize_text(value) for value in (sale.get("ticketIds") or []) if normalize_text(value)]
    primary_ticket_id = normalize_text(sale.get("ticketId"))
    if ticket_ids:
        return 1 if ticket_id in ticket_ids else 0
    if primary_ticket_id == ticket_id:
        return max(parse_int(sale.get("quantity")), 0)
    return 0


def fifa_ticket_available_quantity(ticket, sales, excluded_sale_id=None):
    sold_quantity = sum(
        fifa_sale_quantity_for_ticket(sale, ticket.get("id"))
        for sale in fifa_ticket_sales(sales, ticket.get("id"), excluded_sale_id)
    )
    return max(parse_int(ticket.get("totalQuantity")) - sold_quantity, 0)


def find_fifa_ticket(store, ticket_id):
    for ticket in store.get("tickets", []):
        if ticket.get("id") == ticket_id:
            return ticket
    return None


def normalize_fifa_ticket_match_data(ticket):
    if not isinstance(ticket, dict):
        return ticket
    match_number = normalize_text(ticket.get("matchNumber")).lower()
    team_a = normalize_text(ticket.get("teamA")).upper()
    team_b = normalize_text(ticket.get("teamB")).upper()
    if match_number == "match 53" and team_a == "MEX" and team_b in {"FIFA", "CZECH", "CZECHIA", ""}:
        ticket = {**ticket, "teamB": "CZE"}
        current_label = normalize_text(ticket.get("matchLabel"))
        if not current_label or "FIFA" in current_label.upper():
            ticket["matchLabel"] = "MEX vs CZE"
    return ticket


def enrich_fifa_ticket(ticket, sales):
    ticket = normalize_fifa_ticket_match_data(ticket)
    sold_quantity = sum(
        fifa_sale_quantity_for_ticket(sale, ticket.get("id"))
        for sale in fifa_ticket_sales(sales, ticket.get("id"))
    )
    total_quantity = max(parse_int(ticket.get("totalQuantity")), 0)
    available_quantity = max(total_quantity - sold_quantity, 0)
    public_visible = (
        normalize_text(ticket.get("visibility")).lower() == "public"
        and normalize_text(ticket.get("status")).lower() == "active"
        and available_quantity > 0
    )
    return {
        **ticket,
        "totalQuantity": total_quantity,
        "soldQuantity": sold_quantity,
        "availableQuantity": available_quantity,
        "publicVisible": public_visible,
        "soldOut": available_quantity <= 0,
    }


def enrich_fifa_sale(sale, ticket, store=None):
    ticket_ids = sale.get("ticketIds") or ([sale.get("ticketId")] if sale.get("ticketId") else [])
    linked_tickets = [find_fifa_ticket(store or {"tickets": []}, ticket_id) for ticket_id in ticket_ids]
    linked_tickets = [item for item in linked_tickets if item]
    quantity = max(parse_int(sale.get("quantity")) or len(ticket_ids), 0)
    price_per_ticket = max(parse_number(sale.get("pricePerTicket")), 0)
    discount_amount = max(parse_number(sale.get("discountAmount")), 0)
    block_total_price = sum(max(parse_number(block.get("totalPrice")), 0) for block in (sale.get("ticketBlocks") or []))
    total_price = max(parse_number(sale.get("totalPrice")) or max(block_total_price - discount_amount, 0) or (quantity * price_per_ticket), 0)
    amount_paid = max(parse_number(sale.get("amountPaid")), 0)
    invoice_exchange_rate = max(parse_int(sale.get("invoiceExchangeRate")) or 3600, 1)
    buyer_name = normalize_text(sale.get("buyerName"))
    sold_by = sale.get("soldBy") or {}
    sold_by_email = normalize_text(sold_by.get("email")).lower()
    sold_by_user = find_user_by_email(sold_by_email) if sold_by_email else None
    sold_by_name = (
        normalize_text(sold_by.get("fullName"))
        or normalize_text(sold_by_user.get("fullName") if sold_by_user else "")
        or normalize_text(sold_by_user.get("name") if sold_by_user else "")
        or sold_by_email
    )
    ticket_labels = [
        f"{normalize_text(item.get('matchNumber'))} · {normalize_text(item.get('matchLabel'))}".strip(" ·")
        for item in linked_tickets
        if normalize_text(item.get("matchNumber")) or normalize_text(item.get("matchLabel"))
    ]
    cities = [normalize_text(item.get("city")) for item in linked_tickets if normalize_text(item.get("city"))]
    stages = [normalize_text(item.get("stage")) for item in linked_tickets if normalize_text(item.get("stage"))]
    seat_details = [normalize_text(item.get("seatDetails")) for item in linked_tickets if normalize_text(item.get("seatDetails"))]
    return {
        **sale,
        "ticketIds": ticket_ids,
        "quantity": quantity,
        "pricePerTicket": price_per_ticket,
        "discountAmount": discount_amount,
        "invoiceExchangeRate": invoice_exchange_rate,
        "invoiceBankAccount": normalize_text(sale.get("invoiceBankAccount")) or "state",
        "invoiceDescriptions": sale.get("invoiceDescriptions") or [],
        "invoiceSchedule": sale.get("invoiceSchedule") or [],
        "totalPrice": total_price,
        "amountPaid": amount_paid,
        "balanceDue": max(total_price - amount_paid, 0),
        "isPaid": amount_paid >= total_price and total_price > 0,
        "buyerName": buyer_name,
        "buyerEmail": normalize_text(sale.get("buyerEmail")).lower(),
        "ticket": enrich_fifa_ticket(ticket, []) if ticket else None,
        "ticketLabel": normalize_text(sale.get("buyerTitle")) or ", ".join(dict.fromkeys(ticket_labels)),
        "city": ", ".join(dict.fromkeys(cities)),
        "stage": ", ".join(dict.fromkeys(stages)),
        "seatDetails": " | ".join(seat_details),
        "buyerTitle": normalize_text(sale.get("buyerTitle")),
        "ticketBlocks": sale.get("ticketBlocks") or [],
        "participants": sale.get("participants") or [],
        "soldByName": sold_by_name,
    }


def build_fifa_summary(store):
    tickets = store.get("tickets", [])
    sales = store.get("sales", [])
    enriched_tickets = [enrich_fifa_ticket(ticket, sales) for ticket in tickets]
    active_sales = fifa_active_sales(sales)
    committed_sales = fifa_committed_sales(sales)
    enriched_sales = [enrich_fifa_sale(sale, find_fifa_ticket(store, sale.get("ticketId")), store) for sale in sales]
    public_tickets = [ticket for ticket in enriched_tickets if ticket.get("publicVisible")]
    paid_sales = [sale for sale in active_sales if normalize_text(sale.get("paymentStatus")).lower() == "paid"]
    unpaid_sales = [sale for sale in active_sales if normalize_text(sale.get("paymentStatus")).lower() == "unpaid"]
    partial_sales = [sale for sale in active_sales if normalize_text(sale.get("paymentStatus")).lower() == "partial"]
    total_units = sum(max(parse_int(ticket.get("totalQuantity")), 0) for ticket in tickets)
    sold_units = sum(
        max(
            parse_int(sale.get("quantity"))
            or len([ticket_id for ticket_id in (sale.get("ticketIds") or []) if normalize_text(ticket_id)])
            or (1 if normalize_text(sale.get("ticketId")) else 0),
            0,
        )
        for sale in committed_sales
    )
    available_units = max(total_units - sold_units, 0)
    return {
        "tickets": {
            "total": len(tickets),
            "matches": len({normalize_text(ticket.get("matchNumber")) or normalize_text(ticket.get("matchLabel")) for ticket in tickets}),
            "public": len([ticket for ticket in tickets if normalize_text(ticket.get("visibility")).lower() == "public"]),
            "availableLots": len([ticket for ticket in enriched_tickets if ticket.get("availableQuantity", 0) > 0]),
            "soldOutLots": len([ticket for ticket in enriched_tickets if ticket.get("availableQuantity", 0) <= 0]),
            "availableUnits": available_units,
            "soldUnits": sold_units,
        },
        "sales": {
            "total": len(sales),
            "active": len(active_sales),
            "cancelled": len([sale for sale in sales if normalize_text(sale.get("saleStatus")).lower() == "cancelled"]),
            "paid": len(paid_sales),
            "partial": len(partial_sales),
            "unpaid": len(unpaid_sales),
            "revenue": sum(max(parse_number(sale.get("totalPrice")), 0) for sale in active_sales),
            "collected": sum(max(parse_number(sale.get("amountPaid")), 0) for sale in active_sales),
        },
        "public": {
            "visibleLots": len(public_tickets),
            "visibleUnits": sum(ticket.get("availableQuantity", 0) for ticket in public_tickets),
        },
        "filters": {
            "stages": sorted({normalize_text(ticket.get("stage")) for ticket in tickets if normalize_text(ticket.get("stage"))}),
            "cities": sorted({normalize_text(ticket.get("city")) for ticket in tickets if normalize_text(ticket.get("city"))}),
            "categories": sorted({normalize_text(ticket.get("categoryCode")) for ticket in tickets if normalize_text(ticket.get("categoryCode"))}),
            "matches": sorted({
                f"{normalize_text(ticket.get('matchNumber'))} · {normalize_text(ticket.get('matchLabel'))}".strip(" ·")
                for ticket in tickets
                if normalize_text(ticket.get("matchNumber")) or normalize_text(ticket.get("matchLabel"))
            }),
            "soldBy": sorted({sale.get("soldByName") for sale in enriched_sales if sale.get("soldByName")}),
        },
    }


def build_fifa_ticket(payload):
    timestamp = now_mongolia().isoformat()
    return {
        "id": str(uuid4()),
        "stage": normalize_text(payload.get("stage")),
        "groupLabel": normalize_text(payload.get("groupLabel")),
        "matchNumber": normalize_text(payload.get("matchNumber")),
        "matchLabel": normalize_text(payload.get("matchLabel")),
        "matchDate": normalize_text(payload.get("matchDate")),
        "teamA": normalize_text(payload.get("teamA")),
        "teamB": normalize_text(payload.get("teamB")),
        "city": normalize_text(payload.get("city")),
        "venue": normalize_text(payload.get("venue")),
        "categoryCode": normalize_text(payload.get("categoryCode")) or "1",
        "categoryName": normalize_text(payload.get("categoryName")),
        "seatSection": normalize_text(payload.get("seatSection")),
        "seatDetails": normalize_text(payload.get("seatDetails")),
        "seatAssignedLater": bool(payload.get("seatAssignedLater")),
        "price": max(parse_int(payload.get("price")), 0),
        "currency": normalize_text(payload.get("currency")) or "USD",
        "totalQuantity": max(parse_int(payload.get("totalQuantity")) or 1, 1),
        "visibility": normalize_text(payload.get("visibility")).lower() or "public",
        "status": normalize_text(payload.get("status")).lower() or "active",
        "notes": normalize_text(payload.get("notes")),
        "createdAt": timestamp,
        "updatedAt": timestamp,
    }


def validate_fifa_ticket(ticket):
    if len(ticket.get("matchLabel", "")) < 2:
        return "Match label must be at least 2 characters"
    if not ticket.get("matchDate") or not parse_date_input(ticket.get("matchDate")):
        return "Match date must be in YYYY-MM-DD format"
    if len(ticket.get("city", "")) < 2:
        return "City is required"
    if ticket.get("price", 0) <= 0:
        return "Price must be greater than 0"
    if ticket.get("totalQuantity", 0) <= 0:
        return "Quantity must be greater than 0"
    if ticket.get("categoryCode") not in {"1", "2", "3"}:
        return "Category must be 1, 2, or 3"
    if ticket.get("visibility") not in {"public", "private"}:
        return "Visibility must be public or private"
    if ticket.get("status") not in {"active", "hidden", "archived"}:
        return "Status must be active, hidden, or archived"
    return None


def build_fifa_sale(payload, actor=None):
    ticket_ids = [normalize_text(value) for value in (payload.get("ticketIds") or []) if normalize_text(value)]
    if not ticket_ids and normalize_text(payload.get("ticketId")):
        ticket_ids = [normalize_text(payload.get("ticketId"))]
    ticket_blocks = payload.get("ticketBlocks") if isinstance(payload.get("ticketBlocks"), list) else []
    participants = payload.get("participants") if isinstance(payload.get("participants"), list) else []
    quantity = max(parse_int(payload.get("quantity")) or len(ticket_ids) or 1, 1)
    normalized_blocks = []
    for block in ticket_blocks:
        if not isinstance(block, dict):
            continue
        normalized_block = {
            "matchNumber": normalize_text(block.get("matchNumber")),
            "matchLabel": normalize_text(block.get("matchLabel")),
            "categoryCode": normalize_text(block.get("categoryCode")),
            "quantity": max(parse_int(block.get("quantity")), 0),
            "unitPrice": max(parse_number(block.get("unitPrice")), 0),
            "totalPrice": max(parse_number(block.get("totalPrice")), 0),
            "seatPreview": normalize_text(block.get("seatPreview")),
            "ticketLabels": [normalize_text(item) for item in (block.get("ticketLabels") or []) if normalize_text(item)],
            "ticketIds": [normalize_text(item) for item in (block.get("ticketIds") or []) if normalize_text(item)],
        }
        if normalized_block["quantity"] <= 0 and normalized_block["ticketIds"]:
            normalized_block["quantity"] = len(normalized_block["ticketIds"])
        if normalized_block["totalPrice"] <= 0 and normalized_block["quantity"] > 0 and normalized_block["unitPrice"] > 0:
            normalized_block["totalPrice"] = normalized_block["quantity"] * normalized_block["unitPrice"]
        normalized_blocks.append(normalized_block)
    ticket_blocks = normalized_blocks
    block_total_price = sum(max(parse_number(block.get("totalPrice")), 0) for block in ticket_blocks)
    unique_unit_prices = {max(parse_number(block.get("unitPrice")), 0) for block in ticket_blocks if max(parse_number(block.get("unitPrice")), 0) > 0}
    price_per_ticket = max(parse_number(payload.get("pricePerTicket")), 0)
    if not price_per_ticket and len(unique_unit_prices) == 1:
        price_per_ticket = next(iter(unique_unit_prices))
    discount_amount = max(parse_number(payload.get("discountAmount")), 0)
    invoice_exchange_rate = max(parse_int(payload.get("invoiceExchangeRate")) or 3600, 1)
    invoice_bank_account = normalize_text(payload.get("invoiceBankAccount")) or "state"
    invoice_bank_account_other = normalize_text(payload.get("invoiceBankAccountOther"))
    invoice_schedule = []
    invoice_descriptions = [normalize_text(item) for item in (payload.get("invoiceDescriptions") or []) if normalize_text(item)]
    for row in payload.get("invoiceSchedule") or []:
        if not isinstance(row, dict):
            continue
        amount_mnt = max(parse_int(row.get("amountMnt")), 0)
        invoice_schedule.append(
            {
                "title": normalize_text(row.get("title")),
                "created": normalize_text(row.get("created")),
                "due": normalize_text(row.get("due")),
                "status": normalize_text(row.get("status")).lower() or "waiting",
                "amount": int(round(amount_mnt / invoice_exchange_rate)) if invoice_exchange_rate else 0,
                "amountMnt": amount_mnt,
                "bankAccount": normalize_text(row.get("bankAccount")) or "state",
                "bankAccountOther": normalize_text(row.get("bankAccountOther")),
            }
        )
    total_price = max(parse_number(payload.get("totalPrice")) or max(block_total_price - discount_amount, 0) or (quantity * price_per_ticket), 0)
    amount_paid_mnt = sum(
        max(parse_int(row.get("amountMnt")), 0)
        for row in invoice_schedule
        if normalize_text(row.get("status")).lower() == "paid"
    )
    amount_paid = max(int(round(amount_paid_mnt / invoice_exchange_rate)), 0)
    payment_status = normalize_text(payload.get("paymentStatus")).lower()
    if not payment_status:
        if total_price and amount_paid >= total_price:
            payment_status = "paid"
        elif amount_paid > 0:
            payment_status = "partial"
        else:
            payment_status = "unpaid"
    timestamp = now_mongolia().isoformat()
    return {
        "id": str(uuid4()),
        "ticketId": ticket_ids[0] if ticket_ids else "",
        "ticketIds": ticket_ids,
        "ticketBlocks": ticket_blocks,
        "participants": participants,
        "quantity": quantity,
        "buyerTitle": normalize_text(payload.get("buyerTitle")),
        "buyerName": normalize_text(payload.get("buyerName")),
        "buyerPhone": normalize_text(payload.get("buyerPhone")),
        "buyerEmail": normalize_text(payload.get("buyerEmail")).lower(),
        "buyerPassportNumber": normalize_text(payload.get("buyerPassportNumber")),
        "buyerNationality": normalize_text(payload.get("buyerNationality")),
        "buyerNotes": normalize_text(payload.get("buyerNotes")),
        "pricePerTicket": price_per_ticket,
        "discountAmount": discount_amount,
        "invoiceExchangeRate": invoice_exchange_rate,
        "invoiceBankAccount": invoice_bank_account,
        "invoiceBankAccountOther": invoice_bank_account_other,
        "invoiceDescriptions": invoice_descriptions,
        "invoiceSchedule": invoice_schedule,
        "totalPrice": total_price,
        "amountPaid": amount_paid,
        "paymentStatus": payment_status,
        "paymentMethod": normalize_text(payload.get("paymentMethod")),
        "saleStatus": (
            "confirmed"
            if normalize_text(payload.get("saleStatus")).lower() == "active"
            else normalize_text(payload.get("saleStatus")).lower() or "pending"
        ),
        "soldAt": normalize_text(payload.get("soldAt")) or timestamp,
        "soldBy": actor_snapshot(actor),
        "createdAt": timestamp,
        "updatedAt": timestamp,
    }


def validate_fifa_sale(sale, ticket, sales, store, excluded_sale_id=None):
    ticket_ids = sale.get("ticketIds") or ([sale.get("ticketId")] if sale.get("ticketId") else [])
    if not ticket_ids:
        return "Choose at least one ticket"
    linked_tickets = [find_fifa_ticket(store, ticket_id) for ticket_id in ticket_ids]
    if any(not item for item in linked_tickets):
        return "One or more tickets were not found"
    if any(item.get("status") == "archived" for item in linked_tickets if item):
        return "Archived tickets cannot be sold"
    if len(sale.get("buyerName", "")) < 2:
        return "Buyer name must be at least 2 characters"
    if sale.get("quantity", 0) != len(ticket_ids):
        return "Selected ticket count and quantity must match"
    ticket_blocks = sale.get("ticketBlocks") or []
    if not ticket_blocks and sale.get("pricePerTicket", 0) <= 0:
        return "Sale price must be greater than 0"
    if ticket_blocks:
        if any(max(parse_number(block.get("unitPrice")), 0) <= 0 for block in ticket_blocks):
            return "Each match block must have a valid price per ticket"
        if any(max(parse_number(block.get("totalPrice")), 0) <= 0 for block in ticket_blocks):
            return "Each match block must have a valid total price"
        if sum(max(parse_number(block.get("totalPrice")), 0) for block in ticket_blocks) <= 0:
            return "Total price must be greater than 0"
    if sale.get("paymentStatus") not in {"unpaid", "partial", "paid", "refunded"}:
        return "Payment status is invalid"
    if sale.get("saleStatus") not in {"pending", "confirmed", "cancelled", "active"}:
        return "Sale status is invalid"
    if sale.get("invoiceBankAccount") not in {"state", "golomt", "lkham-erdene", "azjargal", "bayaraa", "azaa", "lkhamaa", "other"}:
        return "Invoice bank account is invalid"
    for row in sale.get("invoiceSchedule") or []:
        if normalize_text(row.get("status")).lower() not in {"paid", "waiting", "overdue"}:
            return "Invoice schedule status is invalid"
        if normalize_text(row.get("bankAccount")) and normalize_text(row.get("bankAccount")) not in {"state", "golomt", "lkham-erdene", "azjargal", "bayaraa", "azaa", "lkhamaa", "other"}:
            return "Invoice schedule bank account is invalid"
    if sale.get("buyerEmail") and "@" not in sale.get("buyerEmail", ""):
        return "Buyer email must be valid"
    if sale.get("soldAt") and not re.fullmatch(r"\d{4}-\d{2}-\d{2}(T\d{2}:\d{2}(:\d{2})?)?", sale.get("soldAt")) and not parse_date_input(sale.get("soldAt")[:10]):
        return "Sold at must be a valid date or date-time"
    if len(sale.get("participants") or []) < sale.get("quantity", 0):
        return "Add all participants for this buyer"
    if sale.get("saleStatus") != "cancelled":
        for linked_ticket in linked_tickets:
            available_quantity = fifa_ticket_available_quantity(linked_ticket, sales, excluded_sale_id)
            if available_quantity < 1:
                return f"Ticket {linked_ticket.get('seatDetails') or linked_ticket.get('id')} is no longer available"
    return None


def update_manager_item(records, item_id, payload, builder, validator):
    for index, item in enumerate(records):
        if item.get("id") != item_id:
            continue
        merged = builder({**item, **payload})
        merged["id"] = item.get("id") or item_id
        merged["createdAt"] = item.get("createdAt") or now_mongolia().isoformat()
        merged["updatedAt"] = now_mongolia().isoformat()
        error = validator(merged)
        if error:
            return None, error
        records[index] = merged
        return merged, None
    return None, "Record not found"


MANAGER_ITEM_LABELS = {
    "tasks": ("task", "Task"),
    "contacts": ("contact", "Contact"),
    "reminders": ("reminder", "Reminder"),
}


def _manager_item_title(record, key):
    if key == "tasks":
        return normalize_text(record.get("title")) or "Untitled task"
    if key == "contacts":
        return normalize_text(record.get("name")) or "Untitled contact"
    return normalize_text(record.get("title")) or normalize_text(record.get("name")) or "Item"


def handle_manager_item_create(environ, start_response, key, builder, validator, response_label):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    record = builder(payload)
    # Stamp the user who *added* this item, so the to-do list can show
    # "Assigned by …" — the manager column is the assignee, this is the
    # one who created the entry.
    record["createdBy"] = actor_snapshot(actor)
    record["createdAt"] = record.get("createdAt") or now_mongolia().isoformat()
    error = validator(record)
    if error:
        return json_response(start_response, "400 Bad Request", {"error": error})
    store = read_manager_dashboard()
    store[key].insert(0, record)
    write_manager_dashboard(store)
    if key == "tasks":
        try:
            send_task_assignment_email(record, actor)
        except Exception as exc:
            print(f"[task-assignment] dispatch failed: {exc}", flush=True)
    try:
        singular, label = MANAGER_ITEM_LABELS.get(key, (response_label, response_label.title()))
        log_notification(
            f"{singular}.created",
            actor,
            f"New {singular} added",
            detail=_manager_item_title(record, key),
            meta={"id": record.get("id"), "key": key},
        )
    except Exception:
        pass
    return json_response(
        start_response,
        "201 Created",
        {
            "ok": True,
            response_label: record,
            "summary": build_manager_summary(store),
        },
    )


def handle_manager_item_update(environ, start_response, key, item_id, builder, validator, response_label):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    store = read_manager_dashboard()
    prior_owners: set[str] = set()
    if key == "tasks":
        existing = next((t for t in store.get(key, []) if t.get("id") == item_id), None)
        if existing:
            prior_owners = {n.lower() for n in _task_owner_names(existing)}
    record, error = update_manager_item(store[key], item_id, payload, builder, validator)
    if error == "Record not found":
        return json_response(start_response, "404 Not Found", {"error": error})
    if error:
        return json_response(start_response, "400 Bad Request", {"error": error})
    write_manager_dashboard(store)
    if key == "tasks":
        new_owners = {n.lower() for n in _task_owner_names(record)}
        added = new_owners - prior_owners
        if added:
            try:
                send_task_assignment_email(record, actor, owner_filter=added)
            except Exception as exc:
                print(f"[task-assignment] dispatch failed: {exc}", flush=True)
    try:
        singular, _ = MANAGER_ITEM_LABELS.get(key, (response_label, response_label.title()))
        if key == "tasks" and normalize_text(record.get("status")) == "done":
            log_notification(
                "task.completed",
                actor,
                "Task completed",
                detail=_manager_item_title(record, key),
                meta={"id": record.get("id"), "key": key},
            )
        else:
            log_notification(
                f"{singular}.updated",
                actor,
                f"{singular.capitalize()} updated",
                detail=_manager_item_title(record, key),
                meta={"id": record.get("id"), "key": key},
            )
    except Exception:
        pass
    return json_response(
        start_response,
        "200 OK",
        {
            "ok": True,
            response_label: record,
            "summary": build_manager_summary(store),
        },
    )


def handle_manager_item_delete(environ, start_response, key, item_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    store = read_manager_dashboard()
    before = len(store[key])
    store[key] = [item for item in store[key] if item.get("id") != item_id]
    if len(store[key]) == before:
        return json_response(start_response, "404 Not Found", {"error": "Record not found"})
    write_manager_dashboard(store)
    # Cascade: a deleted task drops its reference images too. The user
    # explicitly wanted these throwaway, so don't keep orphan files.
    if key == "tasks":
        _delete_all_task_images(item_id)
    return json_response(
        start_response,
        "200 OK",
        {
            "ok": True,
            "deletedId": item_id,
            "summary": build_manager_summary(store),
        },
    )


def handle_dashboard_summary(start_response):
    finance_entries = read_finance_entries()
    bookings = read_bookings()
    reservations = read_reservations()
    ds160_records = read_ds160_applications()
    camp_records = read_camp_reservations()
    return json_response(
        start_response,
        "200 OK",
        {
            "finance": finance_summary(finance_entries),
            "counts": {
                "contracts": len(read_contracts()),
                "ds160": len(ds160_records),
                "bookings": len(bookings),
                "reservations": len(reservations),
                "campReservations": len(camp_records),
            },
            "bookingStatuses": {
                "confirmed": len([entry for entry in bookings if entry["status"] == "confirmed"]),
                "pending": len([entry for entry in bookings if entry["status"] == "pending"]),
                "cancelled": len([entry for entry in bookings if entry["status"] == "cancelled"]),
            },
            "campStatuses": camp_summary(camp_records),
        },
    )


def handle_list_ds160(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    return json_response(start_response, "200 OK", read_ds160_applications())


def handle_get_ds160(environ, start_response, record_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    record = find_ds160_record(record_id=record_id)
    if not record:
        return json_response(start_response, "404 Not Found", {"error": "DS-160 record not found"})
    return json_response(start_response, "200 OK", {"entry": record})


def handle_update_ds160(environ, start_response, record_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})

    records = read_ds160_applications()
    for index, record in enumerate(records):
        if record.get("id") != record_id:
            continue

        merged = {**record}
        for key in ["clientName", "clientEmail", "clientPhone", "managerName", "managerEmail", "managerPhone", "internalNotes", "appId", "appointmentDate", "appointmentTime", "status"]:
            if key in payload:
                value = normalize_text(payload.get(key))
                merged[key] = value.lower() if key in {"clientEmail", "managerEmail"} else value

        merged["status"] = normalize_ds160_status(merged.get("status"))
        merged["updatedAt"] = now_mongolia().isoformat()
        merged["updatedBy"] = actor_snapshot(actor)
        records[index] = normalize_ds160_record(merged)
        write_ds160_applications(records)
        return json_response(start_response, "200 OK", {"ok": True, "entry": records[index]})

    return json_response(start_response, "404 Not Found", {"error": "DS-160 record not found"})


def handle_delete_ds160(environ, start_response, record_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []

    records = read_ds160_applications()
    before = len(records)
    records = [record for record in records if record.get("id") != record_id]
    if len(records) == before:
        return json_response(start_response, "404 Not Found", {"error": "DS-160 record not found"})

    write_ds160_applications(records)
    return json_response(start_response, "200 OK", {"ok": True, "deletedId": record_id})


def handle_create_ds160_invitation(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})

    record = build_ds160_invitation(payload, actor)
    error = validate_ds160_invitation(record)
    if error:
        return json_response(start_response, "400 Bad Request", {"error": error})

    host = environ.get("HTTP_HOST") or environ.get("SERVER_NAME") or "backoffice.travelx.mn"
    scheme = environ.get("HTTP_X_FORWARDED_PROTO") or environ.get("wsgi.url_scheme") or "https"
    record["shareUrl"] = f"{scheme}://{host}/ds160/form/{record['clientToken']}"

    records = read_ds160_applications()
    records.insert(0, record)
    write_ds160_applications(records)
    return json_response(start_response, "201 Created", {"ok": True, "entry": record})


def handle_send_ds160_invitation(environ, start_response, record_id):
    """POST /api/ds160/<id>/send — actually email the share link to the
    client via Resend so managers don't depend on the OS mail client. Body
    can override {to, subject, body}; otherwise a Mongolian boilerplate is
    generated from the stored entry."""
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ) or {}
    records = read_ds160_applications()
    record = next((r for r in records if r.get("id") == record_id), None)
    if not record:
        return json_response(start_response, "404 Not Found", {"error": "DS-160 entry not found"})

    to = (payload.get("to") or record.get("clientEmail") or "").strip()
    if not to:
        return json_response(start_response, "400 Bad Request", {"error": "Client email is missing on this entry — open the entry and set one before sending."})

    share_url = record.get("shareUrl") or ""
    if not share_url:
        host = environ.get("HTTP_HOST") or "backoffice.travelx.mn"
        scheme = environ.get("HTTP_X_FORWARDED_PROTO") or environ.get("wsgi.url_scheme") or "https"
        token = record.get("clientToken") or ""
        if token:
            share_url = f"{scheme}://{host}/ds160/form/{token}"

    client_name = (record.get("clientName") or "").strip()
    given_name = client_name.split()[0] if client_name else ""
    manager_name = (record.get("managerName") or actor.get("fullName") or actor.get("email") or "TravelX").strip()

    subject = (payload.get("subject") or "").strip() or "TravelX DS-160 form"
    if payload.get("body"):
        body = payload["body"]
    else:
        body = (
            f"Сайн байна уу{(', ' + given_name) if given_name else ''}.\n\n"
            "Та доорх холбоосоор DS-160 маягтаа бөглөнө үү.\n\n"
            f"{share_url}\n\n"
            "Хүндэтгэсэн,\n"
            f"{manager_name}\n"
            "Дэлхий Трэвел Икс"
        )

    result = _tool_send_email({
        "to": to,
        "subject": subject,
        "body": body,
        "_company_name": "Дэлхий Трэвел Икс",
    }, actor)
    if result.get("error"):
        return json_response(start_response, "502 Bad Gateway", {"error": result["error"]})

    record["lastEmailedAt"] = datetime.now(timezone.utc).isoformat()
    record["lastEmailedBy"] = actor_snapshot(actor)
    write_ds160_applications(records)
    try:
        log_notification(
            "ds160.emailed",
            actor,
            "DS-160 link sent",
            detail=f"{record.get('clientName', '')} ({to})",
            meta={"id": record.get("id")},
        )
    except Exception:
        pass
    return json_response(start_response, "200 OK", {"ok": True, "to": to, "entry": record})


def handle_create_ds160(environ, start_response):
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})

    record = build_ds160_application(payload)
    error = validate_ds160_application(record)
    if error:
        return json_response(start_response, "400 Bad Request", {"error": error})

    records = read_ds160_applications()
    records.insert(0, record)
    write_ds160_applications(records)
    return json_response(start_response, "201 Created", {"ok": True, "application": record})


def handle_public_get_ds160(start_response, client_token):
    record = find_ds160_record(client_token=client_token)
    if not record:
        return json_response(start_response, "404 Not Found", {"error": "DS-160 form not found"})

    public_entry = {
        "id": record.get("id"),
        "clientToken": record.get("clientToken"),
        "status": record.get("status"),
        "clientName": record.get("clientName"),
        "clientEmail": record.get("clientEmail"),
        "clientPhone": record.get("clientPhone"),
        "managerName": record.get("managerName"),
        "internalNotes": record.get("internalNotes"),
        "submittedAt": record.get("submittedAt"),
        "payload": record.get("payload", {}),
        "officialFlowNote": record.get("officialFlowNote"),
    }
    return json_response(start_response, "200 OK", {"entry": public_entry})


def handle_public_submit_ds160(environ, start_response, client_token):
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})

    records = read_ds160_applications()
    for index, record in enumerate(records):
        if record.get("clientToken") != client_token:
            continue
        cleaned_payload = clean_ds160_payload(payload)
        flattened = flatten_ds160_answers(cleaned_payload)
        error = validate_ds160_application(flattened)
        if error:
            return json_response(start_response, "400 Bad Request", {"error": error})

        now_iso = datetime.now(timezone.utc).isoformat()
        merged = normalize_ds160_record(
            {
                **record,
                "status": "submitted",
                "updatedAt": now_iso,
                "submittedAt": now_iso,
                "payload": cleaned_payload,
                "clientName": record.get("clientName") or flattened.get("applicantName", ""),
                "clientEmail": record.get("clientEmail") or flattened.get("email", "").lower(),
                "clientPhone": record.get("clientPhone") or flattened.get("primaryPhone", ""),
            }
        )
        records[index] = merged
        write_ds160_applications(records)
        return json_response(start_response, "200 OK", {"ok": True, "entry": merged})

    return json_response(start_response, "404 Not Found", {"error": "DS-160 form not found"})


def handle_list_finance(start_response):
    entries = read_finance_entries()
    return json_response(start_response, "200 OK", {"entries": entries, "summary": finance_summary(entries)})


def handle_create_finance(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})

    record = build_finance_entry(payload)
    error = validate_finance_entry(record)
    if error:
        return json_response(start_response, "400 Bad Request", {"error": error})

    records = read_finance_entries()
    record["createdBy"] = actor_snapshot(actor)
    record["updatedBy"] = actor_snapshot(actor)
    records.insert(0, record)
    write_finance_entries(records)
    return json_response(start_response, "201 Created", {"ok": True, "entry": record, "summary": finance_summary(records)})


def handle_list_bookings(start_response):
    return json_response(start_response, "200 OK", read_bookings())


def handle_create_booking(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})

    record = build_booking(payload)
    error = validate_booking(record)
    if error:
        return json_response(start_response, "400 Bad Request", {"error": error})

    records = read_bookings()
    record["createdBy"] = actor_snapshot(actor)
    record["updatedBy"] = actor_snapshot(actor)
    records.insert(0, record)
    write_bookings(records)
    return json_response(start_response, "201 Created", {"ok": True, "booking": record})


def handle_list_reservations(start_response):
    return json_response(start_response, "200 OK", read_reservations())


def handle_create_reservation(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})

    record = build_reservation(payload)
    error = validate_reservation(record)
    if error:
        return json_response(start_response, "400 Bad Request", {"error": error})

    document = save_simple_document(
        "reservation",
        f"Reservation - {record['customerName']}",
        "Printable internal reservation summary for team use.",
        [
            ("Reservation", f"Customer: {record['customerName']}\nService: {record['serviceType']}\nDate: {record['reservationDate']}\nStatus: {record['status']}"),
            ("Finance", f"Amount: {format_money(record['amount'])}"),
            ("Notes", record["notes"] or "No extra notes"),
        ],
    )
    record.update(document)
    record["createdBy"] = actor_snapshot(actor)
    record["updatedBy"] = actor_snapshot(actor)
    records = read_reservations()
    records.insert(0, record)
    write_reservations(records)
    return json_response(start_response, "201 Created", {"ok": True, "reservation": record})


def handle_list_camp_reservations(environ, start_response):
    records = filter_by_company(read_camp_reservations(), environ)
    return json_response(start_response, "200 OK", {"entries": records, "summary": camp_summary(records)})


def handle_get_fifa2026_dashboard(start_response):
    store = ensure_fifa2026_manual_inventory()
    sales = store.get("sales", [])
    tickets = [enrich_fifa_ticket(ticket, sales) for ticket in store.get("tickets", [])]
    enriched_sales = [enrich_fifa_sale(sale, find_fifa_ticket(store, sale.get("ticketId")), store) for sale in sales]
    return json_response(
        start_response,
        "200 OK",
        {
            "tickets": tickets,
            "sales": enriched_sales,
            "summary": build_fifa_summary(store),
        },
    )


def handle_get_fifa2026_public(start_response):
    store = ensure_fifa2026_manual_inventory()
    sales = store.get("sales", [])
    tickets = []
    for ticket in store.get("tickets", []):
        enriched = enrich_fifa_ticket(ticket, sales)
        if enriched.get("publicVisible"):
            tickets.append(enriched)
    return json_response(
        start_response,
        "200 OK",
        {
            "tickets": tickets,
            "summary": {
                **build_fifa_summary(store).get("public", {}),
                "matchCount": len(
                    {
                        normalize_text(ticket.get("matchNumber")) or normalize_text(ticket.get("matchLabel"))
                        for ticket in tickets
                        if normalize_text(ticket.get("matchNumber")) or normalize_text(ticket.get("matchLabel"))
                    }
                ),
            },
        },
    )


def handle_reset_fifa2026_from_seed(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    store = reset_fifa2026_store_from_seed()
    sales = store.get("sales", [])
    tickets = [enrich_fifa_ticket(ticket, sales) for ticket in store.get("tickets", [])]
    return json_response(
        start_response,
        "200 OK",
        {
            "ok": True,
            "tickets": tickets,
            "sales": [],
            "summary": build_fifa_summary(store),
        },
    )


def handle_create_fifa_ticket(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    ticket = build_fifa_ticket(payload)
    error = validate_fifa_ticket(ticket)
    if error:
        return json_response(start_response, "400 Bad Request", {"error": error})
    store = read_fifa2026_store()
    store["tickets"].insert(0, ticket)
    write_fifa2026_store(store)
    return json_response(start_response, "201 Created", {"ok": True, "ticket": ticket, "summary": build_fifa_summary(store)})


def handle_update_fifa_ticket(environ, start_response, ticket_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    store = read_fifa2026_store()
    for index, item in enumerate(store.get("tickets", [])):
        if item.get("id") != ticket_id:
            continue
        updated = build_fifa_ticket({**item, **payload})
        updated["id"] = ticket_id
        updated["createdAt"] = item.get("createdAt") or updated.get("createdAt")
        updated["updatedAt"] = now_mongolia().isoformat()
        error = validate_fifa_ticket(updated)
        if error:
            return json_response(start_response, "400 Bad Request", {"error": error})
        active_sales_quantity = sum(
            fifa_sale_quantity_for_ticket(sale, ticket_id)
            for sale in fifa_ticket_sales(store.get("sales", []), ticket_id)
        )
        if updated.get("totalQuantity", 0) < active_sales_quantity:
            return json_response(
                start_response,
                "400 Bad Request",
                {"error": f"Quantity cannot be less than already sold units ({active_sales_quantity})"},
            )
        store["tickets"][index] = updated
        write_fifa2026_store(store)
        return json_response(start_response, "200 OK", {"ok": True, "ticket": updated, "summary": build_fifa_summary(store)})
    return json_response(start_response, "404 Not Found", {"error": "Ticket not found"})


def handle_delete_fifa_ticket(environ, start_response, ticket_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    store = read_fifa2026_store()
    if fifa_ticket_sales(store.get("sales", []), ticket_id):
        return json_response(start_response, "400 Bad Request", {"error": "Cancel active sales before deleting this ticket lot"})
    before = len(store.get("tickets", []))
    store["tickets"] = [ticket for ticket in store.get("tickets", []) if ticket.get("id") != ticket_id]
    if len(store["tickets"]) == before:
        return json_response(start_response, "404 Not Found", {"error": "Ticket not found"})
    write_fifa2026_store(store)
    return json_response(start_response, "200 OK", {"ok": True, "deletedId": ticket_id, "summary": build_fifa_summary(store)})


def handle_create_fifa_sale(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    store = read_fifa2026_store()
    sale = build_fifa_sale(payload, actor)
    ticket = find_fifa_ticket(store, sale.get("ticketId"))
    error = validate_fifa_sale(sale, ticket, store.get("sales", []), store)
    if error:
        return json_response(start_response, "400 Bad Request", {"error": error})
    store["sales"].insert(0, sale)
    write_fifa2026_store(store)
    return json_response(start_response, "201 Created", {"ok": True, "sale": enrich_fifa_sale(sale, ticket, store), "summary": build_fifa_summary(store)})


def handle_update_fifa_sale(environ, start_response, sale_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    store = read_fifa2026_store()
    for index, item in enumerate(store.get("sales", [])):
        if item.get("id") != sale_id:
            continue
        merged_payload = {**item, **payload}
        sale = build_fifa_sale(merged_payload, actor)
        sale["id"] = sale_id
        sale["createdAt"] = item.get("createdAt") or sale.get("createdAt")
        sale["soldBy"] = item.get("soldBy") or actor_snapshot(actor)
        sale["soldAt"] = normalize_text(payload.get("soldAt")) or item.get("soldAt") or sale.get("soldAt")
        sale["updatedAt"] = now_mongolia().isoformat()
        ticket = find_fifa_ticket(store, sale.get("ticketId"))
        error = validate_fifa_sale(sale, ticket, store.get("sales", []), store, excluded_sale_id=sale_id)
        if error:
            return json_response(start_response, "400 Bad Request", {"error": error})
        store["sales"][index] = sale
        write_fifa2026_store(store)
        return json_response(start_response, "200 OK", {"ok": True, "sale": enrich_fifa_sale(sale, ticket, store), "summary": build_fifa_summary(store)})
    return json_response(start_response, "404 Not Found", {"error": "Sale not found"})


def handle_delete_fifa_sale(environ, start_response, sale_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    store = read_fifa2026_store()
    before = len(store.get("sales", []))
    store["sales"] = [sale for sale in store.get("sales", []) if sale.get("id") != sale_id]
    if len(store["sales"]) == before:
        return json_response(start_response, "404 Not Found", {"error": "Sale not found"})
    write_fifa2026_store(store)
    return json_response(start_response, "200 OK", {"ok": True, "deletedId": sale_id, "summary": build_fifa_summary(store)})


def handle_list_camp_trips(environ, start_response):
    trips = filter_by_company(read_camp_trips(), environ)
    # Enrich each trip with the live tourist count so the GIT pax tile can
    # render "actual / planned". For GIT trips the booked headcount uses
    # each confirmed group's declared headcount (the manager types "5
    # NED vs JPN" before all 5 names are entered — that's the source of
    # truth). Pending/cancelled groups don't add to the count. Tourists
    # without a group fall back to the per-tourist count.
    all_groups = read_tourist_groups()
    confirmed_groups_by_trip = {}
    for g in all_groups:
        status = (g.get("status") or "").lower()
        if status not in ("confirmed", ""):
            continue
        confirmed_groups_by_trip.setdefault(g.get("tripId"), []).append(g)
    confirmed_group_ids = {g.get("id") for groups in confirmed_groups_by_trip.values() for g in groups}
    # Per-trip pax, summed from confirmed groups' declared headcount.
    counts_by_trip = {}
    for tid, groups in confirmed_groups_by_trip.items():
        counts_by_trip[tid] = sum(int(g.get("headcount") or 0) for g in groups)
    # Tourists outside any group (legacy data that pre-dates groups) still
    # count, since their participation isn't otherwise represented.
    for t in read_tourists():
        tid = t.get("tripId")
        if not tid:
            continue
        if t.get("groupId"):
            continue
        counts_by_trip[tid] = counts_by_trip.get(tid, 0) + 1
    enriched = []
    for trip in trips:
        copy = dict(trip)
        copy["actualTouristCount"] = counts_by_trip.get(trip.get("id"), 0)
        enriched.append(copy)
    return json_response(start_response, "200 OK", {"entries": enriched})


def handle_list_camp_settings(start_response):
    return json_response(start_response, "200 OK", {"entry": read_camp_settings()})


def _normalize_camp_details(payload, camp_names):
    """Per-camp price + contract path. Keyed by camp name."""
    details = {}
    if isinstance(payload, dict):
        for name, value in payload.items():
            name_clean = normalize_text(name)
            if not name_clean or not isinstance(value, dict):
                continue
            details[name_clean] = {
                "price": normalize_text(value.get("price")),
                "contractPath": normalize_text(value.get("contractPath")),
            }
    for name in camp_names or []:
        details.setdefault(name, {"price": "", "contractPath": ""})
    return details


def handle_update_camp_settings(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    camp_names = normalize_option_list(payload.get("campNames"))
    existing = read_camp_settings()
    existing_details = existing.get("campDetails") or {}
    incoming_details = payload.get("campDetails")
    if incoming_details is None:
        merged_details = existing_details
    else:
        merged_details = {**existing_details}
        if isinstance(incoming_details, dict):
            for name, value in incoming_details.items():
                key = normalize_text(name)
                if key:
                    merged_details[key] = value
    # Transfer settings keep their previous values whenever the writer doesn't
    # send them — both the chosen-trip page and Settings/Camps tab POST a
    # partial settings payload that drops the new keys.
    incoming_places = payload.get("transferPlaces")
    # Legacy clients may still send the split pickup/dropoff lists; fold them
    # into transferPlaces so the migration completes the next save.
    if incoming_places is None and (payload.get("transferPickups") is not None or payload.get("transferDropoffs") is not None):
        incoming_places = (payload.get("transferPickups") or []) + (payload.get("transferDropoffs") or [])
    incoming_drivers = payload.get("transferDrivers")
    settings = {
        "campNames": camp_names,
        "locationNames": normalize_option_list(payload.get("locationNames")),
        "staffAssignments": normalize_option_list(payload.get("staffAssignments")),
        "roomChoices": normalize_option_list(payload.get("roomChoices")) or DEFAULT_ROOM_CHOICES,
        "campLocations": normalize_camp_location_map(payload.get("campLocations"), camp_names),
        "campDetails": _normalize_camp_details(merged_details, camp_names),
        "transferPlaces": normalize_option_list(incoming_places) if incoming_places is not None else existing.get("transferPlaces", []),
        "transferDrivers": normalize_transfer_drivers(incoming_drivers) if incoming_drivers is not None else existing.get("transferDrivers", []),
    }
    if not settings["campNames"]:
        settings["campNames"] = ["Khustai camp"]
    if not settings["locationNames"]:
        settings["locationNames"] = ["Khustai"]
    if not settings["staffAssignments"]:
        settings["staffAssignments"] = [STEPPE_MANAGER]
    settings["locationNames"] = normalize_option_list(settings["locationNames"] + [value for value in settings["campLocations"].values() if value])
    write_camp_settings(settings)
    return json_response(start_response, "200 OK", {"ok": True, "entry": settings})


def handle_get_settings(environ, start_response):
    if not require_login(environ, start_response):
        return []
    return json_response(start_response, "200 OK", {"entry": read_settings()})


def handle_update_settings_destinations(environ, start_response):
    """POST /api/settings/destinations — set the full destinations list.

    Body: {"destinations": ["...", "..."]}. We replace the saved list verbatim,
    so the page can reorder + remove entries with one round-trip.
    Open to all logged-in users — small team, shared settings.
    """
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    destinations = normalize_option_list(payload.get("destinations"))
    settings = read_settings()
    settings["destinations"] = destinations
    write_settings(settings)
    return json_response(start_response, "200 OK", {"ok": True, "entry": settings})


def handle_update_settings_bank_accounts(environ, start_response):
    """POST /api/settings/bank-accounts — replaces the full list.
    Body: {"bankAccounts": [{label, bankName, accountName, accountNumber,
    currency, swift, notes}, ...]}."""
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    accounts = _normalize_bank_accounts(payload.get("bankAccounts"))
    settings = read_settings()
    settings["bankAccounts"] = accounts
    write_settings(settings)
    return json_response(start_response, "200 OK", {"ok": True, "entry": settings})


def handle_upload_camp_contract(environ, start_response):
    """POST /api/camp-settings/contract — accepts {"campName": "...", "data": "data:application/pdf;base64,..."}
    Saves under camp-contracts/ and returns the relative path to store on the camp record.
    """
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    camp_name = normalize_text(payload.get("campName"))
    data_url = payload.get("data") or ""
    if not camp_name:
        return json_response(start_response, "400 Bad Request", {"error": "campName is required"})
    if "base64," not in data_url:
        return json_response(start_response, "400 Bad Request", {"error": "data must be a base64 data URL"})
    header, encoded = data_url.split("base64,", 1)
    ext = "pdf"
    if "image/jpeg" in header or "image/jpg" in header:
        ext = "jpg"
    elif "image/png" in header:
        ext = "png"
    elif "msword" in header or "officedocument" in header:
        ext = "docx"
    try:
        binary = base64.b64decode(encoded)
    except Exception:
        return json_response(start_response, "400 Bad Request", {"error": "Could not decode upload"})
    if len(binary) > MAX_UPLOAD_BYTES:
        return json_response(start_response, "400 Bad Request", {"error": "File too large (max 10MB)"})
    CAMP_CONTRACTS_DIR.mkdir(parents=True, exist_ok=True)
    safe_slug = re.sub(r"[^a-zA-Z0-9_-]+", "-", camp_name).strip("-").lower() or "camp"
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    filename = f"{safe_slug}-{timestamp}.{ext}"
    target = CAMP_CONTRACTS_DIR / filename
    target.write_bytes(binary)
    return json_response(start_response, "200 OK", {"ok": True, "path": f"/camp-contracts/{filename}"})


def handle_create_camp_trip(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})

    payload["company"] = active_workspace(environ) or DEFAULT_COMPANY
    record = build_camp_trip(payload, actor)
    error = validate_camp_trip(record)
    if error:
        return json_response(start_response, "400 Bad Request", {"error": error})

    trips = read_camp_trips()
    trips.insert(0, record)
    write_camp_trips(trips)
    try:
        log_notification(
            "trip.created",
            actor,
            "New trip created",
            detail=record.get("tripName") or record.get("reservationName") or "",
            meta={"id": record.get("id")},
        )
    except Exception:
        pass
    return json_response(start_response, "201 Created", {"ok": True, "entry": record})


def handle_update_camp_trip(environ, start_response, trip_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    trips = read_camp_trips()
    for index, trip in enumerate(trips):
        if trip["id"] != trip_id:
            continue
        merged = {**trip}
        for key in ["tripName", "reservationName", "startDate", "endDate", "language", "status", "guideName", "driverName", "cookName", "groupName"]:
            if key in payload:
                merged[key] = normalize_text(payload.get(key))
        for key in ["participantCount", "staffCount", "totalDays"]:
            if key in payload:
                merged[key] = parse_int(payload.get(key))
        if "tourists" in payload:
            merged["tourists"] = _normalize_tourist_breakdown(payload.get("tourists"))
            # Keep the total in sync if the client didn't send it explicitly.
            if "participantCount" not in payload:
                merged["participantCount"] = sum(merged["tourists"].values())
        if "staff" in payload:
            merged["staff"] = _normalize_staff_breakdown(payload.get("staff"))
            if "staffCount" not in payload:
                merged["staffCount"] = sum(merged["staff"].values())
        if "currency" in payload:
            ccy = (normalize_text(payload.get("currency")) or "MNT").upper()
            if ccy in {"MNT", "USD", "EUR", "CNY", "JPY", "KRW", "RUB"}:
                merged["currency"] = ccy
        if "tags" in payload:
            merged["tags"] = normalize_tag_list(payload.get("tags"))
        if "tripType" in payload:
            tt = normalize_text(payload.get("tripType")).lower()
            if tt in {"fit", "git"}:
                merged["tripType"] = tt
        if "expenseLines" in payload:
            merged["expenseLines"] = _normalize_trip_template_lines(payload.get("expenseLines"))
        if "marginPct" in payload:
            try:
                merged["marginPct"] = float(payload.get("marginPct") or 0)
            except (TypeError, ValueError):
                pass
        if "exchangeRates" in payload:
            raw_rates = payload.get("exchangeRates") or {}
            fx = {}
            for k in ("USD", "EUR", "CNY", "JPY", "KRW", "RUB"):
                try:
                    v = float(raw_rates.get(k) or 0)
                except (TypeError, ValueError):
                    v = 0
                if v > 0:
                    fx[k] = v
            merged["exchangeRates"] = fx
        error = validate_camp_trip(merged)
        if error:
            return json_response(start_response, "400 Bad Request", {"error": error})
        merged["updatedAt"] = now_mongolia().isoformat()
        merged["updatedBy"] = actor_snapshot(actor)
        trips[index] = merged
        write_camp_trips(trips)
        try:
            log_notification(
                "trip.updated",
                actor,
                "Trip updated",
                detail=merged.get("tripName") or merged.get("reservationName") or "",
                meta={"id": merged.get("id")},
            )
        except Exception:
            pass
        return json_response(start_response, "200 OK", {"ok": True, "entry": merged})
    return json_response(start_response, "404 Not Found", {"error": "Trip not found"})


def ocrspace_scan(file_bytes, filename, content_type):
    """Send the uploaded passport to OCR.space and return the parsed
    text. Free tier: 25k requests/month. On any network/auth error
    raises RuntimeError so the caller can fall back to manual entry.
    """
    if not OCRSPACE_API_KEY:
        raise RuntimeError("OCRSPACE_API_KEY is not configured")
    boundary = "----travelxpassport" + uuid4().hex
    parts = []
    def add_field(name, value):
        parts.append(f"--{boundary}".encode())
        parts.append(f'Content-Disposition: form-data; name="{name}"'.encode())
        parts.append(b"")
        parts.append(value.encode() if isinstance(value, str) else value)
    add_field("language", "eng")
    add_field("isOverlayRequired", "false")
    add_field("OCREngine", "2")
    add_field("scale", "true")
    add_field("detectOrientation", "true")
    parts.append(f"--{boundary}".encode())
    parts.append(
        b'Content-Disposition: form-data; name="file"; filename="'
        + filename.encode()
        + b'"'
    )
    parts.append(f"Content-Type: {content_type}".encode())
    parts.append(b"")
    parts.append(file_bytes)
    parts.append(f"--{boundary}--".encode())
    parts.append(b"")
    body = b"\r\n".join(parts)
    req = urllib.request.Request(
        "https://api.ocr.space/parse/image",
        data=body,
        method="POST",
        headers={
            "apikey": OCRSPACE_API_KEY,
            "Content-Type": f"multipart/form-data; boundary={boundary}",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=45) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        try:
            body_snippet = exc.read().decode("utf-8", errors="replace")[:400]
        except Exception:
            body_snippet = ""
        raise RuntimeError(f"OCR.space HTTP {exc.code}: {body_snippet}") from exc
    except Exception as exc:
        raise RuntimeError(f"OCR.space request failed: {exc}") from exc
    if payload.get("IsErroredOnProcessing"):
        msg = payload.get("ErrorMessage") or payload.get("ErrorDetails") or "OCR failed"
        if isinstance(msg, list):
            msg = " ".join(str(m) for m in msg)
        raise RuntimeError(f"OCR.space error: {msg}")
    results = payload.get("ParsedResults") or []
    if not results:
        raise RuntimeError(f"OCR.space returned no results: {str(payload)[:300]}")
    text = results[0].get("ParsedText") or ""
    return text, payload


def _mrz_yymmdd_to_iso(raw, century_cutoff):
    if len(raw) != 6 or not raw.isdigit():
        return ""
    yy, mm, dd = int(raw[0:2]), raw[2:4], raw[4:6]
    year = (2000 + yy) if yy <= century_cutoff else (1900 + yy)
    return f"{year:04d}-{mm}-{dd}"


_MONTH_MAP = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "SEPT": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}


def _yy_to_full_year(yy):
    """Mongolian passports print the issue date as 2-digit year. Use the
    same century cutoff the MRZ uses for expiry (≤70 → 2000s) so '22' →
    2022 and '99' → 1999."""
    yy = int(yy)
    return 2000 + yy if yy <= 70 else 1900 + yy


def _passport_dates_from_text(text):
    """Extract every date-looking substring from OCR text and return as ISO
    YYYY-MM-DD strings (deduped, in order of appearance). Handles the shapes
    Mongolian/EN passports actually print: '01 MAR 22', '12 JAN 2020',
    '12.01.2020', '12/01/2020', '2020-01-12'."""
    if not text:
        return []
    seen, out = set(), []
    def add(y, m, d):
        try:
            iso = f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
        except Exception:
            return
        if 1900 <= int(y) <= 2100 and 1 <= int(m) <= 12 and 1 <= int(d) <= 31 and iso not in seen:
            seen.add(iso)
            out.append(iso)
    # "01 MAR 22" / "12 JAN 2020" — month-name form. Year may be 2 or 4 digits.
    for m in re.finditer(r"\b(\d{1,2})\s+([A-Za-z]{3,4})\s+(\d{2,4})\b", text):
        mo = _MONTH_MAP.get(m.group(2).upper())
        if not mo:
            continue
        year_raw = m.group(3)
        year = _yy_to_full_year(year_raw) if len(year_raw) == 2 else int(year_raw)
        add(year, mo, m.group(1))
    # DD.MM.YYYY / DD/MM/YYYY / DD-MM-YYYY
    for m in re.finditer(r"\b(\d{1,2})[./-](\d{1,2})[./-](\d{4})\b", text):
        add(m.group(3), m.group(2), m.group(1))
    # YYYY-MM-DD / YYYY/MM/DD / YYYY.MM.DD
    for m in re.finditer(r"\b(\d{4})[./-](\d{1,2})[./-](\d{1,2})\b", text):
        add(m.group(1), m.group(2), m.group(3))
    return out


def parse_passport_visual_zone(text, mrz_fields):
    """Scan the full OCR text for fields that aren't in the MRZ — namely
    Mongolian РД (registration number) and the passport issue date.
    Conservative: only fills if a confident match is found."""
    extras = {}
    if not text:
        return extras

    # Registration number — Mongolian РД is 2 letters + 8 digits. Latin or
    # Cyrillic letters both appear depending on which side of the passport
    # the OCR caught.
    reg_match = re.search(r"\b([A-ZА-ЯҮӨЁ]{2}\d{8})\b", text.upper())
    if reg_match:
        extras["registrationNumber"] = reg_match.group(1)

    # Issue date — the visual zone has DOB, issue, expiry. MRZ already gave
    # us DOB and expiry; the remaining date that's plausibly an issue date
    # (within the last 15 years and earlier than expiry) is what we want.
    dob = (mrz_fields or {}).get("dob") or ""
    expiry = (mrz_fields or {}).get("passportExpiry") or ""
    today = datetime.now(timezone.utc).date().isoformat()
    candidates = []
    for iso in _passport_dates_from_text(text):
        if iso == dob or iso == expiry:
            continue
        if iso > today:
            continue
        if expiry and iso >= expiry:
            continue
        # Issue dates are typically within the validity window (10y for MN passports).
        # Drop obviously wrong dates (older than 30 years).
        try:
            year = int(iso[:4])
            if year < datetime.now(timezone.utc).year - 30:
                continue
        except Exception:
            continue
        candidates.append(iso)
    if candidates:
        # Prefer the latest candidate — passport issue is usually the most
        # recent non-expiry date printed.
        candidates.sort()
        extras["passportIssueDate"] = candidates[-1]

    return extras


def parse_passport_mrz(text):
    """Best-effort parser for TD-3 passport MRZ. Returns a dict with the
    same keys the front-end expects (lastName, firstName, gender, dob,
    nationality, passportNumber, passportExpiry). Missing fields are
    simply omitted. We deliberately keep this lenient — OCR is messy.
    """
    if not text:
        return {}, 0
    lines = []
    for raw_line in text.split("\n"):
        norm = re.sub(r"\s+", "", raw_line).upper()
        norm = norm.replace("«", "<").replace("«", "<")
        if len(norm) >= 30 and norm.count("<") >= 4:
            lines.append(norm)
    line1 = line2 = None
    for i, l in enumerate(lines):
        if l.startswith("P") and i + 1 < len(lines):
            line1 = l
            line2 = lines[i + 1]
            break
    if line1 is None and len(lines) >= 2:
        line1, line2 = lines[-2], lines[-1]
    if not line1 or not line2:
        return {}, 0

    fields = {}
    found = 0

    if line1.startswith("P") and len(line1) >= 5:
        country = line1[2:5].replace("<", "")
        if len(country) == 3 and country.isalpha():
            fields["nationality"] = country
            found += 1
        # Names are encoded as SURNAME<<GIVEN<NAMES<<<…< padding. OCR very
        # often collapses runs of "<" — we've seen "<<" come back as "<" —
        # which made the old split("<<", 1) treat the whole "SURNAME GIVEN"
        # blob as the surname. Use a regex run-of-< splitter instead and
        # treat the first non-empty chunk as the surname, the rest as
        # given names. (Also drops the trailing "<<<<<<" filler cleanly.)
        names_part = line1[5:]
        chunks = [c for c in re.split(r"<+", names_part) if c]
        if chunks:
            fields["lastName"] = chunks[0].strip()
            found += 1
        if len(chunks) > 1:
            given = " ".join(c.strip() for c in chunks[1:] if c.strip())
            if given:
                fields["firstName"] = given
                found += 1

    if len(line2) >= 28:
        passport_no = line2[0:9].replace("<", "").strip()
        if passport_no:
            fields["passportNumber"] = passport_no
            found += 1
        dob = _mrz_yymmdd_to_iso(line2[13:19], century_cutoff=30)
        if dob:
            fields["dob"] = dob
            found += 1
        gender_char = line2[20] if len(line2) > 20 else ""
        if gender_char == "M":
            fields["gender"] = "male"
            found += 1
        elif gender_char == "F":
            fields["gender"] = "female"
            found += 1
        expiry = _mrz_yymmdd_to_iso(line2[21:27], century_cutoff=70)
        if expiry:
            fields["passportExpiry"] = expiry
            found += 1
        # Personal number / registration number lives in positions 28-41
        # (14 chars, '<'-padded). For Mongolian passports this is the РД,
        # e.g. "T<J<<84071401" → "TJ84071401".
        if len(line2) >= 42:
            personal = line2[28:42].replace("<", "").strip()
            if personal:
                fields["registrationNumber"] = personal
                found += 1
    return fields, found


def handle_passport_scan(environ, start_response):
    """POST /api/tourists/passport-scan — accept a passport image, run
    OCR.space + MRZ parsing, save the file to the temp passport area
    keyed by a one-time token, return the parsed fields + token. The
    client sends the token back when it saves the tourist; that's when
    the file gets moved into the trip's documents."""
    actor = require_login(environ, start_response)
    if not actor:
        return []
    fields, files = parse_multipart(environ)
    if "file" not in files:
        return json_response(start_response, "400 Bad Request", {"error": "No file provided"})
    upload = files["file"]
    original_name = upload["filename"] or "passport"
    ext = Path(original_name).suffix.lower() or ".jpg"
    if ext not in {".pdf", ".jpg", ".jpeg", ".png", ".webp"}:
        return json_response(start_response, "400 Bad Request", {"error": f"File type {ext} not supported"})
    data = upload["data"]
    if not data:
        return json_response(start_response, "400 Bad Request", {"error": "Empty file"})
    if len(data) > MAX_UPLOAD_BYTES:
        return json_response(start_response, "400 Bad Request", {"error": "File too large"})

    token = uuid4().hex
    ensure_data_store()
    temp_path = TOURIST_PASSPORT_TEMP_DIR / f"{token}{ext}"
    temp_path.write_bytes(data)

    if not OCRSPACE_API_KEY:
        return json_response(start_response, "200 OK", {
            "ok": True,
            "token": token,
            "fields": {},
            "qualityOk": False,
            "reason": "ocr_disabled",
            "originalName": original_name,
            "contentType": upload["content_type"],
        })

    try:
        text, _raw = ocrspace_scan(data, original_name, upload["content_type"])
    except RuntimeError as exc:
        try:
            print(f"[passport-scan] OCR.space call failed: {exc}", file=sys.stderr, flush=True)
        except Exception:
            pass
        return json_response(start_response, "200 OK", {
            "ok": True,
            "token": token,
            "fields": {},
            "qualityOk": False,
            "reason": "ocr_failed",
            "debug": str(exc)[:400],
            "originalName": original_name,
            "contentType": upload["content_type"],
        })

    parsed, found = parse_passport_mrz(text)
    extras = parse_passport_visual_zone(text, parsed)
    for key, value in extras.items():
        if value and not parsed.get(key):
            parsed[key] = value
    quality_ok = found >= 4
    return json_response(start_response, "200 OK", {
        "ok": True,
        "token": token,
        "fields": parsed,
        "qualityOk": quality_ok,
        "originalName": original_name,
        "contentType": upload["content_type"],
    })


def clone_passport_from_tourist(source_tourist_id, target_trip_id, target_tourist_id, target_tourist_name, actor):
    """When a manager picks an existing tourist on the FIT/GIT 'add participant'
    flow, copy that tourist's most recent passport scan into the new trip's
    documents so they don't have to re-upload it. Returns the new doc dict,
    or None if the source has no passport on file or anything goes wrong."""
    if not source_tourist_id or not target_trip_id or not target_tourist_id:
        return None
    trips = read_camp_trips()
    target_index = next((i for i, t in enumerate(trips) if t.get("id") == target_trip_id), None)
    if target_index is None:
        return None
    source_doc = None
    for trip in trips:
        for doc in (trip.get("documents") or []):
            if doc.get("touristId") != source_tourist_id:
                continue
            if not re.search(r"passport", str(doc.get("category") or ""), re.IGNORECASE):
                continue
            source_path = TRIP_UPLOADS_DIR / trip.get("id") / (doc.get("storedName") or "")
            if not source_path.exists():
                continue
            source_doc = (trip, doc, source_path)
            break
        if source_doc:
            break
    if not source_doc:
        return None
    src_trip, src_doc, src_path = source_doc
    ext = Path(src_doc.get("storedName") or "").suffix.lower() or ".jpg"
    ensure_data_store()
    doc_id = str(uuid4())
    target_dir = TRIP_UPLOADS_DIR / target_trip_id
    target_dir.mkdir(parents=True, exist_ok=True)
    stored_name = doc_id + ext
    dest_path = target_dir / stored_name
    try:
        dest_path.write_bytes(src_path.read_bytes())
    except OSError:
        return None
    safe_name = (target_tourist_name or "passport").strip().replace("/", "-").replace("\\", "-") or "passport"
    new_doc = {
        "id": doc_id,
        "originalName": f"{safe_name} passport{ext}",
        "storedName": stored_name,
        "mimeType": src_doc.get("mimeType") or "application/octet-stream",
        "size": dest_path.stat().st_size,
        "category": "Passports & Visas",
        "touristId": target_tourist_id,
        "touristName": target_tourist_name,
        "uploadedAt": now_mongolia().isoformat(),
        "uploadedBy": actor_snapshot(actor) if actor else {},
    }
    trip = trips[target_index]
    documents = list(trip.get("documents") or [])
    documents.append(new_doc)
    trips[target_index] = {**trip, "documents": documents}
    write_camp_trips(trips)
    return new_doc


def consume_passport_token(token, trip_id, tourist_id, tourist_name, actor):
    """If a tourist record was saved with a passportFileToken from a
    prior scan, move the temp file into the trip's documents area as a
    `Passports & Visas` entry named after the tourist. Returns the new
    document dict, or None if the token was missing/expired/invalid."""
    if not token or not trip_id:
        return None
    safe_token = "".join(ch for ch in token if ch.isalnum())
    if not safe_token:
        return None
    matches = list(TOURIST_PASSPORT_TEMP_DIR.glob(f"{safe_token}.*"))
    if not matches:
        return None
    src = matches[0]
    ext = src.suffix.lower()
    trips = read_camp_trips()
    trip_index = next((i for i, t in enumerate(trips) if t.get("id") == trip_id), None)
    if trip_index is None:
        try:
            src.unlink()
        except Exception:
            pass
        return None
    ensure_data_store()
    doc_id = str(uuid4())
    trip_upload_dir = TRIP_UPLOADS_DIR / trip_id
    trip_upload_dir.mkdir(parents=True, exist_ok=True)
    stored_name = doc_id + ext
    dest_path = trip_upload_dir / stored_name
    try:
        src.rename(dest_path)
    except OSError:
        dest_path.write_bytes(src.read_bytes())
        try:
            src.unlink()
        except Exception:
            pass
    safe_name = (tourist_name or "passport").strip().replace("/", "-").replace("\\", "-")
    if not safe_name:
        safe_name = "passport"
    original_name = f"{safe_name} passport{ext}"
    mime_map = {".pdf": "application/pdf", ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png", ".webp": "image/webp"}
    doc = {
        "id": doc_id,
        "originalName": original_name,
        "storedName": stored_name,
        "mimeType": mime_map.get(ext, "application/octet-stream"),
        "size": dest_path.stat().st_size,
        "category": "Passports & Visas",
        "touristId": tourist_id,
        "touristName": tourist_name,
        "uploadedAt": now_mongolia().isoformat(),
        "uploadedBy": actor_snapshot(actor),
    }
    trip = trips[trip_index]
    documents = list(trip.get("documents") or [])
    documents.append(doc)
    trips[trip_index] = {**trip, "documents": documents}
    write_camp_trips(trips)
    return doc


# ── Mail accounts ───────────────────────────────────────────────────
def handle_list_mail_accounts(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    params = parse_qs(environ.get("QUERY_STRING", ""))
    scope = (params.get("scope", [""])[0] or "").strip().lower()
    accounts = [public_mail_account(a) for a in read_mail_accounts()]
    # The /mail page passes scope=workspace so it only sees mailboxes in the
    # active workspace. /mail-settings (admin) leaves scope unset and gets all.
    if scope == "workspace":
        ws = active_workspace(environ)
        if ws:
            accounts = [a for a in accounts if (a.get("workspace") or "DTX").upper() == ws]
    return json_response(start_response, "200 OK", {"entries": accounts})


def handle_create_mail_account(environ, start_response):
    # Managers can manage mailboxes too — only login/signup approval and
    # user-account CRUD stay admin-only.
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    address = (payload.get("address") or "").strip().lower()
    app_password = re.sub(r"\s+", "", payload.get("appPassword") or "")
    display_name = (payload.get("displayName") or "").strip() or address
    workspace = (payload.get("workspace") or "").strip().upper() or "DTX"
    imap_host = (payload.get("imapHost") or "imap.gmail.com").strip()
    imap_port = int(payload.get("imapPort") or 993)
    smtp_host = (payload.get("smtpHost") or "smtp.gmail.com").strip()
    smtp_port = int(payload.get("smtpPort") or 465)
    if not address or "@" not in address:
        return json_response(start_response, "400 Bad Request", {"error": "Valid email address is required"})
    if not app_password:
        return json_response(start_response, "400 Bad Request", {"error": "App password is required"})
    accounts = read_mail_accounts()
    if any(a.get("address", "").lower() == address for a in accounts):
        return json_response(start_response, "409 Conflict", {"error": "Mailbox already configured"})
    record = {
        "id": uuid4().hex,
        "address": address,
        "displayName": display_name,
        "workspace": workspace,
        "imapHost": imap_host,
        "imapPort": imap_port,
        "smtpHost": smtp_host,
        "smtpPort": smtp_port,
        "appPassword": _mail_obfuscate(app_password),
        "signatureHtml": (payload.get("signatureHtml") or "").strip(),
        "status": "ok",
        "lastError": "",
        "lastSyncedAt": "",
        "createdAt": now_mongolia().isoformat(),
        "updatedAt": now_mongolia().isoformat(),
    }
    accounts.append(record)
    write_mail_accounts(accounts)
    return json_response(start_response, "201 Created", {"ok": True, "entry": public_mail_account(record)})


def handle_update_mail_account(environ, start_response, account_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    accounts = read_mail_accounts()
    for i, account in enumerate(accounts):
        if account["id"] == account_id:
            updated = dict(account)
            for key in ("displayName", "workspace", "imapHost", "smtpHost"):
                if key in payload and payload[key] is not None:
                    updated[key] = str(payload[key]).strip()
            if "signatureHtml" in payload:
                # Trust admin-authored HTML; mail clients sanitize on render.
                updated["signatureHtml"] = str(payload.get("signatureHtml") or "")
            for key in ("imapPort", "smtpPort"):
                if key in payload and payload[key]:
                    try:
                        updated[key] = int(payload[key])
                    except (TypeError, ValueError):
                        pass
            new_password = re.sub(r"\s+", "", payload.get("appPassword") or "")
            if new_password:
                updated["appPassword"] = _mail_obfuscate(new_password)
                updated["status"] = "ok"
                updated["lastError"] = ""
            updated["updatedAt"] = now_mongolia().isoformat()
            accounts[i] = updated
            write_mail_accounts(accounts)
            return json_response(start_response, "200 OK", {"ok": True, "entry": public_mail_account(updated)})
    return json_response(start_response, "404 Not Found", {"error": "Mailbox not found"})


def handle_delete_mail_account(environ, start_response, account_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    accounts = read_mail_accounts()
    accounts2 = [a for a in accounts if a["id"] != account_id]
    if len(accounts2) == len(accounts):
        return json_response(start_response, "404 Not Found", {"error": "Mailbox not found"})
    write_mail_accounts(accounts2)
    msg_path = MAIL_MESSAGES_DIR / f"{account_id}.json"
    try:
        if msg_path.exists():
            msg_path.unlink()
    except Exception:
        pass
    # Clean up all body + attachment files for this account
    safe_acc = re.sub(r"[^A-Za-z0-9_-]", "_", str(account_id))
    try:
        for p in MAIL_MESSAGES_DIR.glob(f"body-{safe_acc}-*"):
            try: p.unlink()
            except Exception: pass
        for p in MAIL_MESSAGES_DIR.glob(f"att-{safe_acc}-*"):
            try: p.unlink()
            except Exception: pass
    except Exception:
        pass
    return json_response(start_response, "200 OK", {"ok": True})


def imap_test_connection(account):
    """Try to log in via IMAP to verify the app password works.
    Returns (ok, error_message)."""
    import imaplib
    host = account.get("imapHost") or "imap.gmail.com"
    port = int(account.get("imapPort") or 993)
    address = account.get("address")
    password = re.sub(r"\s+", "", get_mail_account_password(account) or "")
    if not address or not password:
        return False, "Missing address or password"
    try:
        password.encode("ascii")
    except UnicodeEncodeError:
        return False, "Password contains non-ASCII characters — re-paste it (Update password)"
    try:
        client = imaplib.IMAP4_SSL(host, port, timeout=20)
        try:
            client.login(address, password)
            client.select("INBOX", readonly=True)
            client.logout()
        finally:
            try:
                client.shutdown()
            except Exception:
                pass
    except imaplib.IMAP4.error as exc:
        return False, f"IMAP login failed: {exc}"
    except Exception as exc:
        return False, f"Connection failed: {exc}"
    return True, ""


def handle_test_mail_account(environ, start_response, account_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    accounts = read_mail_accounts()
    account = next((a for a in accounts if a["id"] == account_id), None)
    if not account:
        return json_response(start_response, "404 Not Found", {"error": "Mailbox not found"})
    ok, err = imap_test_connection(account)
    for i, a in enumerate(accounts):
        if a["id"] == account_id:
            accounts[i] = {
                **a,
                "status": "ok" if ok else "error",
                "lastError": err,
                "updatedAt": now_mongolia().isoformat(),
            }
            break
    write_mail_accounts(accounts)
    return json_response(start_response, "200 OK" if ok else "200 OK", {"ok": ok, "error": err})


# ── Mail message fetch + storage ────────────────────────────────────
def _decode_header_value(raw):
    if not raw:
        return ""
    try:
        from email.header import decode_header, make_header
        return str(make_header(decode_header(raw)))
    except Exception:
        return raw


def _parse_email_message(msg, account_id, uid):
    import email
    from email.utils import parseaddr, parsedate_to_datetime

    from_raw = _decode_header_value(msg.get("From", ""))
    to_raw = _decode_header_value(msg.get("To", ""))
    cc_raw = _decode_header_value(msg.get("Cc", ""))
    subject = _decode_header_value(msg.get("Subject", ""))
    date_iso = ""
    try:
        dt = parsedate_to_datetime(msg.get("Date", ""))
        if dt:
            date_iso = dt.isoformat()
    except Exception:
        pass

    body_text = ""
    body_html = ""
    attachments = []  # list of dicts: {idx, filename, contentType, size}
    if msg.is_multipart():
        idx = 0
        for part in msg.walk():
            ctype = part.get_content_type()
            cdisp = (part.get("Content-Disposition") or "").lower()
            fname_raw = part.get_filename()
            fname = _decode_header_value(fname_raw) if fname_raw else ""
            is_attachment = ("attachment" in cdisp) or bool(fname)
            if is_attachment:
                payload = part.get_payload(decode=True) or b""
                attachments.append({
                    "idx": idx,
                    "filename": fname or f"attachment-{idx}",
                    "contentType": ctype,
                    "size": len(payload),
                })
                idx += 1
                continue
            if ctype == "text/plain" and not body_text:
                payload = part.get_payload(decode=True) or b""
                charset = part.get_content_charset() or "utf-8"
                try:
                    body_text = payload.decode(charset, errors="replace")
                except Exception:
                    body_text = payload.decode("utf-8", errors="replace")
            elif ctype == "text/html" and not body_html:
                payload = part.get_payload(decode=True) or b""
                charset = part.get_content_charset() or "utf-8"
                try:
                    body_html = payload.decode(charset, errors="replace")
                except Exception:
                    body_html = payload.decode("utf-8", errors="replace")
    else:
        payload = msg.get_payload(decode=True) or b""
        charset = msg.get_content_charset() or "utf-8"
        try:
            decoded = payload.decode(charset, errors="replace")
        except Exception:
            decoded = payload.decode("utf-8", errors="replace")
        if msg.get_content_type() == "text/html":
            body_html = decoded
        else:
            body_text = decoded

    if not body_text and body_html:
        body_text_strip = re.sub(r"<[^>]+>", " ", body_html)
        body_text_strip = re.sub(r"\s+", " ", body_text_strip).strip()
    else:
        body_text_strip = body_text

    snippet = (body_text_strip or "")[:240].strip()
    from_name, from_email = parseaddr(from_raw)
    return {
        "uid": int(uid),
        "accountId": account_id,
        "messageId": msg.get("Message-ID", ""),
        "from": from_raw,
        "fromName": (from_name or from_email or "").strip(),
        "fromEmail": (from_email or "").strip().lower(),
        "to": to_raw,
        "cc": cc_raw,
        "subject": subject,
        "date": date_iso,
        "snippet": snippet,
        "bodyText": body_text,
        "bodyHtml": body_html,
        "hasAttachment": bool(attachments),
        "attachments": attachments,
        "fetchedAt": now_mongolia().isoformat(),
    }


def _attachment_path(account_id, folder, uid, idx):
    safe_acc = re.sub(r"[^A-Za-z0-9_-]", "_", str(account_id))
    safe_folder = re.sub(r"[^A-Za-z0-9_-]", "_", str(folder))
    return MAIL_MESSAGES_DIR / f"att-{safe_acc}-{safe_folder}-{int(uid)}-{int(idx)}.bin"


def _body_path(account_id, folder, uid):
    """Bodies live as small per-message JSON files on disk so the per-account
    cache stays tiny. This keeps memory pressure low: the cache file is read
    on every poll, but bodies are only loaded when the user opens a message."""
    safe_acc = re.sub(r"[^A-Za-z0-9_-]", "_", str(account_id))
    safe_folder = re.sub(r"[^A-Za-z0-9_-]", "_", str(folder))
    return MAIL_MESSAGES_DIR / f"body-{safe_acc}-{safe_folder}-{int(uid)}.json"


def _save_message_body(account_id, folder, uid, body_text, body_html):
    path = _body_path(account_id, folder, uid)
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        tmp = path.with_suffix(".json.tmp")
        with tmp.open("w", encoding="utf-8") as fh:
            json.dump({"bodyText": body_text or "", "bodyHtml": body_html or ""}, fh, ensure_ascii=False)
        tmp.replace(path)
    except Exception:
        pass


def _load_message_body(account_id, folder, uid):
    path = _body_path(account_id, folder, uid)
    if not path.exists():
        return {"bodyText": "", "bodyHtml": ""}
    try:
        with path.open("r", encoding="utf-8") as fh:
            data = json.load(fh)
        return {
            "bodyText": data.get("bodyText", "") or "",
            "bodyHtml": data.get("bodyHtml", "") or "",
        }
    except Exception:
        return {"bodyText": "", "bodyHtml": ""}


def _delete_message_files(account_id, folder, uid, attachment_count=0):
    """Remove body + attachment files for a message. Best-effort."""
    try:
        bp = _body_path(account_id, folder, uid)
        if bp.exists():
            bp.unlink()
    except Exception:
        pass
    for i in range(max(int(attachment_count or 0), 0)):
        try:
            ap = _attachment_path(account_id, folder, uid, i)
            if ap.exists():
                ap.unlink()
        except Exception:
            pass


def _persist_attachments(account_id, folder, uid, msg_obj, parsed):
    """Walk the parsed MIME message and save each attachment payload to disk
    so it can be served by /api/mail/messages/.../attachments/.../download."""
    if not parsed.get("attachments"):
        return
    try:
        idx = 0
        for part in msg_obj.walk():
            cdisp = (part.get("Content-Disposition") or "").lower()
            fname_raw = part.get_filename()
            is_attachment = ("attachment" in cdisp) or bool(fname_raw)
            if not is_attachment:
                continue
            payload = part.get_payload(decode=True) or b""
            path = _attachment_path(account_id, folder, uid, idx)
            try:
                path.parent.mkdir(parents=True, exist_ok=True)
                with path.open("wb") as fh:
                    fh.write(payload)
            except Exception:
                pass
            idx += 1
    except Exception:
        pass


def _empty_mail_cache():
    """A fresh cache supports multiple folders. Each folder entry tracks
    uidvalidity + lastUid; messages live in a single list and each carries
    a 'folder' field ('inbox' or 'sent')."""
    return {
        "folders": {
            "inbox": {"uidvalidity": None, "lastUid": 0},
            "sent": {"uidvalidity": None, "lastUid": 0},
        },
        "messages": [],
    }


def _read_mail_cache(account_id):
    path = MAIL_MESSAGES_DIR / f"{account_id}.json"
    if not path.exists():
        return _empty_mail_cache()
    try:
        with path.open("r", encoding="utf-8") as fh:
            data = json.load(fh)
        if not isinstance(data, dict):
            return _empty_mail_cache()
        # Migrate old single-folder cache into the new structure.
        if "folders" not in data:
            data = {
                "folders": {
                    "inbox": {
                        "uidvalidity": data.get("uidvalidity"),
                        "lastUid": int(data.get("lastUid") or 0),
                    },
                    "sent": {"uidvalidity": None, "lastUid": 0},
                },
                "messages": [
                    {**m, "folder": m.get("folder") or "inbox"}
                    for m in (data.get("messages") or [])
                ],
            }
        data.setdefault("messages", [])
        data.setdefault("folders", {})
        data["folders"].setdefault("inbox", {"uidvalidity": None, "lastUid": 0})
        data["folders"].setdefault("sent", {"uidvalidity": None, "lastUid": 0})

        # One-time migration: legacy caches stored bodyText/bodyHtml inline
        # on every message. Strip them out and persist to per-message files
        # so the in-memory cache footprint is tiny.
        if not data.get("bodyMigrationDone"):
            for m in data.get("messages", []):
                has_body = ("bodyText" in m) or ("bodyHtml" in m)
                if not has_body:
                    continue
                try:
                    uid_val = int(m.get("uid", 0))
                except Exception:
                    uid_val = 0
                folder_val = m.get("folder") or "inbox"
                if uid_val:
                    _save_message_body(
                        account_id, folder_val, uid_val,
                        m.pop("bodyText", "") or "",
                        m.pop("bodyHtml", "") or "",
                    )
                else:
                    m.pop("bodyText", None)
                    m.pop("bodyHtml", None)
            data["bodyMigrationDone"] = True
            try:
                _write_mail_cache(account_id, data)
            except Exception:
                pass
        return data
    except Exception:
        return _empty_mail_cache()


def _write_mail_cache(account_id, cache):
    path = MAIL_MESSAGES_DIR / f"{account_id}.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(".json.tmp")
    with tmp.open("w", encoding="utf-8") as fh:
        json.dump(cache, fh, ensure_ascii=False, indent=2)
    tmp.replace(path)


FOLDER_IMAP_NAMES = {
    "inbox": ["INBOX"],
    "sent": ['"[Gmail]/Sent Mail"', '"[Gmail]/&BB4EQgQ_BEAEMAQyBDsENQQ9BD0ESwQ1-"', "Sent"],
}


def _select_imap_folder(client, folder_key):
    """Try the conventional folder names for the given logical folder.
    Returns (selected_folder_name, error_or_empty)."""
    candidates = FOLDER_IMAP_NAMES.get(folder_key, [folder_key])
    last_err = ""
    for name in candidates:
        try:
            typ, _ = client.select(name, readonly=True)
            if typ == "OK":
                return name, ""
            last_err = f"select {name}: {typ}"
        except Exception as exc:
            last_err = str(exc)
            continue
    return "", last_err or f"Could not select folder '{folder_key}'"


def imap_sync_account(account, limit=20, folder="inbox"):
    """Fetch most recent messages from the given folder ('inbox' or 'sent')
    and append/update the local cache. Returns (new_message_count, error_string)."""
    import imaplib
    import email as email_mod
    address = account.get("address")
    password = re.sub(r"\s+", "", get_mail_account_password(account) or "")
    host = account.get("imapHost") or "imap.gmail.com"
    port = int(account.get("imapPort") or 993)
    if not address or not password:
        return 0, "Missing address or password"
    try:
        password.encode("ascii")
    except UnicodeEncodeError:
        return 0, "Password contains non-ASCII characters"

    cache = _read_mail_cache(account["id"])
    folder_state = cache["folders"].setdefault(folder, {"uidvalidity": None, "lastUid": 0})
    # Build a uid->message map but only for messages in *this* folder
    cache_messages = {int(m["uid"]): m for m in cache.get("messages", []) if m.get("folder") == folder}
    last_uid = int(folder_state.get("lastUid") or 0)
    cached_uidvalidity = folder_state.get("uidvalidity")

    try:
        client = imaplib.IMAP4_SSL(host, port, timeout=30)
        try:
            client.login(address, password)
            selected_name, sel_err = _select_imap_folder(client, folder)
            if not selected_name:
                return 0, sel_err
            typ, data = client.response("UIDVALIDITY")
            uidvalidity = None
            if typ == "OK" and data and data[0]:
                try:
                    uidvalidity = int(data[0])
                except Exception:
                    uidvalidity = None

            if cached_uidvalidity and uidvalidity and cached_uidvalidity != uidvalidity:
                cache_messages = {}
                last_uid = 0

            if last_uid > 0:
                typ, data = client.uid("search", None, f"UID {last_uid + 1}:*")
            else:
                typ, data = client.uid("search", None, "ALL")
            if typ != "OK" or not data:
                return 0, ""
            uids = [int(u) for u in (data[0] or b"").split() if u.isdigit()]
            uids = sorted(uids)
            if last_uid == 0:
                uids = uids[-limit:]
            else:
                uids = [u for u in uids if u > last_uid]

            new_count = 0
            for uid in uids:
                typ, data = client.uid("fetch", str(uid), "(FLAGS BODY.PEEK[])")
                if typ != "OK" or not data:
                    continue
                raw = None
                flags_str = ""
                for part in data:
                    if isinstance(part, tuple) and len(part) >= 2:
                        if not raw:
                            raw = part[1]
                            try:
                                flags_str = part[0].decode("utf-8", errors="replace")
                            except Exception:
                                flags_str = ""
                if not raw:
                    continue
                msg_obj = email_mod.message_from_bytes(raw)
                parsed = _parse_email_message(msg_obj, account["id"], uid)
                parsed["folder"] = folder
                parsed["isRead"] = "\\Seen" in flags_str
                # Save attachment payloads to disk (so the viewer can offer downloads)
                _persist_attachments(account["id"], folder, uid, msg_obj, parsed)
                # Bodies live on disk; strip them from the metadata kept in memory
                body_text = parsed.pop("bodyText", "") or ""
                body_html = parsed.pop("bodyHtml", "") or ""
                _save_message_body(account["id"], folder, uid, body_text, body_html)
                cache_messages[uid] = parsed
                # Free the raw bytes ASAP so the GC can reclaim them before
                # we fetch the next message in the loop.
                raw = None
                msg_obj = None
                new_count += 1
                if uid > last_uid:
                    last_uid = uid

            # Refresh \Seen flags for already-cached messages
            existing_uids = [u for u in cache_messages.keys() if u not in uids]
            if existing_uids:
                for i in range(0, len(existing_uids), 100):
                    batch = existing_uids[i:i+100]
                    uid_set = ",".join(str(u) for u in batch)
                    typ, data = client.uid("fetch", uid_set, "(FLAGS)")
                    if typ != "OK" or not data:
                        continue
                    for entry in data:
                        if not entry:
                            continue
                        line = entry if isinstance(entry, bytes) else (entry[0] if isinstance(entry, tuple) else b"")
                        if isinstance(line, bytes):
                            line_str = line.decode("utf-8", errors="replace")
                        else:
                            line_str = str(line)
                        m = re.search(r"UID\s+(\d+).*FLAGS\s*\(([^)]*)\)", line_str)
                        if m:
                            u_int = int(m.group(1))
                            flags = m.group(2)
                            if u_int in cache_messages:
                                cache_messages[u_int]["isRead"] = "\\Seen" in flags
        finally:
            try:
                client.logout()
            except Exception:
                pass
            try:
                client.shutdown()
            except Exception:
                pass
    except Exception as exc:
        return 0, f"Sync failed ({folder}): {exc}"

    # Re-merge the per-folder messages back with messages from *other* folders
    other_folders_messages = [m for m in cache.get("messages", []) if m.get("folder") != folder]
    merged = other_folders_messages + list(cache_messages.values())
    merged.sort(key=lambda m: m.get("date") or "", reverse=True)
    # Keep the most recent ~200 across all folders so we don't grow forever
    if len(merged) > 200:
        # Clean up bodies + attachments for messages we're about to drop
        for m in merged[200:]:
            try:
                folder_val = m.get("folder") or "inbox"
                uid_val = int(m.get("uid", 0))
                if uid_val:
                    _delete_message_files(account["id"], folder_val, uid_val, len(m.get("attachments") or []))
            except Exception:
                pass
        merged = merged[:200]

    cache["folders"][folder] = {
        "uidvalidity": uidvalidity if uidvalidity else cached_uidvalidity,
        "lastUid": last_uid,
    }
    cache["messages"] = merged
    _write_mail_cache(account["id"], cache)

    accounts = read_mail_accounts()
    for i, a in enumerate(accounts):
        if a["id"] == account["id"]:
            accounts[i] = {**a, "lastSyncedAt": now_mongolia().isoformat(), "status": "ok", "lastError": ""}
            break
    write_mail_accounts(accounts)
    return new_count, ""


_MAIL_SUBJECT_PREFIX_RE = re.compile(
    r"^(re|fw|fwd|sv|aw|tr|ynt|rv|вн|нэ|fyi)\s*[:\-]\s*",
    re.IGNORECASE,
)

def _normalize_mail_subject(s):
    """Strip Re:/Fwd: (and locale variants) so a thread can be grouped by
    its base subject. Mirrors the JS helper in mail.js."""
    t = (s or "").strip().lower()
    while True:
        new = _MAIL_SUBJECT_PREFIX_RE.sub("", t).strip()
        if new == t:
            break
        t = new
    return re.sub(r"\s+", " ", t)


def handle_mail_thread(environ, start_response, account_id, uid):
    """Return the full thread (inbox + sent) for the given message — same
    accountId, same normalized subject. Used by the conversation viewer
    so reply emails the user has sent appear inline next to the inbound
    side, WhatsApp-style."""
    actor = require_login(environ, start_response)
    if not actor:
        return []
    accounts = read_mail_accounts()
    account = next((a for a in accounts if a["id"] == account_id), None)
    if not account:
        return json_response(start_response, "404 Not Found", {"error": "Mailbox not found"})

    # Pull a fresh slice of Sent so the user's latest outbound replies show
    # up alongside the inbound side. Inbox syncs already happen on the list
    # endpoint; we just need to top up Sent here.
    try:
        imap_sync_account(account, limit=20, folder="sent")
    except Exception:
        pass

    cache = _read_mail_cache(account_id)
    target = next((m for m in cache.get("messages", []) if str(m.get("uid")) == str(uid)), None)
    if not target:
        return json_response(start_response, "404 Not Found", {"error": "Message not found"})

    norm = _normalize_mail_subject(target.get("subject"))
    entries = []
    for msg in cache.get("messages", []):
        if _normalize_mail_subject(msg.get("subject")) != norm:
            continue
        full = dict(msg)
        # Materialize body from disk (cache stores headers only).
        body = _load_message_body(account_id, msg.get("folder") or "inbox", msg.get("uid"))
        if body:
            full.update(body)
        full["accountAddress"] = account["address"]
        full["accountDisplay"] = account.get("displayName") or account["address"]
        full["workspace"] = account.get("workspace") or "DTX"
        entries.append(full)

    entries.sort(key=lambda m: m.get("date") or "")
    return json_response(start_response, "200 OK", {"entries": entries})


def read_mail_followups():
    raw = read_json_list(MAIL_FOLLOWUPS_FILE)
    cleaned = []
    for entry in raw:
        if not isinstance(entry, dict):
            continue
        cleaned.append(entry)
    return cleaned


def write_mail_followups(entries):
    write_json_list(MAIL_FOLLOWUPS_FILE, entries)


def _followup_thread_key(account_id, subject):
    return f"{account_id}::{_normalize_mail_subject(subject)}"


def build_mail_followup(payload, actor):
    days = parse_int(payload.get("days")) or 0
    if days < 1:
        days = 1
    if days > 60:
        days = 60
    now = datetime.now(timezone.utc)
    due = now + timedelta(days=days)
    account_id = normalize_text(payload.get("accountId"))
    subject = normalize_text(payload.get("subject"))
    return {
        "id": uuid4().hex,
        "accountId": account_id,
        "subject": subject,
        "subjectKey": _normalize_mail_subject(subject),
        "threadKey": _followup_thread_key(account_id, subject),
        "uid": normalize_text(payload.get("uid")),  # the message the user was looking at when arming
        "days": days,
        "createdAt": now.isoformat(),
        "dueAt": due.isoformat(),
        "status": "waiting",  # waiting | urgent | replied | cancelled
        "createdBy": actor_snapshot(actor) if actor else {},
        "lastNotifiedAt": "",
        "completedAt": "",
    }


def handle_list_mail_followups(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    try:
        scan_mail_followups()
    except Exception as exc:
        print(f"[mail-followup] scan err: {exc}", flush=True)
    entries = read_mail_followups()
    # Workspace-scope by joining each entry's accountId to its mail account.
    workspace = active_workspace(environ)
    if workspace:
        accounts = read_mail_accounts()
        by_id = {a["id"]: (a.get("workspace") or "DTX").upper() for a in accounts}
        entries = [e for e in entries if by_id.get(e.get("accountId")) == workspace]
    return json_response(start_response, "200 OK", {"entries": entries})


def handle_create_mail_followup(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    if not normalize_text(payload.get("accountId")):
        return json_response(start_response, "400 Bad Request", {"error": "accountId required"})
    if not normalize_text(payload.get("subject")):
        return json_response(start_response, "400 Bad Request", {"error": "subject required"})
    record = build_mail_followup(payload, actor)
    entries = read_mail_followups()
    # If the same thread already has an active follow-up, replace it.
    entries = [
        e for e in entries
        if not (e.get("threadKey") == record["threadKey"] and e.get("status") in {"waiting", "urgent"})
    ]
    entries.insert(0, record)
    write_mail_followups(entries)
    return json_response(start_response, "201 Created", {"ok": True, "entry": record})


def handle_update_mail_followup(environ, start_response, followup_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    method = environ.get("REQUEST_METHOD", "GET").upper()
    entries = read_mail_followups()
    for index, entry in enumerate(entries):
        if entry.get("id") != followup_id:
            continue
        if method == "DELETE":
            entries.pop(index)
            write_mail_followups(entries)
            return json_response(start_response, "200 OK", {"ok": True, "deletedId": followup_id})
        payload = collect_json(environ) or {}
        if "status" in payload:
            new_status = normalize_text(payload.get("status")).lower()
            if new_status in {"waiting", "urgent", "replied", "cancelled"}:
                entry["status"] = new_status
                if new_status in {"replied", "cancelled"}:
                    entry["completedAt"] = datetime.now(timezone.utc).isoformat()
        if "days" in payload:
            new_days = parse_int(payload.get("days")) or 0
            if 1 <= new_days <= 60:
                entry["days"] = new_days
                created = entry.get("createdAt")
                try:
                    base = datetime.fromisoformat(created) if created else datetime.now(timezone.utc)
                except Exception:
                    base = datetime.now(timezone.utc)
                entry["dueAt"] = (base + timedelta(days=new_days)).isoformat()
                if entry.get("status") == "urgent":
                    entry["status"] = "waiting"
        entries[index] = entry
        write_mail_followups(entries)
        return json_response(start_response, "200 OK", {"ok": True, "entry": entry})
    return json_response(start_response, "404 Not Found", {"error": "Follow-up not found"})


_MAIL_FOLLOWUP_SCAN_AT = [0.0]
_MAIL_FOLLOWUP_MIN_INTERVAL_SEC = 60


def scan_mail_followups():
    """For each waiting/urgent follow-up: peek at the thread's latest inbox
    message; if newer than the follow-up's createdAt, mark replied. Otherwise
    if past dueAt, mark urgent + log a one-time bell notification.

    Throttled to once per minute per worker, like scan_task_reminders.
    """
    nowmono = time.monotonic()
    if nowmono - _MAIL_FOLLOWUP_SCAN_AT[0] < _MAIL_FOLLOWUP_MIN_INTERVAL_SEC:
        return
    _MAIL_FOLLOWUP_SCAN_AT[0] = nowmono

    try:
        entries = read_mail_followups()
    except Exception:
        return
    if not entries:
        return

    now_utc = datetime.now(timezone.utc)
    changed = False
    cache_by_account = {}
    for entry in entries:
        if entry.get("status") not in {"waiting", "urgent"}:
            continue
        account_id = entry.get("accountId") or ""
        if not account_id:
            continue
        if account_id not in cache_by_account:
            try:
                cache_by_account[account_id] = _read_mail_cache(account_id)
            except Exception:
                cache_by_account[account_id] = {"messages": []}
        cache = cache_by_account[account_id]
        subject_key = entry.get("subjectKey") or _normalize_mail_subject(entry.get("subject"))
        created_iso = entry.get("createdAt") or ""
        # Latest INBOX message in this thread (i.e. a reply from the other side).
        latest = ""
        for msg in cache.get("messages", []):
            if (msg.get("folder") or "inbox") != "inbox":
                continue
            if _normalize_mail_subject(msg.get("subject")) != subject_key:
                continue
            d = msg.get("date") or ""
            if d > latest:
                latest = d
        if latest and created_iso and latest > created_iso:
            entry["status"] = "replied"
            entry["completedAt"] = now_utc.isoformat()
            changed = True
            continue
        # Past due → urgent + one-shot notification.
        try:
            due_dt = datetime.fromisoformat(entry.get("dueAt"))
        except Exception:
            continue
        if now_utc >= due_dt and entry.get("status") != "urgent":
            entry["status"] = "urgent"
            entry["lastNotifiedAt"] = now_utc.isoformat()
            changed = True
            try:
                log_notification(
                    "mail.followup.urgent",
                    entry.get("createdBy") or {"id": "system", "email": "noreply", "name": "TravelX"},
                    "Follow-up overdue",
                    detail=normalize_text(entry.get("subject")) or "Mail thread",
                    meta={"id": entry.get("id"), "accountId": account_id, "subjectKey": subject_key},
                )
            except Exception:
                pass

    if changed:
        try:
            write_mail_followups(entries)
        except Exception as exc:
            print(f"[mail-followup] write failed: {exc}", flush=True)


# --- Mail dossiers (manual link from a thread to a trip / tourist-group) ---

def read_mail_dossiers():
    # File is created on first write, so the read path has to tolerate the
    # absent / invalid case rather than rely on ensure_data_store.
    if not MAIL_DOSSIERS_FILE.exists():
        return []
    try:
        raw = json.loads(MAIL_DOSSIERS_FILE.read_text(encoding="utf-8"))
    except Exception:
        return []
    if not isinstance(raw, list):
        return []
    cleaned = []
    for entry in raw:
        if isinstance(entry, dict) and entry.get("threadKey"):
            cleaned.append(entry)
    return cleaned


def write_mail_dossiers(entries):
    ensure_data_store()
    MAIL_DOSSIERS_FILE.write_text(
        json.dumps(entries, indent=2, ensure_ascii=False), encoding="utf-8"
    )


def _dossier_thread_key(account_id, subject):
    return f"{account_id}::{_normalize_mail_subject(subject)}"


def _dossier_resolve(kind, target_id):
    """Return (label, extra) for a dossier link so the mail list can show it
    without an extra round-trip per row. `extra` carries the trip id for
    groups (the group page route requires both tripId and groupId)."""
    if not target_id:
        return "", {}
    if kind == "trip":
        for t in read_camp_trips():
            if t.get("id") == target_id:
                serial = t.get("serial") or ""
                name = t.get("tripName") or t.get("reservationName") or ""
                label = f"{serial} {name}".strip() or name or serial
                return label, {}
    elif kind == "group":
        for g in read_tourist_groups():
            if g.get("id") == target_id:
                serial = g.get("serial") or ""
                name = g.get("name") or g.get("groupName") or ""
                label = f"{serial} {name}".strip() or name or serial
                return label, {"tripId": g.get("tripId") or ""}
    return "", {}


def _dossier_label_for(kind, target_id):
    label, _extra = _dossier_resolve(kind, target_id)
    return label


def _build_dossier_index():
    """threadKey → enriched dossier dict for quick lookup during mail list
    enrichment. Re-resolves the label every call so renames flow through."""
    out = {}
    for entry in read_mail_dossiers():
        kind = (entry.get("kind") or "").lower()
        if kind not in ("trip", "group"):
            continue
        tk = entry.get("threadKey") or ""
        if not tk:
            continue
        target_id = entry.get("id") or ""
        label, extra = _dossier_resolve(kind, target_id)
        out[tk] = {
            "threadKey": tk,
            "kind": kind,
            "id": target_id,
            "label": label or entry.get("label") or "",
            "tripId": extra.get("tripId") or entry.get("tripId") or "",
            "linkedAt": entry.get("linkedAt") or "",
        }
    return out


def handle_list_mail_dossiers(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    entries = []
    for tk, d in _build_dossier_index().items():
        entries.append(d)
    return json_response(start_response, "200 OK", {"entries": entries})


def handle_create_mail_dossier(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    account_id = normalize_text(payload.get("accountId"))
    subject = normalize_text(payload.get("subject"))
    kind = (normalize_text(payload.get("kind")) or "").lower()
    target_id = normalize_text(payload.get("id"))
    if not account_id or not subject:
        return json_response(start_response, "400 Bad Request", {"error": "accountId + subject required"})
    if kind not in ("trip", "group"):
        return json_response(start_response, "400 Bad Request", {"error": "kind must be trip|group"})
    if not target_id:
        return json_response(start_response, "400 Bad Request", {"error": "id required"})
    label, extra = _dossier_resolve(kind, target_id)
    if not label:
        return json_response(start_response, "404 Not Found", {"error": f"{kind} not found"})
    tk = _dossier_thread_key(account_id, subject)
    record = {
        "threadKey": tk,
        "accountId": account_id,
        "subject": subject,
        "subjectKey": _normalize_mail_subject(subject),
        "kind": kind,
        "id": target_id,
        "label": label,
        "tripId": extra.get("tripId") or "",
        "linkedAt": datetime.now(timezone.utc).isoformat(),
        "linkedBy": actor_snapshot(actor) if actor else {},
    }
    entries = [e for e in read_mail_dossiers() if e.get("threadKey") != tk]
    entries.insert(0, record)
    write_mail_dossiers(entries)
    enriched = _build_dossier_index().get(tk, record)
    return json_response(start_response, "201 Created", {"ok": True, "entry": enriched})


def handle_delete_mail_dossier(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ) or {}
    params = parse_qs(environ.get("QUERY_STRING", ""))
    account_id = normalize_text(payload.get("accountId")) or (params.get("accountId", [""])[0] or "").strip()
    subject = normalize_text(payload.get("subject")) or (params.get("subject", [""])[0] or "").strip()
    if not account_id or not subject:
        return json_response(start_response, "400 Bad Request", {"error": "accountId + subject required"})
    tk = _dossier_thread_key(account_id, subject)
    entries = read_mail_dossiers()
    new_entries = [e for e in entries if e.get("threadKey") != tk]
    if len(new_entries) == len(entries):
        return json_response(start_response, "404 Not Found", {"error": "Not linked"})
    write_mail_dossiers(new_entries)
    return json_response(start_response, "200 OK", {"ok": True})


def handle_mail_dossier_targets(environ, start_response):
    """Picker payload: every trip + every tourist-group in the active workspace,
    flattened into rows the mail viewer can render without a second request."""
    actor = require_login(environ, start_response)
    if not actor:
        return []
    workspace = active_workspace(environ)
    trips = read_camp_trips()
    if workspace:
        trips = [t for t in trips if normalize_company(t.get("company")) == workspace]
    trip_rows = []
    for t in trips:
        trip_rows.append({
            "id": t.get("id"),
            "kind": "trip",
            "serial": t.get("serial") or "",
            "name": t.get("tripName") or t.get("reservationName") or "",
            "tripType": t.get("tripType") or "",
            "startDate": t.get("startDate") or "",
            "status": t.get("status") or "",
        })
    trip_ids = {t["id"] for t in trips}
    groups = [g for g in read_tourist_groups() if not workspace or g.get("tripId") in trip_ids]
    group_rows = []
    for g in groups:
        group_rows.append({
            "id": g.get("id"),
            "kind": "group",
            "serial": g.get("serial") or "",
            "name": g.get("name") or g.get("groupName") or "",
            "tripId": g.get("tripId") or "",
            "tripSerial": g.get("tripSerial") or "",
        })
    return json_response(start_response, "200 OK", {"trips": trip_rows, "groups": group_rows})


# ── Notes (free-text shared notes with @-mentions) ───────────────────

def read_notes():
    return read_json_list(NOTES_FILE)


def write_notes(records):
    write_json_list(NOTES_FILE, records)


def _parse_note_mentions(body):
    """Find @mentions in note body and resolve to user ids. Format expected:
    @[Full Name] — case-insensitive name match against approved users.
    Returns list of {id, name, email}."""
    if not body:
        return []
    found_names = re.findall(r"@\[([^\]]+)\]", body)
    if not found_names:
        return []
    users = read_users()
    mentions = []
    seen = set()
    for raw in found_names:
        name_clean = (raw or "").strip().lower()
        if not name_clean:
            continue
        for u in users:
            full = (u.get("fullName") or u.get("email") or "").strip()
            if full.lower() == name_clean and u.get("id") not in seen:
                mentions.append({
                    "id": u.get("id"),
                    "name": full,
                    "email": u.get("email") or "",
                })
                seen.add(u.get("id"))
                break
    return mentions


def build_note(payload, actor):
    body = normalize_text(payload.get("body"))
    mentions = _parse_note_mentions(body)
    parent_id = normalize_text(payload.get("parentId"))
    trip_id = normalize_text(payload.get("tripId"))
    group_id = normalize_text(payload.get("groupId"))
    company = normalize_text(payload.get("company")) or ""
    # Replies inherit trip/group/company scope from the parent note so the
    # whole thread stays attached together even if the client only sends parentId.
    if parent_id:
        for existing in read_notes():
            if existing.get("id") == parent_id:
                trip_id = trip_id or normalize_text(existing.get("tripId"))
                group_id = group_id or normalize_text(existing.get("groupId"))
                company = company or normalize_text(existing.get("company")) or ""
                break
    return {
        "id": uuid4().hex,
        "body": body,
        "mentions": mentions,
        "tripId": trip_id,
        "groupId": group_id,
        "parentId": parent_id,
        "company": company,
        "createdAt": datetime.now(timezone.utc).isoformat(),
        "createdBy": actor_snapshot(actor) if actor else {},
        "createdByAvatar": (actor.get("avatarPath") if isinstance(actor, dict) else "") or "",
        "updatedAt": "",
    }


def handle_list_notes(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    params = parse_qs(environ.get("QUERY_STRING", ""))
    trip_id = (params.get("tripId", [""])[0] or "").strip()
    group_id = (params.get("groupId", [""])[0] or "").strip()
    mentioned_me = (params.get("mentionedMe", [""])[0] or "") == "1"
    workspace = active_workspace(environ)
    notes = read_notes()
    if trip_id:
        notes = [n for n in notes if n.get("tripId") == trip_id]
    if group_id:
        notes = [n for n in notes if n.get("groupId") == group_id]
    if mentioned_me:
        my_id = (actor or {}).get("id")
        notes = [
            n for n in notes
            if any((m or {}).get("id") == my_id for m in (n.get("mentions") or []))
        ]
    # Workspace scoping: notes attached to a trip inherit that trip's
    # company; orphan notes are visible to everyone.
    if workspace and not trip_id and not group_id:
        trip_company = {t.get("id"): normalize_company(t.get("company")) for t in read_camp_trips()}
        notes = [
            n for n in notes
            if not n.get("tripId") or trip_company.get(n.get("tripId")) == workspace
        ]
    notes.sort(key=lambda n: n.get("createdAt") or "", reverse=True)
    return json_response(start_response, "200 OK", {"entries": notes})


def handle_create_note(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None or not normalize_text(payload.get("body")):
        return json_response(start_response, "400 Bad Request", {"error": "Note body required"})
    record = build_note(payload, actor)
    notes = read_notes()
    notes.insert(0, record)
    write_notes(notes)
    # Every note creates a bell entry for the team — replies count as note.reply
    # so they're distinguishable. Mentions also fire a personal note.mention
    # entry (used to drive the per-user "mentioning me" filter).
    actor_name = (actor or {}).get("fullName") or (actor or {}).get("email") or "someone"
    try:
        kind = "note.reply" if record.get("parentId") else "note.created"
        title = f"Reply by {actor_name}" if record.get("parentId") else f"New note by {actor_name}"
        log_notification(
            kind,
            actor,
            title,
            detail=record["body"][:120],
            meta={"id": record["id"], "tripId": record.get("tripId"), "groupId": record.get("groupId"), "parentId": record.get("parentId") or ""},
        )
    except Exception:
        pass
    for m in record.get("mentions") or []:
        try:
            log_notification(
                "note.mention",
                actor,
                f"Mentioned by {actor_name}",
                detail=record["body"][:120],
                meta={"id": record["id"], "tripId": record.get("tripId"), "groupId": record.get("groupId"), "mentionedUserId": m.get("id")},
            )
        except Exception:
            pass
    return json_response(start_response, "201 Created", {"ok": True, "entry": record})


def handle_update_note(environ, start_response, note_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    method = environ.get("REQUEST_METHOD", "GET").upper()
    notes = read_notes()
    is_admin = (actor.get("role") or "").lower() == "admin"
    for i, n in enumerate(notes):
        if n.get("id") != note_id:
            continue
        # Only the note's author or an admin may edit/delete.
        owner_id = ((n.get("createdBy") or {}).get("id")) or ""
        if not is_admin and owner_id != actor.get("id"):
            return json_response(start_response, "403 Forbidden", {"error": "Not allowed"})
        if method == "DELETE":
            notes.pop(i)
            write_notes(notes)
            return json_response(start_response, "200 OK", {"ok": True})
        payload = collect_json(environ) or {}
        if "body" in payload:
            n["body"] = normalize_text(payload.get("body"))
            n["mentions"] = _parse_note_mentions(n["body"])
        n["updatedAt"] = datetime.now(timezone.utc).isoformat()
        n["updatedBy"] = actor_snapshot(actor)
        notes[i] = n
        write_notes(notes)
        return json_response(start_response, "200 OK", {"ok": True, "entry": n})
    return json_response(start_response, "404 Not Found", {"error": "Note not found"})


REMINDERS_LAST_READ_FILE = DATA_DIR / "reminders_last_read.json"


def _reminders_last_read():
    try:
        if REMINDERS_LAST_READ_FILE.exists():
            return json.loads(REMINDERS_LAST_READ_FILE.read_text(encoding="utf-8")) or {}
    except Exception:
        return {}
    return {}


def _reminders_save_last_read(data):
    REMINDERS_LAST_READ_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _reminders_for_user(actor):
    """Personal mention inbox. Returns ONLY items that @-mention the
    current user — never tasks they happen to own. Mention form is
    @[Full Name]; case-insensitive whole-token match."""
    my_id = (actor or {}).get("id")
    my_name = ((actor or {}).get("fullName") or (actor or {}).get("email") or "").strip()
    items = []
    if not my_id and not my_name:
        return items
    marker = f"@[{my_name}]" if my_name else ""
    marker_lower = marker.lower()

    # Tasks: title or note must contain @[my name]. Status filter excludes
    # done/cancelled so old completed tasks don't keep pinging.
    if marker_lower:
        store = read_manager_dashboard()
        for t in (store.get("tasks") or []):
            status = (t.get("status") or "pending").lower()
            if status in {"done", "cancelled"}:
                continue
            haystack = ((t.get("title") or "") + " " + (t.get("note") or "")).lower()
            if marker_lower in haystack:
                items.append({
                    "kind": "task",
                    "id": t.get("id"),
                    "title": t.get("title") or "",
                    "note": t.get("note") or "",
                    "dueDate": t.get("dueDate") or "",
                    "dueTime": t.get("dueTime") or "",
                    "status": status,
                    "createdBy": t.get("createdBy") or {},
                    "createdAt": t.get("createdAt") or "",
                })

    # Notes mentioning me by id (parsed at save time).
    for n in read_notes():
        if any((m or {}).get("id") == my_id for m in (n.get("mentions") or [])):
            items.append({
                "kind": "note",
                "id": n.get("id"),
                "body": n.get("body") or "",
                "tripId": n.get("tripId") or "",
                "groupId": n.get("groupId") or "",
                "createdBy": n.get("createdBy") or {},
                "createdByAvatar": n.get("createdByAvatar") or "",
                "createdAt": n.get("createdAt") or "",
            })

    items.sort(key=lambda x: x.get("createdAt") or x.get("dueDate") or "", reverse=True)
    return items


def handle_list_reminders(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    items = _reminders_for_user(actor)
    last_read = (_reminders_last_read().get((actor or {}).get("id") or "") or "")
    unread = sum(1 for it in items if (it.get("createdAt") or "") > last_read)
    return json_response(start_response, "200 OK", {
        "count": unread,
        "total": len(items),
        "items": items[:80],
        "lastReadAt": last_read,
    })


def handle_mark_reminders_read(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    my_id = (actor or {}).get("id") or ""
    if not my_id:
        return json_response(start_response, "400 Bad Request", {"error": "no user id"})
    data = _reminders_last_read()
    data[my_id] = datetime.now(timezone.utc).isoformat()
    _reminders_save_last_read(data)
    return json_response(start_response, "200 OK", {"ok": True, "lastReadAt": data[my_id]})


def handle_mail_unread_summary(environ, start_response):
    """Cache-only summary used by the topbar mail icon. Returns the count of
    unread inbox messages plus a short list of recent ones, scoped to the
    workspace passed in ?workspace=DTX|USM (falls back to all)."""
    actor = require_login(environ, start_response)
    if not actor:
        return []
    params = parse_qs(environ.get("QUERY_STRING", ""))
    workspace = (params.get("workspace", [""])[0] or "").strip().upper()
    limit = 12
    accounts = read_mail_accounts()
    if workspace:
        accounts = [a for a in accounts if (a.get("workspace") or "DTX").upper() == workspace]

    rows = []
    for a in accounts:
        cache = _read_mail_cache(a["id"])
        for msg in cache.get("messages", []):
            if (msg.get("folder") or "inbox") != "inbox":
                continue
            if msg.get("isRead"):
                continue
            rows.append({
                "accountId": a["id"],
                "accountAddress": a["address"],
                "uid": msg.get("uid"),
                "subject": msg.get("subject") or "",
                "fromName": msg.get("fromName") or "",
                "fromEmail": msg.get("fromEmail") or "",
                "snippet": msg.get("snippet") or "",
                "date": msg.get("date") or "",
                "workspace": a.get("workspace") or "DTX",
            })

    rows.sort(key=lambda m: m.get("date") or "", reverse=True)
    return json_response(start_response, "200 OK", {
        "count": len(rows),
        "entries": rows[:limit],
    })


def handle_list_mail_messages(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    params = parse_qs(environ.get("QUERY_STRING", ""))
    do_sync = (params.get("sync", ["1"])[0] or "1") != "0"
    account_filter = (params.get("accountId", [""])[0] or "").strip()
    folder_filter = (params.get("folder", ["inbox"])[0] or "inbox").strip().lower()
    if folder_filter not in ("inbox", "sent"):
        folder_filter = "inbox"

    accounts = read_mail_accounts()
    if account_filter:
        accounts = [a for a in accounts if a["id"] == account_filter]

    sync_errors = []
    if do_sync:
        for a in accounts:
            # Sync the requested folder. Inbox is the hot path; sync sent only
            # when the user is viewing it (saves a Gmail round-trip).
            _, err = imap_sync_account(a, limit=50, folder=folder_filter)
            if err:
                sync_errors.append({"accountId": a["id"], "address": a["address"], "folder": folder_filter, "error": err})

    all_messages = []
    accounts_full = read_mail_accounts()
    if account_filter:
        accounts_full = [a for a in accounts_full if a["id"] == account_filter]
    dossier_index = _build_dossier_index()
    for a in accounts_full:
        cache = _read_mail_cache(a["id"])
        for msg in cache.get("messages", []):
            if (msg.get("folder") or "inbox") != folder_filter:
                continue
            slim = {k: v for k, v in msg.items() if k not in ("bodyText", "bodyHtml")}
            slim["accountAddress"] = a["address"]
            slim["accountDisplay"] = a.get("displayName") or a["address"]
            slim["workspace"] = a.get("workspace") or "DTX"
            tk = _dossier_thread_key(a["id"], msg.get("subject") or "")
            d = dossier_index.get(tk)
            slim["dossier"] = d if d else None
            all_messages.append(slim)

    all_messages.sort(key=lambda m: m.get("date") or "", reverse=True)
    return json_response(start_response, "200 OK", {
        "entries": all_messages,
        "folder": folder_filter,
        "syncErrors": sync_errors,
        "syncedAt": now_mongolia().isoformat(),
    })


def smtp_send_via_account(account, to_list, cc_list, bcc_list, subject, body, reply_to_message_id=None, in_reply_to_subject=None, attachments=None, signature_html="", body_html=""):
    """Send an email through Gmail SMTP using the connected account's
    app password. to_list/cc_list/bcc_list may be either bare emails or
    'Display Name <addr@x>' style strings — only the bare emails are used
    for the SMTP envelope; the headers preserve whatever the caller passed.
    Returns (ok, error_message)."""
    import smtplib
    from email.message import EmailMessage
    from email.utils import getaddresses
    address = account.get("address")
    password = re.sub(r"\s+", "", get_mail_account_password(account) or "")
    smtp_host = account.get("smtpHost") or "smtp.gmail.com"
    smtp_port = int(account.get("smtpPort") or 465)
    if not address or not password:
        return False, "Missing address or password"
    try:
        password.encode("ascii")
    except UnicodeEncodeError:
        return False, "Password contains non-ASCII characters"

    msg = EmailMessage()
    display_name = account.get("displayName") or address
    msg["From"] = f"{display_name} <{address}>"
    msg["To"] = ", ".join(to_list)
    if cc_list:
        msg["Cc"] = ", ".join(cc_list)
    msg["Subject"] = subject or "(no subject)"
    if reply_to_message_id:
        msg["In-Reply-To"] = reply_to_message_id
        msg["References"] = reply_to_message_id

    def html_to_text(html):
        if not html:
            return ""
        t = re.sub(r"<br\s*/?>", "\n", html, flags=re.IGNORECASE)
        t = re.sub(r"</p\s*>", "\n\n", t, flags=re.IGNORECASE)
        t = re.sub(r"</div\s*>", "\n", t, flags=re.IGNORECASE)
        t = re.sub(r"<li\b[^>]*>", "• ", t, flags=re.IGNORECASE)
        t = re.sub(r"</li\s*>", "\n", t, flags=re.IGNORECASE)
        t = re.sub(r"<[^>]+>", "", t)
        t = re.sub(r"&nbsp;", " ", t)
        t = re.sub(r"&amp;", "&", t)
        t = re.sub(r"&lt;", "<", t)
        t = re.sub(r"&gt;", ">", t)
        t = re.sub(r"&quot;", '"', t)
        t = re.sub(r"\n{3,}", "\n\n", t).strip()
        return t

    def escape_for_html(text):
        return (
            (text or "")
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace("\n", "<br>")
        )

    has_html_body = bool(body_html and body_html.strip())
    plain_body_part = html_to_text(body_html) if has_html_body else (body or "")

    if signature_html or has_html_body:
        # Build plain-text fallback
        sig_text = html_to_text(signature_html) if signature_html else ""
        plain_text = plain_body_part
        if sig_text:
            plain_text = f"{plain_text}\n\n--\n{sig_text}" if plain_text else sig_text
        # Build HTML version
        if has_html_body:
            html_main = body_html
        else:
            html_main = escape_for_html(body or "")
        full_html = (
            f"<div style=\"font-family: Arial, sans-serif; font-size: 14px; color: #0f172a;\">"
            f"{html_main}"
            f"{('<br><br>' + signature_html) if signature_html else ''}"
            f"</div>"
        )
        msg.set_content(plain_text or " ")
        msg.add_alternative(full_html, subtype="html")
    else:
        msg.set_content(body or "")

    if attachments:
        for att in attachments:
            try:
                data = att.get("data")
                if isinstance(data, str):
                    data = base64.b64decode(data)
                if not data:
                    continue
                ctype = att.get("contentType") or "application/octet-stream"
                maintype, _, subtype = ctype.partition("/")
                if not subtype:
                    maintype, subtype = "application", "octet-stream"
                fname = att.get("filename") or "attachment"
                msg.add_attachment(data, maintype=maintype, subtype=subtype, filename=fname)
            except Exception:
                pass

    # Build the envelope recipient list using the parsed bare email
    # portion of every header value, deduped case-insensitively.
    pairs = getaddresses(list(to_list) + list(cc_list or []) + list(bcc_list or []))
    seen = set()
    all_recipients = []
    for _name, email_addr in pairs:
        if not email_addr or "@" not in email_addr:
            continue
        key = email_addr.lower().strip()
        if key and key not in seen:
            seen.add(key)
            all_recipients.append(email_addr.strip())
    if not all_recipients:
        return False, "No valid recipients", []
    refused_recipients = []
    try:
        with smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=30) as smtp:
            smtp.login(address, password)
            # send_message returns a dict {refused_addr: (code, msg)} for any
            # recipients the server rejected at RCPT TO time. If ALL fail,
            # SMTPRecipientsRefused is raised instead.
            result = smtp.send_message(msg, from_addr=address, to_addrs=all_recipients) or {}
            for addr, (code, reason) in result.items():
                try:
                    reason_str = reason.decode("utf-8", errors="replace") if isinstance(reason, bytes) else str(reason)
                except Exception:
                    reason_str = str(reason)
                refused_recipients.append({"address": addr, "code": code, "reason": reason_str})
    except smtplib.SMTPAuthenticationError as exc:
        return False, f"SMTP auth failed: {exc}", []
    except smtplib.SMTPRecipientsRefused as exc:
        # ALL recipients refused — surface them all
        for addr, (code, reason) in (exc.recipients or {}).items():
            try:
                reason_str = reason.decode("utf-8", errors="replace") if isinstance(reason, bytes) else str(reason)
            except Exception:
                reason_str = str(reason)
            refused_recipients.append({"address": addr, "code": code, "reason": reason_str})
        return False, "All recipients refused by Gmail", refused_recipients
    except smtplib.SMTPSenderRefused as exc:
        return False, f"Sender refused: {exc.smtp_error}", []
    except smtplib.SMTPDataError as exc:
        return False, f"SMTP data error: {exc.smtp_error}", []
    except Exception as exc:
        return False, f"SMTP send failed: {exc}", []

    # Gmail SMTP automatically saves sent messages to [Gmail]/Sent Mail.
    # We deliberately do NOT IMAP APPEND a copy ourselves — that would
    # produce a duplicate in the sender's Sent folder.
    return True, "", refused_recipients


def handle_send_mail(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    account_id = (payload.get("fromAccountId") or "").strip()
    accounts = read_mail_accounts()
    account = next((a for a in accounts if a["id"] == account_id), None)
    if not account:
        return json_response(start_response, "400 Bad Request", {"error": "Choose a From mailbox"})

    from email.utils import getaddresses
    def split_addrs(value):
        """Parse a recipient string robustly: handles 'a@b, c@d', semicolons,
        and 'Display <addr@x>' forms. Returns a list of header-style strings
        (display name preserved when present, otherwise the bare email)."""
        if not value:
            return []
        if isinstance(value, list):
            raw = ", ".join(str(v) for v in value)
        else:
            raw = str(value)
        raw = raw.replace(";", ",")
        out = []
        for name, email_addr in getaddresses([raw]):
            email_addr = (email_addr or "").strip()
            if not email_addr or "@" not in email_addr:
                continue
            name = (name or "").strip()
            out.append(f"{name} <{email_addr}>" if name else email_addr)
        return out

    to_list = split_addrs(payload.get("to"))
    cc_list = split_addrs(payload.get("cc"))
    bcc_list = split_addrs(payload.get("bcc"))
    subject = (payload.get("subject") or "").strip()
    body = payload.get("body") or ""
    body_html = payload.get("bodyHtml") or ""
    reply_to = (payload.get("replyToMessageId") or "").strip()

    # Signature: prefer the per-user signatureId chosen in compose. Fall back
    # to the per-account signature if includeSignature is set (legacy path).
    signature_id = (payload.get("signatureId") or "").strip()
    if signature_id:
        signature_html = _resolve_user_signature_html(actor, signature_id)
    elif bool(payload.get("includeSignature")):
        signature_html = account.get("signatureHtml") or ""
    else:
        signature_html = ""

    if not to_list:
        return json_response(start_response, "400 Bad Request", {"error": "At least one To recipient is required"})

    ok, err, refused = smtp_send_via_account(
        account, to_list, cc_list, bcc_list, subject, body,
        reply_to_message_id=reply_to or None,
        signature_html=signature_html,
        body_html=body_html,
    )
    if not ok:
        return json_response(start_response, "500 Internal Server Error", {"error": err, "refused": refused})

    try:
        log_notification(
            "mail.sent",
            actor,
            "Mail sent",
            detail=f"{account['address']} → {', '.join(to_list)} · {subject[:80]}",
            meta={"accountId": account["id"], "to": to_list},
        )
    except Exception:
        pass
    return json_response(start_response, "200 OK", {"ok": True, "refused": refused})


# ── Signature image upload + serving ───────────────────────────────
_SIG_IMG_EXT_BY_TYPE = {
    "image/png": ".png",
    "image/jpeg": ".jpg",
    "image/jpg": ".jpg",
    "image/gif": ".gif",
    "image/webp": ".webp",
    "image/svg+xml": ".svg",
}
_SIG_IMG_TYPE_BY_EXT = {v: k for k, v in _SIG_IMG_EXT_BY_TYPE.items()}
_SIG_MAX_BYTES = 4 * 1024 * 1024  # 4MB ceiling per image


def handle_upload_signature_image(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    fields, files = parse_multipart(environ)
    upload = files.get("image")
    if not upload:
        return json_response(start_response, "400 Bad Request", {"error": "No image uploaded"})
    data = upload.get("data") or b""
    if len(data) > _SIG_MAX_BYTES:
        return json_response(start_response, "413 Payload Too Large", {"error": "Image too large (max 4MB)"})
    ctype = (upload.get("content_type") or "").lower().split(";")[0].strip()
    ext = _SIG_IMG_EXT_BY_TYPE.get(ctype)
    if not ext:
        # Fallback: trust the file extension
        fname = upload.get("filename") or ""
        guess = os.path.splitext(fname)[1].lower()
        if guess in _SIG_IMG_TYPE_BY_EXT:
            ext = guess
            ctype = _SIG_IMG_TYPE_BY_EXT[ext]
    if not ext:
        return json_response(start_response, "400 Bad Request", {"error": "Unsupported image type"})
    name = f"sig-{uuid4().hex}{ext}"
    path = MAIL_UPLOADS_DIR / name
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("wb") as fh:
            fh.write(data)
    except Exception as exc:
        return json_response(start_response, "500 Internal Server Error", {"error": f"Save failed: {exc}"})
    return json_response(start_response, "200 OK", {"ok": True, "url": f"/mail-uploads/{name}"})


def handle_serve_mail_upload(environ, start_response, filename):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    safe = re.sub(r"[^A-Za-z0-9._-]", "", filename or "")
    if not safe or safe.startswith(".") or "/" in safe or ".." in safe:
        return json_response(start_response, "400 Bad Request", {"error": "Bad filename"})
    path = MAIL_UPLOADS_DIR / safe
    if not path.exists() or not path.is_file():
        return json_response(start_response, "404 Not Found", {"error": "Not found"})
    ext = os.path.splitext(safe)[1].lower()
    ctype = _SIG_IMG_TYPE_BY_EXT.get(ext) or mimetypes.guess_type(safe)[0] or "application/octet-stream"
    try:
        with path.open("rb") as fh:
            payload = fh.read()
    except Exception:
        return json_response(start_response, "500 Internal Server Error", {"error": "Read failed"})
    headers = [
        ("Content-Type", ctype),
        ("Content-Length", str(len(payload))),
        ("Cache-Control", "private, max-age=86400"),
    ]
    start_response("200 OK", headers)
    return [payload]


# ── Mail templates CRUD ────────────────────────────────────────────
def read_mail_templates():
    if not MAIL_TEMPLATES_PATH.exists():
        return []
    try:
        with MAIL_TEMPLATES_PATH.open("r", encoding="utf-8") as fh:
            data = json.load(fh)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def write_mail_templates(items):
    MAIL_TEMPLATES_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp = MAIL_TEMPLATES_PATH.with_suffix(".json.tmp")
    with tmp.open("w", encoding="utf-8") as fh:
        json.dump(items, fh, ensure_ascii=False, indent=2)
    tmp.replace(MAIL_TEMPLATES_PATH)


def handle_list_mail_templates(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    items = read_mail_templates()
    items.sort(key=lambda t: (t.get("name") or "").lower())
    return json_response(start_response, "200 OK", {"entries": items})


def handle_create_mail_template(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ) or {}
    name = (payload.get("name") or "").strip()
    if not name:
        return json_response(start_response, "400 Bad Request", {"error": "Name is required"})
    record = {
        "id": uuid4().hex,
        "name": name,
        "subject": (payload.get("subject") or "").strip(),
        "bodyHtml": payload.get("bodyHtml") or "",
        "createdAt": now_mongolia().isoformat(),
        "updatedAt": now_mongolia().isoformat(),
    }
    items = read_mail_templates()
    items.append(record)
    write_mail_templates(items)
    return json_response(start_response, "201 Created", {"ok": True, "entry": record})


def handle_update_mail_template(environ, start_response, template_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ) or {}
    items = read_mail_templates()
    for i, t in enumerate(items):
        if t.get("id") == template_id:
            updated = dict(t)
            if "name" in payload:
                name = (payload.get("name") or "").strip()
                if not name:
                    return json_response(start_response, "400 Bad Request", {"error": "Name cannot be empty"})
                updated["name"] = name
            if "subject" in payload:
                updated["subject"] = (payload.get("subject") or "").strip()
            if "bodyHtml" in payload:
                updated["bodyHtml"] = payload.get("bodyHtml") or ""
            updated["updatedAt"] = now_mongolia().isoformat()
            items[i] = updated
            write_mail_templates(items)
            return json_response(start_response, "200 OK", {"ok": True, "entry": updated})
    return json_response(start_response, "404 Not Found", {"error": "Template not found"})


def handle_delete_mail_template(environ, start_response, template_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    items = read_mail_templates()
    new_items = [t for t in items if t.get("id") != template_id]
    if len(new_items) == len(items):
        return json_response(start_response, "404 Not Found", {"error": "Template not found"})
    write_mail_templates(new_items)
    return json_response(start_response, "200 OK", {"ok": True})


# ── Per-user signatures ────────────────────────────────────────────
def read_mail_signatures():
    if not MAIL_SIGNATURES_PATH.exists():
        return []
    try:
        with MAIL_SIGNATURES_PATH.open("r", encoding="utf-8") as fh:
            data = json.load(fh)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def write_mail_signatures(items):
    MAIL_SIGNATURES_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp = MAIL_SIGNATURES_PATH.with_suffix(".json.tmp")
    with tmp.open("w", encoding="utf-8") as fh:
        json.dump(items, fh, ensure_ascii=False, indent=2)
    tmp.replace(MAIL_SIGNATURES_PATH)


def _actor_user_id(actor):
    return (actor or {}).get("id") or (actor or {}).get("email") or ""


def _user_signatures(actor):
    uid = _actor_user_id(actor)
    return [s for s in read_mail_signatures() if s.get("userId") == uid]


def handle_list_my_signatures(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    items = _user_signatures(actor)
    items.sort(key=lambda s: (s.get("name") or "").lower())
    return json_response(start_response, "200 OK", {"entries": items})


def handle_create_my_signature(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ) or {}
    name = (payload.get("name") or "").strip()
    if not name:
        return json_response(start_response, "400 Bad Request", {"error": "Name is required"})
    record = {
        "id": uuid4().hex,
        "userId": _actor_user_id(actor),
        "name": name,
        "html": payload.get("html") or "",
        "createdAt": now_mongolia().isoformat(),
        "updatedAt": now_mongolia().isoformat(),
    }
    items = read_mail_signatures()
    items.append(record)
    write_mail_signatures(items)
    return json_response(start_response, "201 Created", {"ok": True, "entry": record})


def handle_update_my_signature(environ, start_response, sig_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ) or {}
    items = read_mail_signatures()
    uid = _actor_user_id(actor)
    for i, s in enumerate(items):
        if s.get("id") == sig_id:
            if s.get("userId") != uid:
                return json_response(start_response, "403 Forbidden", {"error": "Not your signature"})
            updated = dict(s)
            if "name" in payload:
                name = (payload.get("name") or "").strip()
                if not name:
                    return json_response(start_response, "400 Bad Request", {"error": "Name cannot be empty"})
                updated["name"] = name
            if "html" in payload:
                updated["html"] = payload.get("html") or ""
            updated["updatedAt"] = now_mongolia().isoformat()
            items[i] = updated
            write_mail_signatures(items)
            return json_response(start_response, "200 OK", {"ok": True, "entry": updated})
    return json_response(start_response, "404 Not Found", {"error": "Signature not found"})


def handle_delete_my_signature(environ, start_response, sig_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    items = read_mail_signatures()
    uid = _actor_user_id(actor)
    target = next((s for s in items if s.get("id") == sig_id), None)
    if not target:
        return json_response(start_response, "404 Not Found", {"error": "Signature not found"})
    if target.get("userId") != uid:
        return json_response(start_response, "403 Forbidden", {"error": "Not your signature"})
    new_items = [s for s in items if s.get("id") != sig_id]
    write_mail_signatures(new_items)
    return json_response(start_response, "200 OK", {"ok": True})


def _resolve_user_signature_html(actor, signature_id):
    if not signature_id:
        return ""
    uid = _actor_user_id(actor)
    for s in read_mail_signatures():
        if s.get("id") == signature_id and s.get("userId") == uid:
            return s.get("html") or ""
    return ""


def _imap_set_seen_flag(account, uid, folder, mark_read):
    """Set or clear the \\Seen flag. mark_read=True adds it, False removes it."""
    import imaplib
    address = account.get("address")
    password = re.sub(r"\s+", "", get_mail_account_password(account) or "")
    host = account.get("imapHost") or "imap.gmail.com"
    port = int(account.get("imapPort") or 993)
    if not address or not password:
        return False, "Missing address or password"
    op = "+FLAGS" if mark_read else "-FLAGS"
    try:
        client = imaplib.IMAP4_SSL(host, port, timeout=20)
        try:
            client.login(address, password)
            for name in FOLDER_IMAP_NAMES.get(folder, [folder]):
                try:
                    typ, _ = client.select(name)
                    if typ == "OK":
                        break
                except Exception:
                    continue
            client.uid("STORE", str(uid), op, "(\\Seen)")
        finally:
            try: client.logout()
            except Exception: pass
    except Exception as exc:
        verb = "Mark-read" if mark_read else "Mark-unread"
        return False, f"{verb} failed: {exc}"
    return True, ""


def imap_mark_read(account, uid, folder="inbox"):
    return _imap_set_seen_flag(account, uid, folder, True)


def imap_mark_unread(account, uid, folder="inbox"):
    return _imap_set_seen_flag(account, uid, folder, False)


def _handle_set_read_state(environ, start_response, account_id, uid, mark_read):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    accounts = read_mail_accounts()
    account = next((a for a in accounts if a["id"] == account_id), None)
    if not account:
        return json_response(start_response, "404 Not Found", {"error": "Mailbox not found"})
    try:
        uid_int = int(uid)
    except ValueError:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid uid"})
    params = parse_qs(environ.get("QUERY_STRING", ""))
    folder = (params.get("folder", ["inbox"])[0] or "inbox").strip().lower()
    if folder not in ("inbox", "sent"):
        folder = "inbox"
    ok, err = _imap_set_seen_flag(account, uid_int, folder, mark_read)
    cache = _read_mail_cache(account_id)
    for m in cache.get("messages", []):
        if int(m.get("uid", 0)) == uid_int and (m.get("folder") or "inbox") == folder:
            m["isRead"] = bool(mark_read)
            break
    _write_mail_cache(account_id, cache)
    return json_response(start_response, "200 OK", {"ok": ok, "error": err})


def handle_mark_read_mail_message(environ, start_response, account_id, uid):
    return _handle_set_read_state(environ, start_response, account_id, uid, True)


def handle_mark_unread_mail_message(environ, start_response, account_id, uid):
    return _handle_set_read_state(environ, start_response, account_id, uid, False)


def imap_delete_message(account, uid, folder="inbox"):
    """Move a message to Trash on Gmail. Returns (ok, error)."""
    import imaplib
    address = account.get("address")
    password = re.sub(r"\s+", "", get_mail_account_password(account) or "")
    host = account.get("imapHost") or "imap.gmail.com"
    port = int(account.get("imapPort") or 993)
    if not address or not password:
        return False, "Missing address or password"
    try:
        client = imaplib.IMAP4_SSL(host, port, timeout=20)
        try:
            client.login(address, password)
            # Re-select non-readonly so we can mutate
            candidates = FOLDER_IMAP_NAMES.get(folder, [folder])
            for name in candidates:
                try:
                    typ, _ = client.select(name)
                    if typ == "OK":
                        break
                except Exception:
                    continue
            try:
                client.uid("MOVE", str(uid), '"[Gmail]/Trash"')
            except Exception:
                try:
                    client.uid("COPY", str(uid), '"[Gmail]/Trash"')
                except Exception:
                    pass
                client.uid("STORE", str(uid), "+FLAGS", "(\\Deleted)")
                client.expunge()
        finally:
            try: client.logout()
            except Exception: pass
    except Exception as exc:
        return False, f"Delete failed: {exc}"
    return True, ""


def handle_delete_mail_message(environ, start_response, account_id, uid):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    accounts = read_mail_accounts()
    account = next((a for a in accounts if a["id"] == account_id), None)
    if not account:
        return json_response(start_response, "404 Not Found", {"error": "Mailbox not found"})
    try:
        uid_int = int(uid)
    except ValueError:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid uid"})
    params = parse_qs(environ.get("QUERY_STRING", ""))
    folder = (params.get("folder", ["inbox"])[0] or "inbox").strip().lower()
    if folder not in ("inbox", "sent"):
        folder = "inbox"
    ok, err = imap_delete_message(account, uid_int, folder=folder)
    if not ok:
        return json_response(start_response, "500 Internal Server Error", {"error": err})
    cache = _read_mail_cache(account_id)
    # Find the message first so we know how many attachments to clean up
    deleted_msg = next(
        (m for m in cache.get("messages", [])
         if int(m.get("uid", 0)) == uid_int and (m.get("folder") or "inbox") == folder),
        None,
    )
    cache["messages"] = [
        m for m in cache.get("messages", [])
        if not (int(m.get("uid", 0)) == uid_int and (m.get("folder") or "inbox") == folder)
    ]
    _write_mail_cache(account_id, cache)
    if deleted_msg:
        _delete_message_files(account_id, folder, uid_int, len(deleted_msg.get("attachments") or []))
    return json_response(start_response, "200 OK", {"ok": True})


def handle_get_mail_message(environ, start_response, account_id, uid):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    params = parse_qs(environ.get("QUERY_STRING", ""))
    folder = (params.get("folder", ["inbox"])[0] or "inbox").strip().lower()
    if folder not in ("inbox", "sent"):
        folder = "inbox"
    cache = _read_mail_cache(account_id)
    try:
        uid_int = int(uid)
    except ValueError:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid uid"})
    msg = next(
        (m for m in cache.get("messages", [])
         if int(m.get("uid", 0)) == uid_int and (m.get("folder") or "inbox") == folder),
        None,
    )
    if not msg:
        return json_response(start_response, "404 Not Found", {"error": "Message not found"})
    accounts = read_mail_accounts()
    account = next((a for a in accounts if a["id"] == account_id), None)

    # If this is an older cached message with hasAttachment=true but no
    # attachments array (cached before attachment support was added),
    # re-fetch the raw message once and rebuild metadata + persist payloads.
    if account and msg.get("hasAttachment") and not msg.get("attachments"):
        try:
            refreshed = _refetch_message(account, folder, uid_int)
            if refreshed:
                refreshed["isRead"] = msg.get("isRead", True)
                refreshed["folder"] = folder
                for i, m in enumerate(cache.get("messages", [])):
                    if int(m.get("uid", 0)) == uid_int and (m.get("folder") or "inbox") == folder:
                        cache["messages"][i] = refreshed
                        break
                _write_mail_cache(account_id, cache)
                msg = refreshed
        except Exception:
            pass

    # Bodies live on disk now — load on demand. Fall back to re-fetch if
    # the body file is missing (e.g. message synced before this change).
    body = _load_message_body(account_id, folder, uid_int)
    if not body["bodyText"] and not body["bodyHtml"] and account:
        try:
            refreshed = _refetch_message(account, folder, uid_int)
            if refreshed:
                refreshed["isRead"] = msg.get("isRead", True)
                refreshed["folder"] = folder
                for i, m in enumerate(cache.get("messages", [])):
                    if int(m.get("uid", 0)) == uid_int and (m.get("folder") or "inbox") == folder:
                        cache["messages"][i] = refreshed
                        break
                _write_mail_cache(account_id, cache)
                msg = refreshed
                body = _load_message_body(account_id, folder, uid_int)
        except Exception:
            pass

    body_text = body["bodyText"]
    body_html = body["bodyHtml"]
    if not body_html and body_text and body_text.lstrip()[:1] == "<" and re.search(r"<\s*(html|body|table|div|p|span|td)\b", body_text, re.IGNORECASE):
        body_html, body_text = body_text, ""

    response_msg = {**msg, "bodyText": body_text, "bodyHtml": body_html}
    if account:
        response_msg["accountAddress"] = account["address"]
        response_msg["accountDisplay"] = account.get("displayName") or account["address"]
    return json_response(start_response, "200 OK", {"entry": response_msg})


def _refetch_message(account, folder, uid):
    """Re-download a single message from IMAP, parse it, and persist its
    attachment payloads to disk. Returns the parsed message dict (or None)."""
    import imaplib
    import email as email_mod
    address = account.get("address")
    password = re.sub(r"\s+", "", get_mail_account_password(account) or "")
    host = account.get("imapHost") or "imap.gmail.com"
    port = int(account.get("imapPort") or 993)
    if not address or not password:
        return None
    try:
        client = imaplib.IMAP4_SSL(host, port, timeout=30)
        try:
            client.login(address, password)
            selected, _ = _select_imap_folder(client, folder)
            if not selected:
                return None
            typ, data = client.uid("fetch", str(uid), "(BODY.PEEK[])")
            if typ != "OK" or not data:
                return None
            raw = None
            for part in data:
                if isinstance(part, tuple) and len(part) >= 2:
                    raw = part[1]
                    break
            if not raw:
                return None
            msg_obj = email_mod.message_from_bytes(raw)
            parsed = _parse_email_message(msg_obj, account["id"], uid)
            parsed["folder"] = folder
            _persist_attachments(account["id"], folder, uid, msg_obj, parsed)
            body_text = parsed.pop("bodyText", "") or ""
            body_html = parsed.pop("bodyHtml", "") or ""
            _save_message_body(account["id"], folder, uid, body_text, body_html)
            return parsed
        finally:
            try: client.logout()
            except Exception: pass
    except Exception:
        return None


def handle_download_mail_attachment(environ, start_response, account_id, uid, idx):
    """Stream an attachment payload back to the browser. Falls back to
    re-fetching from IMAP if the on-disk cache file is missing."""
    actor = require_login(environ, start_response)
    if not actor:
        return []
    params = parse_qs(environ.get("QUERY_STRING", ""))
    folder = (params.get("folder", ["inbox"])[0] or "inbox").strip().lower()
    if folder not in ("inbox", "sent"):
        folder = "inbox"
    try:
        uid_int = int(uid)
        idx_int = int(idx)
    except ValueError:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid uid/idx"})

    cache = _read_mail_cache(account_id)
    msg = next(
        (m for m in cache.get("messages", [])
         if int(m.get("uid", 0)) == uid_int and (m.get("folder") or "inbox") == folder),
        None,
    )
    if not msg:
        return json_response(start_response, "404 Not Found", {"error": "Message not found"})
    attachments = msg.get("attachments") or []
    if idx_int < 0 or idx_int >= len(attachments):
        return json_response(start_response, "404 Not Found", {"error": "Attachment not found"})
    att = attachments[idx_int]

    path = _attachment_path(account_id, folder, uid_int, idx_int)
    payload = b""
    if path.exists():
        try:
            with path.open("rb") as fh:
                payload = fh.read()
        except Exception:
            payload = b""
    if not payload:
        # Fallback: re-fetch the message from IMAP and pull the part.
        accounts = read_mail_accounts()
        account = next((a for a in accounts if a["id"] == account_id), None)
        if account:
            payload = _fetch_attachment_from_imap(account, folder, uid_int, idx_int)
            if payload:
                try:
                    path.parent.mkdir(parents=True, exist_ok=True)
                    with path.open("wb") as fh:
                        fh.write(payload)
                except Exception:
                    pass
    if not payload:
        return json_response(start_response, "404 Not Found", {"error": "Attachment payload missing"})

    fname = att.get("filename") or f"attachment-{idx_int}"
    ctype = att.get("contentType") or "application/octet-stream"
    inline = (params.get("inline", ["0"])[0] or "0") == "1"
    disposition = "inline" if inline else "attachment"
    safe_fname = fname.encode("utf-8", errors="replace").decode("ascii", errors="replace")
    headers = [
        ("Content-Type", ctype),
        ("Content-Length", str(len(payload))),
        ("Content-Disposition", f'{disposition}; filename="{safe_fname}"; filename*=UTF-8\'\'{quote(fname)}'),
        ("Cache-Control", "private, max-age=3600"),
    ]
    start_response("200 OK", headers)
    return [payload]


def _fetch_attachment_from_imap(account, folder, uid, idx):
    """Re-download a single attachment by walking the message MIME tree."""
    import imaplib
    import email as email_mod
    address = account.get("address")
    password = re.sub(r"\s+", "", get_mail_account_password(account) or "")
    host = account.get("imapHost") or "imap.gmail.com"
    port = int(account.get("imapPort") or 993)
    if not address or not password:
        return b""
    try:
        client = imaplib.IMAP4_SSL(host, port, timeout=30)
        try:
            client.login(address, password)
            selected, _ = _select_imap_folder(client, folder)
            if not selected:
                return b""
            typ, data = client.uid("fetch", str(uid), "(BODY.PEEK[])")
            if typ != "OK" or not data:
                return b""
            raw = None
            for part in data:
                if isinstance(part, tuple) and len(part) >= 2:
                    raw = part[1]
                    break
            if not raw:
                return b""
            msg_obj = email_mod.message_from_bytes(raw)
            i = 0
            for part in msg_obj.walk():
                cdisp = (part.get("Content-Disposition") or "").lower()
                fname_raw = part.get_filename()
                if not (("attachment" in cdisp) or bool(fname_raw)):
                    continue
                if i == idx:
                    return part.get_payload(decode=True) or b""
                i += 1
        finally:
            try: client.logout()
            except Exception: pass
    except Exception:
        return b""
    return b""


def parse_multipart(environ):
    content_type = environ.get("CONTENT_TYPE", "")
    if "multipart/form-data" not in content_type:
        return {}, {}
    boundary = None
    for segment in content_type.split(";"):
        segment = segment.strip()
        if segment.startswith("boundary="):
            boundary = segment[9:].strip().strip('"')
            break
    if not boundary:
        return {}, {}
    try:
        content_length = int(environ.get("CONTENT_LENGTH") or "0")
    except ValueError:
        return {}, {}
    body = environ["wsgi.input"].read(content_length)
    fields, files = {}, {}
    sep = ("--" + boundary).encode()
    for part in body.split(sep)[1:]:
        if part.startswith(b"--"):
            break
        if part.startswith(b"\r\n"):
            part = part[2:]
        split_pos = part.find(b"\r\n\r\n")
        if split_pos == -1:
            continue
        raw_headers = part[:split_pos].decode("utf-8", errors="replace")
        raw_body = part[split_pos + 4:]
        if raw_body.endswith(b"\r\n"):
            raw_body = raw_body[:-2]
        headers = {}
        for line in raw_headers.split("\r\n"):
            if ":" in line:
                k, v = line.split(":", 1)
                headers[k.strip().lower()] = v.strip()
        disposition = headers.get("content-disposition", "")
        name = filename = None
        for token in disposition.split(";"):
            token = token.strip()
            if token.startswith("name="):
                name = token[5:].strip().strip('"')
            elif token.startswith("filename="):
                filename = token[9:].strip().strip('"')
        if name is None:
            continue
        if filename is not None:
            files[name] = {
                "filename": filename,
                "content_type": headers.get("content-type", "application/octet-stream"),
                "data": raw_body,
            }
        else:
            fields[name] = raw_body.decode("utf-8", errors="replace")
    return fields, files


def handle_upload_trip_document(environ, start_response, trip_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    trips = read_camp_trips()
    trip_index = next((i for i, t in enumerate(trips) if t["id"] == trip_id), None)
    if trip_index is None:
        return json_response(start_response, "404 Not Found", {"error": "Trip not found"})
    fields, files = parse_multipart(environ)
    if "file" not in files:
        return json_response(start_response, "400 Bad Request", {"error": "No file provided"})
    upload = files["file"]
    original_name = upload["filename"] or "file"
    ext = Path(original_name).suffix.lower()
    if ext not in ALLOWED_UPLOAD_EXTENSIONS:
        return json_response(start_response, "400 Bad Request", {"error": f"File type {ext} not allowed"})
    data = upload["data"]
    if len(data) > MAX_UPLOAD_BYTES:
        return json_response(start_response, "400 Bad Request", {"error": "File too large (max 10 MB)"})
    category = fields.get("category", "Other") or "Other"
    tourist_id = (fields.get("touristId") or "").strip()
    group_id = (fields.get("groupId") or "").strip()
    tourist_name = ""
    if tourist_id:
        t = next((x for x in read_tourists() if x.get("id") == tourist_id), None)
        if t:
            tourist_name = (
                (t.get("lastName") or "") + " " + (t.get("firstName") or "")
            ).strip()
            # If uploading from /group, the manager already chose the group;
            # otherwise, fall back to the tourist's own group.
            if not group_id and t.get("groupId"):
                group_id = t["groupId"]
    ensure_data_store()
    doc_id = str(uuid4())
    trip_upload_dir = TRIP_UPLOADS_DIR / trip_id
    trip_upload_dir.mkdir(parents=True, exist_ok=True)
    stored_name = doc_id + ext
    file_path = trip_upload_dir / stored_name
    file_path.write_bytes(data)
    doc = {
        "id": doc_id,
        "originalName": original_name,
        "storedName": stored_name,
        "mimeType": upload["content_type"],
        "size": len(data),
        "category": category,
        "touristId": tourist_id,
        "touristName": tourist_name,
        "groupId": group_id,
        "uploadedAt": now_mongolia().isoformat(),
        "uploadedBy": actor_snapshot(actor),
    }
    trip = trips[trip_index]
    documents = list(trip.get("documents") or [])
    documents.append(doc)
    trips[trip_index] = {**trip, "documents": documents}
    write_camp_trips(trips)
    return json_response(start_response, "201 Created", {"ok": True, "document": doc})


def handle_delete_trip_document(environ, start_response, trip_id, doc_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    trips = read_camp_trips()
    trip_index = next((i for i, t in enumerate(trips) if t["id"] == trip_id), None)
    if trip_index is None:
        return json_response(start_response, "404 Not Found", {"error": "Trip not found"})
    trip = trips[trip_index]
    documents = list(trip.get("documents") or [])
    doc = next((d for d in documents if d["id"] == doc_id), None)
    if doc is None:
        return json_response(start_response, "404 Not Found", {"error": "Document not found"})
    # Paid-document deletes are restricted to admins / accountants —
    # everyone else can still delete their own uploaded trip documents.
    if (doc.get("category") or "") == "Paid documents":
        role = (actor.get("role") or "").lower()
        if role not in ("admin", "accountant"):
            return json_response(start_response, "403 Forbidden", {"error": "Only admin / accountant can delete paid documents."})
    file_path = (TRIP_UPLOADS_DIR / trip_id / doc["storedName"]).resolve()
    if str(file_path).startswith(str(TRIP_UPLOADS_DIR.resolve())) and file_path.exists():
        file_path.unlink()
    trips[trip_index] = {**trip, "documents": [d for d in documents if d["id"] != doc_id]}
    write_camp_trips(trips)
    # Cascade: any payment_request that pointed at this doc now has a
    # dangling paidDocumentId. Blank it so the kebab on /accountant
    # stops offering Open/Download/Rename links that 404.
    requests_data = read_payment_requests()
    requests_changed = False
    for r in requests_data:
        if r.get("paidDocumentId") == doc_id:
            r["paidDocumentId"] = ""
            requests_changed = True
    if requests_changed:
        write_payment_requests(requests_data)
    return json_response(start_response, "200 OK", {"ok": True, "deletedId": doc_id})


def handle_email_trip_documents(environ, start_response, trip_id):
    """POST /api/camp-trips/{tripId}/documents/email — admin selects files and
    sends them to a client with a Mongolian boilerplate body. Reuses
    _tool_send_email so the auto-disclaimer footer + signature stay consistent."""
    actor = require_login(environ, start_response)
    if not actor:
        return []
    try:
        body = environ["wsgi.input"].read(int(environ.get("CONTENT_LENGTH") or "0"))
        data = json.loads(body)
    except Exception:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid JSON"})
    recipient = (data.get("recipientEmail") or "").strip()
    recipient_name = (data.get("recipientName") or "").strip()
    workspace = (data.get("workspace") or "").strip().upper()
    doc_ids = data.get("docIds") or []
    if not recipient or "@" not in recipient:
        return json_response(start_response, "400 Bad Request", {"error": "recipientEmail is required"})
    if not isinstance(doc_ids, list) or not doc_ids:
        return json_response(start_response, "400 Bad Request", {"error": "docIds (non-empty list) is required"})

    trip = next((t for t in read_camp_trips() if t.get("id") == trip_id), None)
    if not trip:
        return json_response(start_response, "404 Not Found", {"error": "Trip not found"})

    docs_by_id = {d.get("id"): d for d in (trip.get("documents") or [])}
    selected = [docs_by_id[i] for i in doc_ids if i in docs_by_id]
    if not selected:
        return json_response(start_response, "400 Bad Request", {"error": "no matching documents on this trip"})

    company_name = "Unlock Steppe Mongolia" if workspace == "USM" else "Дэлхий Трэвел Икс"
    greet = "Сайн байна уу" + (f", {recipient_name}" if recipient_name else "") + ","
    body_text = (
        f"{greet}\n\n"
        "Танд аяллын баримт бичгүүдийг хавсаргаж илгээж байна. "
        "Та хавсралтуудыг хүлээн авч, шаардлагатай бол хадгалаад авна уу."
    )
    subject = f"Аяллын баримт бичгүүд - {company_name}"
    attachments = [{"kind": "trip_document", "tripId": trip_id, "id": d["id"]} for d in selected]
    result = _tool_send_email({
        "to": recipient,
        "subject": subject,
        "body": body_text,
        "attachments": attachments,
        "_company_name": company_name,
    }, actor)

    if result.get("error"):
        return json_response(start_response, "500 Internal Server Error", {"error": result["error"]})
    return json_response(start_response, "200 OK", {
        "ok": True,
        "messageId": result.get("messageId"),
        "sent": len(attachments),
        "warnings": result.get("warnings"),
    })


def handle_rename_trip_document(environ, start_response, trip_id, doc_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    trips = read_camp_trips()
    trip_index = next((i for i, t in enumerate(trips) if t["id"] == trip_id), None)
    if trip_index is None:
        return json_response(start_response, "404 Not Found", {"error": "Trip not found"})
    trip = trips[trip_index]
    documents = list(trip.get("documents") or [])
    doc_index = next((i for i, d in enumerate(documents) if d["id"] == doc_id), None)
    if doc_index is None:
        return json_response(start_response, "404 Not Found", {"error": "Document not found"})
    try:
        body = environ["wsgi.input"].read(int(environ.get("CONTENT_LENGTH") or "0"))
        data = json.loads(body)
    except Exception:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid JSON"})
    updates = {}
    if "name" in data:
        new_name = (data.get("name") or "").strip()
        if not new_name:
            return json_response(start_response, "400 Bad Request", {"error": "Name cannot be empty"})
        updates["originalName"] = new_name
    if "touristId" in data:
        tourist_id = (data.get("touristId") or "").strip()
        tourist_name = ""
        if tourist_id:
            t = next((x for x in read_tourists() if x.get("id") == tourist_id), None)
            if t:
                tourist_name = (
                    (t.get("lastName") or "") + " " + (t.get("firstName") or "")
                ).strip()
        updates["touristId"] = tourist_id
        updates["touristName"] = tourist_name
    if not updates:
        return json_response(start_response, "400 Bad Request", {"error": "No fields to update"})
    documents[doc_index] = {**documents[doc_index], **updates}
    trips[trip_index] = {**trip, "documents": documents}
    write_camp_trips(trips)
    return json_response(start_response, "200 OK", {"ok": True, "document": documents[doc_index]})


def handle_delete_camp_trip(environ, start_response, trip_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    trips = read_camp_trips()
    deleted_trip = next((trip for trip in trips if trip["id"] == trip_id), None)
    if deleted_trip is None:
        return json_response(start_response, "404 Not Found", {"error": "Trip not found"})
    trips = [trip for trip in trips if trip["id"] != trip_id]
    reservations = [record for record in read_camp_reservations() if record.get("tripId") != trip_id]
    flight_reservations = [record for record in read_flight_reservations() if record.get("tripId") != trip_id]
    transfer_reservations = [record for record in read_transfer_reservations() if record.get("tripId") != trip_id]
    write_camp_trips(trips)
    write_camp_reservations(reservations)
    write_flight_reservations(flight_reservations)
    write_transfer_reservations(transfer_reservations)
    try:
        log_notification(
            "trip.deleted",
            actor,
            "Trip deleted",
            detail=normalize_text(deleted_trip.get("tripName")) or normalize_text(deleted_trip.get("serial")) or "",
            meta={"id": trip_id, "serial": deleted_trip.get("serial") or ""},
        )
    except Exception:
        pass
    return json_response(start_response, "200 OK", {"ok": True, "deletedId": trip_id, "summary": camp_summary(reservations)})


def handle_create_camp_reservation(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    guard = _check_payment_status_guard(actor, payload)
    if guard:
        return json_response(start_response, guard[0], guard[1])

    payload["company"] = active_workspace(environ) or DEFAULT_COMPANY
    record = build_camp_reservation(payload, actor)
    trip = find_camp_trip(record["tripId"])
    if trip is None:
        return json_response(start_response, "400 Bad Request", {"error": "Please create or select a trip first"})
    record["tripName"] = trip["tripName"]
    record["reservationName"] = record.get("reservationName") or trip.get("reservationName") or trip["tripName"]
    record["language"] = trip.get("language") or "Other"
    settings = read_camp_settings()
    if record.get("newCampName"):
        record["campName"] = record["newCampName"]
        if record["campName"] not in settings["campNames"]:
            settings["campNames"] = normalize_option_list(settings["campNames"] + [record["campName"]])
    if record.get("campName") and not record.get("locationName"):
        record["locationName"] = settings.get("campLocations", {}).get(record["campName"], "")
    if record.get("campName") and record.get("locationName"):
        settings["campLocations"][record["campName"]] = record["locationName"]
        settings["locationNames"] = normalize_option_list(settings["locationNames"] + [record["locationName"]])
    write_camp_settings(settings)
    error = validate_camp_reservation(record)
    if error:
        return json_response(start_response, "400 Bad Request", {"error": error})

    record.update(save_camp_reservation_document(record))
    records = read_camp_reservations()
    records.insert(0, record)
    write_camp_reservations(records)
    try:
        log_notification(
            "camp_reservation.created",
            actor,
            "Camp reservation created",
            detail=record.get("reservationName") or record.get("tripName") or "",
            meta={"id": record.get("id"), "tripId": record.get("tripId")},
        )
    except Exception:
        pass
    return json_response(start_response, "201 Created", {"ok": True, "entry": record, "summary": camp_summary(records)})


def handle_create_flight_reservation(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    guard = _check_payment_status_guard(actor, payload)
    if guard:
        return json_response(start_response, guard[0], guard[1])

    payload["company"] = active_workspace(environ) or DEFAULT_COMPANY
    record = build_flight_reservation(payload, actor)
    trip = find_camp_trip(record["tripId"])
    if trip is None:
        return json_response(start_response, "400 Bad Request", {"error": "Please create or select a trip first"})
    record["tripName"] = trip["tripName"]
    record["reservationName"] = record.get("reservationName") or trip.get("reservationName") or trip["tripName"]
    error = validate_flight_reservation(record)
    if error:
        return json_response(start_response, "400 Bad Request", {"error": error})
    records = read_flight_reservations()
    records.insert(0, record)
    write_flight_reservations(records)
    try:
        log_notification(
            "flight_reservation.created",
            actor,
            "Flight reservation created",
            detail=f"{record.get('fromCity','')} → {record.get('toCity','')}".strip(" →"),
            meta={"id": record.get("id"), "tripId": record.get("tripId")},
        )
    except Exception:
        pass
    return json_response(start_response, "201 Created", {"ok": True, "entry": record})


def handle_create_transfer_reservation(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    guard = _check_payment_status_guard(actor, payload)
    if guard:
        return json_response(start_response, guard[0], guard[1])

    payload["company"] = active_workspace(environ) or DEFAULT_COMPANY
    record = build_transfer_reservation(payload, actor)
    trip = find_camp_trip(record["tripId"])
    if trip is None:
        return json_response(start_response, "400 Bad Request", {"error": "Please create or select a trip first"})
    record["tripName"] = trip["tripName"]
    record["reservationName"] = record.get("reservationName") or trip.get("reservationName") or trip["tripName"]
    error = validate_transfer_reservation(record)
    if error:
        return json_response(start_response, "400 Bad Request", {"error": error})
    records = read_transfer_reservations()
    records.insert(0, record)
    write_transfer_reservations(records)
    try:
        log_notification(
            "transfer_reservation.created",
            actor,
            "Transfer reservation created",
            detail=record.get("reservationName") or record.get("tripName") or "",
            meta={"id": record.get("id"), "tripId": record.get("tripId")},
        )
    except Exception:
        pass
    return json_response(start_response, "201 Created", {"ok": True, "entry": record})


def handle_list_flight_reservations(environ, start_response):
    return json_response(start_response, "200 OK", {"entries": filter_by_company(read_flight_reservations(), environ)})


def handle_list_transfer_reservations(environ, start_response):
    return json_response(start_response, "200 OK", {"entries": filter_by_company(read_transfer_reservations(), environ)})


def build_flight_reservation(payload, actor=None):
    return {
        "id": str(uuid4()),
        "createdAt": now_mongolia().isoformat(),
        "tripId": normalize_text(payload.get("tripId")),
        "tripName": normalize_text(payload.get("tripName")),
        "reservationName": normalize_text(payload.get("reservationName")) or normalize_text(payload.get("tripName")),
        "flightScope": normalize_text(payload.get("flightScope")).lower() or "domestic",
        "routeType": normalize_text(payload.get("routeType")).lower() or "internal",
        "touristPaidBy": normalize_text(payload.get("touristPaidBy")).lower(),
        "guidePaidBy": normalize_text(payload.get("guidePaidBy")).lower(),
        "airline": normalize_text(payload.get("airline")),
        "flightNumber": normalize_text(payload.get("flightNumber")),
        "fromCity": normalize_text(payload.get("fromCity")),
        "toCity": normalize_text(payload.get("toCity")),
        "departureDate": normalize_text(payload.get("departureDate")),
        "tripDay": normalize_text(payload.get("tripDay")),
        "departureTime": normalize_text(payload.get("departureTime")),
        "arrivalDate": normalize_text(payload.get("arrivalDate")),
        "arrivalTime": normalize_text(payload.get("arrivalTime")),
        "staffCount": parse_int(payload.get("staffCount")),
        "ticketPrice": parse_int(payload.get("ticketPrice")),
        "totalTicketPrice": parse_int(payload.get("totalTicketPrice") or payload.get("amount")),
        "requested": normalize_text(payload.get("requested")).lower() or "no",
        "touristTicketStatus": normalize_text(payload.get("touristTicketStatus") or payload.get("status")).lower() or "waiting_list",
        "guideTicketStatus": normalize_text(payload.get("guideTicketStatus") or payload.get("guideStatus")).lower() or "waiting_list",
        "paidTo": normalize_text(payload.get("paidTo")),
        "paidDate": normalize_text(payload.get("paidDate")),
        "bookingReference": normalize_text(payload.get("bookingReference")),
        "ticketNumber": normalize_text(payload.get("ticketNumber")),
        "passengerCount": parse_int(payload.get("passengerCount")),
        "status": normalize_text(payload.get("status")).lower() or "to_check",
        "boughtDate": normalize_text(payload.get("boughtDate")),
        "paymentStatus": normalize_text(payload.get("paymentStatus")).lower() or "unpaid",
        "amount": parse_int(payload.get("totalTicketPrice") or payload.get("amount")),
        "currency": "MNT",
        "notes": normalize_text(payload.get("notes")),
        "company": normalize_company(payload.get("company")),
        "createdBy": actor_snapshot(actor),
        "updatedAt": "",
        "updatedBy": actor_snapshot(actor),
    }


def validate_flight_reservation(data):
    required = ["tripId", "tripName", "fromCity", "toCity", "departureDate", "touristTicketStatus", "guideTicketStatus", "paymentStatus"]
    missing = [field for field in required if not data.get(field)]
    if missing:
        return f"Missing required fields: {', '.join(missing)}"
    if data.get("passengerCount", 0) <= 0:
        return "Passenger count must be greater than 0"
    if not parse_date_input(data.get("departureDate")):
        return "Departure date must be in YYYY-MM-DD format"
    arrival_date = normalize_text(data.get("arrivalDate"))
    if arrival_date and not parse_date_input(arrival_date):
        return "Arrival date must be in YYYY-MM-DD format"
    return None


def build_transfer_reservation(payload, actor=None):
    return {
        "id": str(uuid4()),
        "createdAt": now_mongolia().isoformat(),
        "tripId": normalize_text(payload.get("tripId")),
        "tripName": normalize_text(payload.get("tripName")),
        "reservationName": normalize_text(payload.get("reservationName")) or normalize_text(payload.get("tripName")),
        "transferType": normalize_text(payload.get("transferType")).lower() or "airport_welcome",
        "pickupLocation": normalize_text(payload.get("pickupLocation")),
        "dropoffLocation": normalize_text(payload.get("dropoffLocation")),
        "serviceDate": normalize_text(payload.get("serviceDate")),
        "serviceTime": normalize_text(payload.get("serviceTime")),
        # New: separate "type time" — flight or train arrival/departure time —
        # distinct from when the driver actually picks the client up.
        "typeTime": normalize_text(payload.get("typeTime")),
        "supplierName": normalize_text(payload.get("supplierName")),
        # Driver is now picked from the Transfer Settings list. The id pins
        # the original record; name/car/plate/phone/salary are snapshotted on
        # the row so deleting/changing the driver later doesn't blank past
        # transfers.
        "driverId": normalize_text(payload.get("driverId")),
        "driverName": normalize_text(payload.get("driverName")),
        "vehicleType": normalize_text(payload.get("vehicleType")),
        "plateNumber": normalize_text(payload.get("plateNumber")),
        "driverPhoneNumber": normalize_text(payload.get("driverPhoneNumber")),
        "passengerCount": parse_int(payload.get("passengerCount")),
        "status": normalize_text(payload.get("status")).lower() or "pending",
        "paymentStatus": normalize_text(payload.get("paymentStatus")).lower() or "unpaid",
        "driverSalary": parse_int(payload.get("driverSalary") or payload.get("amount")),
        "currency": "MNT",
        "notes": normalize_text(payload.get("notes")),
        "company": normalize_company(payload.get("company")),
        "createdBy": actor_snapshot(actor),
        "updatedAt": "",
        "updatedBy": actor_snapshot(actor),
    }


def validate_transfer_reservation(data):
    required = ["tripId", "tripName", "transferType", "serviceDate", "paymentStatus"]
    missing = [field for field in required if not data.get(field)]
    if missing:
        return f"Missing required fields: {', '.join(missing)}"
    if data.get("passengerCount", 0) <= 0:
        return "Passenger count must be greater than 0"
    if not parse_date_input(data.get("serviceDate")):
        return "Service date must be in YYYY-MM-DD format"
    return None


def handle_update_camp_reservation(environ, start_response, reservation_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})

    records = read_camp_reservations()
    settings = read_camp_settings()
    for index, record in enumerate(records):
        if record["id"] != reservation_id:
            continue
        guard = _check_payment_status_guard(actor, payload, existing=record)
        if guard:
            return json_response(start_response, guard[0], guard[1])

        merged = {**record}
        for key in [
            "tripId",
            "tripName",
            "campName",
            "locationName",
            "reservationName",
            "reservationType",
            "createdDate",
            "checkIn",
            "checkOut",
            "roomType",
            "status",
            "staffAssignment",
            "notes",
            "breakfast",
            "lunch",
            "dinner",
            "depositPaidDate",
            "secondPaidDate",
            "thirdPaidDate",
            "fourthPaidDate",
            "paymentStatus",
        ]:
            if key in payload:
                merged[key] = normalize_text(payload.get(key))

        if normalize_text(payload.get("newCampName")):
            merged["campName"] = normalize_text(payload.get("newCampName"))
            if merged["campName"] not in settings["campNames"]:
                settings["campNames"] = normalize_option_list(settings["campNames"] + [merged["campName"]])
        if not normalize_text(merged.get("reservationName")):
            trip = find_camp_trip(merged.get("tripId"))
            merged["reservationName"] = (trip or {}).get("reservationName") or merged.get("tripName") or ""
        if merged.get("campName") and not normalize_text(merged.get("locationName")):
            merged["locationName"] = settings.get("campLocations", {}).get(merged["campName"], "")
        if merged.get("campName") and normalize_text(merged.get("locationName")):
            settings["campLocations"][merged["campName"]] = normalize_text(merged["locationName"])
            settings["locationNames"] = normalize_option_list(settings["locationNames"] + [normalize_text(merged["locationName"])])
        write_camp_settings(settings)

        for key in ["clientCount", "staffCount", "gerCount", "deposit", "secondPayment", "thirdPayment", "fourthPayment", "totalPayment", "balancePayment", "paidAmount"]:
            if key in payload:
                merged[key] = parse_int(payload.get(key))

        merged["checkIn"], merged["nights"], merged["checkOut"] = normalize_stay_fields(
            merged.get("checkIn"),
            payload.get("nights", merged.get("nights")),
            merged.get("checkOut"),
        )

        error = validate_camp_reservation(merged)
        if error:
            return json_response(start_response, "400 Bad Request", {"error": error})

        merged["updatedAt"] = now_mongolia().isoformat()
        merged["updatedBy"] = actor_snapshot(actor)
        merged.update(save_camp_reservation_document(merged))
        records[index] = merged
        write_camp_reservations(records)
        try:
            log_notification(
                "camp_reservation.updated",
                actor,
                "Camp reservation updated",
                detail=f"{merged.get('reservationName', '')} · {merged.get('campName', '')}".strip(" ·"),
                meta={"id": merged.get("id"), "tripId": merged.get("tripId")},
            )
        except Exception:
            pass
        return json_response(start_response, "200 OK", {"ok": True, "entry": merged, "summary": camp_summary(records)})

    return json_response(start_response, "404 Not Found", {"error": "Camp reservation not found"})


def handle_update_flight_reservation(environ, start_response, reservation_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})

    records = read_flight_reservations()
    for index, record in enumerate(records):
        if record["id"] != reservation_id:
            continue
        guard = _check_payment_status_guard(actor, payload, existing=record)
        if guard:
            return json_response(start_response, guard[0], guard[1])
        merged = {**record}
        for key in [
            "tripId",
            "tripName",
            "reservationName",
            "flightScope",
            "routeType",
            "touristPaidBy",
            "guidePaidBy",
            "airline",
            "flightNumber",
            "fromCity",
            "toCity",
            "departureDate",
            "tripDay",
            "departureTime",
            "arrivalDate",
            "arrivalTime",
            "requested",
            "touristTicketStatus",
            "guideTicketStatus",
            "paidTo",
            "paidDate",
            "bookingReference",
            "ticketNumber",
            "status",
            "boughtDate",
            "paymentStatus",
            "currency",
            "notes",
        ]:
            if key in payload:
                value = normalize_text(payload.get(key))
                merged[key] = value.upper() if key == "currency" else value
        for key in ["passengerCount", "staffCount", "ticketPrice", "totalTicketPrice", "amount"]:
            if key in payload:
                target_key = "totalTicketPrice" if key in {"totalTicketPrice", "amount"} else key
                merged[target_key] = parse_int(payload.get(key))
        merged["amount"] = parse_int(merged.get("totalTicketPrice"))
        merged["currency"] = "MNT"
        if normalize_text(merged.get("tripId")):
            trip = find_camp_trip(merged.get("tripId"))
            if trip is None:
                return json_response(start_response, "400 Bad Request", {"error": "Selected trip was not found"})
            merged["tripName"] = trip["tripName"]
            if not normalize_text(merged.get("reservationName")):
                merged["reservationName"] = trip.get("reservationName") or trip["tripName"]
        error = validate_flight_reservation(merged)
        if error:
            return json_response(start_response, "400 Bad Request", {"error": error})
        merged["updatedAt"] = now_mongolia().isoformat()
        merged["updatedBy"] = actor_snapshot(actor)
        records[index] = merged
        write_flight_reservations(records)
        return json_response(start_response, "200 OK", {"ok": True, "entry": merged})
    return json_response(start_response, "404 Not Found", {"error": "Flight reservation not found"})


def handle_update_transfer_reservation(environ, start_response, reservation_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})

    records = read_transfer_reservations()
    for index, record in enumerate(records):
        if record["id"] != reservation_id:
            continue
        guard = _check_payment_status_guard(actor, payload, existing=record)
        if guard:
            return json_response(start_response, guard[0], guard[1])
        merged = {**record}
        for key in [
            "tripId",
            "tripName",
            "reservationName",
            "transferType",
            "pickupLocation",
            "dropoffLocation",
            "serviceDate",
            "serviceTime",
            "typeTime",
            "driverId",
            "driverName",
            "vehicleType",
            "plateNumber",
            "driverPhoneNumber",
            "paymentStatus",
            "currency",
            "notes",
        ]:
            if key in payload:
                value = normalize_text(payload.get(key))
                merged[key] = value.upper() if key == "currency" else value
        for key in ["passengerCount", "amount", "driverSalary"]:
            if key in payload:
                target_key = "driverSalary" if key in {"amount", "driverSalary"} else key
                merged[target_key] = parse_int(payload.get(key))
        merged["currency"] = "MNT"
        if normalize_text(merged.get("tripId")):
            trip = find_camp_trip(merged.get("tripId"))
            if trip is None:
                return json_response(start_response, "400 Bad Request", {"error": "Selected trip was not found"})
            merged["tripName"] = trip["tripName"]
            if not normalize_text(merged.get("reservationName")):
                merged["reservationName"] = trip.get("reservationName") or trip["tripName"]
        error = validate_transfer_reservation(merged)
        if error:
            return json_response(start_response, "400 Bad Request", {"error": error})
        merged["updatedAt"] = now_mongolia().isoformat()
        merged["updatedBy"] = actor_snapshot(actor)
        records[index] = merged
        write_transfer_reservations(records)
        return json_response(start_response, "200 OK", {"ok": True, "entry": merged})
    return json_response(start_response, "404 Not Found", {"error": "Transfer reservation not found"})


def handle_delete_camp_reservation(environ, start_response, reservation_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    records = read_camp_reservations()
    if not any(record["id"] == reservation_id for record in records):
        return json_response(start_response, "404 Not Found", {"error": "Camp reservation not found"})
    records = [record for record in records if record["id"] != reservation_id]
    write_camp_reservations(records)
    return json_response(start_response, "200 OK", {"ok": True, "deletedId": reservation_id, "summary": camp_summary(records)})


def handle_delete_flight_reservation(environ, start_response, reservation_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    records = read_flight_reservations()
    if not any(record["id"] == reservation_id for record in records):
        return json_response(start_response, "404 Not Found", {"error": "Flight reservation not found"})
    records = [record for record in records if record["id"] != reservation_id]
    write_flight_reservations(records)
    return json_response(start_response, "200 OK", {"ok": True, "deletedId": reservation_id})


def handle_delete_transfer_reservation(environ, start_response, reservation_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    records = read_transfer_reservations()
    if not any(record["id"] == reservation_id for record in records):
        return json_response(start_response, "404 Not Found", {"error": "Transfer reservation not found"})
    records = [record for record in records if record["id"] != reservation_id]
    write_transfer_reservations(records)
    return json_response(start_response, "200 OK", {"ok": True, "deletedId": reservation_id})


def handle_export_camp_reservations(environ, start_response):
    if not require_login(environ, start_response):
        return []
    params = parse_qs(environ.get("QUERY_STRING", ""))
    ids = [value for item in params.get("ids", []) for value in item.split(",") if value]
    records = read_camp_reservations()
    selected = [record for record in records if not ids or record["id"] in ids]
    if ids:
        order = {value: index for index, value in enumerate(ids)}
        selected.sort(key=lambda record: order.get(record["id"], len(order)))
    if not selected:
        return json_response(start_response, "400 Bad Request", {"error": "No reservations selected"})
    document = save_camp_reservations_bundle(selected)
    if not str(document.get("pdfPath", "")).endswith(".pdf"):
        return json_response(start_response, "500 Internal Server Error", {"error": "PDF generation failed"})
    return json_response(start_response, "200 OK", {"ok": True, "entry": document})


def handle_create_backup(environ, start_response):
    admin = require_backup_admin(environ, start_response)
    if not admin:
        return []
    archive_path = create_backup_archive()
    return json_response(
        start_response,
        "201 Created",
        {"ok": True, "filename": archive_path.name, "createdAt": now_mongolia().isoformat()},
    )


def handle_list_backups(environ, start_response):
    admin = require_backup_admin(environ, start_response)
    if not admin:
        return []
    return json_response(start_response, "200 OK", {"ok": True, "backups": list_backup_archives()})


def handle_download_backup(environ, start_response, filename):
    admin = require_backup_admin(environ, start_response)
    if not admin:
        return []
    safe_path = (BACKUP_DIR / unquote(filename)).resolve()
    if not str(safe_path).startswith(str(BACKUP_DIR.resolve())) or not safe_path.exists():
        return json_response(start_response, "404 Not Found", {"error": "Backup not found"})
    if safe_path.suffix != ".zip":
        return json_response(start_response, "400 Bad Request", {"error": "Invalid backup file"})
    return file_response(start_response, safe_path, "application/zip")


def handle_camp_reservation_document(environ, start_response, reservation_id):
    if not require_login(environ, start_response):
        return []
    params = parse_qs(environ.get("QUERY_STRING", ""))
    mode = (params.get("mode", ["view"])[0] or "view").strip().lower()
    records = read_camp_reservations()
    record = next((item for item in records if item["id"] == reservation_id), None)
    if not record:
        return json_response(start_response, "404 Not Found", {"error": "Camp reservation not found"})
    document = save_camp_reservation_document(record)
    if mode == "download" and not str(document.get("pdfPath", "")).endswith(".pdf"):
        return json_response(start_response, "500 Internal Server Error", {"error": "PDF generation failed"})
    relative_path = document["pdfViewPath"] if mode == "view" else document["pdfPath"]
    safe_path = (GENERATED_DIR / unquote(relative_path.replace("/generated/", "", 1))).resolve()
    if not str(safe_path).startswith(str(GENERATED_DIR.resolve())) or not safe_path.exists():
        return json_response(start_response, "404 Not Found", {"error": "Document not found"})
    extra_headers = generated_download_headers(safe_path) if mode == "download" else None
    return file_response(start_response, safe_path, extra_headers=extra_headers)


def build_invoice_from_contract(contract, actor):
    """Create a standalone invoice record mirroring a freshly-saved contract.
    Skipped (returns None) if the contract has no trip+group attached, since
    invoices require both. Items mirror the contract line items where possible,
    otherwise fall back to a single line for the total. Two installments
    (deposit / balance) are created from the contract's deposit/balance fields.
    """
    trip_id = (contract.get("tripId") or "").strip()
    group_id = (contract.get("groupId") or "").strip()
    if not trip_id:
        return None
    # FIT trips usually create contracts without an explicit groupId.
    # Fall back to the trip's first existing group; if none exist yet,
    # create a default group named after the trip so the invoice can attach.
    if not group_id:
        groups = [g for g in read_tourist_groups() if g.get("tripId") == trip_id]
        if groups:
            group_id = groups[0]["id"]
        else:
            trip = find_camp_trip(trip_id)
            if trip:
                default = build_tourist_group({
                    "tripId": trip_id,
                    "name": trip.get("tripName") or "Default group",
                    "headcount": trip.get("participantCount") or 0,
                }, actor)
                groups_all = read_tourist_groups()
                groups_all.append(default)
                write_tourist_groups(groups_all)
                group_id = default["id"]
        if not group_id:
            return None

    data = contract.get("data") or {}
    payer = f"{normalize_text(data.get('touristLastName'))} {normalize_text(data.get('touristFirstName'))}".strip()
    if not payer:
        payer = normalize_text(data.get("clientName")) or "Client"

    def _num(v):
        try:
            return float(str(v).replace(",", "").replace("₮", "").strip() or 0)
        except Exception:
            return 0.0

    # Build line items from per-category counts/prices on the contract.
    raw_items = []
    cats = [
        ("adultCount", "adultPrice", "Том хүн"),
        ("childCount", "childPrice", "Хүүхэд"),
        ("infantCount", "infantPrice", "Нярай"),
        ("ticketOnlyCount", "ticketOnlyPrice", "Зөвхөн билет"),
        ("landOnlyCount", "landOnlyPrice", "Газрын үйлчилгээ"),
        ("customCount", "customPrice", "Бусад"),
    ]
    for count_key, price_key, default_label in cats:
        qty = _num(data.get(count_key))
        price = _num(data.get(price_key))
        if qty <= 0 or price <= 0:
            continue
        label = default_label
        if count_key == "customCount":
            label = normalize_text(data.get("customPriceLabel")) or default_label
        raw_items.append({
            "description": label,
            "qty": qty,
            "price": price,
            "total": round(qty * price, 2),
        })

    total = round(sum(it["total"] for it in raw_items), 2)
    if not raw_items:
        # Fallback: a single line with the contract total.
        total = _num(data.get("totalPrice"))
        if total <= 0:
            return None
        raw_items.append({
            "description": normalize_text(data.get("destination")) or "Trip",
            "qty": 1,
            "price": total,
            "total": total,
        })

    deposit = _num(data.get("depositAmount"))
    balance = _num(data.get("balanceAmount"))
    if balance <= 0 and deposit > 0:
        balance = max(total - deposit, 0)
    contract_date = normalize_text(data.get("contractDate")) or now_mongolia().date().isoformat()
    deposit_due = normalize_text(data.get("depositDueDate")) or contract_date
    balance_due = normalize_text(data.get("balanceDueDate")) or ""

    installments = []
    if deposit > 0:
        installments.append({
            "description": "Урьдчилгаа",
            "amount": deposit,
            "issueDate": contract_date,
            "dueDate": deposit_due,
            "status": "pending",
        })
    if balance > 0:
        installments.append({
            "description": "Аяллын үлдэгдэл",
            "amount": balance,
            "issueDate": contract_date,
            "dueDate": balance_due,
            "status": "pending",
        })
    if not installments:
        installments.append({
            "description": "Бүрэн төлбөр",
            "amount": total,
            "issueDate": contract_date,
            "dueDate": balance_due or contract_date,
            "status": "pending",
        })

    trip_for_serial = find_camp_trip(trip_id) if trip_id else None
    # Carry the chosen bank account from the contract through to the auto-
    # generated invoice. The contract document itself stays unchanged — the
    # bank info is only meant to surface on the invoice that goes out to
    # the client.
    bank_id = normalize_text(contract.get("bankAccountId")) or normalize_text((contract.get("data") or {}).get("bankAccountId"))
    bank_snapshot = None
    if bank_id:
        for b in read_settings().get("bankAccounts") or []:
            if b.get("id") == bank_id:
                bank_snapshot = b
                break
    return {
        "id": str(uuid4()),
        "serial": next_invoice_serial_from_group(
            find_tourist_group_serial(group_id),
            (trip_for_serial or {}).get("company"),
        ),
        "tripId": trip_id,
        "groupId": group_id,
        "payerId": "",
        "payerName": payer,
        "participantIds": [],
        "items": raw_items,
        "total": total,
        "installments": installments,
        "currency": "MNT",
        "bankAccountId": bank_id,
        "bankAccount": bank_snapshot,
        "status": "draft",
        "contractId": contract.get("id"),
        "createdAt": now_mongolia().isoformat(),
        "createdBy": actor_snapshot(actor),
        "updatedAt": "",
        "updatedBy": actor_snapshot(actor),
    }


def handle_list_contract_templates(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    return json_response(start_response, "200 OK", {"entries": read_contract_templates()})


def handle_get_default_contract_template(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    return json_response(start_response, "200 OK", get_default_template_payload())


def _normalize_contract_template_big_title(payload):
    raw = (payload or {}).get("bigTitle")
    if raw is None:
        return None
    return (str(raw) or "").strip()


def _normalize_contract_template_payload(payload):
    """Coerce the editor payload into clean storage shape:
    {intro: [html…], sections: [{title, paragraphs:[html…]}]}.
    Each paragraph is sanitised to the allowlist (ul/ol/li/strong/
    em/i/u/b/br/p, no attributes) so stored bodies are safe to
    re-emit at render time without escaping. Drops empty entries.
    """
    intro = []
    for p in ((payload or {}).get("intro") or []):
        cleaned = _sanitize_template_html((p or "").strip()) if isinstance(p, str) else ""
        if cleaned:
            intro.append(cleaned)
    sections = []
    raw_sections = (payload or {}).get("sections") or []
    if not isinstance(raw_sections, list):
        raw_sections = []
    for raw in raw_sections:
        if not isinstance(raw, dict):
            continue
        title = (raw.get("title") or "").strip()
        paragraphs = []
        for p in (raw.get("paragraphs") or []):
            cleaned = _sanitize_template_html((p or "").strip()) if isinstance(p, str) else ""
            if cleaned:
                paragraphs.append(cleaned)
        if title or paragraphs:
            sections.append({"title": title, "paragraphs": paragraphs})
    return {"intro": intro, "sections": sections}


def handle_create_contract_template(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ) or {}
    name = (normalize_text(payload.get("name")) or "").strip()
    if not name:
        return json_response(start_response, "400 Bad Request", {"error": "Template name is required."})
    body = _normalize_contract_template_payload(payload)
    big_title = _normalize_contract_template_big_title(payload) or ""
    record = {
        "id": uuid4().hex,
        "name": name,
        "bigTitle": big_title,
        "intro": body["intro"],
        "sections": body["sections"],
        "createdAt": now_mongolia().isoformat(),
        "createdBy": actor_snapshot(actor),
        "updatedAt": now_mongolia().isoformat(),
        "updatedBy": actor_snapshot(actor),
    }
    records = read_contract_templates()
    records.insert(0, record)
    write_contract_templates(records)
    return json_response(start_response, "201 Created", {"ok": True, "entry": record})


def handle_get_contract_template(environ, start_response, template_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    target = next((r for r in read_contract_templates() if r.get("id") == template_id), None)
    if not target:
        return json_response(start_response, "404 Not Found", {"error": "Template not found."})
    return json_response(start_response, "200 OK", {"entry": target})


def handle_update_contract_template(environ, start_response, template_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ) or {}
    records = read_contract_templates()
    target = next((r for r in records if r.get("id") == template_id), None)
    if not target:
        return json_response(start_response, "404 Not Found", {"error": "Template not found."})
    if "name" in payload:
        name = (normalize_text(payload.get("name")) or "").strip()
        if not name:
            return json_response(start_response, "400 Bad Request", {"error": "Template name is required."})
        target["name"] = name
    if "sections" in payload or "intro" in payload:
        body = _normalize_contract_template_payload(payload)
        target["intro"] = body["intro"]
        target["sections"] = body["sections"]
    if "bigTitle" in payload:
        target["bigTitle"] = _normalize_contract_template_big_title(payload) or ""
    target["updatedAt"] = now_mongolia().isoformat()
    target["updatedBy"] = actor_snapshot(actor)
    for i, r in enumerate(records):
        if r.get("id") == template_id:
            records[i] = target
            break
    write_contract_templates(records)
    return json_response(start_response, "200 OK", {"ok": True, "entry": target})


def handle_delete_contract_template(environ, start_response, template_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    records = read_contract_templates()
    remaining = [r for r in records if r.get("id") != template_id]
    if len(remaining) == len(records):
        return json_response(start_response, "404 Not Found", {"error": "Template not found."})
    write_contract_templates(remaining)
    return json_response(start_response, "200 OK", {"ok": True, "deletedId": template_id})


def handle_generate_contract(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})

    data = build_contract_data(payload)
    error = validate_contract_data(data)
    if error:
        return json_response(start_response, "400 Bad Request", {"error": error})

    try:
        record = save_contract_files(data)
        contracts = read_contracts()
        record["createdBy"] = actor_snapshot(actor)
        record["updatedBy"] = actor_snapshot(actor)
        record["company"] = active_workspace(environ) or "DTX"
        attached_trip = normalize_text(payload.get("attachedTripId"))
        attached_group = normalize_text(payload.get("attachedGroupId"))
        if attached_trip:
            record["tripId"] = attached_trip
        if attached_group:
            record["groupId"] = attached_group
        # Bank account selected on the contract form. We only persist the id
        # on the contract record (NOT into data — keeps it out of the DOCX
        # rendering). Also seed invoiceMeta.bankAccountKey so the rendered
        # invoice picks the same bank by default. The id IS the bank key
        # ("state" / "golomt" for the seeded defaults, uuid otherwise).
        bank_account_id = normalize_text(payload.get("bankAccountId"))
        if bank_account_id:
            record["bankAccountId"] = bank_account_id
            inv_meta = record.get("invoiceMeta") if isinstance(record.get("invoiceMeta"), dict) else {}
            inv_meta["bankAccountKey"] = bank_account_id.lower()
            record["invoiceMeta"] = inv_meta
        contracts.insert(0, record)
        write_contracts(contracts)
        try:
            log_notification(
                "contract.created",
                actor,
                "New contract generated",
                detail=record.get("clientName") or record.get("contractNumber") or "",
                meta={"id": record.get("id")},
            )
        except Exception:
            pass
        # Auto-create a corresponding invoice when the contract is attached to a trip+group.
        try:
            auto_inv = build_invoice_from_contract(record, actor)
            if auto_inv:
                inv_list = read_invoices()
                inv_list.insert(0, auto_inv)
                write_invoices(inv_list)
                record["autoInvoiceId"] = auto_inv["id"]
                # Persist the link back on the contract.
                contracts[0] = record
                write_contracts(contracts)
        except Exception as exc:
            print(f"[contract->invoice] auto-create failed: {exc}")
        return json_response(start_response, "201 Created", {"ok": True, "contract": record})
    except FileNotFoundError as exc:
        return json_response(start_response, "500 Internal Server Error", {"error": str(exc)})
    except Exception as exc:
        return json_response(start_response, "500 Internal Server Error", {"error": f"Could not generate contract: {exc}"})


def handle_list_contracts(environ, start_response):
    workspace = active_workspace(environ)
    contracts = read_contracts()
    if workspace:
        # Legacy contracts written before workspace stamping default to "DTX"
        # (every existing contract was created under DTX per the team's confirmation).
        contracts = [
            c for c in contracts
            if (str(c.get("company") or "DTX")).strip().upper() == workspace
        ]
    return json_response(start_response, "200 OK", contracts)


def find_contract(contract_id):
    for contract in read_contracts():
        if contract.get("id") == contract_id:
            return contract
    return None


def handle_get_contract(start_response, contract_id):
    contract = find_contract(contract_id)
    if not contract:
        return json_response(start_response, "404 Not Found", {"error": "Contract not found"})
    return json_response(start_response, "200 OK", {"contract": contract})


def handle_update_contract(environ, start_response, contract_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    contracts = read_contracts()
    for idx, contract in enumerate(contracts):
        if contract.get("id") == contract_id:
            if contract.get("status") == "signed":
                return json_response(start_response, "400 Bad Request", {"error": "Signed contracts cannot be edited"})
            data = build_contract_data(payload)
            error = validate_contract_data(data)
            if error:
                return json_response(start_response, "400 Bad Request", {"error": error})
            contract["data"] = data
            contract["updatedBy"] = actor_snapshot(actor)
            contract["updatedAt"] = datetime.now(timezone.utc).isoformat()
            docx_path = contract.get("docxPath")
            if docx_path:
                safe_docx = (GENERATED_DIR / unquote(docx_path.replace("/generated/", "", 1))).resolve()
                if str(safe_docx).startswith(str(GENERATED_DIR.resolve())):
                    try:
                        generate_docx(data, safe_docx)
                    except Exception:
                        pass
            view_path = contract.get("pdfViewPath")
            if view_path:
                safe_view = (GENERATED_DIR / unquote(view_path.replace("/generated/", "", 1))).resolve()
                if str(safe_view).startswith(str(GENERATED_DIR.resolve())):
                    try:
                        safe_view.write_text(
                            build_contract_html(
                                data,
                                signature_path=contract.get("signaturePath"),
                                contract_id=contract.get("id"),
                            ),
                            encoding="utf-8",
                        )
                    except Exception:
                        pass
            contract["pdfPath"] = None
            contracts[idx] = contract
            write_contracts(contracts)
            try:
                cdata = contract.get("data") or {}
                log_notification(
                    "contract.updated",
                    actor,
                    "Contract updated",
                    detail=f"{contract.get('contractSerial', '')} {cdata.get('clientName', '')}".strip(),
                    meta={"id": contract.get("id"), "tripId": contract.get("tripId"), "groupId": contract.get("groupId")},
                )
            except Exception:
                pass
            return json_response(start_response, "200 OK", {"ok": True, "contract": contract})
    return json_response(start_response, "404 Not Found", {"error": "Contract not found"})


def handle_delete_contract(environ, start_response, contract_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    contracts = read_contracts()
    deleted = next((c for c in contracts if c.get("id") == contract_id), None)
    remaining = [item for item in contracts if item.get("id") != contract_id]
    if len(remaining) == len(contracts):
        return json_response(start_response, "404 Not Found", {"error": "Contract not found"})
    write_contracts(remaining)
    if deleted:
        try:
            cdata = deleted.get("data") or {}
            log_notification(
                "contract.deleted",
                actor,
                "Contract deleted",
                detail=f"{deleted.get('contractSerial', '')} {cdata.get('clientName', '')}".strip(),
                meta={"id": deleted.get("id"), "tripId": deleted.get("tripId"), "groupId": deleted.get("groupId")},
            )
        except Exception:
            pass
    return json_response(start_response, "200 OK", {"ok": True, "deletedId": contract_id})


def handle_sign_contract(environ, start_response, contract_id):
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})
    signature_data = payload.get("signatureData")
    signer_name = normalize_text(payload.get("signerName"))
    client_phone = normalize_text(payload.get("clientPhone"))
    client_email = normalize_text(payload.get("clientEmail"))
    emergency_name = normalize_text(payload.get("emergencyContactName"))
    emergency_phone = normalize_text(payload.get("emergencyContactPhone"))
    emergency_relation = normalize_text(payload.get("emergencyContactRelation"))
    accepted = bool(payload.get("accepted"))
    if not accepted:
        return json_response(start_response, "400 Bad Request", {"error": "Agreement not accepted"})
    if not client_phone or not client_email or not emergency_name or not emergency_phone or not emergency_relation:
        return json_response(start_response, "400 Bad Request", {"error": "Missing client contact information"})
    if "@" not in client_email or "." not in client_email.split("@")[-1]:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid email address"})

    contracts = read_contracts()
    for idx, contract in enumerate(contracts):
        if contract.get("id") == contract_id:
            signature_path = save_signature_image(signature_data, contract_id)
            if not signature_path:
                return json_response(start_response, "400 Bad Request", {"error": "Invalid signature"})
            data = contract.get("data") or {}
            data["clientPhone"] = client_phone
            data["clientEmail"] = client_email
            data["emergencyContactName"] = emergency_name
            data["emergencyContactPhone"] = emergency_phone
            data["emergencyContactRelation"] = emergency_relation
            contract["data"] = data
            contract["signaturePath"] = signature_path
            contract["signerName"] = signer_name or contract.get("signerName")
            contract["accepted"] = accepted
            contract["status"] = "signed"
            contract["signedAt"] = datetime.now(timezone.utc).isoformat()
            try:
                contract["pdfPath"] = save_contract_pdf(contract)
            except Exception as exc:
                return json_response(
                    start_response,
                    "500 Internal Server Error",
                    {"error": f"Could not generate signed PDF: {exc}"},
                )
            view_path = contract.get("pdfViewPath")
            if view_path:
                safe_view = (GENERATED_DIR / unquote(view_path.replace("/generated/", "", 1))).resolve()
                if str(safe_view).startswith(str(GENERATED_DIR.resolve())):
                    try:
                        safe_view.write_text(
                            build_contract_html(
                                data,
                                signature_path=contract.get("signaturePath"),
                                contract_id=contract.get("id"),
                            ),
                            encoding="utf-8",
                        )
                    except Exception:
                        pass
            contracts[idx] = contract
            write_contracts(contracts)
            email_sent = _send_signed_contract_email(contract)
            return json_response(start_response, "200 OK", {"ok": True, "contract": contract, "emailSent": email_sent})
    return json_response(start_response, "404 Not Found", {"error": "Contract not found"})


def _send_contract_invitation_email(contract, to_list=None, recipient_name="", extra_message=""):
    """Friendly "please sign" email. Caller can override the
    recipient list / name / extra message; defaults to the
    contract's clientEmail when no list is provided.
    """
    data = contract.get("data") or {}
    if not to_list:
        client_email = (data.get("clientEmail") or "").strip()
        if not client_email:
            return {"ok": False, "error": "Үйлчлүүлэгчийн имэйл хаяг байхгүй байна."}
        to_list = [client_email]
    serial = data.get("contractSerial") or contract.get("id")
    signing_link = f"https://www.backoffice.travelx.mn/contract/{contract.get('id')}"
    manager_full = " ".join(filter(None, [data.get("managerLastName"), data.get("managerFirstName")])).strip() or "Дэлхий Трэвел Икс"
    manager_email = (data.get("managerEmail") or "").strip()
    manager_phone = (data.get("managerPhone") or "").strip()
    greeting = f"Сайн байна уу{(' ' + recipient_name) if recipient_name else ''},"
    body_lines = [
        greeting,
        "",
        f"Дэлхий Трэвел Икс ХХК танд аяллын гэрээ {('№ ' + serial) if serial else ''} илгээлээ.".strip(),
        "Та доорх холбоосоор орж гэрээгээ бөглөж, гарын үсгээр баталгаажуулна уу:",
        "",
        signing_link,
    ]
    if extra_message:
        body_lines += ["", extra_message]
    body_lines += ["", "Хүндэтгэсэн,", manager_full]
    if manager_phone:
        body_lines.append(f"Утас: {manager_phone}")
    if manager_email:
        body_lines.append(f"И-мэйл: {manager_email}")
    args = {
        "to": to_list,
        "subject": f"Travelx — Аяллын гэрээ{(' ' + serial) if serial else ''}",
        "body": "\n".join(body_lines),
        "attachments": [{"kind": "contract", "id": contract.get("id")}] if contract.get("pdfPath") else [],
    }
    return _tool_send_email(args, None)


def handle_send_contract_invitation(environ, start_response, contract_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    records = read_contracts()
    target = next((r for r in records if r.get("id") == contract_id), None)
    if not target:
        return json_response(start_response, "404 Not Found", {"error": "Contract not found"})
    payload = collect_json(environ) or {}
    to_raw = payload.get("to")
    if isinstance(to_raw, str):
        to_list = [t.strip() for t in re.split(r"[,;\s]+", to_raw) if t.strip()]
    elif isinstance(to_raw, list):
        to_list = [str(t).strip() for t in to_raw if str(t).strip()]
    else:
        to_list = None
    recipient_name = (normalize_text(payload.get("recipientName")) or "").strip()
    extra_message = (normalize_text(payload.get("message")) or "").strip()
    result = _send_contract_invitation_email(target, to_list=to_list, recipient_name=recipient_name, extra_message=extra_message)
    if result.get("ok"):
        return json_response(start_response, "200 OK", {"ok": True, "to": to_list or [(target.get("data") or {}).get("clientEmail") or ""]})
    return json_response(start_response, "400 Bad Request", {"error": result.get("error") or "Could not send."})


def _send_signed_contract_email(contract):
    """Auto-send after a client signs. Two separate emails:
      1. Friendly confirmation to the client (with auto-disclaimer footer).
      2. Internal notification to info@travelx.mn with full client/contract
         details, no "do not reply" footer (it's self-referential).
    Both carry the contract PDF + invoice PDF as attachments. Failures are
    logged and swallowed — a Resend outage must not break the signing flow."""
    try:
        data = contract.get("data") or {}
        client_email = (data.get("clientEmail") or "").strip()
        if not client_email:
            return False
        serial = data.get("contractSerial") or contract.get("id")
        invoice_id = (contract.get("autoInvoiceId") or "").strip()
        attachments = [{"kind": "contract", "id": contract.get("id")}]
        invoice_serial = None
        if invoice_id:
            attachments.append({"kind": "invoice", "id": invoice_id})
            inv = next((i for i in read_invoices() if i.get("id") == invoice_id), None)
            invoice_serial = inv.get("serial") if inv else None
        client_name = " ".join(filter(None, [data.get("touristLastName"), data.get("touristFirstName")])) or "—"
        client_phone = data.get("clientPhone") or "—"
        destination = data.get("destination") or "—"
        signed_iso = contract.get("signedAt") or ""
        try:
            signed_dt = datetime.fromisoformat(signed_iso.replace("Z", "+00:00")) if signed_iso else None
            signed_str = signed_dt.astimezone(MONGOLIA_TZ).strftime("%Y-%m-%d %H:%M") if signed_dt else "—"
        except Exception:
            signed_str = signed_iso[:16] if signed_iso else "—"

        # Email 1 — to the client
        client_lines = [
            "Сайн байна уу,",
            "",
            "Гэрээ амжилттай байгууллаа. Хавсаргав:",
            f"- Гэрээ {serial}",
        ]
        if invoice_serial:
            client_lines.append(f"- Нэхэмжлэх {invoice_serial}")
        elif invoice_id:
            client_lines.append("- Нэхэмжлэх")
        client_lines += ["", "Баярлалаа"]
        client_args = {
            "to": [client_email],
            "subject": f"Travelx — Гэрээ {serial}",
            "body": "\n".join(client_lines),
            "attachments": attachments,
        }
        client_result = _tool_send_email(client_args, None)

        # Email 2 — to info@travelx.mn (internal notification, no auto-footer)
        internal_lines = [
            "Үйлчлүүлэгч гэрээнд гарын үсэг зурлаа.",
            "",
            f"Гэрээ: {serial}",
            f"Үйлчлүүлэгч: {client_name}",
            f"Утас: {client_phone}",
            f"Имэйл: {client_email}",
            f"Аяллын чиглэл: {destination}",
            f"Гарын үсэг зурсан: {signed_str}",
            "",
            f"Гэрээ болон нэхэмжлэхийг {client_email} хаяг руу автоматаар илгээлээ.",
        ]
        internal_args = {
            "to": ["info@travelx.mn"],
            "subject": f"Шинэ гарын үсэг — Гэрээ {serial} — {client_name}",
            "body": "\n".join(internal_lines),
            "attachments": attachments,
            "_skip_footer": True,
        }
        internal_result = _tool_send_email(internal_args, None)

        ok = bool(client_result.get("ok"))
        print(
            f"[auto-email] sign-confirmation contract={contract.get('id')} "
            f"client_ok={client_result.get('ok')} internal_ok={internal_result.get('ok')} "
            f"client_result={client_result} internal_result={internal_result}",
            flush=True,
        )
        return ok
    except Exception as exc:
        print(f"[auto-email] sign-confirmation crashed: {type(exc).__name__}: {exc}", flush=True)
        return False


def handle_contract_document(environ, start_response, contract_id):
    params = parse_qs(environ.get("QUERY_STRING", ""))
    mode = (params.get("mode", ["view"])[0] or "view").strip().lower()
    paper_mode = (params.get("paper", ["0"])[0] or "0").strip().lower() in ("1", "true", "yes")
    contracts = read_contracts()
    contract = None
    contract_index = None
    for idx, entry in enumerate(contracts):
        if entry.get("id") == contract_id:
            contract = entry
            contract_index = idx
            break
    if not contract:
        return json_response(start_response, "404 Not Found", {"error": "Contract not found"})
    if mode == "download":
        # Paper mode: generate a fresh PDF on the fly with no
        # signatures / stamp so the office can print it for the
        # client to sign by hand. Doesn't touch the stored PDF.
        if paper_mode:
            try:
                from weasyprint import HTML
                from io import BytesIO
                paper_html = build_contract_html(
                    contract.get("data") or {},
                    signature_path=None,
                    asset_mode="file",
                    contract_id=contract.get("id"),
                    paper_mode=True,
                )
                buf = BytesIO()
                HTML(string=paper_html, base_url=str(BASE_DIR)).write_pdf(buf)
                pdf_bytes = buf.getvalue()
            except Exception as exc:
                return json_response(
                    start_response,
                    "500 Internal Server Error",
                    {"error": f"Could not generate paper PDF: {exc}"},
                )
            serial = (contract.get("data") or {}).get("contractSerial") or contract.get("id") or "contract"
            filename = f"{serial}-paper.pdf"
            headers = [
                ("Content-Type", "application/pdf"),
                ("Content-Length", str(len(pdf_bytes))),
                ("Content-Disposition", f'attachment; filename="{filename}"'),
            ]
            start_response("200 OK", headers)
            return [pdf_bytes]
        pdf_path = contract.get("pdfPath")
        if contract.get("status") == "signed":
            try:
                pdf_path = save_contract_pdf(contract)
            except Exception as exc:
                return json_response(
                    start_response,
                    "500 Internal Server Error",
                    {"error": f"Could not generate contract PDF: {exc}"},
                )
            contract["pdfPath"] = pdf_path
            if contract_index is not None:
                contracts[contract_index] = contract
                write_contracts(contracts)
        if not str(pdf_path or "").endswith(".pdf"):
            return json_response(start_response, "409 Conflict", {"error": "PDF not ready"})
        safe_path = (GENERATED_DIR / unquote(pdf_path.replace("/generated/", "", 1))).resolve()
        if not str(safe_path).startswith(str(GENERATED_DIR.resolve())) or not safe_path.exists():
            return json_response(start_response, "404 Not Found", {"error": "Document not found"})
        return file_response(start_response, safe_path, extra_headers=generated_download_headers(safe_path))

    view_path = contract.get("pdfViewPath")
    if not view_path:
        view_path = f"/generated/contract-{contract_id}.html"
        contract["pdfViewPath"] = view_path
    safe_path = (GENERATED_DIR / unquote(view_path.replace("/generated/", "", 1))).resolve()
    if not str(safe_path).startswith(str(GENERATED_DIR.resolve())):
        return json_response(start_response, "404 Not Found", {"error": "Document not found"})
    safe_path.write_text(
        build_contract_html(
            contract.get("data") or {},
            signature_path=contract.get("signaturePath"),
            contract_id=contract.get("id"),
        ),
        encoding="utf-8",
    )
    if contract_index is not None:
        contracts[contract_index] = contract
        write_contracts(contracts)
    return file_response(start_response, safe_path)


def handle_contract_invoice_document(environ, start_response, contract_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    params = parse_qs(environ.get("QUERY_STRING", ""))
    mode = (params.get("mode", ["view"])[0] or "view").strip().lower()
    contracts = read_contracts()
    contract = None
    contract_index = None
    for idx, entry in enumerate(contracts):
        if entry.get("id") == contract_id:
            contract = entry
            contract_index = idx
            break
    if not contract:
        return json_response(start_response, "404 Not Found", {"error": "Contract not found"})
    if mode == "download":
        try:
            invoice_path = save_invoice_pdf(contract)
        except Exception as exc:
            return json_response(
                start_response,
                "500 Internal Server Error",
                {"error": f"Could not generate invoice PDF: {exc}"},
            )
        contract["invoicePath"] = invoice_path
        if contract_index is not None:
            contracts[contract_index] = contract
            write_contracts(contracts)
        safe_path = (GENERATED_DIR / unquote(invoice_path.replace("/generated/", "", 1))).resolve()
        if not str(safe_path).startswith(str(GENERATED_DIR.resolve())) or not safe_path.exists():
            return json_response(start_response, "404 Not Found", {"error": "Invoice not found"})
        return file_response(start_response, safe_path, extra_headers=generated_download_headers(safe_path))

    view_path = contract.get("invoiceViewPath")
    if not view_path:
        view_path = f"/generated/invoice-{contract_id}.html"
        contract["invoiceViewPath"] = view_path
    safe_path = (GENERATED_DIR / unquote(view_path.replace("/generated/", "", 1))).resolve()
    if not str(safe_path).startswith(str(GENERATED_DIR.resolve())):
        return json_response(start_response, "404 Not Found", {"error": "Invoice not found"})
    safe_path.write_text(build_invoice_html(contract), encoding="utf-8")
    if contract_index is not None:
        contracts[contract_index] = contract
        write_contracts(contracts)
    return file_response(start_response, safe_path)


def handle_update_contract_invoice(environ, start_response, contract_id):
    actor = require_login(environ, start_response)
    if not actor:
        return []
    payload = collect_json(environ)
    if payload is None:
        return json_response(start_response, "400 Bad Request", {"error": "Invalid payload"})

    contracts = read_contracts()
    for idx, contract in enumerate(contracts):
        if contract.get("id") != contract_id:
            continue

        items_payload = payload.get("items")
        payments_payload = payload.get("payments")
        if not isinstance(items_payload, list):
            return json_response(start_response, "400 Bad Request", {"error": "Items payload is invalid"})
        if not isinstance(payments_payload, list):
            return json_response(start_response, "400 Bad Request", {"error": "Payments payload is invalid"})

        normalized_items = []
        for item in items_payload:
            if not isinstance(item, dict):
                continue
            description = normalize_text(item.get("description"))
            quantity = parse_int(item.get("quantity"))
            unit_price = parse_int(item.get("unitPrice"))
            total_price = parse_int(item.get("totalPrice")) or quantity * unit_price
            if not description or quantity <= 0:
                continue
            normalized_items.append(
                {
                    "key": normalize_text(item.get("key")) or f"item-{len(normalized_items) + 1}",
                    "description": description,
                    "quantity": quantity,
                    "unitPrice": unit_price,
                    "totalPrice": total_price,
                }
            )
        if not normalized_items:
            return json_response(start_response, "400 Bad Request", {"error": "At least one invoice item is required"})

        normalized_payments = []
        for payment in payments_payload:
            if not isinstance(payment, dict):
                continue
            title = normalize_text(payment.get("title"))
            amount = parse_int(payment.get("amount"))
            if not title or amount <= 0:
                continue
            status = normalize_invoice_status(payment.get("status"))
            normalized_payments.append(
                {
                    "key": normalize_text(payment.get("key")) or f"payment-{len(normalized_payments) + 1}",
                    "title": title,
                    "created": normalize_text(payment.get("created")),
                    "secondaryLabel": normalize_text(payment.get("secondaryLabel")) or "Эцсийн хугацаа",
                    "secondaryValue": normalize_text(payment.get("secondaryValue")),
                    "status": status,
                    "amount": amount,
                }
            )
        if not normalized_payments:
            return json_response(start_response, "400 Bad Request", {"error": "At least one payment row is required"})

        invoice_meta = contract.get("invoiceMeta") if isinstance(contract.get("invoiceMeta"), dict) else {}
        invoice_meta["lineItems"] = normalized_items
        invoice_meta["payments"] = {"rows": normalized_payments}
        invoice_meta["bankAccountKey"] = normalize_invoice_bank_account(
            payload.get("bankAccountKey"),
            fallback=normalize_invoice_bank_account(invoice_meta.get("bankAccountKey")),
        )
        contract["invoiceMeta"] = invoice_meta
        contract["updatedBy"] = actor_snapshot(actor)
        contract["updatedAt"] = datetime.now(timezone.utc).isoformat()
        contracts[idx] = contract
        write_contracts(contracts)
        return json_response(start_response, "200 OK", {"ok": True, "contract": contract})

    return json_response(start_response, "404 Not Found", {"error": "Contract not found"})


# ════════════════════════════════════════════════════════════════════
#  AGENT  (admin-only Claude assistant with tool use over the API)
# ════════════════════════════════════════════════════════════════════

AGENT_CONVERSATIONS_FILE = DATA_DIR / "agent_conversations.json"
AGENT_AUDIT_FILE = DATA_DIR / "agent_audit.json"
AGENT_MEMORY_FILE = DATA_DIR / "agent_memory.json"
AGENT_MODEL = os.environ.get("AGENT_MODEL", "claude-sonnet-4-5-20250929")
AGENT_MAX_TOKENS = 2048  # smaller responses → faster round trips, fewer timeouts
AGENT_MAX_TOOL_LOOPS = 6  # fits comfortably inside Render proxy budget; the agent should split big asks instead
AGENT_HISTORY_TURNS = 8  # smaller history → faster requests, fewer 502s
AGENT_MAX_IMAGES_PER_MESSAGE = 5
AGENT_MAX_IMAGE_BYTES = 5 * 1024 * 1024  # Anthropic per-image limit


def _agent_load_memories():
    if not AGENT_MEMORY_FILE.exists():
        return []
    try:
        data = json.loads(AGENT_MEMORY_FILE.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else []
    except Exception:
        return []


def _agent_save_memories(records):
    AGENT_MEMORY_FILE.parent.mkdir(parents=True, exist_ok=True)
    AGENT_MEMORY_FILE.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")


def _agent_load_conversations():
    if not AGENT_CONVERSATIONS_FILE.exists():
        return {}
    try:
        return json.loads(AGENT_CONVERSATIONS_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _agent_save_conversations(data):
    AGENT_CONVERSATIONS_FILE.parent.mkdir(parents=True, exist_ok=True)
    AGENT_CONVERSATIONS_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _agent_audit(user, tool_name, tool_input, ok, output_summary):
    AGENT_AUDIT_FILE.parent.mkdir(parents=True, exist_ok=True)
    try:
        existing = json.loads(AGENT_AUDIT_FILE.read_text(encoding="utf-8")) if AGENT_AUDIT_FILE.exists() else []
    except Exception:
        existing = []
    existing.append({
        "ts": datetime.now(timezone.utc).isoformat(),
        "user": (user or {}).get("email") or (user or {}).get("id") or "?",
        "tool": tool_name,
        "input": tool_input,
        "ok": ok,
        "summary": str(output_summary)[:400],
    })
    # Keep last 2000 entries.
    if len(existing) > 2000:
        existing = existing[-2000:]
    AGENT_AUDIT_FILE.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")


# ── Tool implementations (thin wrappers over read_/write_/build_) ─────

def _tool_list_trips(args, actor):
    trips = read_camp_trips()
    ws = (args.get("workspace") or args.get("company") or "").upper().strip()
    status = (args.get("status") or "").lower().strip()
    trip_type = (args.get("tripType") or "").lower().strip()
    if ws:
        trips = [t for t in trips if normalize_company(t.get("company")) == ws]
    if status:
        trips = [t for t in trips if (t.get("status") or "").lower() == status]
    if trip_type:
        trips = [t for t in trips if (t.get("tripType") or "").lower() == trip_type]
    items = [{"id": t["id"], "serial": t.get("serial"), "tripName": t.get("tripName"),
              "tripType": t.get("tripType"), "startDate": t.get("startDate"),
              "endDate": t.get("endDate"), "status": t.get("status"),
              "participantCount": t.get("participantCount"), "company": t.get("company")}
             for t in trips]
    return {"count": len(items), "items": items}


def _tool_get_trip(args, actor):
    trip_id = args.get("tripId") or ""
    trip = next((t for t in read_camp_trips() if t.get("id") == trip_id), None)
    return trip or {"error": "Trip not found"}


def _tool_create_trip(args, actor):
    payload = dict(args)
    ws = (payload.get("workspace") or payload.get("company") or "").upper().strip()
    if ws not in ("DTX", "USM"):
        return {"error": "workspace is required and must be 'DTX' (outbound — any non-Mongolia destination) or 'USM' (inbound — inside Mongolia only)."}
    payload["company"] = ws
    payload.pop("workspace", None)
    record = build_camp_trip(payload, actor)
    err = validate_camp_trip(record)
    if err:
        return {"error": err}
    trips = read_camp_trips()
    trips.insert(0, record)
    write_camp_trips(trips)
    return {"ok": True, "trip": {"id": record["id"], "serial": record.get("serial"), "tripName": record.get("tripName")}}


def _tool_update_trip(args, actor):
    trip_id = args.get("tripId") or ""
    fields = args.get("fields") or {}
    trips = read_camp_trips()
    for i, t in enumerate(trips):
        if t.get("id") != trip_id:
            continue
        t.update({k: v for k, v in fields.items() if k != "id"})
        t["updatedAt"] = now_mongolia().isoformat()
        t["updatedBy"] = actor_snapshot(actor)
        trips[i] = t
        write_camp_trips(trips)
        return {"ok": True, "trip": {"id": t["id"], "serial": t.get("serial"), "tripName": t.get("tripName"), "status": t.get("status")}}
    return {"error": "Trip not found"}


def _tool_delete_trip(args, actor):
    trip_id = args.get("tripId") or ""
    trips = read_camp_trips()
    new = [t for t in trips if t.get("id") != trip_id]
    if len(new) == len(trips):
        return {"error": "Trip not found"}
    write_camp_trips(new)
    return {"ok": True}


def _tool_list_groups(args, actor):
    groups = read_tourist_groups()
    trip_id = (args.get("tripId") or "").strip()
    if trip_id:
        groups = [g for g in groups if g.get("tripId") == trip_id]
    return [{"id": g["id"], "serial": g.get("serial"), "name": g.get("name"),
             "tripId": g.get("tripId"), "headcount": g.get("headcount"),
             "status": g.get("status"), "leaderName": g.get("leaderName")}
            for g in groups]


def _tool_create_group(args, actor):
    record = build_tourist_group(dict(args), actor)
    err = validate_tourist_group(record)
    if err:
        return {"error": err}
    groups = read_tourist_groups()
    groups.append(record)
    write_tourist_groups(groups)
    return {"ok": True, "group": {"id": record["id"], "serial": record.get("serial"), "name": record.get("name")}}


def _tool_update_group(args, actor):
    gid = args.get("groupId") or ""
    fields = args.get("fields") or {}
    groups = read_tourist_groups()
    for i, g in enumerate(groups):
        if g.get("id") != gid:
            continue
        g.update({k: v for k, v in fields.items() if k != "id"})
        g["updatedAt"] = now_mongolia().isoformat()
        g["updatedBy"] = actor_snapshot(actor)
        groups[i] = g
        write_tourist_groups(groups)
        return {"ok": True, "group": g}
    return {"error": "Group not found"}


def _tool_delete_group(args, actor):
    gid = args.get("groupId") or ""
    groups = read_tourist_groups()
    new = [g for g in groups if g.get("id") != gid]
    if len(new) == len(groups):
        return {"error": "Group not found"}
    write_tourist_groups(new)
    return {"ok": True}


def _tool_list_tourists(args, actor):
    tourists = read_tourists()
    trip_id = (args.get("tripId") or "").strip()
    group_id = (args.get("groupId") or "").strip()
    if trip_id:
        tourists = [t for t in tourists if t.get("tripId") == trip_id]
    if group_id:
        tourists = [t for t in tourists if t.get("groupId") == group_id]
    return [{"id": t["id"], "serial": t.get("serial"),
             "lastName": t.get("lastName"), "firstName": t.get("firstName"),
             "passportNumber": t.get("passportNumber"), "phone": t.get("phone"),
             "groupId": t.get("groupId"), "tripId": t.get("tripId"),
             "roomType": t.get("roomType"), "roomCode": t.get("roomCode")}
            for t in tourists]


def _tool_create_tourist(args, actor):
    record = build_tourist(dict(args), actor)
    err = validate_tourist(record)
    if err:
        return {"error": err}
    tourists = read_tourists()
    tourists.append(record)
    write_tourists(tourists)
    return {"ok": True, "tourist": {"id": record["id"], "serial": record.get("serial"),
                                     "lastName": record.get("lastName"), "firstName": record.get("firstName")}}


def _tool_update_tourist(args, actor):
    tid = args.get("touristId") or ""
    fields = args.get("fields") or {}
    tourists = read_tourists()
    for i, t in enumerate(tourists):
        if t.get("id") != tid:
            continue
        t.update({k: v for k, v in fields.items() if k != "id"})
        t["updatedAt"] = now_mongolia().isoformat()
        t["updatedBy"] = actor_snapshot(actor)
        tourists[i] = t
        write_tourists(tourists)
        return {"ok": True, "tourist": t}
    return {"error": "Tourist not found"}


def _tool_delete_tourist(args, actor):
    tid = args.get("touristId") or ""
    tourists = read_tourists()
    new = [t for t in tourists if t.get("id") != tid]
    if len(new) == len(tourists):
        return {"error": "Tourist not found"}
    write_tourists(new)
    return {"ok": True}


def _tool_list_invoices(args, actor):
    invoices = read_invoices()
    trip_id = (args.get("tripId") or "").strip()
    if trip_id:
        invoices = [i for i in invoices if i.get("tripId") == trip_id]
    return [{"id": i["id"], "serial": i.get("serial"), "tripId": i.get("tripId"),
             "groupId": i.get("groupId"), "payerName": i.get("payerName"),
             "total": i.get("total"), "status": i.get("status")}
            for i in invoices]


def _tool_get_invoice(args, actor):
    iid = args.get("invoiceId") or ""
    inv = next((i for i in read_invoices() if i.get("id") == iid), None)
    return inv or {"error": "Invoice not found"}


def _tool_create_invoice(args, actor):
    record = build_invoice(dict(args), actor)
    err = validate_invoice(record)
    if err:
        return {"error": err}
    invs = read_invoices()
    invs.insert(0, record)
    write_invoices(invs)
    return {"ok": True, "invoice": {"id": record["id"], "serial": record.get("serial"), "total": record.get("total")}}


def _tool_update_invoice(args, actor):
    iid = args.get("invoiceId") or ""
    fields = args.get("fields") or {}
    invs = read_invoices()
    for i, inv in enumerate(invs):
        if inv.get("id") != iid:
            continue
        inv.update({k: v for k, v in fields.items() if k != "id"})
        inv["updatedAt"] = now_mongolia().isoformat()
        inv["updatedBy"] = actor_snapshot(actor)
        invs[i] = inv
        write_invoices(invs)
        return {"ok": True, "invoice": inv}
    return {"error": "Invoice not found"}


def _tool_delete_invoice(args, actor):
    iid = args.get("invoiceId") or ""
    invs = read_invoices()
    new = [i for i in invs if i.get("id") != iid]
    if len(new) == len(invs):
        return {"error": "Invoice not found"}
    write_invoices(new)
    return {"ok": True}


def _tool_publish_invoice(args, actor):
    iid = args.get("invoiceId") or ""
    invs = read_invoices()
    for i, inv in enumerate(invs):
        if inv.get("id") != iid:
            continue
        inv["status"] = "published"
        inv["publishedAt"] = now_mongolia().isoformat()
        inv["updatedBy"] = actor_snapshot(actor)
        invs[i] = inv
        write_invoices(invs)
        return {"ok": True, "invoice": {"id": inv["id"], "status": inv["status"]}}
    return {"error": "Invoice not found"}


def _tool_register_invoice_payment(args, actor):
    iid = args.get("invoiceId") or ""
    idx = int(args.get("installmentIndex") or 0)
    paid = args.get("paidDate") or now_mongolia().date().isoformat()
    invs = read_invoices()
    for i, inv in enumerate(invs):
        if inv.get("id") != iid:
            continue
        installments = inv.get("installments") or []
        if idx < 0 or idx >= len(installments):
            return {"error": "installmentIndex out of range"}
        installments[idx]["status"] = "paid"
        installments[idx]["paidDate"] = paid
        inv["installments"] = installments
        inv["updatedBy"] = actor_snapshot(actor)
        invs[i] = inv
        write_invoices(invs)
        return {"ok": True, "installment": installments[idx]}
    return {"error": "Invoice not found"}


def _tool_list_contracts(args, actor):
    cs = read_contracts()
    trip_id = (args.get("tripId") or "").strip()
    if trip_id:
        cs = [c for c in cs if c.get("tripId") == trip_id]
    return [{"id": c["id"], "tripId": c.get("tripId"), "groupId": c.get("groupId"),
             "status": c.get("status"), "contractSerial": (c.get("data") or {}).get("contractSerial"),
             "tourist": ((c.get("data") or {}).get("touristLastName") or "") + " " + ((c.get("data") or {}).get("touristFirstName") or "")}
            for c in cs]


def _tool_delete_contract(args, actor):
    cid = args.get("contractId") or ""
    cs = read_contracts()
    new = [c for c in cs if c.get("id") != cid]
    if len(new) == len(cs):
        return {"error": "Contract not found"}
    write_contracts(new)
    return {"ok": True}


def _tool_list_camp_reservations(args, actor):
    rs = read_camp_reservations()
    trip_id = (args.get("tripId") or "").strip()
    if trip_id:
        rs = [r for r in rs if r.get("tripId") == trip_id]
    return rs


def _tool_list_flight_reservations(args, actor):
    rs = read_flight_reservations()
    trip_id = (args.get("tripId") or "").strip()
    if trip_id:
        rs = [r for r in rs if r.get("tripId") == trip_id]
    return rs


def _tool_list_transfer_reservations(args, actor):
    rs = read_transfer_reservations()
    trip_id = (args.get("tripId") or "").strip()
    if trip_id:
        rs = [r for r in rs if r.get("tripId") == trip_id]
    return rs


def _tool_list_users(args, actor):
    return [{"id": u.get("id"), "email": u.get("email"), "fullName": u.get("fullName"),
             "role": u.get("role"), "status": u.get("status")} for u in read_users()]


def _tool_list_ds160(args, actor):
    return [{"id": a.get("id"), "email": a.get("email"), "status": a.get("status"),
             "createdAt": a.get("createdAt")} for a in read_ds160_applications()]


def _tool_list_notifications(args, actor):
    return read_notifications()[-50:]


def _tool_list_fifa_inventory(args, actor):
    store = read_fifa2026_store()
    tickets = store.get("tickets") if isinstance(store, dict) else None
    return tickets[:200] if isinstance(tickets, list) else (tickets or [])


# ── New read tools (Bataa coverage expansion) ──────────────────────

def _tool_get_tourist(args, actor):
    tid = (args.get("touristId") or "").strip()
    if not tid:
        return {"error": "touristId is required"}
    t = next((x for x in read_tourists() if x.get("id") == tid), None)
    if not t:
        return {"error": "Tourist not found"}
    return t


def _tool_get_group(args, actor):
    gid = (args.get("groupId") or "").strip()
    if not gid:
        return {"error": "groupId is required"}
    g = next((x for x in read_tourist_groups() if x.get("id") == gid), None)
    if not g:
        return {"error": "Group not found"}
    return g


def _tool_list_documents(args, actor):
    """Flatten every document attached to every trip's documents[]. Optional
    tripId / category / touristId filter so the agent can answer 'what files
    does T-0007 have?' or 'all passport scans'."""
    trip_id = (args.get("tripId") or "").strip()
    category = (args.get("category") or "").strip()
    tourist_id = (args.get("touristId") or "").strip()
    rows = []
    for trip in read_camp_trips():
        if trip_id and trip.get("id") != trip_id:
            continue
        for doc in (trip.get("documents") or []):
            if category and doc.get("category") != category:
                continue
            if tourist_id and doc.get("touristId") != tourist_id:
                continue
            rows.append({
                "id": doc.get("id"),
                "tripId": trip.get("id"),
                "tripSerial": trip.get("serial"),
                "tripName": trip.get("tripName"),
                "category": doc.get("category"),
                "originalName": doc.get("originalName"),
                "uploadedAt": doc.get("uploadedAt"),
                "uploadedBy": (doc.get("uploadedBy") or {}).get("name") or (doc.get("uploadedBy") or {}).get("email"),
                "touristId": doc.get("touristId"),
                "size": doc.get("size"),
            })
    return {"count": len(rows), "items": rows[:200]}


def _tool_list_tasks(args, actor):
    store = read_manager_dashboard()
    items = store.get("tasks") or []
    status = (args.get("status") or "").strip().lower()
    if status:
        items = [t for t in items if (t.get("status") or "").lower() == status]
    return {"count": len(items), "items": items}


def _tool_create_task(args, actor):
    payload = {
        "title": normalize_text(args.get("title")),
        "owner": normalize_text(args.get("owner")),
        "owners": args.get("owners"),
        "priority": (normalize_text(args.get("priority")) or "medium").lower(),
        "status": (normalize_text(args.get("status")) or "pending").lower(),
        "dueDate": normalize_text(args.get("dueDate")),
        "dueTime": normalize_text(args.get("dueTime")),
        "destinations": args.get("destinations"),
        "note": normalize_text(args.get("note")),
    }
    record = build_manager_task(payload)
    record["createdBy"] = actor_snapshot(actor)
    record["createdAt"] = record.get("createdAt") or now_mongolia().isoformat()
    err = validate_manager_task(record)
    if err:
        return {"error": err}
    store = read_manager_dashboard()
    store["tasks"].insert(0, record)
    write_manager_dashboard(store)
    return {"ok": True, "task": record}


def _tool_update_task(args, actor):
    tid = (args.get("taskId") or "").strip()
    fields = args.get("fields") or {}
    store = read_manager_dashboard()
    for i, t in enumerate(store.get("tasks") or []):
        if t.get("id") != tid:
            continue
        merged = {**t}
        for k, v in fields.items():
            if k != "id":
                merged[k] = v
        merged["updatedBy"] = actor_snapshot(actor)
        merged["updatedAt"] = now_mongolia().isoformat()
        store["tasks"][i] = merged
        write_manager_dashboard(store)
        return {"ok": True, "task": merged}
    return {"error": "Task not found"}


def _tool_delete_task(args, actor):
    tid = (args.get("taskId") or "").strip()
    store = read_manager_dashboard()
    new_items = [t for t in (store.get("tasks") or []) if t.get("id") != tid]
    if len(new_items) == len(store.get("tasks") or []):
        return {"error": "Task not found"}
    store["tasks"] = new_items
    write_manager_dashboard(store)
    return {"ok": True}


def _tool_list_contacts(args, actor):
    store = read_manager_dashboard()
    items = store.get("contacts") or []
    return {"count": len(items), "items": items}


def _tool_create_contact(args, actor):
    payload = {
        "name": normalize_text(args.get("name")),
        "phone": normalize_text(args.get("phone")),
        "type": (normalize_text(args.get("type")) or "client").lower(),
        "status": (normalize_text(args.get("status")) or "new").lower(),
        "lastContacted": normalize_text(args.get("lastContacted")),
        "destinations": args.get("destinations"),
        "note": normalize_text(args.get("note")),
    }
    record = build_manager_contact(payload)
    record["createdBy"] = actor_snapshot(actor)
    record["createdAt"] = record.get("createdAt") or now_mongolia().isoformat()
    err = validate_manager_contact(record)
    if err:
        return {"error": err}
    store = read_manager_dashboard()
    store["contacts"].insert(0, record)
    write_manager_dashboard(store)
    return {"ok": True, "contact": record}


def _tool_update_contact(args, actor):
    cid = (args.get("contactId") or "").strip()
    fields = args.get("fields") or {}
    store = read_manager_dashboard()
    for i, c in enumerate(store.get("contacts") or []):
        if c.get("id") != cid:
            continue
        merged = {**c}
        for k, v in fields.items():
            if k != "id":
                merged[k] = v
        merged["updatedBy"] = actor_snapshot(actor)
        merged["updatedAt"] = now_mongolia().isoformat()
        store["contacts"][i] = merged
        write_manager_dashboard(store)
        return {"ok": True, "contact": merged}
    return {"error": "Contact not found"}


def _tool_delete_contact(args, actor):
    cid = (args.get("contactId") or "").strip()
    store = read_manager_dashboard()
    new_items = [c for c in (store.get("contacts") or []) if c.get("id") != cid]
    if len(new_items) == len(store.get("contacts") or []):
        return {"error": "Contact not found"}
    store["contacts"] = new_items
    write_manager_dashboard(store)
    return {"ok": True}


def _tool_get_settings(args, actor):
    s = read_settings()
    camp_settings = read_camp_settings() if "read_camp_settings" in globals() else {}
    return {
        "destinations": s.get("destinations") or [],
        "bankAccounts": s.get("bankAccounts") or [],
        "campNames": camp_settings.get("campNames") or [],
        "campDetails": camp_settings.get("campDetails") or {},
    }


def _tool_update_destinations(args, actor):
    items = normalize_option_list(args.get("destinations"))
    s = read_settings()
    s["destinations"] = items
    write_settings(s)
    return {"ok": True, "destinations": items}


def _tool_update_bank_accounts(args, actor):
    accounts = _normalize_bank_accounts(args.get("bankAccounts"))
    s = read_settings()
    s["bankAccounts"] = accounts
    write_settings(s)
    return {"ok": True, "bankAccounts": accounts}


def _tool_list_team_members(args, actor):
    out = []
    for u in read_users():
        if u.get("status") != "approved":
            continue
        out.append({
            "id": u.get("id"),
            "email": u.get("email"),
            "fullName": u.get("fullName") or u.get("email"),
            "role": u.get("role") or "staff",
            "phone": u.get("contractPhone"),
        })
    return {"count": len(out), "items": out}


def _tool_list_mail_messages(args, actor):
    """Recent mail headers (no body) across every connected mailbox in the
    requested workspace. Defaults to inbox; pass folder='sent' for sent."""
    folder = (args.get("folder") or "inbox").strip().lower()
    if folder not in ("inbox", "sent"):
        folder = "inbox"
    workspace = (args.get("workspace") or "").strip().upper()
    accounts = read_mail_accounts()
    if workspace:
        accounts = [a for a in accounts if (a.get("workspace") or "DTX").upper() == workspace]
    rows = []
    for a in accounts:
        cache = _read_mail_cache(a["id"])
        for msg in cache.get("messages", []):
            if (msg.get("folder") or "inbox") != folder:
                continue
            slim = {k: v for k, v in msg.items() if k not in ("bodyText", "bodyHtml")}
            slim["accountAddress"] = a["address"]
            slim["workspace"] = a.get("workspace") or "DTX"
            rows.append(slim)
    rows.sort(key=lambda m: m.get("date") or "", reverse=True)
    limit = int(args.get("limit") or 50)
    return {"count": len(rows), "items": rows[:limit]}


def _tool_list_mail_followups(args, actor):
    return {"items": read_mail_followups()}


def _tool_list_mail_dossiers(args, actor):
    return {"items": list(_build_dossier_index().values())}


def _tool_send_ds160_invitation(args, actor):
    rid = (args.get("ds160Id") or "").strip()
    if not rid:
        return {"error": "ds160Id is required"}
    records = read_ds160_applications()
    record = next((r for r in records if r.get("id") == rid), None)
    if not record:
        return {"error": "DS-160 record not found"}
    to = (args.get("to") or record.get("clientEmail") or "").strip()
    if not to:
        return {"error": "Client email missing — set it on the entry first"}
    share_url = record.get("shareUrl") or ""
    given = (record.get("clientName") or "").split()[0] if record.get("clientName") else ""
    body = (
        f"Сайн байна уу{(', ' + given) if given else ''}.\n\n"
        "Та доорх холбоосоор DS-160 маягтаа бөглөнө үү.\n\n"
        f"{share_url}\n\nХүндэтгэсэн,\nДэлхий Трэвел Икс"
    )
    res = _tool_send_email({
        "to": to,
        "subject": "TravelX DS-160 form",
        "body": body,
        "_company_name": "Дэлхий Трэвел Икс",
    }, actor)
    if res.get("error"):
        return res
    record["lastEmailedAt"] = datetime.now(timezone.utc).isoformat()
    record["lastEmailedBy"] = actor_snapshot(actor)
    write_ds160_applications(records)
    return {"ok": True, "to": to}


def _tool_register_invoice_payment_with_note(args, actor):
    """Wrapper that forwards a free-text note to handle_invoice_payment so
    the agent can record context like 'paid only for the father'."""
    iid = (args.get("invoiceId") or "").strip()
    idx = args.get("installmentIndex")
    if not iid or idx is None:
        return {"error": "invoiceId + installmentIndex required"}
    records = read_invoices()
    for i, record in enumerate(records):
        if record.get("id") != iid:
            continue
        installments = record.get("installments") or []
        if idx < 0 or idx >= len(installments):
            return {"error": "installmentIndex out of range"}
        installments[idx]["status"] = (args.get("status") or "paid")
        if args.get("paidDate"):
            installments[idx]["paidDate"] = args.get("paidDate")
        if "note" in args:
            installments[idx]["note"] = normalize_text(args.get("note"))
        record["installments"] = installments
        record["updatedAt"] = now_mongolia().isoformat()
        record["updatedBy"] = actor_snapshot(actor)
        records[i] = record
        write_invoices(records)
        return {"ok": True, "invoice": record}
    return {"error": "Invoice not found"}


def _tool_get_contract(args, actor):
    cid = (args.get("contractId") or "").strip()
    if not cid:
        return {"error": "contractId is required"}
    contract = next((c for c in read_contracts() if c.get("id") == cid), None)
    if not contract:
        return {"error": "Contract not found"}
    data = contract.get("data") or {}
    return {
        "id": contract.get("id"),
        "contractSerial": data.get("contractSerial"),
        "tripId": contract.get("tripId"),
        "groupId": contract.get("groupId"),
        "destination": data.get("destination"),
        "tripStartDate": data.get("tripStartDate"),
        "tripEndDate": data.get("tripEndDate"),
        "touristLastName": data.get("touristLastName"),
        "touristFirstName": data.get("touristFirstName"),
        "adultCount": data.get("adultCount"),
        "adultPrice": data.get("adultPrice"),
        "childCount": data.get("childCount"),
        "childPrice": data.get("childPrice"),
        "totalPrice": data.get("totalPrice"),
        "depositAmount": data.get("depositAmount"),
        "depositDueDate": data.get("depositDueDate"),
        "balanceDueDate": data.get("balanceDueDate"),
        "autoInvoiceId": contract.get("autoInvoiceId"),
        "pdfPath": contract.get("pdfPath"),
    }


def _tool_create_contract(args, actor):
    payload = dict(args)

    # Auto-fill manager fields from the logged-in admin if not provided.
    if actor:
        payload.setdefault("managerLastName", actor.get("contractLastName") or "")
        payload.setdefault("managerFirstName",
                            actor.get("contractFirstName") or (actor.get("fullName") or "").split(" ")[0] or "Admin")
        payload.setdefault("managerPhone", actor.get("contractPhone") or "")
        payload.setdefault("managerEmail", actor.get("contractEmail") or actor.get("email") or "")

    # Auto-fill from trip.
    trip_id = payload.get("tripId") or payload.get("attachedTripId")
    if trip_id:
        trip = next((t for t in read_camp_trips() if t.get("id") == trip_id), None)
        if trip:
            payload.setdefault("destination", trip.get("tripName") or "")
            payload.setdefault("tripStartDate", trip.get("startDate") or "")
            payload.setdefault("tripEndDate", trip.get("endDate") or "")
            payload.setdefault("tripDuration", str(trip.get("totalDays") or ""))
            payload["attachedTripId"] = trip_id

    # Auto-fill from tourist.
    tid = payload.get("touristId")
    if tid:
        tourist = next((t for t in read_tourists() if t.get("id") == tid), None)
        if tourist:
            payload.setdefault("touristLastName", tourist.get("lastName") or "")
            payload.setdefault("touristFirstName", tourist.get("firstName") or "")
            payload.setdefault("touristRegister", tourist.get("registrationNumber") or tourist.get("passportNumber") or "")
            payload.setdefault("touristPhone", tourist.get("phone") or "")
            payload.setdefault("touristEmail", tourist.get("email") or "")

    gid = payload.get("groupId")
    if gid:
        payload["attachedGroupId"] = gid

    data = build_contract_data(payload)
    err = validate_contract_data(data)
    if err:
        return {"error": err}
    try:
        record = save_contract_files(data)
    except Exception as exc:
        return {"error": f"save_contract_files failed: {type(exc).__name__}: {exc}"}
    record["createdBy"] = actor_snapshot(actor)
    record["updatedBy"] = actor_snapshot(actor)
    if trip_id:
        record["tripId"] = trip_id
    if gid:
        record["groupId"] = gid
    contracts = read_contracts()
    contracts.insert(0, record)
    write_contracts(contracts)

    # Auto-create the matching invoice (mirrors handle_generate_contract).
    auto_invoice = None
    try:
        auto_inv = build_invoice_from_contract(record, actor)
        if auto_inv:
            inv_list = read_invoices()
            inv_list.insert(0, auto_inv)
            write_invoices(inv_list)
            record["autoInvoiceId"] = auto_inv["id"]
            contracts[0] = record
            write_contracts(contracts)
            auto_invoice = {"id": auto_inv["id"], "serial": auto_inv.get("serial")}
    except Exception as exc:
        print(f"[agent contract->invoice] auto-create failed: {exc}", file=sys.stderr, flush=True)

    return {"ok": True, "contract": {
        "id": record["id"],
        "contractSerial": (record.get("data") or {}).get("contractSerial"),
        "pdfPath": record.get("pdfPath") or "",
        "docxPath": record.get("docxPath") or "",
        "autoInvoice": auto_invoice,
    }}


def _strip_image_bytes_from_history(history):
    """Replace image / document content blocks with text placeholders so
    the conversation stays small after the first round-trip. Used between
    tool loops — after Claude has already read the image once, sending it
    again wastes proxy budget."""
    out = []
    for turn in history:
        content = turn.get("content")
        if isinstance(content, list):
            new_content = []
            for block in content:
                if isinstance(block, dict) and block.get("type") in ("image", "document"):
                    new_content.append({"type": "text", "text": "[image/file already read in earlier turn — see notes above]"})
                else:
                    new_content.append(block)
            out.append({**turn, "content": new_content})
        else:
            out.append(turn)
    return out


def _agent_save_uploaded_doc(raw_bytes, original_name, ext_hint):
    """Persist an arbitrary uploaded file to the generated dir and return public path."""
    if not raw_bytes:
        return ""
    safe_ext = re.sub(r"[^a-z0-9]", "", (ext_hint or "bin").lower())[:6] or "bin"
    GENERATED_DIR.mkdir(parents=True, exist_ok=True)
    base = re.sub(r"[^A-Za-z0-9._-]", "_", original_name or "doc")[:40]
    filename = f"agent-doc-{uuid4().hex[:10]}-{base}"
    if not filename.lower().endswith("." + safe_ext):
        filename = f"{filename}.{safe_ext}"
    (GENERATED_DIR / filename).write_bytes(raw_bytes)
    return f"/generated/{filename}"


def _agent_extract_xlsx_text(raw_bytes):
    """Return a compact text dump of an xlsx file (first sheet, first 200 rows)."""
    try:
        from openpyxl import load_workbook
        from io import BytesIO
        wb = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
        out = []
        for ws in wb.worksheets[:3]:
            out.append(f"--- Sheet: {ws.title} ---")
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx >= 200:
                    out.append("(... rows truncated ...)")
                    break
                cells = ["" if v is None else str(v) for v in row]
                out.append("\t".join(cells)[:500])
        return "\n".join(out)[:15000]
    except Exception as exc:
        return f"(xlsx parse failed: {type(exc).__name__}: {exc})"


def _agent_save_uploaded_image(b64, media_type):
    """Persist a base64 image to the generated dir and return its public path."""
    if not b64:
        return ""
    ext_map = {"image/jpeg": "jpg", "image/jpg": "jpg", "image/png": "png",
               "image/gif": "gif", "image/webp": "webp"}
    ext = ext_map.get((media_type or "").lower(), "jpg")
    try:
        raw = base64.b64decode(b64)
    except Exception:
        return ""
    if not raw or len(raw) > AGENT_MAX_IMAGE_BYTES:
        return ""
    GENERATED_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"agent-upload-{uuid4().hex[:12]}.{ext}"
    (GENERATED_DIR / filename).write_bytes(raw)
    return f"/generated/{filename}"


def _agent_resolve_attachment(att):
    """Translate {kind, id} or {kind, path} into (filename, bytes). Returns
    (None, error_message) on failure."""
    if not isinstance(att, dict):
        return None, "attachment must be an object"
    kind = (att.get("kind") or "").lower().strip()

    if kind == "file":
        rel = (att.get("path") or "").strip()
        if not rel.startswith("/generated/"):
            return None, "file path must start with /generated/"
        p = GENERATED_DIR / rel.replace("/generated/", "", 1)
        if not p.exists():
            return None, f"file not found: {rel}"
        return p.name, p.read_bytes()

    if kind == "invoice":
        iid = (att.get("id") or "").strip()
        inv = next((i for i in read_invoices() if i.get("id") == iid), None)
        if not inv:
            return None, f"invoice {iid} not found"
        try:
            pdf_url = save_standalone_invoice_pdf(inv)
        except Exception as exc:
            return None, f"invoice PDF render failed: {exc}"
        p = GENERATED_DIR / pdf_url.replace("/generated/", "", 1)
        return f"invoice-{inv.get('serial') or iid}.pdf", p.read_bytes()

    if kind == "contract":
        cid = (att.get("id") or "").strip()
        contracts = read_contracts()
        contract = None
        contract_index = None
        for idx, c in enumerate(contracts):
            if c.get("id") == cid:
                contract = c
                contract_index = idx
                break
        if not contract:
            return None, f"contract {cid} not found"
        # Always attach as PDF. Render on demand if not already saved (or if the
        # saved pdfPath points at the html fallback / a missing file). DOCX is
        # never used as the email attachment — clients expect PDFs.
        rel = contract.get("pdfPath") or ""
        p = GENERATED_DIR / rel.replace("/generated/", "", 1) if rel.startswith("/generated/") else None
        if not (rel.endswith(".pdf") and p and p.exists()):
            try:
                rel = save_contract_pdf(contract)
            except Exception as exc:
                return None, f"contract PDF render failed: {exc}"
            contract["pdfPath"] = rel
            if contract_index is not None:
                contracts[contract_index] = contract
                write_contracts(contracts)
            p = GENERATED_DIR / rel.replace("/generated/", "", 1)
        serial = (contract.get("data") or {}).get("contractSerial") or cid
        return f"contract-{serial}.pdf", p.read_bytes()

    if kind == "tourist_passport":
        tid = (att.get("id") or "").strip()
        tourist = next((t for t in read_tourists() if t.get("id") == tid), None)
        if not tourist:
            return None, f"tourist {tid} not found"
        rel = tourist.get("passportScanPath") or ""
        if not rel.startswith("/generated/"):
            return None, "tourist has no saved passport scan"
        p = GENERATED_DIR / rel.replace("/generated/", "", 1)
        if not p.exists():
            return None, f"passport file not found on disk: {rel}"
        name = ((tourist.get("lastName") or "") + "-" + (tourist.get("firstName") or "")).strip("-") or tid
        return f"passport-{name}{p.suffix}", p.read_bytes()

    if kind == "trip_document":
        trip_id = (att.get("tripId") or "").strip()
        doc_id = (att.get("id") or "").strip()
        if not trip_id or not doc_id:
            return None, "trip_document needs tripId and id"
        trip = next((t for t in read_camp_trips() if t.get("id") == trip_id), None)
        if not trip:
            return None, f"trip {trip_id} not found"
        doc = next((d for d in (trip.get("documents") or []) if d.get("id") == doc_id), None)
        if not doc:
            return None, f"document {doc_id} not found in trip {trip_id}"
        fp = (TRIP_UPLOADS_DIR / trip_id / (doc.get("storedName") or "")).resolve()
        if not str(fp).startswith(str(TRIP_UPLOADS_DIR.resolve())) or not fp.exists():
            return None, "document file missing on disk"
        return doc.get("originalName") or fp.name, fp.read_bytes()

    return None, f"unknown attachment kind: {kind}"


def _tool_send_email(args, actor):
    api_key = os.environ.get("RESEND_API_KEY", "").strip()
    if not api_key:
        return {"error": "Email is not configured. Sign up at resend.com (free, 100 emails/day) and add RESEND_API_KEY to Render → Environment."}
    sender = os.environ.get("EMAIL_FROM", "").strip() or "Travelx <onboarding@resend.dev>"
    # Per-call friendly-name override: when the caller passes _company_name
    # (DTX or USM workspace), the inbox shows "Дэлхий Трэвел Икс" /
    # "Unlock Steppe Mongolia" as the sender name even though the underlying
    # address (e.g. noreply@travelx.mn) doesn't change.
    company_name_override = (args.get("_company_name") or "").strip()
    if company_name_override:
        m = re.search(r"<([^>]+)>", sender)
        addr = m.group(1) if m else sender
        sender = f"{company_name_override} <{addr}>"
    to_raw = args.get("to") or ""
    if isinstance(to_raw, str):
        to_list = [x.strip() for x in to_raw.replace(";", ",").split(",") if x.strip()]
    elif isinstance(to_raw, list):
        to_list = [str(x).strip() for x in to_raw if str(x).strip()]
    else:
        to_list = []
    if not to_list:
        return {"error": "to (recipient email) is required"}

    subject = (args.get("subject") or "").strip() or "Travelx документ"
    body = args.get("body") or ""
    # Callers (e.g. the promo modal with a rich editor) can pass pre-rendered
    # HTML via _body_html_override and we use it verbatim. Otherwise we treat
    # plain newlines as <br> so the email renders as the agent intended.
    body_html_override = args.get("_body_html_override") or ""
    if body_html_override:
        body_html = body_html_override
    else:
        body_html = "<p>" + html.escape(body).replace("\n\n", "</p><p>").replace("\n", "<br>") + "</p>"
    # Auto-appended footer: small gray italic disclaimer + signature. Server-owned
    # so Bataa can't forget it and styling stays consistent across all sends.
    # Internal callers can pass _skip_footer=True (e.g. when emailing info@travelx.mn
    # itself, where the "contact info@travelx.mn" line would be self-referential).
    if not args.get("_skip_footer"):
        company_name = (args.get("_company_name") or "").strip() or "Дэлхий Трэвел Икс"
        body_html += (
            '<p style="color:#888;font-style:italic;font-size:12px;margin-top:20px">'
            'Энэ имэйл нь автомат илгээгдсэн тул буцааж хариу бичихгүй байхыг хүсье. '
            'Шаардлагатай бол <a href="mailto:info@travelx.mn" style="color:#888">info@travelx.mn</a> '
            'эсвэл <a href="tel:+97672007722" style="color:#888">72007722</a> дугаараар холбогдоорой.'
            '</p>'
            f'<p style="margin-top:8px">Хүндэтгэсэн,<br>{html.escape(company_name)}</p>'
        )

    raw_attachments = args.get("attachments") or []
    if not isinstance(raw_attachments, list):
        raw_attachments = []
    if len(raw_attachments) > 35:
        return {"error": "too many attachments (max 35 per email)"}

    api_attachments = []
    failed = []
    for att in raw_attachments:
        # Pre-encoded attachment (already has filename + content base64) —
        # just pass through, including optional content_id/type for inline
        # rendering. Otherwise resolve from a path/URL.
        if isinstance(att, dict) and att.get("content") and att.get("filename"):
            entry = {
                "filename": att["filename"],
                "content": att["content"],
            }
            if att.get("content_id"):
                entry["content_id"] = att["content_id"]
            if att.get("type"):
                entry["type"] = att["type"]
            api_attachments.append(entry)
            continue
        filename, payload = _agent_resolve_attachment(att)
        if filename is None:
            failed.append(payload)
            continue
        api_attachments.append({
            "filename": filename,
            "content": base64.b64encode(payload).decode("ascii"),
        })

    if failed and not api_attachments:
        return {"error": "All attachments failed: " + "; ".join(failed)}

    payload = {
        "from": sender,
        "to": to_list,
        "subject": subject,
        "html": body_html,
    }
    if api_attachments:
        payload["attachments"] = api_attachments

    req = urllib.request.Request(
        "https://api.resend.com/emails",
        data=json.dumps(payload).encode("utf-8"),
        method="POST",
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "User-Agent": "Travelx-Bataa/1.0 (+https://www.backoffice.travelx.mn)",
            "Accept": "application/json",
        },
    )
    print(f"[send_email] from={sender!r} to={to_list!r} subject={subject!r} attachments={[a['filename'] for a in api_attachments]}", flush=True)
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        try:
            err_body = e.read().decode("utf-8")
        except Exception:
            err_body = str(e)
        print(f"[send_email] HTTPError {e.code} body={err_body!r}", flush=True)
        return {"error": f"Resend API error {e.code}: {err_body[:600]}"}
    except Exception as exc:
        print(f"[send_email] {type(exc).__name__}: {exc}", flush=True)
        return {"error": f"Email send failed: {type(exc).__name__}: {exc}"}

    return {
        "ok": True,
        "messageId": data.get("id"),
        "to": to_list,
        "attachments": [a["filename"] for a in api_attachments],
        "warnings": failed if failed else None,
    }


def _tool_attach_document_to_trip(args, actor):
    """Attach a previously-uploaded file (from /generated/agent-upload-...
    or /generated/agent-doc-...) to a trip's Documents section."""
    trip_id = (args.get("tripId") or "").strip()
    src_path = (args.get("filePath") or "").strip()
    category = (args.get("category") or "Other").strip() or "Other"
    rename_to = (args.get("name") or "").strip()
    if not trip_id:
        return {"error": "tripId is required"}
    if not src_path.startswith("/generated/"):
        return {"error": "filePath must be a /generated/... path returned in an [Uploaded ...] note"}
    src_file = GENERATED_DIR / src_path.replace("/generated/", "", 1)
    if not src_file.exists() or not src_file.is_file():
        return {"error": f"Source file not found: {src_path}"}

    trips = read_camp_trips()
    idx = next((i for i, t in enumerate(trips) if t.get("id") == trip_id), None)
    if idx is None:
        return {"error": "Trip not found"}

    doc_id = str(uuid4())
    ext = src_file.suffix.lower()
    original_name = rename_to or src_file.name
    if not original_name.lower().endswith(ext):
        original_name = original_name + ext
    trip_upload_dir = TRIP_UPLOADS_DIR / trip_id
    trip_upload_dir.mkdir(parents=True, exist_ok=True)
    stored_name = doc_id + ext
    dest = trip_upload_dir / stored_name
    dest.write_bytes(src_file.read_bytes())

    mime, _ = mimetypes.guess_type(src_file.name)
    doc = {
        "id": doc_id,
        "originalName": original_name,
        "storedName": stored_name,
        "mimeType": mime or "application/octet-stream",
        "size": dest.stat().st_size,
        "category": category,
        "uploadedAt": now_mongolia().isoformat(),
        "uploadedBy": actor_snapshot(actor),
    }
    trip = trips[idx]
    documents = list(trip.get("documents") or [])
    documents.append(doc)
    trips[idx] = {**trip, "documents": documents}
    write_camp_trips(trips)
    return {"ok": True, "document": {"id": doc_id, "originalName": original_name, "category": category}}


def _tool_attach_image_to_tourist(args, actor):
    tid = (args.get("touristId") or "").strip()
    image_path = (args.get("imagePath") or "").strip()
    field = (args.get("field") or "passport").strip().lower()
    if not tid:
        return {"error": "touristId is required"}
    if not image_path:
        return {"error": "imagePath is required (use a path returned in the [Uploaded image paths] note)"}
    if field not in ("passport", "photo"):
        return {"error": "field must be 'passport' or 'photo'"}
    target_field = "passportScanPath" if field == "passport" else "photoPath"
    tourists = read_tourists()
    for i, t in enumerate(tourists):
        if t.get("id") != tid:
            continue
        t[target_field] = image_path
        t["updatedAt"] = now_mongolia().isoformat()
        t["updatedBy"] = actor_snapshot(actor)
        tourists[i] = t
        write_tourists(tourists)
        return {"ok": True, "tourist": {"id": t["id"], "serial": t.get("serial"),
                                          "lastName": t.get("lastName"), "firstName": t.get("firstName"),
                                          target_field: image_path}}
    return {"error": "Tourist not found"}


def _tool_save_memory(args, actor):
    text = (args.get("text") or "").strip()
    if not text:
        return {"error": "text is required"}
    if len(text) > 1000:
        return {"error": "memory text too long (max 1000 chars)"}
    category = (args.get("category") or "general").strip().lower() or "general"
    mems = _agent_load_memories()
    record = {
        "id": str(uuid4()),
        "text": text,
        "category": category,
        "createdAt": now_mongolia().isoformat(),
        "createdBy": (actor or {}).get("email") or (actor or {}).get("id") or "admin",
    }
    mems.append(record)
    if len(mems) > 200:
        mems = mems[-200:]
    _agent_save_memories(mems)
    return {"ok": True, "memory": {"id": record["id"], "category": category, "text": text[:80]}}


def _tool_list_memories(args, actor):
    cat = (args.get("category") or "").strip().lower()
    mems = _agent_load_memories()
    if cat:
        mems = [m for m in mems if (m.get("category") or "") == cat]
    return mems


def _tool_delete_memory(args, actor):
    mid = (args.get("memoryId") or "").strip()
    if not mid:
        return {"error": "memoryId is required"}
    mems = _agent_load_memories()
    new = [m for m in mems if m.get("id") != mid]
    if len(new) == len(mems):
        return {"error": "Memory not found"}
    _agent_save_memories(new)
    return {"ok": True}


# ── Tool registry (Claude-facing JSON schemas) ────────────────────────

AGENT_TOOLS = [
    {"name": "list_trips", "description": "List trips. Optional filters: workspace (DTX or USM), status (planning, confirmed, travelling, completed, cancelled, offer), tripType (fit or git). Returns {count, items: [...]} — ALWAYS use the returned `count` field for any 'how many' answer; never count the items array yourself.",
     "input_schema": {"type": "object", "properties": {
         "workspace": {"type": "string"}, "status": {"type": "string"}, "tripType": {"type": "string"}}},
     "handler": _tool_list_trips},
    {"name": "get_trip", "description": "Get one trip by id.",
     "input_schema": {"type": "object", "required": ["tripId"], "properties": {"tripId": {"type": "string"}}},
     "handler": _tool_get_trip},
    {"name": "create_trip", "description": "Create a trip. workspace MUST be set explicitly: 'DTX' for any destination outside Mongolia (Solongos/Korea, Dubai, Japan, Turkey, Singapore, Russia, China, Europe, etc.) and 'USM' ONLY for trips inside Mongolia (Khustai, Gobi, Khuvsgul, Terelj, etc.). Never let workspace default. tripType is 'fit' or 'git'. Dates ISO yyyy-mm-dd.",
     "input_schema": {"type": "object", "required": ["tripName", "startDate", "tripType", "workspace"], "properties": {
         "workspace": {"type": "string", "enum": ["DTX", "USM"], "description": "DTX = outbound (any non-Mongolia destination); USM = inbound (inside Mongolia only). REQUIRED."},
         "tripName": {"type": "string"}, "tripType": {"type": "string", "enum": ["fit", "git"]},
         "startDate": {"type": "string"}, "endDate": {"type": "string"},
         "participantCount": {"type": "integer"}, "staffCount": {"type": "integer"},
         "totalDays": {"type": "integer"},
         "status": {"type": "string", "enum": ["offer", "confirmed", "cancelled", "ignored"]},
         "tags": {"type": "string"}}},
     "handler": _tool_create_trip},
    {"name": "update_trip", "description": "Update fields on an existing trip. Pass fields object with keys to change.",
     "input_schema": {"type": "object", "required": ["tripId", "fields"], "properties": {
         "tripId": {"type": "string"}, "fields": {"type": "object"}}},
     "handler": _tool_update_trip},
    {"name": "delete_trip", "description": "Delete a trip.",
     "input_schema": {"type": "object", "required": ["tripId"], "properties": {"tripId": {"type": "string"}}},
     "handler": _tool_delete_trip},

    {"name": "list_groups", "description": "List groups, optionally filtered by tripId.",
     "input_schema": {"type": "object", "properties": {"tripId": {"type": "string"}}},
     "handler": _tool_list_groups},
    {"name": "create_group", "description": "Create a tourist group on a trip.",
     "input_schema": {"type": "object", "required": ["tripId", "name"], "properties": {
         "tripId": {"type": "string"}, "name": {"type": "string"}, "headcount": {"type": "integer"},
         "leaderName": {"type": "string"}, "leaderEmail": {"type": "string"},
         "leaderPhone": {"type": "string"}, "leaderNationality": {"type": "string"},
         "status": {"type": "string", "enum": ["pending", "confirmed", "cancelled"]}}},
     "handler": _tool_create_group},
    {"name": "update_group", "description": "Update group fields.",
     "input_schema": {"type": "object", "required": ["groupId", "fields"], "properties": {
         "groupId": {"type": "string"}, "fields": {"type": "object"}}},
     "handler": _tool_update_group},
    {"name": "delete_group", "description": "Delete a group.",
     "input_schema": {"type": "object", "required": ["groupId"], "properties": {"groupId": {"type": "string"}}},
     "handler": _tool_delete_group},

    {"name": "list_tourists", "description": "List participants, optionally filtered by tripId or groupId.",
     "input_schema": {"type": "object", "properties": {"tripId": {"type": "string"}, "groupId": {"type": "string"}}},
     "handler": _tool_list_tourists},
    {"name": "create_tourist", "description": "Add a participant to a group.",
     "input_schema": {"type": "object", "required": ["tripId", "groupId", "lastName", "firstName"], "properties": {
         "tripId": {"type": "string"}, "groupId": {"type": "string"},
         "lastName": {"type": "string"}, "firstName": {"type": "string"},
         "gender": {"type": "string"}, "dob": {"type": "string"},
         "nationality": {"type": "string"}, "passportNumber": {"type": "string"},
         "passportIssueDate": {"type": "string"}, "passportExpiry": {"type": "string"},
         "passportIssuePlace": {"type": "string"}, "registrationNumber": {"type": "string"},
         "phone": {"type": "string"}, "email": {"type": "string"},
         "roomType": {"type": "string"}, "roomCode": {"type": "string"},
         "notes": {"type": "string"}}},
     "handler": _tool_create_tourist},
    {"name": "update_tourist", "description": "Update a participant.",
     "input_schema": {"type": "object", "required": ["touristId", "fields"], "properties": {
         "touristId": {"type": "string"}, "fields": {"type": "object"}}},
     "handler": _tool_update_tourist},
    {"name": "delete_tourist", "description": "Delete a participant.",
     "input_schema": {"type": "object", "required": ["touristId"], "properties": {"touristId": {"type": "string"}}},
     "handler": _tool_delete_tourist},

    {"name": "list_invoices", "description": "List invoices, optionally filtered by tripId.",
     "input_schema": {"type": "object", "properties": {"tripId": {"type": "string"}}},
     "handler": _tool_list_invoices},
    {"name": "get_invoice", "description": "Get one invoice by id.",
     "input_schema": {"type": "object", "required": ["invoiceId"], "properties": {"invoiceId": {"type": "string"}}},
     "handler": _tool_get_invoice},
    {"name": "create_invoice", "description": "Create an invoice. Items are [{description, qty, price}]. Installments are [{description, amount, issueDate, dueDate, status}].",
     "input_schema": {"type": "object", "required": ["tripId", "groupId", "payerName", "items"], "properties": {
         "tripId": {"type": "string"}, "groupId": {"type": "string"},
         "payerName": {"type": "string"}, "payerId": {"type": "string"},
         "participantIds": {"type": "array", "items": {"type": "string"}},
         "items": {"type": "array", "items": {"type": "object"}},
         "installments": {"type": "array", "items": {"type": "object"}},
         "currency": {"type": "string"}}},
     "handler": _tool_create_invoice},
    {"name": "update_invoice", "description": "Update invoice fields.",
     "input_schema": {"type": "object", "required": ["invoiceId", "fields"], "properties": {
         "invoiceId": {"type": "string"}, "fields": {"type": "object"}}},
     "handler": _tool_update_invoice},
    {"name": "delete_invoice", "description": "Delete an invoice.",
     "input_schema": {"type": "object", "required": ["invoiceId"], "properties": {"invoiceId": {"type": "string"}}},
     "handler": _tool_delete_invoice},
    {"name": "publish_invoice", "description": "Publish a draft invoice.",
     "input_schema": {"type": "object", "required": ["invoiceId"], "properties": {"invoiceId": {"type": "string"}}},
     "handler": _tool_publish_invoice},
    {"name": "register_invoice_payment", "description": "Mark an installment as paid.",
     "input_schema": {"type": "object", "required": ["invoiceId", "installmentIndex"], "properties": {
         "invoiceId": {"type": "string"}, "installmentIndex": {"type": "integer"},
         "paidDate": {"type": "string"}}},
     "handler": _tool_register_invoice_payment},

    {"name": "list_contracts", "description": "List contracts, optionally filtered by tripId.",
     "input_schema": {"type": "object", "properties": {"tripId": {"type": "string"}}},
     "handler": _tool_list_contracts},
    {"name": "delete_contract", "description": "Delete a contract.",
     "input_schema": {"type": "object", "required": ["contractId"], "properties": {"contractId": {"type": "string"}}},
     "handler": _tool_delete_contract},

    {"name": "list_camp_reservations", "description": "List camp reservations, optionally for a trip.",
     "input_schema": {"type": "object", "properties": {"tripId": {"type": "string"}}},
     "handler": _tool_list_camp_reservations},
    {"name": "list_flight_reservations", "description": "List flight reservations, optionally for a trip.",
     "input_schema": {"type": "object", "properties": {"tripId": {"type": "string"}}},
     "handler": _tool_list_flight_reservations},
    {"name": "list_transfer_reservations", "description": "List transfer reservations, optionally for a trip.",
     "input_schema": {"type": "object", "properties": {"tripId": {"type": "string"}}},
     "handler": _tool_list_transfer_reservations},

    {"name": "list_users", "description": "List back-office users.",
     "input_schema": {"type": "object", "properties": {}},
     "handler": _tool_list_users},
    {"name": "list_ds160_applications", "description": "List DS-160 applications.",
     "input_schema": {"type": "object", "properties": {}},
     "handler": _tool_list_ds160},
    {"name": "list_notifications", "description": "List recent notifications (last 50).",
     "input_schema": {"type": "object", "properties": {}},
     "handler": _tool_list_notifications},
    {"name": "list_fifa_inventory", "description": "List FIFA 2026 ticket inventory entries (first 200).",
     "input_schema": {"type": "object", "properties": {}},
     "handler": _tool_list_fifa_inventory},
    {"name": "save_memory", "description": "Save a long-term note that you (Батаа) should remember in every future conversation. Use this when the admin says 'remember that…', 'санаж аваач', or teaches you a business rule, naming convention, recurring price, or fact that is not obvious from the data. Categories: 'rule' for business rules, 'fact' for facts, 'preference' for stylistic preferences, 'general' otherwise. Keep text short and self-contained.",
     "input_schema": {"type": "object", "required": ["text"], "properties": {
         "text": {"type": "string", "description": "The note to remember (max 1000 chars)."},
         "category": {"type": "string", "description": "rule | fact | preference | general"}}},
     "handler": _tool_save_memory},
    {"name": "list_memories", "description": "List long-term notes saved via save_memory. Optional category filter.",
     "input_schema": {"type": "object", "properties": {"category": {"type": "string"}}},
     "handler": _tool_list_memories},
    {"name": "delete_memory", "description": "Delete a long-term note by id. Use when the admin says 'forget that' or 'мартчих'.",
     "input_schema": {"type": "object", "required": ["memoryId"], "properties": {"memoryId": {"type": "string"}}},
     "handler": _tool_delete_memory},
    {"name": "get_contract", "description": "Read back a contract by id. Use this AFTER create_contract to verify what was actually saved (totalPrice, adultCount, depositAmount, dates) before reporting success to the admin.",
     "input_schema": {"type": "object", "required": ["contractId"], "properties": {"contractId": {"type": "string"}}},
     "handler": _tool_get_contract},
    {"name": "create_contract", "description": "Create a tour contract (гэрээ) and automatically generate the matching invoice. Pass tripId and touristId so destination, dates, and tourist info auto-fill from the existing records — you only need to provide pricing fields and counts. Required pricing fields: adultCount, adultPrice (per-adult price), depositAmount, depositDueDate (yyyy-mm-dd), balanceDueDate (yyyy-mm-dd). Optional: childCount, childPrice, infantCount, infantPrice. The tool fills in destination/dates from the trip and tourist info from the tourist record. Returns {contract: {id, contractSerial, pdfPath, autoInvoice}}.",
     "input_schema": {"type": "object", "required": ["tripId", "touristId", "adultCount", "adultPrice", "depositAmount", "depositDueDate", "balanceDueDate"],
       "properties": {
         "tripId": {"type": "string"},
         "groupId": {"type": "string"},
         "touristId": {"type": "string"},
         "adultCount": {"type": "number"},
         "adultPrice": {"type": "number"},
         "childCount": {"type": "number"},
         "childPrice": {"type": "number"},
         "depositAmount": {"type": "number"},
         "depositDueDate": {"type": "string"},
         "balanceDueDate": {"type": "string"},
         "totalPrice": {"type": "number"},
         "destination": {"type": "string"},
         "tripStartDate": {"type": "string"},
         "tripEndDate": {"type": "string"}}},
     "handler": _tool_create_contract},
    {"name": "send_email", "description": "Send an email with attached files (contract PDF, invoice PDF, tourist passport scans, or arbitrary uploaded files). Requires RESEND_API_KEY env var. Attachments is an array of objects: {kind: 'invoice'|'contract'|'tourist_passport'|'file', id?: '<recordId>', path?: '/generated/...'}. Use kind='invoice' with the invoice id, kind='contract' with the contract id, kind='tourist_passport' with the tourist id, kind='file' with a /generated/... path.",
     "input_schema": {"type": "object", "required": ["to", "subject"], "properties": {
         "to": {"type": "string", "description": "Recipient email address(es), comma-separated for multiple"},
         "subject": {"type": "string"},
         "body": {"type": "string", "description": "Plain text body; newlines become line breaks"},
         "attachments": {"type": "array", "items": {"type": "object", "properties": {
             "kind": {"type": "string", "enum": ["invoice", "contract", "tourist_passport", "file"]},
             "id": {"type": "string"},
             "path": {"type": "string"}}}}}},
     "handler": _tool_send_email},
    {"name": "attach_document_to_trip", "description": "Attach a previously-uploaded file (image OR document) to a trip's Documents section. Use the /generated/... path from the [Uploaded image paths] or [Uploaded document paths] note. Categories: 'Passport', 'Visa', 'Ticket', 'Hotel voucher', 'Insurance', 'Itinerary', 'Other'. Optional `name` to rename for display.",
     "input_schema": {"type": "object", "required": ["tripId", "filePath"], "properties": {
         "tripId": {"type": "string"},
         "filePath": {"type": "string", "description": "/generated/agent-upload-XXX.jpg or /generated/agent-doc-XXX.pdf path from the upload note"},
         "category": {"type": "string"},
         "name": {"type": "string", "description": "Display filename (extension auto-added if missing)"}}},
     "handler": _tool_attach_document_to_trip},
    {"name": "attach_image_to_tourist", "description": "Attach an image (passport scan or portrait photo) to an existing tourist record. Use this AFTER the admin uploads an image in chat — every uploaded image is saved to disk and its path is given to you in a system note like '[Uploaded image paths] img-1=/generated/agent-upload-XXX.jpg'. Pass that path as imagePath. field='passport' fills passportScanPath; field='photo' fills photoPath.",
     "input_schema": {"type": "object", "required": ["touristId", "imagePath"], "properties": {
         "touristId": {"type": "string"},
         "imagePath": {"type": "string", "description": "Path returned in the [Uploaded image paths] note, e.g. /generated/agent-upload-abc123.jpg"},
         "field": {"type": "string", "description": "'passport' (default) or 'photo'"}}},
     "handler": _tool_attach_image_to_tourist},

    # ── Read tools (single-record) ───────────────────────────────
    {"name": "get_tourist", "description": "Get one tourist by id. Returns the full record (passport, contact, group + trip ids, marketing status, tags).",
     "input_schema": {"type": "object", "required": ["touristId"], "properties": {"touristId": {"type": "string"}}},
     "handler": _tool_get_tourist},
    {"name": "get_group", "description": "Get one group by id. Returns name, leader, headcount, status, tripId.",
     "input_schema": {"type": "object", "required": ["groupId"], "properties": {"groupId": {"type": "string"}}},
     "handler": _tool_get_group},

    # ── Documents (global storage view) ──────────────────────────
    {"name": "list_documents", "description": "Flatten every uploaded file across every trip in the workspace. Optional filters: tripId, category (Passport, Visa, Ticket, Hotel voucher, Insurance, Itinerary, Invoices, Contracts, Other), touristId.",
     "input_schema": {"type": "object", "properties": {"tripId": {"type": "string"}, "category": {"type": "string"}, "touristId": {"type": "string"}}},
     "handler": _tool_list_documents},

    # ── Manager-dashboard tasks + contacts (the To-Do page) ──────
    {"name": "list_tasks", "description": "List manager-dashboard tasks. Optional status filter (pending, in-progress, done, cancelled, overdue).",
     "input_schema": {"type": "object", "properties": {"status": {"type": "string"}}},
     "handler": _tool_list_tasks},
    {"name": "create_task", "description": "Create a to-do task. Pass `owners` (list of manager names) for multi-assign, or `owner` (single name) for backwards compat. createdBy stamps automatically.",
     "input_schema": {"type": "object", "required": ["title"], "properties": {
         "title": {"type": "string"}, "owner": {"type": "string"},
         "owners": {"type": "array", "items": {"type": "string"}},
         "priority": {"type": "string", "enum": ["low", "medium", "high"]},
         "status": {"type": "string", "enum": ["pending", "in-progress", "done", "cancelled"]},
         "dueDate": {"type": "string"}, "dueTime": {"type": "string"},
         "destinations": {"type": "array", "items": {"type": "string"}},
         "note": {"type": "string"}}},
     "handler": _tool_create_task},
    {"name": "update_task", "description": "Update task fields by id.",
     "input_schema": {"type": "object", "required": ["taskId", "fields"], "properties": {"taskId": {"type": "string"}, "fields": {"type": "object"}}},
     "handler": _tool_update_task},
    {"name": "delete_task", "description": "Delete a task.",
     "input_schema": {"type": "object", "required": ["taskId"], "properties": {"taskId": {"type": "string"}}},
     "handler": _tool_delete_task},

    {"name": "list_contacts", "description": "List manager-dashboard contacts (clients, leads, vendors).",
     "input_schema": {"type": "object", "properties": {}},
     "handler": _tool_list_contacts},
    {"name": "create_contact", "description": "Create a contact entry.",
     "input_schema": {"type": "object", "required": ["name"], "properties": {
         "name": {"type": "string"}, "phone": {"type": "string"},
         "type": {"type": "string", "enum": ["client", "lead", "vendor", "guide", "driver"]},
         "status": {"type": "string", "enum": ["new", "warm", "priority", "cold"]},
         "lastContacted": {"type": "string"},
         "destinations": {"type": "array", "items": {"type": "string"}},
         "note": {"type": "string"}}},
     "handler": _tool_create_contact},
    {"name": "update_contact", "description": "Update contact fields.",
     "input_schema": {"type": "object", "required": ["contactId", "fields"], "properties": {"contactId": {"type": "string"}, "fields": {"type": "object"}}},
     "handler": _tool_update_contact},
    {"name": "delete_contact", "description": "Delete a contact.",
     "input_schema": {"type": "object", "required": ["contactId"], "properties": {"contactId": {"type": "string"}}},
     "handler": _tool_delete_contact},

    # ── Settings ────────────────────────────────────────────────
    {"name": "get_settings", "description": "Read shared workspace settings: destinations list, bank accounts (for invoice picker), camp names + per-camp prices/contracts.",
     "input_schema": {"type": "object", "properties": {}},
     "handler": _tool_get_settings},
    {"name": "update_destinations", "description": "Replace the full destinations list. Pass destinations as an array of strings.",
     "input_schema": {"type": "object", "required": ["destinations"], "properties": {"destinations": {"type": "array", "items": {"type": "string"}}}},
     "handler": _tool_update_destinations},
    {"name": "update_bank_accounts", "description": "Replace the full bank-accounts list shown in the contract bank-picker + rendered invoice. Each account: {id?, label, bankName, accountName, accountNumber, currency, swift?, notes?}.",
     "input_schema": {"type": "object", "required": ["bankAccounts"], "properties": {"bankAccounts": {"type": "array", "items": {"type": "object"}}}},
     "handler": _tool_update_bank_accounts},

    # ── Team ────────────────────────────────────────────────────
    {"name": "list_team_members", "description": "List approved team members (managers + admins) with email, role, phone.",
     "input_schema": {"type": "object", "properties": {}},
     "handler": _tool_list_team_members},

    # ── Mail ────────────────────────────────────────────────────
    {"name": "list_mail_messages", "description": "Recent mail headers from every connected mailbox in the workspace. folder='inbox' (default) or 'sent'. Optional workspace + limit (default 50).",
     "input_schema": {"type": "object", "properties": {"folder": {"type": "string"}, "workspace": {"type": "string"}, "limit": {"type": "integer"}}},
     "handler": _tool_list_mail_messages},
    {"name": "list_mail_followups", "description": "All mail follow-ups (waiting, urgent, replied, cancelled).",
     "input_schema": {"type": "object", "properties": {}},
     "handler": _tool_list_mail_followups},
    {"name": "list_mail_dossiers", "description": "Every mail-thread → trip / group manual link.",
     "input_schema": {"type": "object", "properties": {}},
     "handler": _tool_list_mail_dossiers},

    # ── DS-160 actions ──────────────────────────────────────────
    {"name": "send_ds160_invitation", "description": "Email the share link for an existing DS-160 invitation to the client (uses Resend, same plumbing as Send via TravelX).",
     "input_schema": {"type": "object", "required": ["ds160Id"], "properties": {"ds160Id": {"type": "string"}, "to": {"type": "string"}}},
     "handler": _tool_send_ds160_invitation},

    # ── Invoice payment with note ───────────────────────────────
    {"name": "register_invoice_payment_with_note", "description": "Mark an installment as paid AND record a free-text note (context like 'paid only for the father, not the mother'). Prefer this over register_invoice_payment when the admin gives any context.",
     "input_schema": {"type": "object", "required": ["invoiceId", "installmentIndex"], "properties": {
         "invoiceId": {"type": "string"}, "installmentIndex": {"type": "integer"},
         "paidDate": {"type": "string"}, "status": {"type": "string"}, "note": {"type": "string"}}},
     "handler": _tool_register_invoice_payment_with_note},
]
AGENT_TOOL_BY_NAME = {t["name"]: t for t in AGENT_TOOLS}


def _agent_system_prompt(actor):
    name = (actor or {}).get("fullName") or (actor or {}).get("email") or "admin"
    today = now_mongolia().date().isoformat()
    memories = _agent_load_memories()
    if memories:
        mem_lines = "\n".join(
            f"- [{m.get('category', 'general')}] {m.get('text', '')}"
            for m in memories[-80:]
        )
        memory_block = (
            "\n\nLong-term memory (notes saved by the admin team — treat as authoritative business rules):\n"
            + mem_lines + "\n"
        )
    else:
        memory_block = ""
    return f"""You are "Батаа" (Bataa) — the in-app admin assistant for travelx.mn back-office. Today is {today}. You are talking to {name} (admin). Introduce yourself as Батаа when greeted; reply in Mongolian by default unless the user writes in another language.{memory_block}

Domain context:
- Two workspaces: DTX (Delkhii Travel Ix — outbound tours, Монгол хүн гадаад руу) and USM (Unlock Steppe Mongolia — inbound, гадаад жуулчин Монгол руу). Records carry a `company` field with value "DTX" or "USM"; tool inputs use `workspace` as a friendly synonym for the same thing.
- WORKSPACE INFERENCE (very important): if the destination is ANY country other than Mongolia (Dubai, Japan, Turkey, Singapore, Korea, China, Russia, etc.), this is a DTX trip. Only trips whose destination is inside Mongolia (Khustai, Gobi, Khuvsgul, Terelj, etc.) belong to USM. NEVER create a DTX-style trip in USM by mistake. If the user mentions "Dubai", "Japan", "Korea" etc. without naming a workspace, default to DTX.
- TRIP SERIAL PREFIXES: DTX trips use prefix "T-" (e.g., T-0007), USM trips use prefix "S-" (e.g., S-0007). When the admin asks about "T-0007", search DTX; "S-0007" → USM. Some legacy USM trips still carry T- prefix from before the change, so when an exact serial collides across workspaces ask the admin which workspace.
- LANGUAGE: when working in DTX (Mongolian outbound), reply in Mongolian. When working in USM (foreign inbound clients), generated documents (invoice descriptions, contract destinations) should be in English even though you speak Mongolian to the admin.
- Trip types: FIT (individual / family booking, usually no group needed) and GIT (group tour, has named group).
- Trip status values: planning, offer, confirmed, travelling, completed, cancelled. Lowercase only.
- Each trip can have groups, participants (tourists), camp/flight/transfer reservations, contracts, invoices, documents.
- Invoices have items + installments. Each item has fields description, qty, price; total = qty × price. Each installment has issueDate, dueDate, amount, status (pending/paid/overdue).
- IMPORTANT for invoices with multiple people: if the admin says "4 хүн × 2,500,000₮", create ONE item with qty=4 and price=2500000 (total auto-calculates to 10,000,000). Do NOT create one item with qty=1 and price=2500000.
- Mongolian-language naming is normal: payerName, descriptions, etc. may be in Mongolian.

Mongolian number words — read these EXACTLY, no rounding, no estimating:
- "сая" / "say" = 1,000,000 (million). "2 сая" / "2 say" / "2sayiig" / "2sayig" = 2,000,000 (NOT 20 million).
- "мянга" / "myanga" = 1,000 (thousand). "500 мянга" = 500,000.
- "тэрбум" / "terbum" / "tervum" = 1,000,000,000 (billion).
- When the admin gives a specific deposit/balance amount in words, USE THAT EXACT NUMBER. Do NOT recompute as a percentage of the total. Example: total 50,000,000₮, admin says "урьдчилгаа 2 сая" → depositAmount = 2,000,000, balance = 48,000,000. NEVER infer 20,000,000.
- After parsing any money figure, double-check: "2 сая" must equal 2,000,000 — verify the digit count before saving.

How you work:
- Use the provided tools to read and write the system. Don't make up data — call list_/get_ first when you need a fact.
- When the user asks about "DTX-н аяллууд" / "USM-н аяллууд" / "this workspace's X", always pass workspace="DTX" or "USM" to the list tool — never assume.
- Before destructive actions (delete_*, update_invoice fields wiping data, deleting trips/groups), briefly confirm with the user in your reply, then call the tool only if they assent.
- BE EFFICIENT: if the admin gave you a clear instruction (e.g. "Дубайд явах FIT аялал, 4 хүн × 2,500,000₮, гэрээ ба нэхэмжлэх хий"), execute the whole chain in this turn — do NOT ask for confirmation between create_trip → create_group → create_tourist → create_contract. Confirm the FINAL result, not each step.
- One short status sentence, then act. Avoid 3-paragraph plans before doing anything.
- After you do something, summarize the result in 1-2 sentences and include the relevant id/serial. Reply in Mongolian unless the user writes in another language.
- If a tool returns {{"error": ...}}, explain what went wrong and propose a fix. If a list returns 0 items, double-check by calling the same tool without filters before telling the user the data is empty — the filter may be wrong.
- For dates, use yyyy-mm-dd. For money, use plain numbers (no commas) when calling tools.
- DATE INFERENCE: when the admin gives only a month and day (e.g. "12 sariin 1-s 9", "11.01-11.08"), use the CURRENT year (today is shown above). If the resulting date is in the past relative to today, use NEXT year. NEVER use a year more than 1 year away from today. Example with today 2026-04-25: "12 сарын 1-9" → 2026-12-01 to 2026-12-09 (NOT 2016, NOT 2027). Verify the year is within ±1 of today before saving.
- Never invent IDs; only use ones returned by tools.

Counting:
- When a tool returns an object with a `count` field, that count is authoritative — never re-count the items array yourself, and never enumerate items just to get a number. Quote `count` directly.
- Do NOT infer that a serial is "missing" from gaps in returned data unless you have explicitly listed every record and confirmed the gap. Gaps in serial numbers are normal (e.g., a trip was deleted, or status filter excluded it) and do not affect the count.

Honesty (very important):
- Never make up data, IDs, names, prices, dates, or counts. If you don't have a tool for what was asked, say "Уучлаарай, надад тэр үйлдлийг хийх tool алга" (or in English: "I don't have a tool for that") and suggest the closest thing you CAN do.
- If a tool returns nothing or an error, say so plainly: "Олдсонгүй", "Алдаа гарлаа: ...". Do not invent placeholder data to fill the gap.
- If you are unsure (e.g., the user's question is ambiguous, or two records look like they could match), ask a clarifying question instead of guessing.
- Never claim a record exists, a price is X, or a status is Y unless that fact came back from a tool call you just made in this turn.
- It is far better to say "Би мэдэхгүй байна" or "Энэ функц одоогоор боломжгүй" than to give a confident wrong answer.

Verify-after-write (mandatory):
- After EVERY create_* or update_* call that involves money, counts, or dates, you MUST call the matching get_* (get_invoice, get_contract, get_trip) and quote the returned numbers verbatim in your reply. Do not summarize from memory — read the saved values back from the tool result.
- When the admin asks "did you do X correctly?" or "show me what you saved", call get_* first; never answer from your own recollection of what you intended to save.
- If the saved totals do not match what the admin asked for (e.g. admin said 4 × 2,500,000 = 10,000,000 but the saved invoice total is 4,000,000), say so explicitly: "Уучлаарай, би буруу үүсгэсэн байна. Жинхэнэ дүн: 4,000,000₮, таны хэлсэн дүн: 10,000,000₮. Засъя?" — then offer to update or delete-and-recreate. NEVER claim the wrong value is correct.

Long-term memory:
- Use save_memory whenever the admin teaches you something durable — a business rule, a recurring price, a naming convention, a person's role, a recurring instruction. Confirm in your reply (e.g., "Тэмдэглэн авлаа").
- Use list_memories when you want to recall what's been saved (the most recent ~80 are also injected into your system prompt automatically).
- Use delete_memory when the admin says forget / мартчих / устга.

Images and files:
- The admin can attach images (passport pages, contracts, ID photos, receipts), PDF documents, Excel spreadsheets, and plain-text files. Drag-and-drop into the chat works too.
- Images (jpg/png/gif/webp) are sent to you as vision blocks — read them carefully.
- PDF files are sent as document blocks — read every page.
- Excel (.xlsx) files are pre-extracted server-side and the table content is appended to the user message inside a "[Хавсаргасан Excel: filename]" block — use that text directly.
- Every uploaded image's saved path is given to you in a "[Uploaded image paths] img-1=/generated/...; img-2=..." note. Every uploaded document's saved path is in a "[Uploaded document paths] doc-1=/generated/...; doc-2=..." note. Use these exact paths when calling attach_image_to_tourist.
- Supported formats: JPG, PNG, GIF, WebP, PDF, XLSX, TXT/CSV/MD. HEIC/HEIF and Word .docx are NOT supported yet — if those are uploaded, ask the admin to convert (HEIC → JPG, DOCX → PDF) before retrying.
- For passports: read fields exactly as printed (Latin letters), then offer to create a tourist record via create_tourist with the extracted fields. After the tourist is created, ALSO call attach_image_to_tourist with arguments touristId=<the new tourist id>, imagePath=<the img-1 path from the upload note>, field="passport" so the passport scan is saved to the tourist's record. Mention "Паспортын зургийг хавсаргалаа" in the reply so the admin knows.
- For attaching files to a TRIP's Documents section (е.g., admin says "trip document hesegt upload hii", "энэ файлыг аяллын баримт бичигт нэм"), use attach_document_to_trip with the /generated/... path from the upload note and an appropriate category (Passport, Visa, Ticket, Hotel voucher, Insurance, Itinerary, Other).

Email:
- Use send_email when the admin asks to mail something ("mail-r yavuulj og", "имэйлээр илгээ", "send to client@example.com"). Requires RESEND_API_KEY in the env; if it's missing, send_email returns a clear error and you should pass that error to the admin verbatim.
- Attach files by reference. For each attachment pass an object with a `kind` field plus the right id/path: kind="invoice" with id=<invoice id>; kind="contract" with id=<contract id>; kind="tourist_passport" with id=<tourist id>; kind="file" with path="/generated/...". The server fetches and renders the right file — you do not need to know the underlying path.
- Default subject in Mongolian unless admin specified otherwise. Always confirm to the admin which addresses received the mail and which files were attached.
- If a field is unclear or partially obscured, say so rather than guessing.
- DO NOT write a closing signature ("Хүндэтгэсэн", "Дэлхий Трэвел Икс", "Travelx team", company name, contact info, "автомат имэйл" disclaimer, etc.) at the end of `body`. The server appends a fixed gray-italic disclaimer and the company signature automatically. Just write greeting → content → "Баярлалаа" (or similar short close), and stop.

The whole back-office — every page, every workflow you can help with:

1. Home / Trip page (/backoffice or /trip-detail)
   - Trip creator (modal): tripType (FIT/GIT), dates, pax, status, destinations.
   - Active trip card shows pax counter (FIT: planned; GIT: actual/planned).
   - Sections under a chosen trip: Tourists, Groups (GIT only), Camp/Flight/Transfer reservations, Contracts, Invoices, Documents, Rooming.
   - + Add tourist auto-creates the implicit group on FIT only; GIT requires + Add group first.
   - Bank account chosen on contract form flows to the auto-invoice via invoiceMeta.bankAccountKey.

2. Tourist Directory (/tourist)
   - Every tourist across every trip + promo-only contacts (no trip).
   - Filters: search by name/serial/destination, trip, group, DOB range, age range, status pills.
   - + Add tourist: full passport-aware form with OCR autofill via passport-scan dropzone; saves with no group/trip → it becomes a "PR-XXXX" promo contact.
   - Mark-all marks every eligible tourist across all pages of the filtered set.
   - + Send promo: bulk-mails the selected tourists; skips children (<18 auto), do_not_contact, and rows without email.

3. Contracts (/contracts)
   - Contract maker: 2-step (count setup → details). Saves contract DOCX + PDF.
   - Bank dropdown: sources from Settings → Bank accounts. Doesn't appear in the contract document, only on the auto-generated invoice.
   - Auto-creates a matching invoice with two installments (deposit + balance).

4. Invoices (/invoices)
   - Global list of every invoice in the workspace.
   - Each invoice has items + installments + payments. Installments accept a free-text note via register_invoice_payment_with_note.

5. Documents (/documents)
   - Global flat view of every uploaded file across every trip. Filter by file name, person name, destination, trip, category, file type, upload date range.

6. Mail (/mail)
   - Inbox + Sent for every connected mailbox in the workspace.
   - Compose with rich editor + signature picker + template loader.
   - Per-thread bell (follow-up timer 1/3/5/7/custom days). Per-thread "Linked dossier" lets the admin manually attach a thread to a trip or group; the chip on the row navigates there.

7. To Do (/todo)
   - Mixed list: tasks (what to do) + contacts (who to call). Same table.
   - Tasks: title, owner (assignee), priority, status, due date+time, destinations, note. Assigned by column shows the creator.
   - Notifications fire 6 hours before due (email + bell, idempotent).

8. Reservations (/camp-reservations, /flight-reservations, /transfer-reservations)
   - Each is a global list + + Add modal scoped to a trip.

9. DS-160 (/ds160)
   - Create invitation → unique share URL for the client. "Send via TravelX" delivers the link via Resend so it works without a desktop mail client. Once filled, status flips to Submitted and the answers are viewable.

10. FIFA 2026 admin (/fifa2026/admin)
    - Per-match ticket inventory: load tickets by match number + city, mark sold, attach buyer info.

11. Settings (/settings)
    - Tabs: Destinations (used in every dropdown), Camps (camp names + prices + contracts used in camp reservations), Bank accounts (used in contract → invoice picker; seeded with Төрийн Банк + Голомт Банк).

12. Team / Admin (/admin) — admin-only
    - Approve signups, assign roles, deactivate accounts. Everything else (mail accounts, bank accounts, settings, agent chat) is open to managers.

Workspace separation:
- The cookie `activeWorkspace` (DTX or USM) scopes every list endpoint. When the admin clicks a notification linking to a trip in the OTHER workspace, the workspace cookie is auto-switched before navigation.
- When in doubt about which workspace, infer from destination (Mongolia → USM, anything else → DTX).
"""



def _agent_call_anthropic(system, messages, tools):
    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        return {"error": "ANTHROPIC_API_KEY is not set on the server. Add it in Render → Environment → ANTHROPIC_API_KEY."}
    url = "https://api.anthropic.com/v1/messages"
    body = {
        "model": AGENT_MODEL,
        "max_tokens": AGENT_MAX_TOKENS,
        "system": system,
        "messages": messages,
        "tools": [{"name": t["name"], "description": t["description"], "input_schema": t["input_schema"]} for t in tools],
    }
    try:
        data = json.dumps(body).encode("utf-8")
    except Exception as e:
        print(f"[agent] payload serialization failed: {e}", file=sys.stderr, flush=True)
        return {"error": f"Could not serialize request: {e}"}
    req = urllib.request.Request(url, data=data, method="POST", headers={
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    })
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        try:
            err_body = e.read().decode("utf-8")
        except Exception:
            err_body = str(e)
        print(f"[agent] Anthropic HTTP {e.code}: {err_body[:1000]}", file=sys.stderr, flush=True)
        return {"error": f"Anthropic API error {e.code}: {err_body[:500]}"}
    except urllib.error.URLError as e:
        print(f"[agent] URLError: {e}", file=sys.stderr, flush=True)
        return {"error": f"Network error contacting Anthropic: {e}"}
    except Exception as e:
        print(f"[agent] {type(e).__name__}: {e}", file=sys.stderr, flush=True)
        return {"error": f"{type(e).__name__}: {e}"}


def handle_agent_chat(environ, start_response):
    try:
        return _handle_agent_chat_impl(environ, start_response)
    except Exception as exc:
        import traceback
        tb = traceback.format_exc()
        print(f"[agent] handler crashed: {tb}", file=sys.stderr, flush=True)
        return json_response(start_response, "500 Internal Server Error",
                             {"error": f"{type(exc).__name__}: {exc}"})


def _handle_agent_chat_impl(environ, start_response):
    actor = require_login(environ, start_response)
    if not actor:
        return []

    method = environ["REQUEST_METHOD"]
    convs = _agent_load_conversations()
    user_key = actor.get("id") or actor.get("email") or "anon"
    history = convs.get(user_key, [])
    # Hard cap loaded history so a long-running conversation doesn't drag
    # every new request down with 80 stale turns.
    if len(history) > AGENT_HISTORY_TURNS * 2:
        history = history[-AGENT_HISTORY_TURNS * 2:]

    if method == "GET":
        # Return history (for hydrating the chat panel on page load).
        params = parse_qs(environ.get("QUERY_STRING", ""))
        if (params.get("history", [""])[0] or "").strip() == "1":
            return json_response(start_response, "200 OK", {"messages": history[-AGENT_HISTORY_TURNS * 2:]})
        return json_response(start_response, "200 OK", {"ok": True})

    if method == "DELETE":
        convs[user_key] = []
        _agent_save_conversations(convs)
        return json_response(start_response, "200 OK", {"ok": True, "messages": []})

    if method != "POST":
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    payload = collect_json(environ) or {}
    user_msg = (payload.get("message") or "").strip()
    raw_images = payload.get("images") or []
    if not isinstance(raw_images, list):
        raw_images = []
    raw_images = raw_images[:AGENT_MAX_IMAGES_PER_MESSAGE]
    raw_docs = payload.get("documents") or []
    if not isinstance(raw_docs, list):
        raw_docs = []
    raw_docs = raw_docs[:AGENT_MAX_IMAGES_PER_MESSAGE]

    image_blocks = []
    doc_blocks = []
    extra_text_blocks = []  # for xlsx/text content extracted server-side
    saved_paths = []  # public paths for images (so agent can attach them later)
    saved_doc_paths = []  # public paths for documents

    for img in raw_images:
        data_url = (img.get("dataUrl") or "") if isinstance(img, dict) else ""
        if not data_url.startswith("data:"):
            continue
        try:
            header, b64 = data_url.split(";base64,", 1)
        except ValueError:
            continue
        media_type = header.replace("data:", "", 1).strip().lower() or "image/jpeg"
        if media_type not in ("image/jpeg", "image/png", "image/gif", "image/webp"):
            continue
        if len(b64) * 3 // 4 > AGENT_MAX_IMAGE_BYTES:
            continue
        image_blocks.append({
            "type": "image",
            "source": {"type": "base64", "media_type": media_type, "data": b64},
        })
        saved_paths.append(_agent_save_uploaded_image(b64, media_type))

    for doc in raw_docs:
        if not isinstance(doc, dict):
            continue
        data_url = doc.get("dataUrl") or ""
        name = (doc.get("name") or "").strip() or "document"
        if not data_url.startswith("data:"):
            continue
        try:
            header, b64 = data_url.split(";base64,", 1)
        except ValueError:
            continue
        media_type = header.replace("data:", "", 1).strip().lower()
        try:
            raw = base64.b64decode(b64)
        except Exception:
            continue
        if not raw or len(raw) > 30 * 1024 * 1024:  # 30 MB cap per document
            continue

        if media_type == "application/pdf" or name.lower().endswith(".pdf"):
            doc_blocks.append({
                "type": "document",
                "source": {"type": "base64", "media_type": "application/pdf", "data": b64},
                "title": name[:120],
            })
            saved_doc_paths.append(_agent_save_uploaded_doc(raw, name, "pdf"))
        elif name.lower().endswith(".xlsx") or media_type in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ):
            extracted = _agent_extract_xlsx_text(raw)
            extra_text_blocks.append(f"\n\n[Хавсаргасан Excel: {name}]\n{extracted}")
            saved_doc_paths.append(_agent_save_uploaded_doc(raw, name, "xlsx"))
        elif media_type.startswith("text/") or name.lower().endswith((".txt", ".csv", ".md")):
            try:
                txt = raw.decode("utf-8", errors="replace")[:20000]
            except Exception:
                txt = ""
            extra_text_blocks.append(f"\n\n[Хавсаргасан текст файл: {name}]\n{txt}")
            saved_doc_paths.append(_agent_save_uploaded_doc(raw, name, "txt"))
        else:
            # Unknown type — just save and tell agent the file name.
            ext = (name.rsplit(".", 1)[-1] if "." in name else "bin").lower()[:6]
            saved = _agent_save_uploaded_doc(raw, name, ext)
            saved_doc_paths.append(saved)
            extra_text_blocks.append(
                f"\n\n[Хавсаргасан файл: {name} — энэ форматыг шууд унших боломжгүй, зөвхөн файлын зам хадгалагдсан: {saved}]"
            )

    if not user_msg and not image_blocks and not doc_blocks and not extra_text_blocks:
        return json_response(start_response, "400 Bad Request", {"error": "Empty message"})

    # Build the user content blocks: images, then PDFs, then text. Always
    # include at least a text block so Claude has something to respond to.
    content_blocks = list(image_blocks) + list(doc_blocks)
    text_parts = [user_msg or "(файл хавсаргав — танилцана уу)"]
    upload_lines = [f"img-{i+1}={p}" for i, p in enumerate(saved_paths) if p]
    if upload_lines:
        text_parts.append("\n\n[Uploaded image paths] " + "; ".join(upload_lines))
    doc_lines = [f"doc-{i+1}={p}" for i, p in enumerate(saved_doc_paths) if p]
    if doc_lines:
        text_parts.append("\n\n[Uploaded document paths] " + "; ".join(doc_lines))
    text_parts.extend(extra_text_blocks)
    content_blocks.append({"type": "text", "text": "\n".join(text_parts)})
    history.append({"role": "user", "content": content_blocks})

    system = _agent_system_prompt(actor)
    actions = []  # human-readable summary of tool calls performed
    final_text = ""
    loops = 0

    import time
    loop_start = time.time()
    while loops < AGENT_MAX_TOOL_LOOPS:
        loops += 1
        # Strip image bytes from earlier turns before subsequent calls — Claude
        # has already extracted what it needs from the image in loop 1, and
        # re-uploading multi-MB base64 on every loop blows past Render's proxy
        # timeout. Keep the placeholder so the model knows an image was sent.
        if loops > 1:
            history = _strip_image_bytes_from_history(history)
        elapsed = time.time() - loop_start
        if elapsed > 80:
            print(f"[agent] aborting at loop {loops} after {elapsed:.1f}s to stay under proxy timeout", file=sys.stderr, flush=True)
            final_text = (final_text or "") + "\n\n⏱ Хэт удаан үргэлжиллээ. Үлдсэн алхмуудыг дараагийн мессежээр үргэлжлүүлээрэй."
            break
        t0 = time.time()
        resp = _agent_call_anthropic(system, history, AGENT_TOOLS)
        print(f"[agent] loop {loops} anthropic call took {time.time()-t0:.1f}s", file=sys.stderr, flush=True)
        if "error" in resp:
            history.pop()  # Don't persist the failed turn.
            return json_response(start_response, "502 Bad Gateway", {"error": resp["error"]})

        assistant_blocks = resp.get("content") or []
        # Append assistant turn to history (full content, including tool_use blocks).
        history.append({"role": "assistant", "content": assistant_blocks})

        tool_uses = [b for b in assistant_blocks if b.get("type") == "tool_use"]
        for b in assistant_blocks:
            if b.get("type") == "text" and b.get("text"):
                final_text += b["text"]

        if not tool_uses or resp.get("stop_reason") == "end_turn":
            break

        # Execute every tool_use, then feed the results back as a single user turn.
        results_blocks = []
        for tu in tool_uses:
            name = tu.get("name") or ""
            tool = AGENT_TOOL_BY_NAME.get(name)
            tool_input = tu.get("input") or {}
            if not tool:
                out = {"error": f"Unknown tool: {name}"}
                ok = False
            else:
                try:
                    # Tool dispatch uses the Bataa actor — every record this
                    # tool stamps (createdBy) and every notification it logs
                    # shows "Бата" with the panda avatar instead of the human
                    # admin who happened to be chatting. The audit log still
                    # carries the human actor so we can trace who asked.
                    out = tool["handler"](tool_input, _bataa_agent_actor(actor))
                    ok = "error" not in (out if isinstance(out, dict) else {})
                except Exception as exc:
                    out = {"error": f"{type(exc).__name__}: {exc}"}
                    ok = False
            _agent_audit(actor, name, tool_input, ok, out if isinstance(out, (str, dict)) else "ok")
            actions.append({"tool": name, "ok": ok, "input": tool_input, "summary": _summarize_tool_output(name, out)})
            results_blocks.append({
                "type": "tool_result",
                "tool_use_id": tu.get("id"),
                "content": json.dumps(out, ensure_ascii=False)[:8000],
                "is_error": (not ok),
            })
        history.append({"role": "user", "content": results_blocks})

    # Trim history to keep it bounded but preserve assistant/tool pairs.
    if len(history) > AGENT_HISTORY_TURNS * 4:
        history = history[-AGENT_HISTORY_TURNS * 4:]
    # Strip base64 image bytes before persisting — keep a placeholder so the
    # assistant's later turns still know an image was attached, but don't bloat
    # the JSON file with megabytes of base64.
    persistable = []
    for turn in history:
        content = turn.get("content")
        if isinstance(content, list):
            new_content = []
            for block in content:
                if isinstance(block, dict) and block.get("type") == "image":
                    new_content.append({"type": "text", "text": "[image attached, omitted from history]"})
                else:
                    new_content.append(block)
            persistable.append({**turn, "content": new_content})
        else:
            persistable.append(turn)
    convs[user_key] = persistable
    _agent_save_conversations(convs)

    return json_response(start_response, "200 OK", {
        "ok": True,
        "reply": final_text or "(done)",
        "actions": actions,
    })


def _summarize_tool_output(name, out):
    if isinstance(out, dict):
        if "error" in out:
            return f"error: {out['error']}"
        if "count" in out:
            return f"{out['count']} item(s)"
        if "ok" in out:
            extras = []
            for k in ("trip", "group", "tourist", "invoice", "installment"):
                if k in out and isinstance(out[k], dict):
                    s = out[k].get("serial") or out[k].get("id") or ""
                    if s:
                        extras.append(f"{k}={s}")
            return "✓ " + (" ".join(extras) if extras else "done")
    if isinstance(out, list):
        return f"{len(out)} item(s)"
    return "done"


def app(environ, start_response):
    try:
        return _dispatch(environ, start_response)
    except Exception as exc:
        try:
            print(f"[server] unhandled error on {environ.get('REQUEST_METHOD')} {environ.get('PATH_INFO')}: {exc}", file=sys.stderr, flush=True)
            import traceback as _tb
            _tb.print_exc()
        except Exception:
            pass
        try:
            return json_response(start_response, "500 Internal Server Error", {"error": f"Server error: {type(exc).__name__}: {str(exc)[:200]}"})
        except Exception:
            start_response("500 Internal Server Error", [("Content-Type", "text/plain; charset=utf-8")])
            return [b"Server error"]


def _dispatch(environ, start_response):
    ensure_data_store()

    method = environ["REQUEST_METHOD"]
    path = environ.get("PATH_INFO", "/")
    host = request_host(environ)

    if method == "GET" and path == "/health":
        return text_response(start_response, "200 OK", "ok")

    # Per-store probe: which JSON file (if any) is unreadable on the
    # persistent disk. Not authenticated — only reports file size + parse
    # status, never contents. Pass ?heal=1 to also push each file through
    # read_json_list so the auto-recovery path runs (peels off trailing
    # garbage from a concurrent-writer collision and rewrites the clean copy).
    if method == "GET" and path == "/api/debug/storage":
        params = parse_qs(environ.get("QUERY_STRING", ""))
        heal = (params.get("heal", [""])[0] or "") == "1"
        targets = [
            ("camp_trips", CAMP_TRIPS_FILE),
            ("tourists", TOURISTS_FILE),
            ("tourist_groups", GROUPS_FILE),
            ("invoices", INVOICES_FILE),
            ("camp_reservations", CAMP_RESERVATIONS_FILE),
            ("flight_reservations", FLIGHT_RESERVATIONS_FILE),
            ("transfer_reservations", TRANSFER_RESERVATIONS_FILE),
            ("contracts", CONTRACTS_FILE),
            ("ds160", DS160_FILE),
            ("notifications", NOTIFICATIONS_FILE),
            ("mail_followups", MAIL_FOLLOWUPS_FILE),
            ("mail_dossiers", MAIL_DOSSIERS_FILE),
        ]
        report = []
        for name, path_obj in targets:
            row = {"name": name, "path": str(path_obj)}
            if heal:
                try:
                    healed = read_json_list(path_obj)
                    row["healed_count"] = len(healed)
                except Exception as exc:
                    row["heal_error"] = f"{type(exc).__name__}: {str(exc)[:120]}"
            try:
                if not path_obj.exists():
                    row["status"] = "missing"
                else:
                    raw = path_obj.read_text(encoding="utf-8")
                    row["bytes"] = len(raw)
                    try:
                        parsed = json.loads(raw) if raw.strip() else []
                        row["status"] = "ok" if isinstance(parsed, list) else f"not-a-list ({type(parsed).__name__})"
                        row["count"] = len(parsed) if isinstance(parsed, list) else 0
                    except Exception as parse_exc:
                        row["status"] = f"parse-error: {type(parse_exc).__name__}: {str(parse_exc)[:120]}"
            except Exception as exc:
                row["status"] = f"read-error: {type(exc).__name__}: {str(exc)[:120]}"
            report.append(row)
        return json_response(start_response, "200 OK", {"data_dir": str(DATA_DIR), "files": report})

    if path == "/api/auth/register" and method == "POST":
        return handle_auth_register(environ, start_response)

    if path == "/api/auth/bootstrap-admin" and method == "POST":
        return handle_auth_bootstrap_admin(environ, start_response)

    if path == "/api/auth/login" and method == "POST":
        return handle_auth_login(environ, start_response)

    if path == "/api/auth/request-reset" and method == "POST":
        return handle_auth_request_reset(environ, start_response)

    if path == "/api/auth/logout" and method == "POST":
        return handle_auth_logout(environ, start_response)

    if path == "/api/auth/me" and method == "GET":
        return handle_auth_me(environ, start_response)

    if path == "/api/auth/profile" and method == "POST":
        return handle_auth_profile_update(environ, start_response)

    if path == "/api/users":
        if method == "GET":
            return handle_list_users(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/team-members":
        if method == "GET":
            return handle_list_team_members(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/announcements/active":
        if method == "GET":
            return handle_list_announcements_active(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/announcements":
        if method == "GET":
            return handle_list_announcements(environ, start_response)
        if method == "POST":
            return handle_create_announcement(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/announcements/") and path.endswith("/dismiss"):
        announcement_id = path.replace("/api/announcements/", "", 1).rsplit("/dismiss", 1)[0].strip("/")
        if method == "POST" and announcement_id:
            return handle_dismiss_announcement(environ, start_response, announcement_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/announcements/") and path.endswith("/archive"):
        announcement_id = path.replace("/api/announcements/", "", 1).rsplit("/archive", 1)[0].strip("/")
        if method == "POST" and announcement_id:
            return handle_archive_announcement(environ, start_response, announcement_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/announcements/") and path.endswith("/attachment"):
        announcement_id = path.replace("/api/announcements/", "", 1).rsplit("/attachment", 1)[0].strip("/")
        if method == "GET" and announcement_id:
            return handle_download_announcement_attachment(environ, start_response, announcement_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/trip-creators/"):
        trip_id = path.replace("/api/trip-creators/", "", 1).strip("/")
        if not trip_id:
            return json_response(start_response, "400 Bad Request", {"error": "tripId required"})
        if method == "GET":
            return handle_get_trip_creator(environ, start_response, trip_id)
        if method == "POST":
            return handle_save_trip_creator(environ, start_response, trip_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    # Public read-only brochure — clients hit this from /trip/<id>.
    if path.startswith("/api/public/trips/"):
        trip_id = path.replace("/api/public/trips/", "", 1).strip("/")
        if method == "GET" and trip_id:
            return handle_get_public_trip_creator(environ, start_response, trip_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    # Public content lookup by slug, used by trip-public popups.
    if path.startswith("/api/public/content/"):
        slug = path.replace("/api/public/content/", "", 1).strip("/")
        if method == "GET" and slug:
            return handle_get_public_content(environ, start_response, slug)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    # Gallery (image upload + listing). Image bytes are public via /file
    # because trip brochures embed them.
    if path == "/api/gallery":
        if method == "GET":
            return handle_list_gallery(environ, start_response)
        if method == "POST":
            return handle_upload_gallery_image(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})
    if path == "/api/gallery/video":
        if method == "POST":
            return handle_create_gallery_video(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})
    if path == "/api/gallery/folders":
        if method == "POST":
            return handle_create_gallery_folder(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})
    if path.startswith("/api/gallery/folders/") and path.endswith("/rename"):
        old_name = path.replace("/api/gallery/folders/", "", 1).rsplit("/rename", 1)[0]
        old_name = unquote(old_name)
        if method == "POST" and old_name:
            return handle_rename_gallery_folder(environ, start_response, old_name)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})
    if path.startswith("/api/gallery/folders/"):
        name = unquote(path.replace("/api/gallery/folders/", "", 1).strip("/"))
        if method == "DELETE" and name:
            return handle_delete_gallery_folder(environ, start_response, name)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})
    if path.startswith("/api/gallery/") and path.endswith("/file"):
        item_id = path.replace("/api/gallery/", "", 1).rsplit("/file", 1)[0].strip("/")
        if method == "GET" and item_id:
            return handle_serve_gallery_image(environ, start_response, item_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})
    if path.startswith("/api/gallery/"):
        item_id = path.replace("/api/gallery/", "", 1).strip("/")
        if method == "DELETE" and item_id:
            return handle_delete_gallery_item(environ, start_response, item_id)
        if method == "POST" and item_id:
            return handle_update_gallery_item(environ, start_response, item_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    # Content library CRUD.
    if path == "/api/content":
        if method == "GET":
            return handle_list_content(environ, start_response)
        if method == "POST":
            return handle_create_content(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})
    if path.startswith("/api/content/") and path.endswith("/video"):
        item_id = path.replace("/api/content/", "", 1).replace("/video", "", 1).strip("/")
        if not item_id:
            return json_response(start_response, "400 Bad Request", {"error": "id required"})
        if method == "POST":
            return handle_upload_content_video(environ, start_response, item_id)
        if method == "DELETE":
            return handle_delete_content_video(environ, start_response, item_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})
    if path.startswith("/api/content/"):
        item_id = path.replace("/api/content/", "", 1).strip("/")
        if not item_id:
            return json_response(start_response, "400 Bad Request", {"error": "id required"})
        if method == "GET":
            return handle_get_content(environ, start_response, item_id)
        if method == "POST":
            return handle_update_content(environ, start_response, item_id)
        if method == "DELETE":
            return handle_delete_content(environ, start_response, item_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/meal-templates":
        if method == "GET":
            return handle_list_meal_templates(environ, start_response)
        if method == "POST":
            return handle_create_meal_template(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})
    if path.startswith("/api/meal-templates/"):
        item_id = path.replace("/api/meal-templates/", "", 1).strip("/")
        if not item_id:
            return json_response(start_response, "400 Bad Request", {"error": "id required"})
        if method == "POST":
            return handle_update_meal_template(environ, start_response, item_id)
        if method == "DELETE":
            return handle_delete_meal_template(environ, start_response, item_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/locations":
        if method == "GET":
            return handle_list_locations(environ, start_response)
        if method == "POST":
            return handle_create_location(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})
    if path.startswith("/api/locations/"):
        item_id = path.replace("/api/locations/", "", 1).strip("/")
        if not item_id:
            return json_response(start_response, "400 Bad Request", {"error": "id required"})
        if method == "GET":
            return handle_get_location(environ, start_response, item_id)
        if method == "POST":
            return handle_update_location(environ, start_response, item_id)
        if method == "DELETE":
            return handle_delete_location(environ, start_response, item_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/notifications":
        if method == "GET":
            return handle_list_notifications(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/notifications/read":
        if method == "POST":
            return handle_mark_notifications_read(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/agent/chat":
        if method in ("GET", "POST", "DELETE"):
            return handle_agent_chat(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/tourist-groups":
        if method == "GET":
            return handle_list_tourist_groups(environ, start_response)
        if method == "POST":
            return handle_create_tourist_group(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/tourist-groups/"):
        group_id = path.replace("/api/tourist-groups/", "", 1).strip("/")
        if method == "POST" and group_id:
            return handle_update_tourist_group(environ, start_response, group_id)
        if method == "DELETE" and group_id:
            return handle_delete_tourist_group(environ, start_response, group_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/tourists":
        if method == "GET":
            return handle_list_tourists(environ, start_response)
        if method == "POST":
            return handle_create_tourist(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/tourists/email-rooming":
        if method == "POST":
            return handle_email_rooming(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})
    if path == "/api/tourists/export":
        if method == "POST":
            return handle_export_tourists(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/tourists/passport-scan":
        if method == "POST":
            return handle_passport_scan(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/tourists/promo-email":
        if method == "POST":
            return handle_promo_email(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/invoices":
        if method == "GET":
            return handle_list_invoices(environ, start_response)
        if method == "POST":
            return handle_create_invoice(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/documents":
        if method == "GET":
            return handle_list_documents(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/invoices/") and path.endswith("/publish"):
        invoice_id = path.replace("/api/invoices/", "", 1).replace("/publish", "", 1).strip("/")
        if method == "POST" and invoice_id:
            return handle_publish_invoice(environ, start_response, invoice_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/invoices/") and path.endswith("/payment"):
        invoice_id = path.replace("/api/invoices/", "", 1).replace("/payment", "", 1).strip("/")
        if method == "POST" and invoice_id:
            return handle_invoice_payment(environ, start_response, invoice_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/invoices/") and path.endswith("/pdf"):
        invoice_id = path.replace("/api/invoices/", "", 1).replace("/pdf", "", 1).strip("/")
        if method == "GET" and invoice_id:
            return handle_standalone_invoice_pdf(environ, start_response, invoice_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/invoices/"):
        invoice_id = path.replace("/api/invoices/", "", 1).strip("/")
        if method == "POST" and invoice_id:
            return handle_update_invoice(environ, start_response, invoice_id)
        if method == "DELETE" and invoice_id:
            return handle_delete_invoice(environ, start_response, invoice_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/payment-requests":
        if method == "GET":
            return handle_list_payment_requests(environ, start_response)
        if method == "POST":
            return handle_create_payment_request(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/payment-requests/count":
        if method == "GET":
            return handle_payment_request_count(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/trip-templates":
        if method == "GET":
            return handle_list_trip_templates(environ, start_response)
        if method == "POST":
            return handle_create_trip_template(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/trip-templates/"):
        template_id = path.replace("/api/trip-templates/", "", 1).strip("/")
        if method == "POST" and template_id:
            return handle_update_trip_template(environ, start_response, template_id)
        if method == "DELETE" and template_id:
            return handle_delete_trip_template(environ, start_response, template_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/service-templates":
        if method == "GET":
            return handle_list_service_templates(environ, start_response)
        if method == "POST":
            return handle_create_service_template(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/service-templates/"):
        template_id = path.replace("/api/service-templates/", "", 1).strip("/")
        if method == "POST" and template_id:
            return handle_update_service_template(environ, start_response, template_id)
        if method == "DELETE" and template_id:
            return handle_delete_service_template(environ, start_response, template_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/translate-prose":
        if method == "POST":
            return handle_translate_prose(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/accountant/paid":
        if method == "GET":
            return handle_list_accountant_paid(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/accountant/download-zip":
        if method == "POST":
            return handle_accountant_download_zip(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/payment-requests/") and path.endswith("/approve"):
        request_id = path.replace("/api/payment-requests/", "", 1).replace("/approve", "", 1).strip("/")
        if method == "POST" and request_id:
            return handle_approve_payment_request(environ, start_response, request_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/payment-requests/") and path.endswith("/document"):
        request_id = path.replace("/api/payment-requests/", "", 1).replace("/document", "", 1).strip("/")
        if method == "POST" and request_id:
            return handle_attach_payment_document(environ, start_response, request_id)
        if method == "DELETE" and request_id:
            return handle_delete_payment_document(environ, start_response, request_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    # /api/invoices/<id>/installments/<idx>/document — backfill a
    # receipt onto a paid installment from the invoice side panel.
    if "/installments/" in path and path.endswith("/document") and path.startswith("/api/invoices/"):
        rest = path.replace("/api/invoices/", "", 1).replace("/document", "", 1).strip("/")
        try:
            invoice_id, _, idx_str = rest.partition("/installments/")
            inst_index = int(idx_str)
        except (ValueError, AttributeError):
            return json_response(start_response, "400 Bad Request", {"error": "Bad path"})
        if method == "POST" and invoice_id:
            return handle_attach_invoice_receipt(environ, start_response, invoice_id, inst_index)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/payment-requests/") and path.endswith("/reject"):
        request_id = path.replace("/api/payment-requests/", "", 1).replace("/reject", "", 1).strip("/")
        if method == "POST" and request_id:
            return handle_reject_payment_request(environ, start_response, request_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/payment-requests/"):
        request_id = path.replace("/api/payment-requests/", "", 1).strip("/")
        if method == "DELETE" and request_id:
            return handle_delete_payment_request(environ, start_response, request_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/tourists/"):
        tourist_id = path.replace("/api/tourists/", "", 1).strip("/")
        if method == "POST" and tourist_id:
            return handle_update_tourist(environ, start_response, tourist_id)
        if method == "DELETE" and tourist_id:
            return handle_delete_tourist(environ, start_response, tourist_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/users/"):
        user_id = path.replace("/api/users/", "", 1).strip("/")
        if method == "POST" and user_id:
            return handle_update_user(environ, start_response, user_id)
        if method == "DELETE" and user_id:
            return handle_delete_user(environ, start_response, user_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/mail/accounts":
        if method == "GET":
            return handle_list_mail_accounts(environ, start_response)
        if method == "POST":
            return handle_create_mail_account(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/mail/accounts/"):
        tail = path.replace("/api/mail/accounts/", "", 1).strip("/")
        if tail.endswith("/test"):
            account_id = tail[:-len("/test")]
            if method == "POST":
                return handle_test_mail_account(environ, start_response, account_id)
            return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})
        account_id = tail
        if method == "PATCH" and account_id:
            return handle_update_mail_account(environ, start_response, account_id)
        if method == "DELETE" and account_id:
            return handle_delete_mail_account(environ, start_response, account_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/mail/messages":
        if method == "GET":
            return handle_list_mail_messages(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/mail/unread-summary":
        if method == "GET":
            return handle_mail_unread_summary(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/mail/followups":
        if method == "GET":
            return handle_list_mail_followups(environ, start_response)
        if method == "POST":
            return handle_create_mail_followup(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/mail/followups/"):
        followup_id = path.replace("/api/mail/followups/", "", 1).strip("/")
        if followup_id and method in ("POST", "PATCH", "DELETE"):
            return handle_update_mail_followup(environ, start_response, followup_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/mail/dossiers":
        if method == "GET":
            return handle_list_mail_dossiers(environ, start_response)
        if method == "POST":
            return handle_create_mail_dossier(environ, start_response)
        if method == "DELETE":
            return handle_delete_mail_dossier(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/mail/dossier-targets":
        if method == "GET":
            return handle_mail_dossier_targets(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/notes":
        if method == "GET":
            return handle_list_notes(environ, start_response)
        if method == "POST":
            return handle_create_note(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/notes/"):
        note_id = path.replace("/api/notes/", "", 1).strip("/")
        if note_id and method in {"POST", "PATCH", "DELETE"}:
            return handle_update_note(environ, start_response, note_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/reminders":
        if method == "GET":
            return handle_list_reminders(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/reminders/mark-read":
        if method == "POST":
            return handle_mark_reminders_read(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/mail/send":
        if method == "POST":
            return handle_send_mail(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/mail/signature-image":
        if method == "POST":
            return handle_upload_signature_image(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/mail-uploads/"):
        if method == "GET":
            filename = path.replace("/mail-uploads/", "", 1).strip("/")
            return handle_serve_mail_upload(environ, start_response, filename)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/mail/templates":
        if method == "GET":
            return handle_list_mail_templates(environ, start_response)
        if method == "POST":
            return handle_create_mail_template(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/mail/templates/"):
        template_id = path.replace("/api/mail/templates/", "", 1).strip("/")
        if method == "PATCH" and template_id:
            return handle_update_mail_template(environ, start_response, template_id)
        if method == "DELETE" and template_id:
            return handle_delete_mail_template(environ, start_response, template_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/mail/my-signatures":
        if method == "GET":
            return handle_list_my_signatures(environ, start_response)
        if method == "POST":
            return handle_create_my_signature(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/mail/my-signatures/"):
        sig_id = path.replace("/api/mail/my-signatures/", "", 1).strip("/")
        if method == "PATCH" and sig_id:
            return handle_update_my_signature(environ, start_response, sig_id)
        if method == "DELETE" and sig_id:
            return handle_delete_my_signature(environ, start_response, sig_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/mail/messages/"):
        tail = path.replace("/api/mail/messages/", "", 1).strip("/")
        if "/" in tail:
            parts = tail.split("/")
            acc_id = parts[0]
            uid = parts[1]
            sub = parts[2] if len(parts) > 2 else ""
            sub_arg = parts[3] if len(parts) > 3 else ""
            if sub == "read" and method == "POST":
                return handle_mark_read_mail_message(environ, start_response, acc_id, uid)
            if sub == "unread" and method == "POST":
                return handle_mark_unread_mail_message(environ, start_response, acc_id, uid)
            if sub == "attachments" and sub_arg and method == "GET":
                return handle_download_mail_attachment(environ, start_response, acc_id, uid, sub_arg)
            if sub == "thread" and method == "GET":
                return handle_mail_thread(environ, start_response, acc_id, uid)
            if not sub and method == "GET":
                return handle_get_mail_message(environ, start_response, acc_id, uid)
            if not sub and method == "DELETE":
                return handle_delete_mail_message(environ, start_response, acc_id, uid)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/backoffice/summary":
        if method == "GET":
            if not require_login(environ, start_response):
                return []
            return handle_dashboard_summary(start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/manager-dashboard":
        if method == "GET":
            if not require_login(environ, start_response):
                return []
            return handle_get_manager_dashboard(start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/manager-dashboard/tasks":
        if method == "POST":
            return handle_manager_item_create(environ, start_response, "tasks", build_manager_task, validate_manager_task, "task")
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/manager-dashboard/tasks/"):
        rest = path.replace("/api/manager-dashboard/tasks/", "", 1).strip("/")
        # /tasks/<id>/image[/imgId] — multiple throwaway reference images per task.
        if "/image" in rest:
            head, _, tail = rest.partition("/image")
            item_id = head.strip("/")
            image_id = tail.strip("/") or None
            if not item_id:
                return json_response(start_response, "400 Bad Request", {"error": "id required"})
            if method == "POST":
                return handle_task_image_upload(environ, start_response, item_id)
            if method == "GET":
                return handle_task_image_serve(environ, start_response, item_id, image_id)
            if method == "DELETE":
                return handle_task_image_delete(environ, start_response, item_id, image_id)
            return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})
        item_id = rest
        if method == "POST" and item_id:
            return handle_manager_item_update(environ, start_response, "tasks", item_id, build_manager_task, validate_manager_task, "task")
        if method == "DELETE" and item_id:
            return handle_manager_item_delete(environ, start_response, "tasks", item_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/manager-dashboard/reminders":
        if method == "POST":
            return handle_manager_item_create(
                environ,
                start_response,
                "reminders",
                build_manager_reminder,
                validate_manager_reminder,
                "reminder",
            )
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/manager-dashboard/reminders/"):
        item_id = path.replace("/api/manager-dashboard/reminders/", "", 1).strip("/")
        if method == "POST" and item_id:
            return handle_manager_item_update(
                environ,
                start_response,
                "reminders",
                item_id,
                build_manager_reminder,
                validate_manager_reminder,
                "reminder",
            )
        if method == "DELETE" and item_id:
            return handle_manager_item_delete(environ, start_response, "reminders", item_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/manager-dashboard/contacts":
        if method == "POST":
            return handle_manager_item_create(
                environ,
                start_response,
                "contacts",
                build_manager_contact,
                validate_manager_contact,
                "contact",
            )
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/manager-dashboard/contacts/"):
        item_id = path.replace("/api/manager-dashboard/contacts/", "", 1).strip("/")
        if method == "POST" and item_id:
            return handle_manager_item_update(
                environ,
                start_response,
                "contacts",
                item_id,
                build_manager_contact,
                validate_manager_contact,
                "contact",
            )
        if method == "DELETE" and item_id:
            return handle_manager_item_delete(environ, start_response, "contacts", item_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/backups":
        if method == "GET":
            return handle_list_backups(environ, start_response)
        if method == "POST":
            return handle_create_backup(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/backups/") and method == "GET":
        filename = path.replace("/api/backups/", "", 1).strip("/")
        if filename:
            return handle_download_backup(environ, start_response, filename)
        return json_response(start_response, "404 Not Found", {"error": "Backup not found"})

    if path == "/api/contract-templates":
        if method == "GET":
            return handle_list_contract_templates(environ, start_response)
        if method == "POST":
            return handle_create_contract_template(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/contract-templates/default":
        if method == "GET":
            return handle_get_default_contract_template(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/contract-templates/"):
        template_id = path.replace("/api/contract-templates/", "", 1).strip("/")
        if method == "GET" and template_id:
            return handle_get_contract_template(environ, start_response, template_id)
        if method == "POST" and template_id:
            return handle_update_contract_template(environ, start_response, template_id)
        if method == "DELETE" and template_id:
            return handle_delete_contract_template(environ, start_response, template_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/contracts":
        if method == "GET":
            if not require_login(environ, start_response):
                return []
            return handle_list_contracts(environ, start_response)
        if method == "POST":
            if not require_login(environ, start_response):
                return []
            return handle_generate_contract(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/contracts/"):
        tail = path.replace("/api/contracts/", "", 1).strip("/")
        if not tail:
            return json_response(start_response, "404 Not Found", {"error": "Contract not found"})
        if tail.endswith("/sign"):
            contract_id = tail.replace("/sign", "", 1).strip("/")
            if method == "POST":
                return handle_sign_contract(environ, start_response, contract_id)
            return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})
        if tail.endswith("/invite"):
            contract_id = tail.replace("/invite", "", 1).strip("/")
            if method == "POST":
                return handle_send_contract_invitation(environ, start_response, contract_id)
            return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})
        if tail.endswith("/document"):
            contract_id = tail.replace("/document", "", 1).strip("/")
            if method == "GET":
                return handle_contract_document(environ, start_response, contract_id)
            return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})
        if tail.endswith("/invoice"):
            contract_id = tail.replace("/invoice", "", 1).strip("/")
            if method == "GET":
                return handle_contract_invoice_document(environ, start_response, contract_id)
            if method == "POST":
                return handle_update_contract_invoice(environ, start_response, contract_id)
            return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})
        contract_id = tail
        if method == "GET":
            return handle_get_contract(start_response, contract_id)
        if method == "POST":
            if not require_login(environ, start_response):
                return []
            return handle_update_contract(environ, start_response, contract_id)
        if method == "DELETE":
            if not require_login(environ, start_response):
                return []
            return handle_delete_contract(environ, start_response, contract_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/ds160":
        if method == "GET":
            return handle_list_ds160(environ, start_response)
        if method == "POST":
            return handle_create_ds160(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/ds160/invitations":
        if method == "POST":
            return handle_create_ds160_invitation(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/ds160/public/"):
        client_token = path.replace("/api/ds160/public/", "", 1).strip("/")
        if not client_token:
            return json_response(start_response, "404 Not Found", {"error": "DS-160 form not found"})
        if method == "GET":
            return handle_public_get_ds160(start_response, client_token)
        if method == "POST":
            return handle_public_submit_ds160(environ, start_response, client_token)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/ds160/"):
        record_id = path.replace("/api/ds160/", "", 1).strip("/")
        # /api/ds160/<id>/send — server-side email send (Resend) so a manager
        # without a desktop mail client can still deliver the share link.
        if record_id.endswith("/send") and method == "POST":
            inner_id = record_id[: -len("/send")]
            return handle_send_ds160_invitation(environ, start_response, inner_id)
        if method == "GET" and record_id:
            return handle_get_ds160(environ, start_response, record_id)
        if method in {"PATCH", "PUT"} and record_id:
            return handle_update_ds160(environ, start_response, record_id)
        if method == "DELETE" and record_id:
            return handle_delete_ds160(environ, start_response, record_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/finance":
        if method == "GET":
            if not require_login(environ, start_response):
                return []
            return handle_list_finance(start_response)
        if method == "POST":
            if not require_login(environ, start_response):
                return []
            return handle_create_finance(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/bookings":
        if method == "GET":
            if not require_login(environ, start_response):
                return []
            return handle_list_bookings(start_response)
        if method == "POST":
            if not require_login(environ, start_response):
                return []
            return handle_create_booking(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/reservations":
        if method == "GET":
            if not require_login(environ, start_response):
                return []
            return handle_list_reservations(start_response)
        if method == "POST":
            if not require_login(environ, start_response):
                return []
            return handle_create_reservation(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/camp-reservations":
        if method == "GET":
            if not require_login(environ, start_response):
                return []
            return handle_list_camp_reservations(environ, start_response)
        if method == "POST":
            if not require_login(environ, start_response):
                return []
            return handle_create_camp_reservation(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/flight-reservations":
        if method == "GET":
            if not require_login(environ, start_response):
                return []
            return handle_list_flight_reservations(environ, start_response)
        if method == "POST":
            if not require_login(environ, start_response):
                return []
            return handle_create_flight_reservation(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/transfer-reservations":
        if method == "GET":
            if not require_login(environ, start_response):
                return []
            return handle_list_transfer_reservations(environ, start_response)
        if method == "POST":
            if not require_login(environ, start_response):
                return []
            return handle_create_transfer_reservation(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/camp-trips":
        if method == "GET":
            if not require_login(environ, start_response):
                return []
            return handle_list_camp_trips(environ, start_response)
        if method == "POST":
            if not require_login(environ, start_response):
                return []
            return handle_create_camp_trip(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/camp-settings":
        if method == "GET":
            if not require_login(environ, start_response):
                return []
            return handle_list_camp_settings(start_response)
        if method == "POST":
            return handle_update_camp_settings(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/camp-settings/contract":
        if method == "POST":
            return handle_upload_camp_contract(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/settings":
        if method == "GET":
            return handle_get_settings(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/settings/destinations":
        if method == "POST":
            return handle_update_settings_destinations(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/settings/bank-accounts":
        if method == "POST":
            return handle_update_settings_bank_accounts(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/fifa2026":
        if method == "GET":
            if not require_login(environ, start_response):
                return []
            return handle_get_fifa2026_dashboard(start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/fifa2026/public":
        if method == "GET":
            return handle_get_fifa2026_public(start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/fifa2026/reset-from-seed":
        if method == "POST":
            return handle_reset_fifa2026_from_seed(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/fifa2026/tickets":
        if method == "POST":
            return handle_create_fifa_ticket(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/fifa2026/sales":
        if method == "POST":
            return handle_create_fifa_sale(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/fifa2026/tickets/"):
        ticket_id = path.replace("/api/fifa2026/tickets/", "", 1).strip("/")
        if method == "POST" and ticket_id:
            return handle_update_fifa_ticket(environ, start_response, ticket_id)
        if method == "DELETE" and ticket_id:
            return handle_delete_fifa_ticket(environ, start_response, ticket_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/fifa2026/sales/"):
        sale_id = path.replace("/api/fifa2026/sales/", "", 1).strip("/")
        if method == "POST" and sale_id:
            return handle_update_fifa_sale(environ, start_response, sale_id)
        if method == "DELETE" and sale_id:
            return handle_delete_fifa_sale(environ, start_response, sale_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/camp-payment-groups/") and path.endswith("/invoice"):
        group_key = path.replace("/api/camp-payment-groups/", "", 1).replace("/invoice", "", 1).strip("/")
        if method == "POST" and group_key:
            return handle_upload_camp_group_invoice(environ, start_response, group_key)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path == "/api/camp-reservations/export":
        if method == "GET":
            return handle_export_camp_reservations(environ, start_response)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/camp-reservations/"):
        reservation_id = path.replace("/api/camp-reservations/", "", 1).strip("/")
        if method == "GET" and reservation_id.endswith("/document"):
            reservation_id = reservation_id[:-len("/document")].strip("/")
            return handle_camp_reservation_document(environ, start_response, reservation_id)
        if method == "POST" and reservation_id:
            if not require_login(environ, start_response):
                return []
            return handle_update_camp_reservation(environ, start_response, reservation_id)
        if method == "DELETE" and reservation_id:
            return handle_delete_camp_reservation(environ, start_response, reservation_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/flight-reservations/"):
        reservation_id = path.replace("/api/flight-reservations/", "", 1).strip("/")
        if method == "POST" and reservation_id:
            if not require_login(environ, start_response):
                return []
            return handle_update_flight_reservation(environ, start_response, reservation_id)
        if method == "DELETE" and reservation_id:
            return handle_delete_flight_reservation(environ, start_response, reservation_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/transfer-reservations/"):
        reservation_id = path.replace("/api/transfer-reservations/", "", 1).strip("/")
        if method == "POST" and reservation_id:
            if not require_login(environ, start_response):
                return []
            return handle_update_transfer_reservation(environ, start_response, reservation_id)
        if method == "DELETE" and reservation_id:
            return handle_delete_transfer_reservation(environ, start_response, reservation_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if path.startswith("/api/camp-trips/"):
        tail = path.replace("/api/camp-trips/", "", 1).strip("/")
        # /api/camp-trips/{id}/info — workspace-agnostic single-trip lookup,
        # used so a notification opened in the wrong workspace can prompt the
        # user to switch instead of misreporting the trip as deleted.
        if tail.endswith("/info") and method == "GET":
            trip_id = tail[: -len("/info")]
            actor = require_login(environ, start_response)
            if not actor:
                return []
            for t in read_camp_trips():
                if t.get("id") == trip_id:
                    return json_response(start_response, "200 OK", {
                        "id": t.get("id"),
                        "company": (t.get("company") or "").upper(),
                        "tripName": t.get("tripName") or "",
                        "serial": t.get("serial") or "",
                        "startDate": t.get("startDate") or "",
                        "tripType": t.get("tripType") or "",
                    })
            return json_response(start_response, "404 Not Found", {"error": "Trip not found"})
        # /api/camp-trips/{id}/documents  — upload
        if tail.endswith("/documents") and method == "POST":
            trip_id = tail[: -len("/documents")]
            return handle_upload_trip_document(environ, start_response, trip_id)
        # /api/camp-trips/{id}/documents/email  — bulk email selected docs to a client
        if tail.endswith("/documents/email") and method == "POST":
            trip_id = tail[: -len("/documents/email")]
            return handle_email_trip_documents(environ, start_response, trip_id)
        # /api/camp-trips/{id}/documents/{doc_id}  — delete or rename
        if "/documents/" in tail and method == "DELETE":
            trip_id, doc_id = tail.split("/documents/", 1)
            return handle_delete_trip_document(environ, start_response, trip_id, doc_id)
        if "/documents/" in tail and method == "PATCH":
            trip_id, doc_id = tail.split("/documents/", 1)
            return handle_rename_trip_document(environ, start_response, trip_id, doc_id)
        trip_id = tail
        if method == "POST" and trip_id:
            return handle_update_camp_trip(environ, start_response, trip_id)
        if method == "DELETE" and trip_id:
            return handle_delete_camp_trip(environ, start_response, trip_id)
        return json_response(start_response, "405 Method Not Allowed", {"error": "Method not allowed"})

    if method != "GET":
        return json_response(start_response, "404 Not Found", {"error": "Not found"})

    if path == "/login":
        if current_user(environ):
            if not active_workspace(environ):
                start_response("302 Found", [("Location", "/workspace")])
                return [b""]
            return file_response(start_response, PUBLIC_DIR / "backoffice.html")
        return file_response(start_response, PUBLIC_DIR / "login.html")

    if path == "/workspace":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        return file_response(start_response, PUBLIC_DIR / "workspace.html")

    if path == "/":
        if host == "camp.travelx.mn":
            if not current_user(environ):
                return file_response(start_response, PUBLIC_DIR / "login.html")
            return file_response(start_response, PUBLIC_DIR / "camp.html")
        if host in {"backoffice.travelx.mn", "www.backoffice.travelx.mn"}:
            if not current_user(environ):
                return file_response(start_response, PUBLIC_DIR / "login.html")
            if not active_workspace(environ):
                start_response("302 Found", [("Location", "/workspace")])
                return [b""]
            return file_response(start_response, PUBLIC_DIR / "camp.html")
        return file_response(start_response, PUBLIC_DIR / "index.html")

    if path == "/backoffice":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        if not active_workspace(environ):
            start_response("302 Found", [("Location", "/workspace")])
            return [b""]
        return file_response(start_response, PUBLIC_DIR / "camp.html")

    if path == "/todo":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        return file_response(start_response, PUBLIC_DIR / "backoffice.html")

    if path == "/contacts":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        # Same UI as /todo but the script default-filters to contacts.
        return file_response(start_response, PUBLIC_DIR / "backoffice.html")

    if path == "/notes":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        return file_response(start_response, PUBLIC_DIR / "notes.html")

    if path == "/settings":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        return file_response(start_response, PUBLIC_DIR / "settings.html")

    if path == "/contracts":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        return file_response(start_response, PUBLIC_DIR / "contracts.html")

    if path == "/invoices":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        return file_response(start_response, PUBLIC_DIR / "invoices.html")

    if path == "/documents":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        return file_response(start_response, PUBLIC_DIR / "documents.html")

    if path == "/pdf-viewer":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        return file_response(start_response, PUBLIC_DIR / "pdf-viewer.html")

    if path.startswith("/contract/"):
        contract_id = path.replace("/contract/", "", 1).strip("/")
        if contract_id:
            return file_response(start_response, PUBLIC_DIR / "contract-sign.html")

    if path == "/ds160":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        if active_workspace(environ) == "USM":
            start_response("302 Found", [("Location", "/backoffice")])
            return [b""]
        return file_response(start_response, PUBLIC_DIR / "ds160.html")

    if path.startswith("/ds160/form/"):
        client_token = path.replace("/ds160/form/", "", 1).strip("/")
        if client_token:
            return file_response(start_response, PUBLIC_DIR / "ds160-client.html")

    if path == "/camp":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        return file_response(start_response, PUBLIC_DIR / "camp.html")

    if path == "/camp-reservations":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        return file_response(start_response, PUBLIC_DIR / "camp-reservations.html")

    if path == "/flight-reservations":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        return file_response(start_response, PUBLIC_DIR / "flight-reservations.html")

    if path == "/transfer-reservations":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        return file_response(start_response, PUBLIC_DIR / "transfer-reservations.html")

    if path == "/trip-detail":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        return file_response(start_response, PUBLIC_DIR / "trip-detail.html")

    if path == "/trip-creator":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        return file_response(start_response, PUBLIC_DIR / "trip-creator.html")

    if path == "/gallery":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        return file_response(start_response, PUBLIC_DIR / "gallery.html")

    if path == "/content":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        return file_response(start_response, PUBLIC_DIR / "content.html")

    if path == "/templates":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        return file_response(start_response, PUBLIC_DIR / "templates.html")

    # /trip/<id> — public brochure for clients. No auth gate. We inline the
    # trip-creator doc as a <script type="application/json"> so the page
    # paints without a fetch round-trip; the JS still falls back to fetching
    # if the inline data is missing (e.g. a 404 placeholder served as HTML).
    if path.startswith("/trip/"):
        trip_id = path.replace("/trip/", "", 1).strip("/")
        public = build_public_trip_view(trip_id) if trip_id else None
        return _serve_html_with_inline_data(start_response, PUBLIC_DIR / "trip-public.html", public)

    # /c/<slug> — standalone public view of one content item. Same trick.
    if path.startswith("/c/"):
        slug = path.replace("/c/", "", 1).strip("/")
        public = build_public_content_view(slug, environ=environ) if slug else None
        return _serve_html_with_inline_data(start_response, PUBLIC_DIR / "content-view.html", public)

    if path == "/tourist":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        return file_response(start_response, PUBLIC_DIR / "tourist.html")

    if path == "/group":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        return file_response(start_response, PUBLIC_DIR / "group.html")

    if path == "/invoice-view":
        return file_response(start_response, PUBLIC_DIR / "invoice-view.html")

    if path == "/admin":
        user = current_user(environ)
        if not user:
            return file_response(start_response, PUBLIC_DIR / "login.html")
        if user.get("role") != "admin":
            return text_response(start_response, "403 Forbidden", "Admin access required")
        return file_response(start_response, PUBLIC_DIR / "admin.html")

    if path == "/accountant":
        user = current_user(environ)
        if not user:
            return file_response(start_response, PUBLIC_DIR / "login.html")
        # Everyone can read the Accountant page (paid-doc archive); only
        # accountants / admins approve, which is gated server-side on
        # /api/payment-requests/<id>/approve.
        return file_response(start_response, PUBLIC_DIR / "accountant.html")

    if path == "/mail":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        return file_response(start_response, PUBLIC_DIR / "mail.html")

    if path == "/mail-settings":
        user = current_user(environ)
        if not user:
            return file_response(start_response, PUBLIC_DIR / "login.html")
        if user.get("role") != "admin":
            return text_response(start_response, "403 Forbidden", "Admin access required")
        return file_response(start_response, PUBLIC_DIR / "mail-settings.html")

    if path == "/fifa2026-admin":
        if not current_user(environ):
            return file_response(start_response, PUBLIC_DIR / "login.html")
        if active_workspace(environ) == "USM":
            start_response("302 Found", [("Location", "/backoffice")])
            return [b""]
        return file_response(start_response, PUBLIC_DIR / "fifa2026-admin.html")

    if path == "/fifa2026":
        return file_response(start_response, PUBLIC_DIR / "fifa2026.html")

    if path.startswith("/generated/"):
        safe_path = (GENERATED_DIR / unquote(path.replace("/generated/", "", 1))).resolve()
        if not str(safe_path).startswith(str(GENERATED_DIR.resolve())) or not safe_path.exists():
            return json_response(start_response, "404 Not Found", {"error": "Not found"})
        params = parse_qs(environ.get("QUERY_STRING", ""))
        extra_headers = generated_download_headers(safe_path) if params.get("download", ["0"])[0] == "1" else None
        return file_response(start_response, safe_path, extra_headers=extra_headers)

    if path.startswith("/trip-uploads/"):
        if not current_user(environ):
            return json_response(start_response, "401 Unauthorized", {"error": "Login required"})
        rel = unquote(path.replace("/trip-uploads/", "", 1))
        safe_path = (TRIP_UPLOADS_DIR / rel).resolve()
        if not str(safe_path).startswith(str(TRIP_UPLOADS_DIR.resolve())) or not safe_path.exists():
            return json_response(start_response, "404 Not Found", {"error": "Not found"})
        params = parse_qs(environ.get("QUERY_STRING", ""))
        extra_headers = generated_download_headers(safe_path) if params.get("download", ["0"])[0] == "1" else None
        return file_response(start_response, safe_path, extra_headers=extra_headers)

    # Public content videos — no login gate, since public popups
    # need to play them. Pathname is /content-videos/<id>.<ext>.
    if path.startswith("/content-videos/"):
        rel = unquote(path.replace("/content-videos/", "", 1))
        safe_path = (CONTENT_VIDEO_DIR / rel).resolve()
        if not str(safe_path).startswith(str(CONTENT_VIDEO_DIR.resolve())) or not safe_path.exists():
            return json_response(start_response, "404 Not Found", {"error": "Not found"})
        return file_response(start_response, safe_path)

    if path.startswith("/camp-contracts/"):
        if not current_user(environ):
            return json_response(start_response, "401 Unauthorized", {"error": "Login required"})
        rel = unquote(path.replace("/camp-contracts/", "", 1))
        safe_path = (CAMP_CONTRACTS_DIR / rel).resolve()
        if not str(safe_path).startswith(str(CAMP_CONTRACTS_DIR.resolve())) or not safe_path.exists():
            return json_response(start_response, "404 Not Found", {"error": "Not found"})
        params = parse_qs(environ.get("QUERY_STRING", ""))
        extra_headers = generated_download_headers(safe_path) if params.get("download", ["0"])[0] == "1" else None
        return file_response(start_response, safe_path, extra_headers=extra_headers)

    safe_path = (PUBLIC_DIR / path.lstrip("/")).resolve()
    if not str(safe_path).startswith(str(PUBLIC_DIR.resolve())) or not safe_path.exists():
        return json_response(start_response, "404 Not Found", {"error": "Not found"})

    return file_response(start_response, safe_path)


# Run-once migrations at module import. Render restarts the worker on every
# deploy, so this fires automatically; the marker file inside DATA_DIR keeps
# it idempotent if the worker is restarted without a redeploy.
try:
    migrate_usm_serials_t_to_s_once()
except Exception:
    pass


if __name__ == "__main__":
    print(f"Client intake app running at http://{HOST}:{PORT}")
    with make_server(HOST, PORT, app) as server:
        server.serve_forever()
